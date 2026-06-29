# -*- coding: utf-8 -*-
"""
维修商贸一体化管理系统 - Flask后端主应用 (增强版 V2)
包含完整的工单管理、接件管理、进销存、财务管理等功能
"""
import os
import io
import json
import time
import logging
import re
import openpyxl
from datetime import datetime, timedelta
from decimal import Decimal
from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from flask_cors import CORS
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
import bcrypt

# 设置时区为亚洲/上海（北京时间）
os.environ.setdefault('TZ', 'Asia/Shanghai')
import time as _time
if hasattr(_time, 'tzset'):
    _time.tzset()
from functools import wraps

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# === 早期注入：让 backend/ 子模块（包括 extensions、models、config）可被 import ===
# 注意必须在 app 创建之前，否则 config 加载会失败
import sys as _early_sys
_early_backend_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backend')
if _early_backend_dir not in _early_sys.path:
    _early_sys.path.insert(0, _early_backend_dir)
_early_source_dir = os.path.dirname(os.path.abspath(__file__))
if _early_source_dir not in _early_sys.path:
    _early_sys.path.insert(0, _early_source_dir)

# 创建Flask应用
app = Flask(__name__)

# 强制JSON响应使用UTF-8编码
app.config['JSON_AS_ASCII'] = False
app.config['RESTFUL_JSON'] = {'ensure_ascii': False, 'charset': 'utf-8'}

# 强制所有JSON响应添加charset=utf-8
@app.after_request
def set_charset(response):
    if response.content_type and 'application/json' in response.content_type:
        if 'charset' not in response.content_type:
            response.content_type = 'application/json; charset=utf-8'
    return response

# 配置（敏感配置必须从环境变量读取，无硬编码 fallback）
# 优先从 backend/app/config.py 加载；缺失 JWT_SECRET_KEY/DATABASE_URL 直接抛错
try:
    import importlib as _il_cfg
    import sys as _sys_cfg
    _here_cfg = os.path.dirname(os.path.abspath(__file__))
    if _here_cfg not in _sys_cfg.path:
        _sys_cfg.path.insert(0, _here_cfg)
    _cfg_mod = _il_cfg.import_module('backend.app.config')
    _get_config = _cfg_mod.get_config
    app.config.from_object(_get_config(os.environ.get('FLASK_ENV', 'production')))
    logger.info('配置已从 backend/app/config.py 加载')
except Exception as _cfg_err:
    import traceback as _tb_cfg
    logger.warning('加载 backend/app/config.py 失败，使用 fallback 配置: %s\n%s', _cfg_err, _tb_cfg.format_exc())
    app.config['SECRET_KEY'] = os.environ['JWT_SECRET_KEY']
    app.config['JWT_SECRET_KEY'] = os.environ['JWT_SECRET_KEY']
    app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=7)
    app.config['SQLALCHEMY_DATABASE_URI'] = os.environ['DATABASE_URL']
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', '/app/uploads')
    if app.config['SQLALCHEMY_DATABASE_URI'].startswith('mysql'):
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
            'connect_args': {
                'charset': 'utf8mb4',
                'init_command': "SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci",
            },
            'pool_pre_ping': True,
            'pool_recycle': 3600,
        }

# 初始化扩展
db = SQLAlchemy(app)
jwt = JWTManager(app)
migrate = Migrate(app, db)

# CORS 白名单（从环境变量读取，生产环境应严格限制）
_cors_origins = os.environ.get('CORS_ORIGINS', '')
_cors_list = [o.strip() for o in _cors_origins.split(',') if o.strip()] or ['http://localhost']
CORS(app, resources={r"/api/*": {"origins": _cors_list}}, supports_credentials=True)

# 注册全局错误处理器（阶段 2 拆分后由 backend/app/errors.py 接管）
@app.errorhandler(400)
def _err_400(e): return jsonify({'code': 400, 'message': '请求参数错误'}), 400
@app.errorhandler(401)
def _err_401(e): return jsonify({'code': 401, 'message': '未登录或登录已过期'}), 401
@app.errorhandler(403)
def _err_403(e): return jsonify({'code': 403, 'message': '无权限'}), 403
@app.errorhandler(404)
def _err_404(e): return jsonify({'code': 404, 'message': '资源不存在'}), 404
@app.errorhandler(405)
def _err_405(e): return jsonify({'code': 405, 'message': '请求方法不允许'}), 405
@app.errorhandler(413)
def _err_413(e): return jsonify({'code': 413, 'message': '文件过大'}), 413
@app.errorhandler(Exception)
def _err_500(e):
    import traceback as _tb
    logger.error('unhandled: %s\n%s', e, _tb.format_exc())
    return jsonify({'code': 500, 'message': '服务异常'}), 500

# 批量操作API已直接内嵌在app.py中（见文件底部），无需额外注册

# JWT token 黑名单（阶段 0 已建立基础，阶段 2 拆分后接 redis）
_revoked_jtis = set()
@jwt.token_in_blocklist_loader
def check_if_token_revoked(jwt_header, jwt_payload):
    return jwt_payload.get('jti') in _revoked_jtis

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ============================================
# 工具函数
# ============================================

def generate_pinyin_code(text):
    """生成拼音首字母简码"""
    if not text:
        return ''
    # 简单实现：提取首字母
    result = ''
    for char in text:
        if char.isalpha():
            result += char.upper()
        elif char.isdigit():
            result += char
    return result[:20]

def generate_code(prefix, last_id=None):
    """生成业务编码"""
    date_str = datetime.now().strftime("%Y%m%d")
    seq = str(last_id + 1 if last_id else 1).zfill(4)
    return f'{prefix}{date_str}{seq}'

def to_dict(model, exclude=None):
    """将SQLAlchemy模型转换为字典"""
    if model is None:
        return None
    exclude = exclude or []
    result = {}
    for column in model.__table__.columns:
        if column.name in exclude:
            continue
        value = getattr(model, column.name)
        if isinstance(value, datetime):
            value = value.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(value, Decimal):
            value = float(value)
        result[column.name] = value
    return result

def get_current_user_id():
    """获取当前登录用户ID"""
    return get_jwt_identity()

def get_current_user_name():
    """获取当前登录用户姓名"""
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    if user:
        return user.real_name or user.username
    return ''

# ============================================
# 数据模型定义（已迁移到 backend/models/）
# ============================================
# 注意：项目根有 app.py，与 backend/app 同名冲突。这里用绝对路径导入 backend.models
import importlib as _il
import sys as _sys_for_models
def _load_models():
    """从 backend.models 加载所有模型到当前命名空间。"""
    _here = os.path.dirname(os.path.abspath(__file__))
    if _here not in _sys_for_models.path:
        _sys_for_models.path.insert(0, _here)
    # backend/ 也必须加入 sys.path，models 子包内部用 `from extensions import db`
    _backend_dir = os.path.join(_here, 'backend')
    if _backend_dir not in _sys_for_models.path:
        _sys_for_models.path.insert(0, _backend_dir)
    for _domain in ['system', 'product', 'customer', 'supplier', 'workorder',
                    'receive', 'dispatch', 'inventory', 'purchase',
                    'finance', 'asset', 'printer']:
        _mod = _il.import_module(f'backend.models.{_domain}')
        for _name in dir(_mod):
            if _name[0].isupper() and hasattr(getattr(_mod, _name), '__tablename__'):
                globals()[_name] = getattr(_mod, _name)
_load_models()
del _load_models, _il, _sys_for_models

# ============================================
# API路由 - 认证模块
# ============================================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'code': 400, 'message': '用户名和密码不能为空'}), 400
    
    user = SysUser.query.filter_by(username=username, status=1, is_deleted=0).first()
    
    # 支持 bcrypt 和 werkzeug 两种密码哈希格式
    password_valid = False
    if user:
        if user.password.startswith('$2'):
            # bcrypt 格式
            password_valid = bcrypt.checkpw(password.encode('utf-8'), user.password.encode('utf-8'))
        else:
            # werkzeug pbkdf2 格式
            password_valid = check_password_hash(user.password, password)
    
    if not user or not password_valid:
        return jsonify({'code': 401, 'message': '用户名或密码错误'}), 401
    
    # 更新登录信息
    user.last_login_time = datetime.now()
    user.last_login_ip = request.remote_addr
    db.session.commit()
    
    # 生成JWT令牌
    access_token = create_access_token(identity=str(user.id))
    
    # 获取角色和权限信息
    role_name = ''
    role_code = ''
    permissions = []
    if user.role_id:
        role = SysRole.query.get(user.role_id)
        if role:
            role_name = role.role_name
            role_code = role.role_code
            perms = role.permissions if role.permissions else []
            if isinstance(perms, list) and '*' in perms:
                # 超级管理员，返回所有权限code
                all_perms = SysPermission.query.filter_by(status=1).all()
                permissions = [p.code for p in all_perms]
            else:
                permissions = perms

    return jsonify({
        'code': 200,
        'message': '登录成功',
        'data': {
            'token': access_token,
            'userInfo': {
                'id': user.id,
                'username': user.username,
                'real_name': user.real_name,
                'avatar': user.avatar,
                'role_id': user.role_id,
                'role_name': role_name,
                'role_code': role_code,
                'permissions': permissions,
                'role': {
                    'role_name': role_name,
                    'role_code': role_code,
                    'permissions': permissions
                }
            }
        }
    })

@app.route('/api/auth/register', methods=['POST'])
@jwt_required()
def register():
    """用户注册（需要管理员权限）"""
    claims = get_jwt()
    user_perms = claims.get('permissions', []) if hasattr(claims, 'get') else []
    if '*' not in user_perms and 'user:add' not in user_perms:
        return jsonify({'code': 403, 'message': '无权限创建用户'}), 403
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    real_name = data.get('real_name')

    if not username or not password:
        return jsonify({'code': 400, 'message': '用户名和密码不能为空'}), 400
    if len(password) < 8:
        return jsonify({'code': 400, 'message': '密码至少 8 个字符'}), 400

    existing_user = SysUser.query.filter_by(username=username).first()
    if existing_user:
        return jsonify({'code': 400, 'message': '用户名已存在'}), 400

    hashed_password = generate_password_hash(password)
    new_user = SysUser(
        username=username,
        password=hashed_password,
        real_name=real_name
    )

    db.session.add(new_user)
    db.session.commit()

    return jsonify({'code': 200, 'message': '注册成功'})

@app.route('/api/user/info', methods=['GET'])
@jwt_required()
def get_user_info():
    """获取当前用户信息"""
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    
    if not user:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404
    
    # 获取角色和权限信息
    role_name = ''
    role_code = ''
    permissions = []
    if user.role_id:
        role = SysRole.query.get(user.role_id)
        if role:
            role_name = role.role_name
            role_code = role.role_code
            perms = role.permissions if role.permissions else []
            if isinstance(perms, list) and '*' in perms:
                # 超级管理员，返回所有权限code
                all_perms = SysPermission.query.filter_by(status=1).all()
                permissions = [p.code for p in all_perms]
            else:
                permissions = perms
    
    # 构建菜单树（type=1的权限，用户有权限的菜单）
    menus = []
    if permissions:
        all_menu_perms = SysPermission.query.filter_by(type=1, status=1).order_by(SysPermission.sort_order).all()
        if isinstance(permissions, list) and '*' in (SysRole.query.get(user.role_id).permissions or []):
            # 超级管理员返回所有菜单
            menu_list = all_menu_perms
        else:
            menu_list = [p for p in all_menu_perms if p.code in permissions]
        # 构建树形结构
        menu_dict = {}
        for m in menu_list:
            menu_dict[m.id] = {
                'id': m.id,
                'name': m.name,
                'code': m.code,
                'path': m.path,
                'icon': m.icon,
                'sort_order': m.sort_order,
                'parent_id': m.parent_id,
                'children': []
            }
        for m in menu_list:
            if m.parent_id and m.parent_id in menu_dict:
                menu_dict[m.parent_id]['children'].append(menu_dict[m.id])
        menus = [menu_dict[m.id] for m in menu_list if m.parent_id == 0 or m.parent_id not in menu_dict]
        menus.sort(key=lambda x: x.get('sort_order', 0))
    
    return jsonify({
        'code': 200,
        'data': {
            'id': user.id,
            'username': user.username,
            'real_name': user.real_name,
            'phone': user.phone,
            'email': user.email,
            'avatar': user.avatar,
            'role_id': user.role_id,
            'department': user.department,
            'position': user.position,
            'role_name': role_name,
            'role_code': role_code,
            'permissions': permissions,
            'menus': menus
        }
    })

# ============================================
# API路由 - 客户管理
# ============================================

@app.route('/api/customers', methods=['GET'])
@jwt_required()
def get_customers():
    """获取客户列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    customer_type = request.args.get('customer_type', type=int)
    
    query = BaseCustomer.query.filter_by(status=1)
    
    if keyword:
        query = query.filter(
            db.or_(
                BaseCustomer.customer_name.contains(keyword),
                BaseCustomer.pinyin_code.contains(keyword.upper()),
                BaseCustomer.phone.contains(keyword),
                BaseCustomer.customer_code.contains(keyword)
            )
        )
    
    if customer_type:
        query = query.filter_by(customer_type=customer_type)
    
    total = query.count()
    customers = query.order_by(BaseCustomer.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': c.id,
                'customer_code': c.customer_code,
                'customer_name': c.customer_name,
                'customer_type': c.customer_type,
                'contact_name': c.contact_name,
                'phone': c.phone,
                'phone2': c.phone2,
                'address': c.address,
                'discount_rate': float(c.discount_rate) if c.discount_rate else 100.00,
                'total_sales_amount': float(c.total_sales_amount) if c.total_sales_amount else 0.00,
                'total_sales_count': c.total_sales_count,
                'created_at': c.created_at.strftime('%Y-%m-%d %H:%M:%S') if c.created_at else None
            } for c in customers.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/customers/<int:id>', methods=['GET'])
@jwt_required()
def get_customer(id):
    """获取客户详情"""
    customer = BaseCustomer.query.get(id)
    if not customer:
        return jsonify({'code': 404, 'message': '客户不存在'}), 404
    
    return jsonify({'code': 200, 'data': to_dict(customer)})

@app.route('/api/customers', methods=['POST'])
@jwt_required()
def create_customer():
    """创建客户"""
    data = request.get_json()
    user_id = get_jwt_identity()
    
    # 生成客户编码
    last_customer = BaseCustomer.query.order_by(BaseCustomer.id.desc()).first()
    customer_code = generate_code('C', last_customer.id if last_customer else 0)
    
    customer = BaseCustomer(
        customer_code=customer_code,
        customer_name=data.get('customer_name'),
        customer_type=data.get('customer_type', 1),
        pinyin_code=generate_pinyin_code(data.get('customer_name', '')),
        contact_name=data.get('contact_name'),
        phone=data.get('phone'),
        phone2=data.get('phone2'),
        email=data.get('email'),
        address=data.get('address'),
        discount_rate=data.get('discount_rate', 100.00),
        credit_limit=data.get('credit_limit', 0.00),
        tax_number=data.get('tax_number'),
        bank_name=data.get('bank_name'),
        bank_account=data.get('bank_account'),
        remark=data.get('remark')
    )
    
    db.session.add(customer)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': customer.id, 'customer_code': customer_code}})

@app.route('/api/customers/<int:id>', methods=['PUT'])
@jwt_required()
def update_customer(id):
    """更新客户"""
    customer = BaseCustomer.query.get(id)
    if not customer:
        return jsonify({'code': 404, 'message': '客户不存在'}), 404
    
    data = request.get_json()
    
    for field in ['customer_name', 'customer_type', 'contact_name', 'phone', 'phone2', 
                  'email', 'address', 'discount_rate', 'credit_limit', 'tax_number', 
                  'bank_name', 'bank_account', 'remark']:
        if field in data:
            setattr(customer, field, data[field])
    
    if 'customer_name' in data:
        customer.pinyin_code = generate_pinyin_code(data['customer_name'])
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/customers/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_customer(id):
    """删除客户"""
    customer = BaseCustomer.query.get(id)
    if not customer:
        return jsonify({'code': 404, 'message': '客户不存在'}), 404
    
    customer.status = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# API路由 - 供应商管理
# ============================================

@app.route('/api/suppliers', methods=['GET'])
@jwt_required()
def get_suppliers():
    """获取供应商列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    is_repair_partner = request.args.get('is_repair_partner', type=int)
    
    query = BaseSupplier.query.filter_by(status=1)
    
    if keyword:
        query = query.filter(
            db.or_(
                BaseSupplier.supplier_name.contains(keyword),
                BaseSupplier.pinyin_code.contains(keyword.upper()),
                BaseSupplier.phone.contains(keyword),
                BaseSupplier.supplier_code.contains(keyword)
            )
        )
    
    if is_repair_partner is not None:
        query = query.filter_by(is_repair_partner=is_repair_partner)
    
    total = query.count()
    suppliers = query.order_by(BaseSupplier.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': s.id,
                'supplier_code': s.supplier_code,
                'supplier_name': s.supplier_name,
                'contact_name': s.contact_name,
                'phone': s.phone,
                'is_repair_partner': s.is_repair_partner,
                'total_purchase_amount': float(s.total_purchase_amount) if s.total_purchase_amount else 0.00,
                'total_repair_amount': float(s.total_repair_amount) if s.total_repair_amount else 0.00,
                'created_at': s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else None
            } for s in suppliers.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/suppliers/<int:id>', methods=['GET'])
@jwt_required()
def get_supplier(id):
    """获取供应商详情"""
    supplier = BaseSupplier.query.get(id)
    if not supplier:
        return jsonify({'code': 404, 'message': '供应商不存在'}), 404
    return jsonify({'code': 200, 'data': to_dict(supplier)})

@app.route('/api/suppliers', methods=['POST'])
@jwt_required()
def create_supplier():
    """创建供应商"""
    data = request.get_json()
    
    last_supplier = BaseSupplier.query.order_by(BaseSupplier.id.desc()).first()
    supplier_code = generate_code('S', last_supplier.id if last_supplier else 0)
    
    supplier = BaseSupplier(
        supplier_code=supplier_code,
        supplier_name=data.get('supplier_name'),
        pinyin_code=generate_pinyin_code(data.get('supplier_name', '')),
        contact_name=data.get('contact_name'),
        phone=data.get('phone'),
        email=data.get('email'),
        address=data.get('address'),
        tax_number=data.get('tax_number'),
        bank_name=data.get('bank_name'),
        bank_account=data.get('bank_account'),
        is_repair_partner=data.get('is_repair_partner', 0),
        repair_types=data.get('repair_types'),
        remark=data.get('remark')
    )
    
    db.session.add(supplier)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': supplier.id, 'supplier_code': supplier_code}})

@app.route('/api/suppliers/<int:id>', methods=['PUT'])
@jwt_required()
def update_supplier(id):
    """更新供应商"""
    supplier = BaseSupplier.query.get(id)
    if not supplier:
        return jsonify({'code': 404, 'message': '供应商不存在'}), 404
    
    data = request.get_json()
    
    for field in ['supplier_name', 'contact_name', 'phone', 'email', 'address', 
                  'tax_number', 'bank_name', 'bank_account', 'is_repair_partner', 
                  'repair_types', 'remark']:
        if field in data:
            setattr(supplier, field, data[field])
    
    if 'supplier_name' in data:
        supplier.pinyin_code = generate_pinyin_code(data['supplier_name'])
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/suppliers/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_supplier(id):
    """删除供应商"""
    supplier = BaseSupplier.query.get(id)
    if not supplier:
        return jsonify({'code': 404, 'message': '供应商不存在'}), 404
    
    supplier.status = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# API路由 - 商品管理
# ============================================

@app.route('/api/products', methods=['GET'])
@jwt_required()
def get_products():
    """获取商品列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    category_id = request.args.get('category_id', type=int)
    
    query = ProductInfo.query.filter_by(status=1)
    
    if keyword:
        query = query.filter(
            db.or_(
                ProductInfo.product_name.contains(keyword),
                ProductInfo.pinyin_code.contains(keyword.upper()),
                ProductInfo.barcode.contains(keyword),
                ProductInfo.product_code.contains(keyword)
            )
        )
    
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    total = query.count()
    products = query.order_by(ProductInfo.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    # 获取库存信息（从InventoryStock表）
    product_ids = [p.id for p in products.items]
    stocks = InventoryStock.query.filter(InventoryStock.product_id.in_(product_ids)).all()
    stock_map = {}
    for s in stocks:
        stock_map[int(s.product_id)] = s
    
    # 调试：打印前3个商品的库存
    print(f"[DEBUG] 商品数量: {len(products.items)}, 库存记录数: {len(stocks)}")
    for p in products.items[:3]:
        stock = stock_map.get(int(p.id))
        if stock:
            print(f"[DEBUG] 商品 {p.product_name} (ID:{p.id}): available_quantity={stock.available_quantity}")
        else:
            print(f"[DEBUG] 商品 {p.product_name} (ID:{p.id}): 无库存记录")
    
    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': p.id,
                'product_code': p.product_code,
                'barcode': p.barcode,
                'product_name': p.product_name,
                'category_id': p.category_id,
                'category_name': p.category_name,
                'brand': p.brand,
                'specification': p.specification,
                'unit_name': p.unit_name or '个',
                'purchase_price': float(p.purchase_price) if p.purchase_price else 0.00,
                'sale_price': float(p.sale_price) if p.sale_price else 0.00,
                'member_price': float(p.member_price) if p.member_price else 0.00,
                'wholesale_price': float(p.wholesale_price) if p.wholesale_price else 0.00,
                'cost_price': float(p.cost_price) if p.cost_price else 0.00,
                'current_stock': float(stock_map.get(int(p.id)).available_quantity) if stock_map.get(int(p.id)) and stock_map.get(int(p.id)).available_quantity else 0,
                'min_stock': p.min_stock,
                'is_serial': p.is_serial,
                'is_batch': p.is_batch,
                'is_assembly': p.is_assembly,
                'image_url': p.image_url,
                'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None
            } for p in products.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/debug/stock', methods=['GET'])
def debug_stock():
    """调试库存数据（无需认证）"""
    stocks = InventoryStock.query.limit(20).all()
    result = []
    for s in stocks:
        result.append({
            'id': s.id,
            'product_id': s.product_id,
            'product_name': s.product_name,
            'product_code': s.product_code,
            'quantity': float(s.quantity) if s.quantity else 0,
            'available_quantity': float(s.available_quantity) if s.available_quantity else 0
        })
    return jsonify({'code': 200, 'data': result, 'total': len(result)})

@app.route('/api/products/<int:id>', methods=['GET'])
@jwt_required()
def get_product(id):
    """获取商品详情"""
    product = ProductInfo.query.get(id)
    if not product:
        return jsonify({'code': 404, 'message': '商品不存在'}), 404
    
    # 获取库存信息
    stock = InventoryStock.query.filter_by(product_id=id).first()
    stock_info = None
    if stock:
        stock_info = {
            'quantity': float(stock.quantity) if stock.quantity else 0,
            'frozen_quantity': float(stock.frozen_quantity) if stock.frozen_quantity else 0,
            'available_quantity': float(stock.available_quantity) if stock.available_quantity else 0
        }
    
    # 获取多单位信息
    unit_rels = ProductUnitRel.query.filter_by(product_id=id).all()
    units = []
    for rel in unit_rels:
        unit = ProductUnit.query.get(rel.unit_id)
        if unit:
            units.append({
                'id': rel.id,
                'unit_id': rel.unit_id,
                'unit_name': unit.unit_name,
                'unit_code': unit.unit_code,
                'conversion_rate': float(rel.conversion_rate) if rel.conversion_rate else 1.0000,
                'is_default': rel.is_default
            })
    
    result = to_dict(product)
    result['stock'] = stock_info
    result['units'] = units
    return jsonify({'code': 200, 'data': result})

@app.route('/api/products', methods=['POST'])
@jwt_required()
def create_product():
    """创建商品"""
    data = request.get_json()
    
    last_product = ProductInfo.query.order_by(ProductInfo.id.desc()).first()
    product_code = generate_code('P', last_product.id if last_product else 0)
    
    # 处理barcode：空字符串转为None，避免唯一索引冲突
    barcode = data.get('barcode')
    if barcode is not None and str(barcode).strip() == '':
        barcode = None
    
    # 处理单位信息：从units中获取默认单位名称
    units = data.get('units', [])
    unit_name = data.get('unit_name', '')
    unit_id = data.get('unit_id')
    if units:
        for u in units:
            if u.get('is_default') == 1:
                unit_id = u.get('unit_id')
                unit_name = u.get('unit_name', '')
                break
        # 如果没有默认单位，取第一个
        if not unit_name and units:
            unit_id = units[0].get('unit_id')
            unit_name = units[0].get('unit_name', '')
    
    product = ProductInfo(
        product_code=product_code,
        barcode=barcode,
        product_name=data.get('product_name'),
        pinyin_code=generate_pinyin_code(data.get('product_name', '')),
        category_id=data.get('category_id'),
        category_name=data.get('category_name'),
        brand=data.get('brand'),
        specification=data.get('specification'),
        unit_id=unit_id,
        unit_name=unit_name,
        sub_unit_id=data.get('sub_unit_id'),
        sub_unit_rate=data.get('sub_unit_rate'),
        purchase_price=data.get('purchase_price', 0),
        sale_price=data.get('sale_price', 0),
        member_price=data.get('member_price', 0),
        wholesale_price=data.get('wholesale_price', 0),
        customer_price=data.get('customer_price', 0),
        cost_price=data.get('cost_price', 0),
        min_stock=data.get('min_stock', 0),
        max_stock=data.get('max_stock', 0),
        is_serial=data.get('is_serial', 0),
        is_batch=data.get('is_batch', 0),
        is_assembly=data.get('is_assembly', 0),
        is_gift=data.get('is_gift', 0),
        no_cost=data.get('no_cost', 0),
        no_stock=data.get('no_stock', 0),
        remark=data.get('remark')
    )
    
    db.session.add(product)
    db.session.commit()
    
    # 创建库存记录
    stock = InventoryStock(
        product_id=product.id,
        product_code=product.product_code,
        product_name=product.product_name,
        quantity=0,
        frozen_quantity=0,
        available_quantity=0,
        cost_price=product.cost_price
    )
    db.session.add(stock)
    
    # 处理多单位关联
    units = data.get('units', [])
    if units:
        for unit_data in units:
            unit_rel = ProductUnitRel(
                product_id=product.id,
                unit_id=unit_data.get('unit_id'),
                conversion_rate=unit_data.get('conversion_rate', 1.0000),
                is_default=unit_data.get('is_default', 0)
            )
            db.session.add(unit_rel)
    
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': product.id, 'product_code': product_code}})

@app.route('/api/products/<int:id>', methods=['PUT'])
@jwt_required()
def update_product(id):
    """更新商品"""
    product = ProductInfo.query.get(id)
    if not product:
        return jsonify({'code': 404, 'message': '商品不存在'}), 404
    
    data = request.get_json()
    
    for field in ['product_name', 'barcode', 'category_id', 'category_name', 'brand', 
                  'specification', 'unit_id', 'unit_name', 'sub_unit_id', 'sub_unit_rate',
                  'purchase_price', 'sale_price', 'member_price', 'wholesale_price', 
                  'customer_price', 'cost_price', 'min_stock', 'max_stock',
                  'is_serial', 'is_batch', 'is_assembly', 'is_gift', 'no_cost', 'no_stock', 'remark']:
        if field in data:
            setattr(product, field, data[field])
    
    if 'product_name' in data:
        product.pinyin_code = generate_pinyin_code(data['product_name'])
    
    # 处理多单位关联更新
    if 'units' in data:
        # 删除旧的单位关联
        ProductUnitRel.query.filter_by(product_id=id).delete()
        
        # 添加新的单位关联
        for unit_data in data['units']:
            unit_rel = ProductUnitRel(
                product_id=id,
                unit_id=unit_data.get('unit_id'),
                conversion_rate=unit_data.get('conversion_rate', 1.0000),
                is_default=unit_data.get('is_default', 0)
            )
            db.session.add(unit_rel)
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/products/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_product(id):
    """删除商品"""
    product = ProductInfo.query.get(id)
    if not product:
        return jsonify({'code': 404, 'message': '商品不存在'}), 404
    
    product.status = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# API路由 - 商品单位管理
# ============================================

@app.route('/api/product/units', methods=['GET'])
@jwt_required()
def get_product_units():
    """获取单位列表"""
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    
    query = ProductUnit.query
    
    if keyword:
        query = query.filter(ProductUnit.unit_name.contains(keyword))
    
    if status is not None:
        query = query.filter_by(status=status)
    
    units = query.order_by(ProductUnit.sort_order.asc(), ProductUnit.created_at.desc()).all()
    
    return jsonify({
        'code': 200,
        'data': [{
            'id': u.id,
            'unit_name': u.unit_name,
            'unit_code': u.unit_code,
            'conversion_rate': float(u.conversion_rate) if u.conversion_rate else 1.0000,
            'is_base': u.is_base,
            'sort_order': u.sort_order,
            'status': u.status,
            'created_at': u.created_at.strftime('%Y-%m-%d %H:%M:%S') if u.created_at else None
        } for u in units]
    })

@app.route('/api/product/units', methods=['POST'])
@jwt_required()
def create_product_unit():
    """创建单位"""
    data = request.get_json()
    
    if not data.get('unit_name'):
        return jsonify({'code': 400, 'message': '单位名称不能为空'}), 400
    
    # 检查单位名称是否已存在
    existing = ProductUnit.query.filter_by(unit_name=data.get('unit_name')).first()
    if existing:
        return jsonify({'code': 400, 'message': '单位名称已存在'}), 400
    
    # 处理 unit_code，如果为空则使用 unit_name
    unit_code = data.get('unit_code')
    if not unit_code or str(unit_code).strip() == '':
        unit_code = data.get('unit_name')
    
    # 检查 unit_code 是否已存在（如果有值）
    if unit_code:
        existing_code = ProductUnit.query.filter_by(unit_code=unit_code).first()
        if existing_code:
            # 如果 unit_code 已存在，追加数字后缀
            import re
            base_code = re.sub(r'\d+$', '', unit_code)
            counter = 1
            new_code = f"{base_code}{counter}"
            while ProductUnit.query.filter_by(unit_code=new_code).first():
                counter += 1
                new_code = f"{base_code}{counter}"
            unit_code = new_code
    
    unit = ProductUnit(
        unit_name=data.get('unit_name'),
        unit_code=unit_code,
        conversion_rate=data.get('conversion_rate', 1.0000),
        is_base=data.get('is_base', 0),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    
    db.session.add(unit)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': unit.id}})

@app.route('/api/product/units/<int:id>', methods=['PUT'])
@jwt_required()
def update_product_unit(id):
    """更新单位"""
    unit = ProductUnit.query.get(id)
    if not unit:
        return jsonify({'code': 404, 'message': '单位不存在'}), 404
    
    data = request.get_json()
    
    # 检查单位名称是否与其他单位重复
    if 'unit_name' in data and data['unit_name'] != unit.unit_name:
        existing = ProductUnit.query.filter_by(unit_name=data['unit_name']).first()
        if existing:
            return jsonify({'code': 400, 'message': '单位名称已存在'}), 400
        unit.unit_name = data['unit_name']
    
    if 'unit_code' in data:
        unit_code = data['unit_code']
        if not unit_code or str(unit_code).strip() == '':
            unit_code = unit.unit_name
        # 检查是否与其他单位重复
        if unit_code != unit.unit_code:
            existing_code = ProductUnit.query.filter_by(unit_code=unit_code).first()
            if existing_code:
                import re
                base_code = re.sub(r'\d+$', '', unit_code)
                counter = 1
                new_code = f"{base_code}{counter}"
                while ProductUnit.query.filter_by(unit_code=new_code).first():
                    counter += 1
                    new_code = f"{base_code}{counter}"
                unit_code = new_code
        unit.unit_code = unit_code
    if 'conversion_rate' in data:
        unit.conversion_rate = data['conversion_rate']
    if 'is_base' in data:
        unit.is_base = data['is_base']
    if 'sort_order' in data:
        unit.sort_order = data['sort_order']
    if 'status' in data:
        unit.status = data['status']
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/product/units/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_product_unit(id):
    """删除单位"""
    unit = ProductUnit.query.get(id)
    if not unit:
        return jsonify({'code': 404, 'message': '单位不存在'}), 404
    
    # 检查是否有关联的商品
    rel_count = ProductUnitRel.query.filter_by(unit_id=id).count()
    if rel_count > 0:
        return jsonify({'code': 400, 'message': '该单位已被商品使用，无法删除'}), 400
    
    db.session.delete(unit)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# API路由 - 商品分类管理
# ============================================

def build_category_tree(categories, parent_id=0):
    """构建分类树形结构"""
    tree = []
    for cat in categories:
        if cat.parent_id == parent_id:
            node = {
                'id': cat.id,
                'category_name': cat.category_name,
                'parent_id': cat.parent_id,
                'sort_order': cat.sort_order,
                'status': cat.status,
                'created_at': cat.created_at.strftime('%Y-%m-%d %H:%M:%S') if cat.created_at else None,
                'children': build_category_tree(categories, cat.id)
            }
            tree.append(node)
    return tree

@app.route('/api/product/categories', methods=['GET'])
@jwt_required()
def get_product_categories():
    """获取分类列表（树形结构）"""
    status = request.args.get('status', type=int)
    
    query = ProductCategory.query
    
    if status is not None:
        query = query.filter_by(status=status)
    
    categories = query.order_by(ProductCategory.sort_order.asc(), ProductCategory.created_at.desc()).all()
    
    # 构建树形结构
    tree = build_category_tree(categories)
    
    return jsonify({
        'code': 200,
        'data': tree
    })

@app.route('/api/product/categories', methods=['POST'])
@jwt_required()
def create_product_category():
    """创建分类"""
    data = request.get_json()
    
    if not data.get('category_name'):
        return jsonify({'code': 400, 'message': '分类名称不能为空'}), 400
    
    category = ProductCategory(
        category_name=data.get('category_name'),
        parent_id=data.get('parent_id', 0),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    
    db.session.add(category)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': category.id}})

@app.route('/api/product/categories/<int:id>', methods=['PUT'])
@jwt_required()
def update_product_category(id):
    """更新分类"""
    category = ProductCategory.query.get(id)
    if not category:
        return jsonify({'code': 404, 'message': '分类不存在'}), 404
    
    data = request.get_json()
    
    if 'category_name' in data:
        category.category_name = data['category_name']
    if 'parent_id' in data:
        # 防止循环引用
        if data['parent_id'] == id:
            return jsonify({'code': 400, 'message': '不能将自己设为上级分类'}), 400
        category.parent_id = data['parent_id']
    if 'sort_order' in data:
        category.sort_order = data['sort_order']
    if 'status' in data:
        category.status = data['status']
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/product/categories/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_product_category(id):
    """删除分类"""
    category = ProductCategory.query.get(id)
    if not category:
        return jsonify({'code': 404, 'message': '分类不存在'}), 404
    
    # 检查是否有子分类
    child_count = ProductCategory.query.filter_by(parent_id=id).count()
    if child_count > 0:
        return jsonify({'code': 400, 'message': '该分类下有子分类，无法删除'}), 400
    
    # 检查是否有关联的商品
    product_count = ProductInfo.query.filter_by(category_id=id).count()
    if product_count > 0:
        return jsonify({'code': 400, 'message': '该分类下有商品，无法删除'}), 400
    
    db.session.delete(category)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# API路由 - 工单管理
# ============================================

# 工单状态映射（8个核心状态）
WO_STATUS_MAP = {
    0: '待派单',      # 工单已创建，等待派单
    1: '待接单',      # 已派单，等待工程师接单
    2: '待备件',      # 备件未就绪（需采购或二次补领）
    3: '待上门',      # 备件已就绪，等待上门
    4: '处理中',      # 正在维修处理
    5: '待结算',      # 维修完成，等待结算
    6: '已完成',      # 结算完成
    7: '已取消'       # 工单取消
}

# 工单类型名称映射（七大固定类型）
WO_TYPE_MAP = {
    'maintenance': '维护服务',
    'inspection': '检测服务',
    'delivery': '送货服务',
    'installation': '安装服务',
    'repair': '维修服务',
    'purchase': '产品代购',
    'survey': '现场勘察'
}

# 工单二级分类映射
WO_SUB_TYPE_MAP = {
    # 维修服务二级分类
    'repair_monitor': '监控维修',
    'repair_network': '网络维修',
    'repair_printer': '打印维修',
    'repair_computer': '电脑维修',
    'repair_other': '其他办公设备维修',
    # 维护服务二级分类
    'maintenance_monitor': '监控维护',
    'maintenance_network': '网络维护',
    'maintenance_printer': '打印维护',
    'maintenance_computer': '电脑维护',
    'maintenance_other': '其他办公设备维护',
    # 检测服务二级分类
    'inspection_monitor': '监控检测',
    'inspection_network': '网络检测',
    'inspection_printer': '打印检测',
    'inspection_computer': '电脑检测',
    'inspection_other': '其他办公设备检测',
    # 安装服务二级分类
    'installation_monitor': '监控安装',
    'installation_network': '网络安装',
    'installation_printer': '打印安装',
    'installation_computer': '电脑安装',
    'installation_other': '其他办公设备安装'
}

# 工单类型与二级分类关联
WO_TYPE_SUBTYPES = {
    'repair': ['repair_monitor', 'repair_network', 'repair_printer', 'repair_computer', 'repair_other'],
    'maintenance': ['maintenance_monitor', 'maintenance_network', 'maintenance_printer', 'maintenance_computer', 'maintenance_other'],
    'inspection': ['inspection_monitor', 'inspection_network', 'inspection_printer', 'inspection_computer', 'inspection_other'],
    'installation': ['installation_monitor', 'installation_network', 'installation_printer', 'installation_computer', 'installation_other'],
    'delivery': [],
    'purchase': [],
    'survey': []
}

# 工单状态合法流转规则（8个核心状态）
WO_STATUS_TRANSITIONS = {
    0: [1, 7],       # 待派单 -> 待接单/已取消
    1: [2, 3, 7],    # 待接单 -> 待备件/待上门/已取消
    2: [3, 7],       # 待备件 -> 待上门/已取消
    3: [4, 7],       # 待上门 -> 处理中/已取消
    4: [5, 7],       # 处理中 -> 待结算/已取消
    5: [6],          # 待结算 -> 已完成
    6: [],           # 已完成 -> 终态
    7: []            # 已取消 -> 终态
}

def auto_dispatch_engineer(wo_type, required_skills=None):
    """自动派单：根据工单类型匹配最优工程师
    1. 查询所有在线工程师(online_status=1)
    2. 按技能匹配过滤（skills字段包含wo_type相关关键词）
    3. 按工作量排序（today_count < max_daily）
    4. 按好评率排序（rating降序）
    5. 返回最优工程师列表
    """
    # 工单类型与技能关键词的映射
    skill_keywords = {
        'repair': ['维修', 'repair', '打印机', 'printer', '电脑', 'computer', '笔记本', '硬件', '监控', '摄像头', 'camera', 'NVR', 'DVR', '海康', '大华', '网络', 'network', '宽带', '路由', '交换机', 'WiFi', '布线'],
        'maintenance': ['维护', 'maintenance', '保养', '打印机', 'printer', '电脑', 'computer', '监控', '摄像头', 'camera', '网络', 'network', '宽带', '路由'],
        'inspection': ['检测', 'inspection', '打印机', 'printer', '电脑', 'computer', '监控', '摄像头', 'camera', '网络', 'network'],
        'installation': ['安装', 'installation', '监控', '摄像头', 'camera', '网络', 'network', '打印机', 'printer', '电脑', 'computer'],
        'delivery': ['配送', '送货', 'delivery'],
        'purchase': ['代购', '采购', 'purchase'],
        'survey': ['勘察', 'survey', '现场']
    }

    keywords = skill_keywords.get(wo_type, [])
    if required_skills:
        keywords.extend(required_skills if isinstance(required_skills, list) else [required_skills])

    # 查询所有在线工程师
    engineers = StaffStatus.query.filter_by(online_status=1).all()

    # 按技能匹配过滤
    matched = []
    for eng in engineers:
        if not keywords:
            matched.append(eng)
            continue
        eng_skills = (eng.skills or '').lower()
        for kw in keywords:
            if kw.lower() in eng_skills:
                matched.append(eng)
                break

    # 按工作量过滤（今日接单数 < 每日最大接单量）
    available = [e for e in matched if e.today_count < e.max_daily]

    # 按好评率降序、今日接单数升序排序
    available.sort(key=lambda x: (-float(x.rating or 0), x.today_count or 0))

    # 返回推荐列表
    result = []
    for eng in available[:5]:  # 最多返回5个推荐
        result.append({
            'staff_id': eng.staff_id,
            'staff_name': eng.staff_name,
            'skills': eng.skills,
            'rating': float(eng.rating or 0),
            'today_count': eng.today_count or 0,
            'max_daily': eng.max_daily or 10,
            'location': eng.location
        })

    return result

def add_wo_log(wo_id, action, old_status, new_status, content, operator_id=None, operator_name=None):
    """添加工单操作日志的辅助函数"""
    log = WorkOrderLog(
        wo_id=wo_id,
        action=action,
        old_status=old_status,
        new_status=new_status,
        content=content,
        operator_id=operator_id,
        operator_name=operator_name
    )
    db.session.add(log)
    return log

@app.route('/api/workorders', methods=['GET'])
@jwt_required()
def get_workorders():
    """获取工单列表 - 支持多条件筛选和分页"""
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)
        keyword = request.args.get('keyword', '')
        wo_type = request.args.get('wo_type', '')
        status = request.args.get('status', type=int)
        assigned_user_id = request.args.get('assigned_user_id', type=int)
        customer_id = request.args.get('customer_id', type=int)
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        min_amount = request.args.get('min_amount', type=float)
        max_amount = request.args.get('max_amount', type=float)

        query = WorkOrder.query

        # 权限控制：工程师仅查看个人工单
        user_id = get_jwt_identity()
        current_user = SysUser.query.get(user_id)
        if current_user and current_user.role_id:
            role = SysRole.query.get(current_user.role_id)
            if role and 'engineer' in (role.role_code or '').lower():
                query = query.filter(WorkOrder.assigned_user_id == user_id)

        # 关键字搜索
        if keyword:
            query = query.filter(
                db.or_(
                    WorkOrder.wo_no.contains(keyword),
                    WorkOrder.customer_name.contains(keyword),
                    WorkOrder.customer_phone.contains(keyword),
                    WorkOrder.customer_company.contains(keyword),
                    WorkOrder.device_sn.contains(keyword),
                    WorkOrder.device_model.contains(keyword),
                    WorkOrder.fault_desc.contains(keyword)
                )
            )

        # 工单类型筛选
        if wo_type:
            query = query.filter_by(wo_type=wo_type)

        # 状态筛选
        if status is not None:
            query = query.filter_by(status=status)

        # 指派工程师筛选
        if assigned_user_id:
            query = query.filter_by(assigned_user_id=assigned_user_id)

        # 客户筛选
        if customer_id:
            query = query.filter_by(customer_id=customer_id)

        # 日期范围筛选
        if date_start:
            query = query.filter(WorkOrder.created_at >= date_start)
        if date_end:
            query = query.filter(WorkOrder.created_at <= date_end + ' 23:59:59')

        # 金额范围筛选
        if min_amount is not None:
            query = query.filter(WorkOrder.total_cost >= min_amount)
        if max_amount is not None:
            query = query.filter(WorkOrder.total_cost <= max_amount)

        total = query.count()
        orders = query.order_by(WorkOrder.created_at.desc()).paginate(
            page=page, per_page=page_size, error_out=False
        )

        return jsonify({
            'code': 200,
            'data': {
                'list': [{
                    'id': o.id,
                    'wo_no': o.wo_no,
                    'wo_type': o.wo_type,
                    'wo_type_text': WO_TYPE_MAP.get(o.wo_type, o.wo_type or '未知'),
                    'wo_sub_type': o.wo_sub_type,
                    'sub_type_name': WO_SUB_TYPE_MAP.get(o.wo_sub_type, '') if o.wo_sub_type else '',
                    'customer_name': o.customer_name,
                    'customer_phone': o.customer_phone,
                    'customer_company': o.customer_company,
                    'device_type': o.device_type,
                    'device_brand': o.device_brand,
                    'device_model': o.device_model,
                    'fault_desc': o.fault_desc,
                    'status': o.status,
                    'status_text': WO_STATUS_MAP.get(o.status, '未知'),
                    'priority': o.priority,
                    'assigned_user_id': o.assigned_user_id,
                    'assigned_user_name': o.assigned_user_name,
                    'labor_cost': float(o.labor_cost) if o.labor_cost else 0.00,
                    'parts_cost': float(o.parts_cost) if o.parts_cost else 0.00,
                    'total_cost': float(o.total_cost) if o.total_cost else 0.00,
                    'settlement_status': o.settlement_status,
                    'customer_address': o.customer_address,
                    'order_source': '',  # 从关联表查询
                    'service_type': '',  # 从关联表查询
                    'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else None
                } for o in orders.items],
                'total': total,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        logger.error(f'获取工单列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取工单列表失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>', methods=['GET'])
@jwt_required()
def get_workorder(id):
    """获取工单详情 - 包含配件明细和操作日志"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        # 获取配件明细
        parts = WorkOrderPart.query.filter_by(wo_id=id).all()

        # 获取操作日志
        logs = WorkOrderLog.query.filter_by(wo_id=id).order_by(WorkOrderLog.created_at.desc()).all()

        # 获取报价配件明细
        quote_items = WorkOrderQuoteItem.query.filter_by(work_order_id=id).all()

        # 获取工单扩展信息（关联表）
        extend = WorkOrderExtend.query.filter_by(wo_id=id).first()

        result = to_dict(order)
        result['status_text'] = WO_STATUS_MAP.get(order.status, '未知')
        result['wo_type_text'] = WO_TYPE_MAP.get(order.wo_type, order.wo_type or '未知')
        result['sub_type_name'] = WO_SUB_TYPE_MAP.get(order.wo_sub_type, '') if order.wo_sub_type else ''
        # 字段映射（兼容前端字段名）
        result['reception_user_id'] = order.receiver_id
        result['reception_user_name'] = order.receiver_name
        result['expected_finish_time'] = order.estimated_time.strftime('%Y-%m-%d %H:%M:%S') if order.estimated_time else ''
        result['parts'] = [to_dict(p) for p in parts]
        result['logs'] = [to_dict(l) for l in logs]
        result['quote_items'] = [to_dict(i) for i in quote_items]
        
        # 获取客户需求配件
        customer_parts = WoCustomerPart.query.filter_by(wo_id=id).all()
        result['customer_parts'] = [to_dict(cp) for cp in customer_parts]
        
        # 从关联表获取扩展信息
        result['order_source'] = extend.order_source if extend else ''
        result['service_type'] = extend.service_type if extend else ''

        # 获取动态字段
        dynamic_fields = WoDynamicField.query.filter_by(wo_id=id).all()
        result['dynamic_fields'] = [{
            'id': df.id,
            'field_key': df.field_key,
            'field_value': df.field_value,
            'field_label': df.field_label
        } for df in dynamic_fields]

        # 解析配送产品JSON
        if result.get('delivery_products'):
            try:
                result['delivery_products'] = json.loads(result['delivery_products'])
            except:
                result['delivery_products'] = []
        else:
            result['delivery_products'] = []

        # 根据wo_type返回可选的二级分类列表
        if order.wo_type:
            sub_types = WoSubType.query.filter_by(parent_type=order.wo_type, status=1).order_by(WoSubType.sort_order).all()
            result['sub_types'] = [{
                'sub_type_code': s.sub_type_code,
                'sub_type_name': s.sub_type_name,
                'device_category': s.device_category
            } for s in sub_types]
        else:
            result['sub_types'] = []

        return jsonify({'code': 200, 'data': result})
    except Exception as e:
        logger.error(f'获取工单详情失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取工单详情失败: {str(e)}'}), 500

@app.route('/api/workorders/init-subtypes', methods=['POST'])
@jwt_required()
def init_wo_subtypes():
    """初始化工单二级分类数据"""
    # 定义所有二级分类
    subtypes = [
        # 维修服务
        {'parent_type': 'repair', 'sub_type_code': 'repair_monitor', 'sub_type_name': '监控维修', 'device_category': 'monitor', 'sort_order': 1},
        {'parent_type': 'repair', 'sub_type_code': 'repair_network', 'sub_type_name': '网络维修', 'device_category': 'network', 'sort_order': 2},
        {'parent_type': 'repair', 'sub_type_code': 'repair_printer', 'sub_type_name': '打印维修', 'device_category': 'printer', 'sort_order': 3},
        {'parent_type': 'repair', 'sub_type_code': 'repair_computer', 'sub_type_name': '电脑维修', 'device_category': 'computer', 'sort_order': 4},
        {'parent_type': 'repair', 'sub_type_code': 'repair_other', 'sub_type_name': '其他办公设备维修', 'device_category': 'other', 'sort_order': 5},
        # 维护服务
        {'parent_type': 'maintenance', 'sub_type_code': 'maintenance_monitor', 'sub_type_name': '监控维护', 'device_category': 'monitor', 'sort_order': 1},
        {'parent_type': 'maintenance', 'sub_type_code': 'maintenance_network', 'sub_type_name': '网络维护', 'device_category': 'network', 'sort_order': 2},
        {'parent_type': 'maintenance', 'sub_type_code': 'maintenance_printer', 'sub_type_name': '打印维护', 'device_category': 'printer', 'sort_order': 3},
        {'parent_type': 'maintenance', 'sub_type_code': 'maintenance_computer', 'sub_type_name': '电脑维护', 'device_category': 'computer', 'sort_order': 4},
        {'parent_type': 'maintenance', 'sub_type_code': 'maintenance_other', 'sub_type_name': '其他办公设备维护', 'device_category': 'other', 'sort_order': 5},
        # 检测服务
        {'parent_type': 'inspection', 'sub_type_code': 'inspection_monitor', 'sub_type_name': '监控检测', 'device_category': 'monitor', 'sort_order': 1},
        {'parent_type': 'inspection', 'sub_type_code': 'inspection_network', 'sub_type_name': '网络检测', 'device_category': 'network', 'sort_order': 2},
        {'parent_type': 'inspection', 'sub_type_code': 'inspection_printer', 'sub_type_name': '打印检测', 'device_category': 'printer', 'sort_order': 3},
        {'parent_type': 'inspection', 'sub_type_code': 'inspection_computer', 'sub_type_name': '电脑检测', 'device_category': 'computer', 'sort_order': 4},
        {'parent_type': 'inspection', 'sub_type_code': 'inspection_other', 'sub_type_name': '其他办公设备检测', 'device_category': 'other', 'sort_order': 5},
        # 安装服务
        {'parent_type': 'installation', 'sub_type_code': 'installation_monitor', 'sub_type_name': '监控安装', 'device_category': 'monitor', 'sort_order': 1},
        {'parent_type': 'installation', 'sub_type_code': 'installation_network', 'sub_type_name': '网络安装', 'device_category': 'network', 'sort_order': 2},
        {'parent_type': 'installation', 'sub_type_code': 'installation_printer', 'sub_type_name': '打印安装', 'device_category': 'printer', 'sort_order': 3},
        {'parent_type': 'installation', 'sub_type_code': 'installation_computer', 'sub_type_name': '电脑安装', 'device_category': 'computer', 'sort_order': 4},
        {'parent_type': 'installation', 'sub_type_code': 'installation_other', 'sub_type_name': '其他办公设备安装', 'device_category': 'other', 'sort_order': 5},
    ]

    count = 0
    for st in subtypes:
        existing = WoSubType.query.filter_by(sub_type_code=st['sub_type_code']).first()
        if not existing:
            obj = WoSubType(**st)
            db.session.add(obj)
            count += 1

    db.session.commit()
    return jsonify({'code': 200, 'message': f'初始化完成，新增 {count} 条二级分类'})

@app.route('/api/workorders/subtypes', methods=['GET'])
@jwt_required()
def get_wo_subtypes():
    """获取工单二级分类列表"""
    parent_type = request.args.get('parent_type', '')
    query = WoSubType.query.filter_by(status=1)
    if parent_type:
        query = query.filter_by(parent_type=parent_type)
    subtypes = query.order_by(WoSubType.sort_order).all()
    return jsonify({
        'code': 200,
        'data': [{
            'id': s.id,
            'parent_type': s.parent_type,
            'sub_type_code': s.sub_type_code,
            'sub_type_name': s.sub_type_name,
            'device_category': s.device_category
        } for s in subtypes]
    })

@app.route('/api/workorders', methods=['POST'])
@jwt_required()
def create_workorder():
    """创建工单 - 根据wo_type接收不同字段，自动生成工单号"""
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 统一类型转换：防止前端传字符串导致数据库报错
        _int_fields = ['priority', 'customer_id', 'receiver_id', 'assigned_user_id',
                       'need_bring_back', 'auto_dispatch', 'device_need_door',
                       'net_need_device', 'goods_quantity', 'goods_need_install',
                       'goods_floor', 'monitor_need_record', 'record_days',
                       'camera_count', 'purchase_qty', 'print_count',
                       'test_result', 'customer_acceptance', 'settlement_status']
        _float_fields = ['labor_hours', 'labor_unit_price', 'service_fee',
                         'estimated_cost', 'purchase_price',
                         'labor_cost', 'parts_cost', 'material_cost',
                         'transport_cost', 'total_cost']
        for f in _int_fields:
            if f in data and data[f] is not None and data[f] != '':
                try: data[f] = int(data[f])
                except (ValueError, TypeError): data[f] = 0 if f == 'priority' else None
        for f in _float_fields:
            if f in data and data[f] is not None and data[f] != '':
                try: data[f] = float(data[f])
                except (ValueError, TypeError): data[f] = 0

        # 统一类型转换：防止前端传字符串导致数据库报错
        _int_fields = ['priority', 'customer_id', 'receiver_id', 'assigned_user_id',
                       'need_bring_back', 'auto_dispatch', 'device_need_door',
                       'net_need_device', 'goods_quantity', 'goods_need_install',
                       'goods_floor', 'monitor_need_record', 'record_days',
                       'camera_count', 'purchase_qty', 'print_count',
                       'test_result', 'customer_acceptance', 'settlement_status']
        _float_fields = ['labor_hours', 'labor_unit_price', 'service_fee',
                         'estimated_cost', 'purchase_price',
                         'labor_cost', 'parts_cost', 'material_cost',
                         'transport_cost', 'total_cost']
        for f in _int_fields:
            if f in data and data[f] is not None and data[f] != '':
                try:
                    data[f] = int(data[f])
                except (ValueError, TypeError):
                    if f == 'priority':
                        data[f] = 0  # normal->0
                    elif f in ('need_bring_back', 'auto_dispatch', 'device_need_door',
                               'net_need_device', 'goods_need_install', 'monitor_need_record'):
                        data[f] = 0
                    else:
                        data[f] = None
        for f in _float_fields:
            if f in data and data[f] is not None and data[f] != '':
                try:
                    data[f] = float(data[f])
                except (ValueError, TypeError):
                    data[f] = 0

        # 生成工单号：WO + 日期 + 序号
        today_str = datetime.now().strftime('%Y%m%d')
        today_prefix = f'WO{today_str}'
        last_order = WorkOrder.query.filter(WorkOrder.wo_no.like(f'{today_prefix}%')).order_by(WorkOrder.id.desc()).first()
        if last_order:
            seq = int(last_order.wo_no[-4:]) + 1
        else:
            seq = 1
        wo_no = f'{today_prefix}{str(seq).zfill(4)}'

        # 创建工单对象
        order = WorkOrder(
            wo_no=wo_no,
            wo_type=data.get('wo_type'),
            wo_sub_type=data.get('wo_sub_type'),
            customer_id=data.get('customer_id'),
            customer_name=data.get('customer_name'),
            customer_phone=data.get('customer_phone'),
            customer_address=data.get('customer_address'),
            # 设备类字段
            device_type=data.get('device_type'),
            device_brand=data.get('device_brand'),
            device_model=data.get('device_model'),
            device_sn=data.get('device_sn'),
            device_imei=data.get('device_imei'),
            device_password=data.get('device_password'),
            device_need_door=data.get('device_need_door', 0),
            # 网络类字段
            net_type=data.get('net_type'),
            net_operator=data.get('net_operator'),
            net_need_device=data.get('net_need_device', 0),
            # 配送/监控/代购类字段
            goods_type=data.get('goods_type'),
            goods_quantity=data.get('goods_quantity', 1),
            goods_need_install=data.get('goods_need_install', 0),
            goods_floor_type=data.get('goods_floor_type'),
            goods_floor=data.get('goods_floor', 1),
            # 监控类字段
            monitor_brand=data.get('monitor_brand'),
            camera_count=data.get('camera_count', 0),
            camera_location=data.get('camera_location'),
            monitor_need_record=data.get('monitor_need_record', 0),
            record_days=data.get('record_days', 7),
            # 通用字段
            fault_desc=data.get('fault_desc'),
            appearance_desc=data.get('appearance_desc'),
            accessories=data.get('accessories'),
            remark=data.get('remark'),
            priority=data.get('priority', 0),
            # 新增通用字段
            customer_company=data.get('customer_company'),
            customer_office=data.get('customer_office'),
            receiver_id=data.get('receiver_id'),
            receiver_name=data.get('receiver_name'),
            need_bring_back=data.get('need_bring_back', 0),
            auto_dispatch=data.get('auto_dispatch', 0),
            dispatch_rule=data.get('dispatch_rule'),
            labor_hours=data.get('labor_hours'),
            labor_unit_price=data.get('labor_unit_price'),
            service_fee=data.get('service_fee', 0),
            delivery_address=data.get('delivery_address'),
            install_position=data.get('install_position'),
            install_material=data.get('install_material'),
            acceptance_standard=data.get('acceptance_standard'),
            customer_confirm_items=data.get('customer_confirm_items'),
            survey_address=data.get('survey_address'),
            site_environment=data.get('site_environment'),
            device_status_desc=data.get('device_status_desc'),
            problem_summary=data.get('problem_summary'),
            construction_plan=data.get('construction_plan'),
            required_parts=json.dumps(data.get('required_parts')) if isinstance(data.get('required_parts'), (list, dict)) else data.get('required_parts'),
            estimated_duration=data.get('estimated_duration'),
            estimated_cost=data.get('estimated_cost'),
            customer_device_model=data.get('customer_device_model'),
            device_source=data.get('device_source'),
            install_requirement=data.get('install_requirement'),
            consumable_usage=data.get('consumable_usage'),
            purchase_product=data.get('purchase_product'),
            purchase_brand=data.get('purchase_brand'),
            purchase_spec=data.get('purchase_spec'),
            purchase_qty=data.get('purchase_qty'),
            customer_demand=data.get('customer_demand'),
            expected_arrival_date=data.get('expected_arrival_date'),
            purchase_price=data.get('purchase_price'),
            # 网络维修专属字段
            net_topology=data.get('net_topology'),
            fault_location=data.get('fault_location'),
            net_ip=data.get('net_ip'),
            device_port=data.get('device_port'),
            line_type=data.get('line_type'),
            test_items=data.get('test_items'),
            net_speed_data=data.get('net_speed_data'),
            maintenance_cycle=data.get('maintenance_cycle'),
            restart_record=data.get('restart_record'),
            debug_content=data.get('debug_content'),
            # 设备维修专属字段
            device_config=data.get('device_config'),
            os_version=data.get('os_version'),
            error_code=data.get('error_code'),
            repair_part=data.get('repair_part'),
            maintenance_items=data.get('maintenance_items'),
            replaced_parts=json.dumps(data.get('replaced_parts')) if isinstance(data.get('replaced_parts'), (list, dict)) else data.get('replaced_parts'),
            retest_result=data.get('retest_result'),
            # 监控维修专属字段
            channel_no=data.get('channel_no'),
            nvr_model=data.get('nvr_model'),
            disk_capacity=data.get('disk_capacity'),
            recording_status=data.get('recording_status'),
            screen_fault=data.get('screen_fault'),
            infrared_status=data.get('infrared_status'),
            power_status=data.get('power_status'),
            line_inspection=data.get('line_inspection'),
            point_debug_record=data.get('point_debug_record'),
            # 监控安装专属字段
            install_points=data.get('install_points'),
            camera_model=data.get('camera_model'),
            storage_config=data.get('storage_config'),
            cable_length=data.get('cable_length'),
            consumable_qty=data.get('consumable_qty'),
            debug_result=data.get('debug_result'),
            picture_clarity=data.get('picture_clarity'),
            recording_settings=data.get('recording_settings'),
            # 新增字段
            delivery_products=json.dumps(data.get('delivery_products')) if data.get('delivery_products') else None,
            repair_camera_count=data.get('repair_camera_count', 0),
            # 状态
            status=0,
            status_name='待派单',
            created_by=user_id
        )

        db.session.add(order)
        db.session.flush()

        # 保存扩展信息到关联表
        if data.get('order_source') or data.get('service_type'):
            extend = WorkOrderExtend(
                wo_id=order.id,
                order_source=data.get('order_source'),
                service_type=data.get('service_type')
            )
            db.session.add(extend)

        # 保存动态字段
        dynamic_fields = data.get('dynamic_fields', [])
        for df in dynamic_fields:
            df_obj = WoDynamicField(
                wo_id=order.id,
                field_key=df.get('field_key'),
                field_value=str(df.get('field_value', '')),
                field_label=df.get('field_label', '')
            )
            db.session.add(df_obj)

        # 保存客户需求配件（只做信息记录，不扣库存）
        customer_parts = data.get('parts', [])
        for cp in customer_parts:
            cp_obj = WoCustomerPart(
                wo_id=order.id,
                product_id=cp.get('product_id'),
                product_name=cp.get('product_name') or cp.get('part_name', ''),
                specification=cp.get('specification', ''),
                quantity=cp.get('quantity', 1),
                unit_price=cp.get('unit_price', 0),
                remark=cp.get('remark', '客户需求配件')
            )
            db.session.add(cp_obj)

        # 如果 need_bring_back=1，自动创建接件单
        if data.get('need_bring_back') == 1:
            try:
                last_ro = ReceiveOrder.query.order_by(ReceiveOrder.id.desc()).first()
                ro_no = generate_code('RO', last_ro.id if last_ro else 0)
                receive_order = ReceiveOrder(
                    receive_no=ro_no,
                    customer_id=data.get('customer_id'),
                    customer_name=data.get('customer_name'),
                    customer_phone=data.get('customer_phone'),
                    receive_type=1,
                    receiver_id=user_id,
                    receiver_name=user_name,
                    remark=f'工单{wo_no}自动创建接件单',
                    created_by=user_id
                )
                db.session.add(receive_order)
                db.session.flush()
                order.receive_order_id = receive_order.id
            except Exception as e:
                logger.warning(f'自动创建接件单失败: {str(e)}')

        # 记录操作日志
        add_wo_log(
            wo_id=order.id,
            action='创建工单',
            old_status=None,
            new_status=0,
            content=f'创建工单，工单号：{wo_no}，类型：{WO_TYPE_MAP.get(data.get("wo_type", ""), data.get("wo_type", ""))}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()

        return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'wo_no': wo_no}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'创建工单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'创建工单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>', methods=['PUT'])
@jwt_required()
def update_workorder(id):
    """更新工单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 统一类型转换
        _int_fields = ['priority', 'customer_id', 'receiver_id', 'assigned_user_id',
                       'need_bring_back', 'auto_dispatch', 'device_need_door',
                       'net_need_device', 'goods_quantity', 'goods_need_install',
                       'goods_floor', 'monitor_need_record', 'record_days',
                       'camera_count', 'purchase_qty', 'print_count',
                       'test_result', 'customer_acceptance', 'settlement_status']
        _float_fields = ['labor_hours', 'labor_unit_price', 'service_fee',
                         'estimated_cost', 'purchase_price',
                         'labor_cost', 'parts_cost', 'material_cost',
                         'transport_cost', 'total_cost']
        for f in _int_fields:
            if f in data and data[f] is not None and data[f] != '':
                try: data[f] = int(data[f])
                except (ValueError, TypeError): data[f] = 0 if f == 'priority' else None
        for f in _float_fields:
            if f in data and data[f] is not None and data[f] != '':
                try: data[f] = float(data[f])
                except (ValueError, TypeError): data[f] = 0

        # 可更新的字段列表
        updatable_fields = [
            'wo_type', 'wo_sub_type', 'customer_id', 'customer_name', 'customer_phone',
            'customer_address', 'customer_company', 'customer_office',
            'device_type', 'device_brand', 'device_model', 'device_sn',
            'device_imei', 'device_password', 'device_need_door',
            'net_type', 'net_operator', 'net_need_device',
            'goods_type', 'goods_quantity', 'goods_need_install', 'goods_floor_type', 'goods_floor',
            'monitor_brand', 'camera_count', 'camera_location', 'monitor_need_record', 'record_days',
            'fault_desc', 'appearance_desc', 'accessories', 'remark', 'priority',
            'receiver_id', 'receiver_name', 'auto_dispatch', 'dispatch_rule',
            'labor_hours', 'labor_unit_price', 'service_fee',
            'delivery_address', 'install_position', 'arrival_time',
            'install_material', 'acceptance_standard', 'customer_confirm_items',
            'survey_address', 'site_environment', 'device_status_desc',
            'problem_summary', 'construction_plan', 'required_parts',
            'estimated_duration', 'estimated_cost',
            'customer_device_model', 'device_source', 'install_requirement',
            'consumable_usage', 'purchase_product', 'purchase_brand', 'purchase_spec',
            'purchase_qty', 'customer_demand', 'expected_arrival_date', 'purchase_price',
            'net_topology', 'fault_location', 'net_ip', 'device_port', 'line_type',
            'test_items', 'net_speed_data', 'maintenance_cycle', 'restart_record', 'debug_content',
            'device_config', 'os_version', 'error_code', 'repair_part',
            'maintenance_items', 'replaced_parts', 'retest_result',
            'channel_no', 'nvr_model', 'disk_capacity', 'recording_status',
            'screen_fault', 'infrared_status', 'power_status',
            'line_inspection', 'point_debug_record',
            'install_points', 'camera_model', 'storage_config', 'cable_length',
            'consumable_qty', 'debug_result', 'picture_clarity', 'recording_settings',
            'labor_cost', 'parts_cost', 'material_cost', 'transport_cost', 'total_cost',
            'assigned_user_id', 'assigned_user_name',
            'delivery_products', 'repair_camera_count',
            'estimated_time',
        ]

        # 前端字段名到后端字段名映射
        field_mapping = {
            'reception_user_id': 'receiver_id',
            'reception_user_name': 'receiver_name',
            'expected_finish_time': 'estimated_time'
        }

        for field in updatable_fields:
            if field in data:
                value = data[field]
                # JSON字段自动序列化
                if field in ('required_parts', 'replaced_parts', 'finish_photos') and isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                # delivery_products 序列化为JSON
                if field == 'delivery_products' and isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                setattr(order, field, value)

        # 处理前端字段名映射
        for frontend_field, backend_field in field_mapping.items():
            if frontend_field in data:
                value = data[frontend_field]
                if backend_field == 'estimated_time' and value:
                    try:
                        from datetime import datetime
                        if isinstance(value, str):
                            value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                    except:
                        pass
                setattr(order, backend_field, value)

        # 更新动态字段
        dynamic_fields = data.get('dynamic_fields', [])
        if dynamic_fields:
            # 先删除旧的动态字段
            WoDynamicField.query.filter_by(wo_id=id).delete()
            # 重新添加
            for df in dynamic_fields:
                df_obj = WoDynamicField(
                    wo_id=id,
                    field_key=df.get('field_key'),
                    field_value=str(df.get('field_value', '')),
                    field_label=df.get('field_label', '')
                )
                db.session.add(df_obj)

        # 更新扩展信息到关联表
        if 'order_source' in data or 'service_type' in data:
            extend = WorkOrderExtend.query.filter_by(wo_id=id).first()
            if extend:
                if 'order_source' in data:
                    extend.order_source = data.get('order_source')
                if 'service_type' in data:
                    extend.service_type = data.get('service_type')
            else:
                extend = WorkOrderExtend(
                    wo_id=id,
                    order_source=data.get('order_source'),
                    service_type=data.get('service_type')
                )
                db.session.add(extend)

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='更新工单',
            old_status=order.status,
            new_status=order.status,
            content=f'更新工单信息',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新工单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'更新工单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_workorder(id):
    """删除工单 - 只有状态为待派单(0)或已取消(7)的工单才能删除"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        if order.status not in (0, 7):
            return jsonify({'code': 400, 'message': f'当前状态为【{WO_STATUS_MAP.get(order.status, "未知")}】，只有待派单或已取消的工单才能删除'}), 400

        # 删除关联数据
        WorkOrderPart.query.filter_by(wo_id=id).delete()
        WorkOrderQuoteItem.query.filter_by(work_order_id=id).delete()
        WorkOrderLog.query.filter_by(wo_id=id).delete()
        WorkOrderExtend.query.filter_by(wo_id=id).delete()
        WoCustomerPart.query.filter_by(wo_id=id).delete()

        db.session.delete(order)
        db.session.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除工单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'删除工单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/status', methods=['POST'])
@jwt_required()
def change_workorder_status(id):
    """通用状态变更接口 - 验证状态流转合法性"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        new_status = data.get('status')
        content = data.get('content', '')

        if new_status is None:
            return jsonify({'code': 400, 'message': '状态不能为空'}), 400

        old_status = order.status

        # 验证状态流转合法性
        allowed = WO_STATUS_TRANSITIONS.get(old_status, [])
        if new_status not in allowed:
            return jsonify({'code': 400, 'message': f'不允许从【{WO_STATUS_MAP.get(old_status, "未知")}】变更为【{WO_STATUS_MAP.get(new_status, "未知")}】'}), 400

        order.status = new_status
        order.status_name = WO_STATUS_MAP.get(new_status, '')

        # 特殊状态处理
        if new_status == 3 and data.get('assigned_user_id'):  # 开始处理
            order.assigned_user_id = data['assigned_user_id']
            user_obj = SysUser.query.get(data['assigned_user_id'])
            if user_obj:
                order.assigned_user_name = user_obj.real_name or user_obj.username
            order.assigned_time = datetime.now()
        elif new_status == 8:  # 待结算
            # 计算总费用
            parts = WorkOrderPart.query.filter_by(wo_id=id, status=1).all()
            parts_cost = sum(float(p.total_price or 0) for p in parts)
            order.parts_cost = parts_cost
            order.total_cost = float(order.labor_cost or 0) + parts_cost + float(order.material_cost or 0) + float(order.transport_cost or 0)
        elif new_status == 9:  # 已完成
            order.actual_time = datetime.now()

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='状态变更',
            old_status=old_status,
            new_status=new_status,
            content=content or f'状态从【{WO_STATUS_MAP.get(old_status, "未知")}】变更为【{WO_STATUS_MAP.get(new_status, "未知")}】',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '状态变更成功', 'data': {'status': new_status, 'status_text': WO_STATUS_MAP.get(new_status, '未知')}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'状态变更失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'状态变更失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/dispatch', methods=['POST'])
@jwt_required()
def dispatch_workorder(id):
    """派单 - 支持手动选择工程师和自动派单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 验证当前状态（只能对待派单状态的工单进行派单）
        if order.status != 0:
            return jsonify({'code': 400, 'message': f'当前状态【{WO_STATUS_MAP.get(order.status, "未知")}】不允许派单'}), 400

        dispatch_type = data.get('dispatch_type', 'manual')  # manual/auto
        staff_id = data.get('staff_id') or data.get('assigned_user_id')
        staff_name = data.get('staff_name')

        # 自动派单：仅当明确选择自动派单且未指定工程师时
        if dispatch_type == 'auto' and not staff_id:
            recommendations = auto_dispatch_engineer(order.wo_type)
            if not recommendations:
                return jsonify({'code': 400, 'message': '没有可用的在线工程师'}), 400
            staff_id = recommendations[0]['staff_id']
            staff_name = recommendations[0]['staff_name']
            dispatch_type = 'auto'

        # 手动派单验证
        if dispatch_type == 'manual' and not staff_id:
            return jsonify({'code': 400, 'message': '请选择工程师'}), 400

        # 验证工程师
        staff_user = SysUser.query.get(staff_id)
        if not staff_user:
            return jsonify({'code': 404, 'message': '工程师不存在'}), 404
        if not staff_name:
            staff_name = staff_user.real_name or staff_user.username

        # 创建派单记录
        dispatch_record = DispatchRecord(
            wo_id=id,
            dispatch_type=dispatch_type,
            dispatcher_id=user_id,
            dispatcher_name=user_name,
            staff_id=staff_id,
            staff_name=staff_name,
            staff_phone=staff_user.phone,
            accept_status=1  # 直接接单
        )
        db.session.add(dispatch_record)

        # 更新工程师状态
        staff_status = StaffStatus.query.filter_by(staff_id=staff_id).first()
        if staff_status:
            staff_status.current_wo_id = id
            staff_status.today_count = (staff_status.today_count or 0) + 1
        else:
            # 如果没有状态记录，创建一条
            new_staff_status = StaffStatus(
                staff_id=staff_id,
                staff_name=staff_name,
                online_status=1,
                current_wo_id=id,
                today_count=1
            )
            db.session.add(new_staff_status)

        # 更新工单状态
        old_status = order.status
        order.status = 1  # 已派单
        order.status_name = '已派单'
        order.assigned_user_id = staff_id
        order.assigned_user_name = staff_name
        order.assigned_time = datetime.now()
        order.auto_dispatch = 1 if dispatch_type == 'auto' else 0

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='派单',
            old_status=old_status,
            new_status=1,
            content=f'{dispatch_type == "auto" and "自动" or "手动"}派单给工程师【{staff_name}】',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '派单成功', 'data': {
            'staff_id': staff_id,
            'staff_name': staff_name,
            'dispatch_type': dispatch_type,
            'status': 1,
            'status_text': '已派单'
        }})
    except Exception as e:
        db.session.rollback()
        logger.error(f'派单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'派单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/accept', methods=['POST'])
@jwt_required()
def accept_workorder(id):
    """工程师接单 - 状态从1(已派单)变为2(处理中)"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 验证当前状态必须是已派单(1)
        if order.status != 1:
            return jsonify({'code': 400, 'message': f'当前状态【{WO_STATUS_MAP.get(order.status, "未知")}】不允许接单，需为已派单'}), 400

        # 状态变更：1(已派单) -> 2(处理中)
        old_status = order.status
        order.status = 2
        order.status_name = '处理中'
        order.accept_time = datetime.now()  # 记录接单时间

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='接单',
            old_status=old_status,
            new_status=2,
            content=f'工程师【{user_name}】已接单，开始处理',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '接单成功', 'data': {
            'status': 2,
            'status_text': '处理中',
            'accept_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }})
    except Exception as e:
        db.session.rollback()
        logger.error(f'接单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'接单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/allocate-parts', methods=['POST'])
@jwt_required()
def allocate_parts(id):
    """领用配件 - 检查库存，有库存出库，无库存创建采购预订单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        parts_list = data.get('parts', [])

        if not parts_list:
            return jsonify({'code': 400, 'message': '配件列表不能为空'}), 400

        allocated_parts = []
        need_purchase_parts = []

        for item in parts_list:
            product_id = item.get('product_id')
            quantity = item.get('quantity', 1)

            if not product_id:
                continue

            product = ProductInfo.query.get(product_id)
            if not product:
                continue

            # 检查库存
            stock = InventoryStock.query.filter_by(product_id=product_id).first()
            available_qty = float(stock.available_quantity or 0) if stock else 0

            if available_qty >= quantity:
                # 有库存：创建出库单(out_type=2维修领料)
                last_out = InventoryOut.query.order_by(InventoryOut.id.desc()).first()
                out_no = generate_code('CK', last_out.id if last_out else 0)

                out_order = InventoryOut(
                    out_no=out_no,
                    out_type=2,  # 维修领料
                    customer_id=order.customer_id,
                    customer_name=order.customer_name,
                    total_quantity=quantity,
                    total_amount=float(product.sale_price or 0) * quantity,
                    status=1,  # 直接审核通过
                    related_order_id=id,
                    related_order_no=order.wo_no,
                    remark=f'工单{order.wo_no}维修领料',
                    created_by=user_id
                )
                db.session.add(out_order)
                db.session.flush()

                # 创建出库明细
                out_item = InventoryOutItem(
                    out_id=out_order.id,
                    product_id=product_id,
                    product_code=product.product_code,
                    product_name=product.product_name,
                    specification=product.specification,
                    unit_name=product.unit_name,
                    quantity=quantity,
                    unit_price=product.sale_price,
                    total_price=float(product.sale_price or 0) * quantity,
                    cost_price=product.cost_price
                )
                db.session.add(out_item)

                # 锁定库存（增加冻结数量）
                if stock:
                    stock.frozen_quantity = float(stock.frozen_quantity or 0) + quantity
                    stock.available_quantity = float(stock.quantity or 0) - float(stock.frozen_quantity or 0)

                # 更新工单配件
                wo_part = WorkOrderPart(
                    wo_id=id,
                    product_id=product_id,
                    product_name=product.product_name,
                    product_code=product.product_code,
                    specification=product.specification,
                    quantity=quantity,
                    unit_price=product.sale_price,
                    total_price=float(product.sale_price or 0) * quantity,
                    cost_price=product.cost_price,
                    is_own=1,
                    status=1,  # 已用
                    remark='维修领料'
                )
                db.session.add(wo_part)

                allocated_parts.append({'product_name': product.product_name, 'quantity': quantity, 'source': '库存'})
            else:
                # 无库存：创建采购预订单
                need_purchase_parts.append({
                    'product_id': product_id,
                    'product_name': product.product_name,
                    'product_code': product.product_code,
                    'quantity': quantity,
                    'available': available_qty
                })

                # 更新工单配件（待采购状态）
                wo_part = WorkOrderPart(
                    wo_id=id,
                    product_id=product_id,
                    product_name=product.product_name,
                    product_code=product.product_code,
                    specification=product.specification,
                    quantity=quantity,
                    unit_price=product.purchase_price,
                    total_price=float(product.purchase_price or 0) * quantity,
                    cost_price=product.cost_price,
                    is_own=1,
                    status=0,  # 待用
                    remark='待采购'
                )
                db.session.add(wo_part)

        # 如果有需要采购的配件，创建采购预订单
        if need_purchase_parts:
            last_pre = PreOrder.query.order_by(PreOrder.id.desc()).first()
            pre_no = generate_code('YD', last_pre.id if last_pre else 0)

            total_amount = 0
            pre_order = PreOrder(
                pre_no=pre_no,
                pre_type=1,  # 采购预定
                total_quantity=sum(p['quantity'] for p in need_purchase_parts),
                total_amount=0,
                status=0,
                remark=f'工单{order.wo_no}配件采购预订单',
                created_by=user_id
            )
            db.session.add(pre_order)
            db.session.flush()

            for p in need_purchase_parts:
                product = ProductInfo.query.get(p['product_id'])
                price = float(product.purchase_price or 0)
                total_amount += price * p['quantity']

                pre_item = PreOrderItem(
                    pre_id=pre_order.id,
                    product_id=p['product_id'],
                    product_code=p['product_code'],
                    product_name=p['product_name'],
                    quantity=p['quantity'],
                    unit_price=price,
                    total_price=price * p['quantity'],
                    remark=f'工单{order.wo_no}需要'
                )
                db.session.add(pre_item)

            pre_order.total_amount = total_amount
            order.related_purchase_id = pre_order.id

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='领用配件',
            old_status=order.status,
            new_status=order.status,
            content=f'领用配件{len(allocated_parts)}项（库存），{len(need_purchase_parts)}项需采购',
            operator_id=user_id,
            operator_name=user_name
        )

        # 如果没有需要采购的配件，自动变更为"待上门"
        if not need_purchase_parts:
            old_status = order.status
            order.status = 3
            order.status_name = WO_STATUS_MAP.get(3, '待上门')
            add_wo_log(
                wo_id=id,
                action='状态变更',
                old_status=old_status,
                new_status=3,
                content=f'配件已全部从库存领取，状态变更为【待上门】',
                operator_id=user_id,
                operator_name=user_name
            )

        db.session.commit()
        return jsonify({
            'code': 200,
            'message': '配件分配完成',
            'data': {
                'allocated': allocated_parts,
                'need_purchase': need_purchase_parts
            }
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'领用配件失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'领用配件失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/finish', methods=['POST'])
@jwt_required()
def finish_workorder(id):
    """完工提交 - 提交完工报告、照片、测试结果，状态从2(处理中)变为3(待结算)"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        if order.status != 2:
            return jsonify({'code': 400, 'message': f'当前状态【{WO_STATUS_MAP.get(order.status, "未知")}】不允许完工提交，需为处理中'}), 400

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 更新完工信息
        if 'finish_report' in data:
            order.finish_report = data['finish_report']
        if 'finish_photos' in data:
            photos = data['finish_photos']
            order.finish_photos = json.dumps(photos, ensure_ascii=False) if isinstance(photos, (list, dict)) else photos
        if 'test_result' in data:
            order.test_result = data['test_result']
        if 'test_remark' in data:
            order.test_remark = data['test_remark']

        # 状态变更：2(处理中) -> 3(待结算)
        old_status = order.status
        order.status = 3
        order.status_name = '待结算'

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='完工提交',
            old_status=old_status,
            new_status=3,
            content=f'完工提交，测试结果：{["待测试", "通过", "未通过"][data.get("test_result", 0)]}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '完工提交成功', 'data': {'status': 3, 'status_text': '待结算'}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'完工提交失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'完工提交失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/settle', methods=['POST'])
@jwt_required()
def settle_workorder(id):
    """结算 - 计算费用、创建销售单和财务/应收记录，状态从4(处理中)或5(待结算)变为6(已完成)
    自动计算已用备件和未用备件，未用备件自动退回库存
    支持现金结算（直接入账）和签单结算（生成应收）"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        # 允许状态4(处理中)和5(待结算)进行结算
        if order.status not in [4, 5]:
            return jsonify({'code': 400, 'message': f'当前状态【{WO_STATUS_MAP.get(order.status, "未知")}】不允许结算，需为处理中或待结算'}), 400

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        
        # 结算方式：cash=现金结算(直接入账), credit=签单结算(生成应收)
        settle_type = data.get('settle_type', 'cash')

        # 获取工单已领备件
        parts = WorkOrderPart.query.filter_by(wo_id=id).all()
        
        # 自动计算已用备件和未用备件
        used_parts = []  # 已用备件清单
        unused_parts = []  # 未用备件清单
        parts_cost_total = 0.0  # 配件费合计
        
        for part in parts:
            # 如果前端传了 used_quantity，使用前端的值；否则默认为领用数量
            used_qty = float(data.get(f'used_qty_{part.id}', part.used_quantity or part.quantity or 0))
            unused_qty = float(part.quantity or 0) - used_qty
            
            # 更新已用数量
            part.used_quantity = used_qty
            
            if used_qty > 0:
                # 标记为已用
                part.status = 1  # 已用
                used_parts.append({
                    'product_name': part.product_name,
                    'specification': part.specification,
                    'quantity': used_qty,
                    'unit_price': float(part.unit_price or 0),
                    'total': used_qty * float(part.unit_price or 0)
                })
                parts_cost_total += used_qty * float(part.unit_price or 0)
            
            if unused_qty > 0:
                # 未用备件自动退回（增加库存）
                part.status = 2  # 已退
                
                # 如果是本店配件，退回库存
                if part.is_own == 1 and part.product_id:
                    try:
                        product = Product.query.get(part.product_id)
                        if product:
                            product.stock = float(product.stock or 0) + unused_qty
                    except Exception as e:
                        logger.warning(f'退回库存失败: {str(e)}')
                
                unused_parts.append({
                    'product_name': part.product_name,
                    'specification': part.specification,
                    'quantity': unused_qty
                })

        # 更新费用信息
        labor_hours = data.get('labor_hours')
        labor_unit_price = data.get('labor_unit_price')
        labor_cost = data.get('labor_cost')
        material_cost = data.get('material_cost')
        other_cost = data.get('other_cost', 0)
        account_id = data.get('account_id')

        if labor_hours is not None:
            order.labor_hours = labor_hours
        if labor_unit_price is not None:
            order.labor_unit_price = labor_unit_price
        if labor_cost is not None:
            order.labor_cost = labor_cost
        elif labor_hours and labor_unit_price:
            order.labor_cost = float(labor_hours) * float(labor_unit_price)
        # 如果前端没有传配件费，自动用计算出的配件费
        if data.get('parts_cost') is not None:
            order.parts_cost = data.get('parts_cost')
        else:
            order.parts_cost = parts_cost_total
        if material_cost is not None:
            order.material_cost = material_cost

        # 计算总费用
        total = (float(order.labor_cost or 0) + float(order.parts_cost or 0) +
                 float(order.material_cost or 0) + float(order.transport_cost or 0) +
                 float(other_cost or 0) + float(order.service_fee or 0))
        order.total_cost = total

        # 创建销售单关联
        last_so = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
        so_no = generate_code('XS', last_so.id if last_so else 0)
        sales_order = SalesOrder(
            order_no=so_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            customer_phone=order.customer_phone,
            customer_address=order.customer_address,
            order_date=datetime.now().date(),
            total_amount=total,
            total_quantity=1,
            actual_amount=total,
            paid_amount=0,
            salesperson_id=user_id,
            salesperson_name=user_name,
            status=1,  # 已审核
            remark=f'工单{order.wo_no}结算生成',
            created_by=user_id
        )
        db.session.add(sales_order)
        db.session.flush()
        order.related_sales_id = sales_order.id

        # 根据结算方式处理财务
        finance_record = None
        receivable_record = None
        
        if settle_type == 'cash' and account_id:
            # 现金结算：直接入账
            account = FinanceAccount.query.get(account_id)
            if account:
                balance_before = float(account.balance or 0)
                account.balance = balance_before + total

                finance_record = FinanceRecord(
                    account_id=account.id,
                    account_name=account.account_name,
                    record_type=1,  # 收入
                    amount=total,
                    balance_before=balance_before,
                    balance_after=balance_before + total,
                    related_type='work_order',
                    related_id=id,
                    related_no=order.wo_no,
                    remark=f'工单现金结算：{order.wo_no}',
                    created_by=user_id
                )
                db.session.add(finance_record)
                db.session.flush()
                order.related_finance_id = finance_record.id
        elif settle_type == 'credit':
            # 签单结算：生成应收账款
            today = datetime.now().strftime('%Y%m%d')
            prefix_ys = 'YS' + today
            last_receivable = FinanceReceivable.query.filter(FinanceReceivable.receivable_no.like(prefix_ys + '%')).order_by(FinanceReceivable.receivable_no.desc()).first()
            if last_receivable and last_receivable.receivable_no and len(last_receivable.receivable_no) > len(prefix_ys):
                seq_ys = int(last_receivable.receivable_no[len(prefix_ys):]) + 1
            else:
                seq_ys = 1
            receivable_no = prefix_ys + str(seq_ys).zfill(4)
            
            receivable_record = FinanceReceivable(
                receivable_no=receivable_no,
                related_type='work_order',
                related_id=id,
                related_no=order.wo_no,
                customer_id=order.customer_id,
                customer_name=order.customer_name,
                total_amount=total,
                received_amount=0,
                remaining_amount=total,
                status=0,  # 0待收款
                remark=f'工单签单结算：{order.wo_no}',
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.session.add(receivable_record)
            db.session.flush()

        # 更新结算信息
        order.settlement_status = 1
        order.settlement_account_id = account_id if settle_type == 'cash' else None
        order.settlement_time = datetime.now()
        order.settlement_type = settle_type  # 记录结算方式
        # 状态变更：4(处理中)或5(待结算) -> 6(已完成)
        old_status = order.status
        order.status = 6  # 已完成
        order.status_name = '已完成'
        order.actual_time = datetime.now()

        # 记录日志
        settle_type_text = '现金结算' if settle_type == 'cash' else '签单结算'
        log_content = f'工单{settle_type_text}，总费用：{total}元，人工费：{order.labor_cost}，配件费：{order.parts_cost}，材料费：{order.material_cost}'
        if used_parts:
            used_str = '、'.join([f'{p["product_name"]}x{p["quantity"]}' for p in used_parts])
            log_content += f'。已用配件：{used_str}'
        if unused_parts:
            unused_str = '、'.join([f'{p["product_name"]}x{p["quantity"]}' for p in unused_parts])
            log_content += f'。已退回配件：{unused_str}'
        if settle_type == 'credit' and receivable_record:
            log_content += f'。生成应收单：{receivable_no}'
        
        add_wo_log(
            wo_id=id,
            action='结算',
            old_status=old_status,
            new_status=6,
            content=log_content,
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '结算成功', 'data': {
            'status': 6,
            'status_text': '已完成',
            'total_cost': total,
            'parts_cost': order.parts_cost,
            'used_parts': used_parts,
            'unused_parts': unused_parts,
            'sales_order_id': sales_order.id,
            'finance_record_id': finance_record.id if finance_record else None
        }})
    except Exception as e:
        db.session.rollback()
        logger.error(f'结算失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'结算失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/to-quote', methods=['POST'])
@jwt_required()
def workorder_to_quote(id):
    """转报价单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json() if request.is_json else {}
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 创建报价单
        last_qo = QuoteOrder.query.order_by(QuoteOrder.id.desc()).first()
        quote_no = generate_code('QT', last_qo.id if last_qo else 0)

        quote_order = QuoteOrder(
            quote_no=quote_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            customer_phone=order.customer_phone,
            address=order.customer_address,
            quote_date=datetime.now().date(),
            total_amount=data.get('total_amount', float(order.estimated_cost or 0)),
            remark=f'从工单{order.wo_no}转来',
            related_type='work_order',
            related_id=id,
            created_by=user_id
        )
        db.session.add(quote_order)
        db.session.flush()
        order.related_quote_id = quote_order.id

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='转报价单',
            old_status=order.status,
            new_status=order.status,
            content=f'转报价单，报价单号：{quote_no}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '转报价单成功', 'data': {'quote_id': quote_order.id, 'quote_no': quote_no}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'转报价单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'转报价单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/quote', methods=['POST'])
@jwt_required()
def workorder_quote(id):
    """工单报价 - 提交报价费用和客户确认状态"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404
        
        if order.status not in [3, 4]:  # 只允许处理中或待配件状态报价
            return jsonify({'code': 400, 'message': '当前工单状态不允许报价'}), 400
        
        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        
        # 获取报价数据
        labor_cost = float(data.get('labor_cost', 0) or 0)
        parts_cost = float(data.get('parts_cost', 0) or 0)
        other_cost = float(data.get('other_cost', 0) or 0)
        total_cost = float(data.get('total_cost', 0) or 0)
        items = data.get('items', [])
        customer_confirm = data.get('customer_confirm', 1)
        reject_reason = data.get('reject_reason', '')
        
        # 更新工单费用
        order.labor_cost = labor_cost
        order.parts_cost = parts_cost
        order.other_cost = other_cost
        order.total_cost = total_cost
        
        # 保存报价配件明细
        # 先删除旧的报价明细
        WorkOrderQuoteItem.query.filter_by(work_order_id=id).delete()
        
        # 添加新的报价明细
        for item in items:
            quote_item = WorkOrderQuoteItem(
                work_order_id=id,
                product_name=item.get('product_name', ''),
                spec=item.get('spec', ''),
                unit=item.get('unit', ''),
                quantity=float(item.get('quantity', 1) or 1),
                unit_price=float(item.get('unit_price', 0) or 0),
                subtotal=float(item.get('subtotal', 0) or 0)
            )
            db.session.add(quote_item)
        
        # 根据客户确认状态更新工单状态
        if customer_confirm == 1:
            # 客户确认，状态变为待审核
            order.status = 5
            add_workorder_log(order.id, '客户确认报价', f'报价总额: ¥{total_cost:.2f}', user_id, user_name)
        else:
            # 客户拒绝，状态变为待重新报价
            order.status = 7
            add_workorder_log(order.id, '客户拒绝报价', f'拒绝原因: {reject_reason}', user_id, user_name)
        
        db.session.commit()
        
        return jsonify({
            'code': 200, 
            'message': '报价提交成功',
            'data': {
                'status': order.status,
                'status_text': WO_STATUS_MAP.get(order.status, '未知'),
                'total_cost': total_cost
            }
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'工单报价失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'报价失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/to-purchase', methods=['POST'])
@jwt_required()
def workorder_to_purchase(id):
    """转采购预订单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json() if request.is_json else {}
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 创建采购预订单
        last_pre = PreOrder.query.order_by(PreOrder.id.desc()).first()
        pre_no = generate_code('YD', last_pre.id if last_pre else 0)

        items = data.get('items', [])
        total_amount = 0
        total_quantity = 0

        pre_order = PreOrder(
            pre_no=pre_no,
            pre_type=1,  # 采购预定
            total_quantity=0,
            total_amount=0,
            status=0,
            remark=f'从工单{order.wo_no}转来',
            created_by=user_id
        )
        db.session.add(pre_order)
        db.session.flush()

        for item in items:
            product_id = item.get('product_id')
            quantity = item.get('quantity', 1)
            unit_price = item.get('unit_price', 0)
            product = ProductInfo.query.get(product_id) if product_id else None

            total_amount += float(unit_price) * int(quantity)
            total_quantity += int(quantity)

            pre_item = PreOrderItem(
                pre_id=pre_order.id,
                product_id=product_id,
                product_code=product.product_code if product else '',
                product_name=item.get('product_name', product.product_name if product else ''),
                quantity=quantity,
                unit_price=unit_price,
                total_price=float(unit_price) * int(quantity),
                remark=item.get('remark', '')
            )
            db.session.add(pre_item)

        pre_order.total_quantity = total_quantity
        pre_order.total_amount = total_amount
        order.related_purchase_id = pre_order.id

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='转采购预订单',
            old_status=order.status,
            new_status=order.status,
            content=f'转采购预订单，预订单号：{pre_no}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '转采购预订单成功', 'data': {'pre_order_id': pre_order.id, 'pre_no': pre_no}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'转采购预订单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'转采购预订单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/to-sales', methods=['POST'])
@jwt_required()
def workorder_to_sales(id):
    """转销售单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json() if request.is_json else {}
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        # 创建销售单
        last_so = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
        so_no = generate_code('XS', last_so.id if last_so else 0)

        sales_order = SalesOrder(
            order_no=so_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            customer_phone=order.customer_phone,
            customer_address=order.customer_address,
            order_date=datetime.now().date(),
            total_amount=data.get('total_amount', float(order.total_cost or 0)),
            total_quantity=data.get('total_quantity', 1),
            actual_amount=data.get('actual_amount', float(order.total_cost or 0)),
            paid_amount=data.get('paid_amount', 0),
            payment_method=data.get('payment_method'),
            delivery_method=data.get('delivery_method'),
            salesperson_id=user_id,
            salesperson_name=user_name,
            status=0,  # 待审核
            remark=f'从工单{order.wo_no}转来',
            created_by=user_id
        )
        db.session.add(sales_order)
        db.session.flush()
        order.related_sales_id = sales_order.id

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='转销售单',
            old_status=order.status,
            new_status=order.status,
            content=f'转销售单，销售单号：{so_no}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '转销售单成功', 'data': {'sales_order_id': sales_order.id, 'order_no': so_no}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'转销售单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'转销售单失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/return-visit', methods=['POST'])
@jwt_required()
def workorder_return_visit(id):
    """上门送回"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        return_visit_time = data.get('return_visit_time')
        return_visit_result = data.get('return_visit_result')

        if return_visit_time:
            if isinstance(return_visit_time, str):
                order.return_visit_time = datetime.strptime(return_visit_time, '%Y-%m-%d %H:%M:%S')
            else:
                order.return_visit_time = return_visit_time
        if return_visit_result is not None:
            order.return_visit_result = return_visit_result

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='上门送回',
            old_status=order.status,
            new_status=order.status,
            content=f'上门送回，时间：{return_visit_time}，结果：{return_visit_result or ""}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '上门送回记录成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'上门送回记录失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'上门送回记录失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/acceptance', methods=['POST'])
@jwt_required()
def workorder_acceptance(id):
    """客户验收"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()

        acceptance = data.get('customer_acceptance')
        sign = data.get('customer_acceptance_sign')

        if acceptance not in (1, 2):
            return jsonify({'code': 400, 'message': '验收结果必须是1(通过)或2(未通过)'}), 400

        order.customer_acceptance = acceptance
        order.customer_acceptance_time = datetime.now()
        if sign:
            order.customer_acceptance_sign = sign

        # 记录日志
        acceptance_text = '通过' if acceptance == 1 else '未通过'
        add_wo_log(
            wo_id=id,
            action='客户验收',
            old_status=order.status,
            new_status=order.status,
            content=f'客户验收{acceptance_text}',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': f'客户验收{acceptance_text}', 'data': {'customer_acceptance': acceptance}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'客户验收失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'客户验收失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:id>/cancel', methods=['POST'])
@jwt_required()
def cancel_workorder(id):
    """取消工单"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        if order.status in (9, 10):
            return jsonify({'code': 400, 'message': f'当前状态【{WO_STATUS_MAP.get(order.status, "未知")}】不允许取消'}), 400

        data = request.get_json() if request.is_json else {}
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        reason = data.get('reason', '')

        old_status = order.status
        order.status = 10
        order.status_name = '已取消'

        # 记录日志
        add_wo_log(
            wo_id=id,
            action='取消工单',
            old_status=old_status,
            new_status=10,
            content=f'取消工单，原因：{reason}' if reason else '取消工单',
            operator_id=user_id,
            operator_name=user_name
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '工单已取消', 'data': {'status': 10, 'status_text': '已取消'}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'取消工单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'取消工单失败: {str(e)}'}), 500

def add_workorder_log(wo_id, action, content, operator_id, operator_name):
    """添加工单操作日志"""
    log = WorkOrderLog(
        wo_id=wo_id,
        action=action,
        content=content,
        operator_id=operator_id,
        operator_name=operator_name
    )
    db.session.add(log)
    db.session.commit()

@app.route('/api/workorders/<int:id>/logs', methods=['GET'])
@jwt_required()
def get_workorder_logs(id):
    """获取工单操作日志"""
    try:
        order = WorkOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        logs = WorkOrderLog.query.filter_by(wo_id=id).order_by(WorkOrderLog.created_at.desc()).all()

        return jsonify({
            'code': 200,
            'data': [{
                'id': l.id,
                'action': l.action,
                'old_status': l.old_status,
                'old_status_text': WO_STATUS_MAP.get(l.old_status, '') if l.old_status is not None else '',
                'new_status': l.new_status,
                'new_status_text': WO_STATUS_MAP.get(l.new_status, '') if l.new_status is not None else '',
                'content': l.content,
                'operator_id': l.operator_id,
                'operator_name': l.operator_name,
                'created_at': l.created_at.strftime('%Y-%m-%d %H:%M:%S') if l.created_at else None
            } for l in logs]
        })
    except Exception as e:
        logger.error(f'获取操作日志失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取操作日志失败: {str(e)}'}), 500

@app.route('/api/workorders/auto-dispatch', methods=['GET'])
@jwt_required()
def get_auto_dispatch_recommendation():
    """获取自动派单推荐 - 根据工单类型匹配工程师"""
    try:
        wo_type = request.args.get('wo_type', '')
        if not wo_type:
            return jsonify({'code': 400, 'message': '请提供工单类型(wo_type)'}), 400

        recommendations = auto_dispatch_engineer(wo_type)

        return jsonify({
            'code': 200,
            'data': {
                'wo_type': wo_type,
                'wo_type_text': WO_TYPE_MAP.get(wo_type, wo_type),
                'recommendations': recommendations
            }
        })
    except Exception as e:
        logger.error(f'获取自动派单推荐失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取自动派单推荐失败: {str(e)}'}), 500

# 工单配件管理
@app.route('/api/workorders/<int:wo_id>/parts', methods=['GET'])
@jwt_required()
def get_workorder_parts(wo_id):
    """获取工单配件列表"""
    try:
        parts = WorkOrderPart.query.filter_by(wo_id=wo_id).all()
        return jsonify({
            'code': 200,
            'data': [{
                'id': p.id,
                'product_id': p.product_id,
                'product_name': p.product_name,
                'product_code': p.product_code,
                'specification': p.specification,
                'quantity': float(p.quantity) if p.quantity else 0,
                'used_quantity': float(p.used_quantity) if p.used_quantity else 0,
                'unit_price': float(p.unit_price) if p.unit_price else 0,
                'total_price': float(p.total_price) if p.total_price else 0,
                'cost_price': float(p.cost_price) if p.cost_price else 0,
                'is_own': p.is_own,
                'status': p.status,
                'remark': p.remark,
                'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None
            } for p in parts]
        })
    except Exception as e:
        logger.error(f'获取工单配件列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取工单配件列表失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:wo_id>/parts', methods=['POST'])
@jwt_required()
def add_workorder_part(wo_id):
    """添加工单配件"""
    try:
        order = WorkOrder.query.get(wo_id)
        if not order:
            return jsonify({'code': 404, 'message': '工单不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()

        part = WorkOrderPart(
            wo_id=wo_id,
            product_id=data.get('product_id'),
            product_name=data.get('product_name'),
            product_code=data.get('product_code'),
            quantity=data.get('quantity', 1),
            unit_price=data.get('unit_price', 0),
            total_price=float(data.get('quantity', 1)) * float(data.get('unit_price', 0)),
            cost_price=data.get('cost_price', 0),
            is_own=data.get('is_own', 1),
            status=0,
            remark=data.get('remark')
        )

        db.session.add(part)
        db.session.flush()

        # 记录日志
        add_wo_log(
            wo_id=wo_id,
            action='添加配件',
            old_status=order.status,
            new_status=order.status,
            content=f'添加配件：{data.get("product_name", "")} x {data.get("quantity", 1)}',
            operator_id=user_id,
            operator_name=get_current_user_name()
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '添加成功', 'data': {'id': part.id}})
    except Exception as e:
        db.session.rollback()
        logger.error(f'添加工单配件失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'添加工单配件失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:wo_id>/parts/<int:part_id>', methods=['PUT'])
@jwt_required()
def update_workorder_part(wo_id, part_id):
    """更新工单配件"""
    try:
        part = WorkOrderPart.query.get(part_id)
        if not part or part.wo_id != wo_id:
            return jsonify({'code': 404, 'message': '配件不存在'}), 404

        data = request.get_json()
        user_id = get_jwt_identity()

        for field in ['quantity', 'unit_price', 'is_own', 'status', 'remark', 'product_name', 'product_code', 'cost_price']:
            if field in data:
                setattr(part, field, data[field])

        if 'quantity' in data or 'unit_price' in data:
            part.total_price = float(part.quantity or 0) * float(part.unit_price or 0)

        # 记录日志
        add_wo_log(
            wo_id=wo_id,
            action='更新配件',
            old_status=None,
            new_status=None,
            content=f'更新配件ID:{part_id}',
            operator_id=user_id,
            operator_name=get_current_user_name()
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新工单配件失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'更新工单配件失败: {str(e)}'}), 500

@app.route('/api/workorders/<int:wo_id>/parts/<int:part_id>', methods=['DELETE'])
@jwt_required()
def delete_workorder_part(wo_id, part_id):
    """删除工单配件"""
    try:
        part = WorkOrderPart.query.get(part_id)
        if not part or part.wo_id != wo_id:
            return jsonify({'code': 404, 'message': '配件不存在'}), 404

        part_name = part.product_name or ''
        db.session.delete(part)

        # 记录日志
        user_id = get_jwt_identity()
        add_wo_log(
            wo_id=wo_id,
            action='删除配件',
            old_status=None,
            new_status=None,
            content=f'删除配件：{part_name}',
            operator_id=user_id,
            operator_name=get_current_user_name()
        )

        db.session.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除工单配件失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'删除工单配件失败: {str(e)}'}), 500

# ============================================
# API路由 - 接件管理
# ============================================

@app.route('/api/receiveorders', methods=['GET'])
@jwt_required()
def get_receiveorders():
    """获取接件单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    receive_type = request.args.get('receive_type', type=int)
    
    query = ReceiveOrder.query
    
    if keyword:
        query = query.filter(
            db.or_(
                ReceiveOrder.receive_no.contains(keyword),
                ReceiveOrder.customer_name.contains(keyword),
                ReceiveOrder.customer_phone.contains(keyword)
            )
        )
    
    if status is not None:
        query = query.filter_by(status=status)
    
    if receive_type is not None:
        query = query.filter_by(receive_type=receive_type)
    
    total = query.count()
    orders = query.order_by(ReceiveOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    # 关联设备明细信息
    result_list = []
    for order in orders.items:
        order_dict = to_dict(order)
        # 获取所有设备明细
        devices = ReceiveOrderDevice.query.filter_by(receive_order_id=order.id).all()
        order_dict['devices'] = [to_dict(d) for d in devices]
        # 获取第一个设备的品牌型号（兼容扁平字段）
        if devices:
            order_dict['device_type'] = devices[0].device_type
            order_dict['device_brand'] = devices[0].device_brand
            order_dict['device_model'] = devices[0].device_model
            order_dict['device_sn'] = devices[0].device_sn
            order_dict['fault_desc'] = devices[0].fault_desc
        else:
            order_dict['device_type'] = ''
            order_dict['device_brand'] = ''
            order_dict['device_model'] = ''
            order_dict['device_sn'] = ''
            order_dict['fault_desc'] = ''
        # 附加接待员工和维修工程师名称
        order_dict['reception_user_name'] = order.receiver_name or ''
        order_dict['engineer_user_name'] = order.engineer_name or ''
        # 附加状态文本
        order_dict['status_text'] = RO_STATUS_MAP.get(order.status, '未知')
        result_list.append(order_dict)
    
    return jsonify({
        'code': 200,
        'data': {
            'list': result_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/receiveorders/<int:id>', methods=['GET'])
@jwt_required()
def get_receiveorder(id):
    """获取接件单详情"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404
    
    # 获取设备明细
    devices = ReceiveOrderDevice.query.filter_by(receive_order_id=id).all()
    # 获取配件/领料明细
    parts = ReceiveOrderPart.query.filter_by(receive_order_id=id).all()
    
    result = to_dict(order)
    result['status_text'] = RO_STATUS_MAP.get(order.status, '未知')
    result['devices'] = [to_dict(d) for d in devices]
    result['parts'] = [to_dict(p) for p in parts]
    
    # 获取报价明细
    quote_order = QuoteOrder.query.filter_by(related_type='receive_order', related_id=id).first()
    if quote_order:
        quote_items = QuoteOrderItem.query.filter_by(quote_id=quote_order.id).all()
        result['quote_items'] = [to_dict(item) for item in quote_items]
    else:
        result['quote_items'] = []
    
    return jsonify({'code': 200, 'data': result})

@app.route('/api/receiveorders', methods=['POST'])
@jwt_required()
def create_receiveorder():
    """创建接件单"""
    data = request.get_json()
    
    # 调试日志
    print(f"[DEBUG] 创建接件单 - 接收数据: {data}")
    print(f"[DEBUG] devices: {data.get('devices', [])}")
    for i, d in enumerate(data.get('devices', [])):
        print(f"[DEBUG] 设备 {i}: parts = {d.get('parts', [])}")
    
    user_id = get_jwt_identity()
    user_name = get_current_user_name()
    
    last_order = ReceiveOrder.query.order_by(ReceiveOrder.id.desc()).first()
    receive_no = generate_code('RO', last_order.id if last_order else 0)
    
    # 查询接待员工和维修工程师名称
    reception_user_id = data.get('reception_user_id')
    engineer_user_id = data.get('engineer_user_id')
    
    reception_user_name = user_name
    engineer_user_name = ''
    
    if reception_user_id:
        reception_user = SysUser.query.get(reception_user_id)
        if reception_user:
            reception_user_name = reception_user.real_name or reception_user.username
    
    if engineer_user_id:
        engineer_user = SysUser.query.get(engineer_user_id)
        if engineer_user:
            engineer_user_name = engineer_user.real_name or engineer_user.username
    
    order = ReceiveOrder(
        receive_no=receive_no,
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        customer_phone=data.get('customer_phone'),
        receive_type=data.get('receive_type', 1),
        external_shop_id=data.get('external_shop_id'),
        external_shop_name=data.get('external_shop_name'),
        remark=data.get('remark'),
        created_by=user_id,
        receiver_id=reception_user_id or user_id,
        receiver_name=reception_user_name,
        engineer_id=engineer_user_id,
        engineer_name=engineer_user_name
    )
    
    db.session.add(order)
    db.session.flush()
    
    # 添加设备明细（支持多设备）
    devices = data.get('devices', [])
    for device_data in devices:
        # 字段映射：前端用 brand/model/serial_number，后端用 device_brand/device_model/device_sn
        device_brand = device_data.get('device_brand') or device_data.get('brand') or ''
        device_model = device_data.get('device_model') or device_data.get('model') or ''
        device_sn = device_data.get('device_sn') or device_data.get('serial_number') or ''
        device = ReceiveOrderDevice(
            receive_order_id=order.id,
            device_type=device_data.get('device_type'),
            device_brand=device_brand,
            device_model=device_model,
            device_sn=device_sn,
            device_imei=device_data.get('device_imei'),
            fault_desc=device_data.get('fault_desc'),
            appearance_desc=device_data.get('appearance_desc'),
            accessories=device_data.get('accessories'),
            device_name=device_data.get('device_name'),
            cpu=device_data.get('cpu'),
            memory=device_data.get('memory'),
            disk=device_data.get('disk'),
            os=device_data.get('os'),
            os_version=device_data.get('os_version'),
            toner_model=device_data.get('toner_model'),
            drum_model=device_data.get('drum_model'),
            ip_address=device_data.get('ip_address'),
            port=device_data.get('port'),
            camera_count=device_data.get('camera_count'),
            monitor_brand=device_data.get('monitor_brand'),
            firmware_version=device_data.get('firmware_version'),
            port_count=device_data.get('port_count')
        )
        db.session.add(device)
        
        # 保存设备关联的配件（嵌套在设备中的parts）
        device_parts = device_data.get('parts', [])
        print(f"[DEBUG] 接件单 {order.id} - 设备类型: {device_data.get('device_type')}, 配件数量: {len(device_parts)}")
        for part in device_parts:
            part_status = part.get('status', 0)
            print(f"[DEBUG] 配件: name={part.get('name')}, quantity={part.get('quantity')}, status={part_status}")
            p = ReceiveOrderPart(
                receive_order_id=order.id,
                product_name=part.get('name') or part.get('product_name', ''),
                specification=part.get('specification', ''),
                unit_name=part.get('unit', '个'),
                quantity=int(part.get('quantity', 1)) if part.get('quantity') else 1,
                status=int(part_status) if str(part_status).isdigit() else 0,
                remark=str(part.get('remark', '')),
                source=1
            )
            db.session.add(p)
    
    db.session.commit()
    
    # 验证配件是否保存成功
    saved_parts = ReceiveOrderPart.query.filter_by(receive_order_id=order.id).all()
    print(f"[DEBUG] 接件单 {order.id} 保存后，数据库中配件数量: {len(saved_parts)}")
    log = ReceiveOrderLog(
        receive_order_id=order.id,
        action='创建接件单',
        old_status=None,
        new_status=0,
        content=f'创建接件单 {receive_no}，客户：{order.customer_name}',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'receive_no': receive_no}})

# ============================================
# API路由 - 库存管理
# ============================================

@app.route('/api/inventory/stock', methods=['GET'])
@jwt_required()
def get_inventory_stock():
    """获取库存列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    warehouse_id = request.args.get('warehouse_id', type=int)
    product_id = request.args.get('product_id', type=int)
    
    query = InventoryStock.query
    
    # 按商品ID过滤（前端选择商品后查询库存）
    if product_id:
        query = query.filter_by(product_id=product_id)
    
    if keyword:
        query = query.filter(
            db.or_(
                InventoryStock.product_name.contains(keyword),
                InventoryStock.product_code.contains(keyword)
            )
        )
    
    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)
    
    total = query.count()
    stocks = query.order_by(InventoryStock.updated_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    stock_list = []
    for s in stocks.items:
        product = ProductInfo.query.get(s.product_id)
        purchase_price = float(product.purchase_price) if product and product.purchase_price else 0
        safe_stock = product.min_stock if product else 0
        qty = float(s.quantity) if s.quantity else 0

        if safe_stock and safe_stock > 0:
            if qty <= 0:
                stock_status = '缺货'
            elif qty < safe_stock:
                stock_status = '不足'
            elif qty > safe_stock * 3:
                stock_status = '过剩'
            else:
                stock_status = '正常'
        else:
            stock_status = '正常'

        stock_list.append({
            'id': s.id,
            'product_id': s.product_id,
            'product_code': s.product_code,
            'product_name': s.product_name,
            'warehouse_id': s.warehouse_id,
            'warehouse_name': s.warehouse_name,
            'quantity': qty,
            'frozen_quantity': float(s.frozen_quantity) if s.frozen_quantity else 0,
            'available_quantity': float(s.available_quantity) if s.available_quantity else 0,
            'cost_price': float(s.cost_price) if s.cost_price else 0,
            'batch_no': s.batch_no,
            'serial_no': s.serial_no,
            'purchase_price': purchase_price,
            'safe_stock': safe_stock,
            'stock_status': stock_status
        })

    return jsonify({
        'code': 200,
        'data': {
            'list': stock_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/inventory/in', methods=['GET'])
@jwt_required()
def get_inventory_in_list():
    """获取入库单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    in_type = request.args.get('in_type', type=int)
    status = request.args.get('status', type=int)
    
    query = InventoryIn.query
    
    if keyword:
        query = query.filter(InventoryIn.in_no.contains(keyword))
    
    if in_type is not None:
        query = query.filter_by(in_type=in_type)
    
    if status is not None:
        query = query.filter_by(status=status)
    
    total = query.count()
    orders = query.order_by(InventoryIn.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/inventory/in/<int:id>', methods=['GET'])
@jwt_required()
def get_inventory_in(id):
    """获取入库单详情"""
    order = InventoryIn.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '入库单不存在'}), 404
    
    items = InventoryInItem.query.filter_by(in_id=id).all()
    
    result = to_dict(order)
    result['items'] = [to_dict(i) for i in items]
    
    return jsonify({'code': 200, 'data': result})

@app.route('/api/inventory/in', methods=['POST'])
@jwt_required()
def create_inventory_in():
    """创建入库单"""
    data = request.get_json()
    user_id = get_jwt_identity()
    
    last_order = InventoryIn.query.order_by(InventoryIn.id.desc()).first()
    in_no = generate_code('IN', last_order.id if last_order else 0)
    
    order = InventoryIn(
        in_no=in_no,
        in_type=data.get('in_type', 1),
        supplier_id=data.get('supplier_id'),
        supplier_name=data.get('supplier_name'),
        warehouse_id=data.get('warehouse_id', 1),
        warehouse_name=data.get('warehouse_name', '主仓库'),
        remark=data.get('remark'),
        related_order_id=data.get('related_order_id'),
        related_order_no=data.get('related_order_no'),
        created_by=user_id
    )
    
    db.session.add(order)
    db.session.flush()
    
    # 添加明细
    items = data.get('items', [])
    total_quantity = 0
    total_amount = 0
    
    for item_data in items:
        item = InventoryInItem(
            in_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit_name=item_data.get('unit_name'),
            quantity=item_data.get('quantity', 0),
            unit_price=item_data.get('unit_price', 0),
            total_price=float(item_data.get('quantity', 0)) * float(item_data.get('unit_price', 0)),
            cost_price=item_data.get('cost_price', 0),
            batch_no=item_data.get('batch_no'),
            serial_no=item_data.get('serial_no'),
            remark=item_data.get('remark')
        )
        db.session.add(item)
        total_quantity += float(item_data.get('quantity', 0))
        total_amount += float(item.quantity) * float(item.unit_price or 0)
    
    order.total_quantity = total_quantity
    order.total_amount = total_amount
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'in_no': in_no}})

@app.route('/api/inventory/in/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_inventory_in(id):
    """审核入库单"""
    order = InventoryIn.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '入库单不存在'}), 404
    
    if order.status != 0:
        return jsonify({'code': 400, 'message': '入库单状态不正确'}), 400
    
    data = request.get_json()
    user_id = get_jwt_identity()
    
    # 更新入库单状态
    order.status = 2
    order.auditor_id = user_id
    order.auditor_name = data.get('auditor_name', '')
    order.audit_time = datetime.now()
    
    # 更新库存
    items = InventoryInItem.query.filter_by(in_id=id).all()
    user_name = get_current_user_name()

    # 入库类型映射
    in_type_map = {1: '采购入库', 2: '退货入库', 3: '调拨入库', 4: '组装入库', 5: '其他入库'}
    order_type_text = in_type_map.get(order.in_type, '入库')

    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.warehouse_id
        ).first()
        
        before_qty = 0
        if stock:
            before_qty = float(stock.quantity or 0)
            stock.quantity = before_qty + float(item.quantity or 0)
            stock.available_quantity = float(stock.available_quantity or 0) + float(item.quantity or 0)
            if item.cost_price:
                stock.cost_price = item.cost_price
        else:
            stock = InventoryStock(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=order.warehouse_id,
                warehouse_name=order.warehouse_name,
                quantity=item.quantity,
                frozen_quantity=0,
                available_quantity=item.quantity,
                cost_price=item.cost_price,
                batch_no=item.batch_no,
                serial_no=item.serial_no
            )
            db.session.add(stock)
        
        after_qty = float(stock.quantity or 0)
        cost = float(item.cost_price or stock.cost_price or 0)
        qty = float(item.quantity or 0)

        # 写入 InventoryLog
        log = InventoryLog(
            product_id=item.product_id,
            product_code=item.product_code,
            product_name=item.product_name,
            warehouse_id=order.warehouse_id,
            warehouse_name=order.warehouse_name,
            change_type='in',
            order_type=order_type_text,
            order_id=order.id,
            order_no=order.in_no,
            quantity_change=qty,
            before_quantity=before_qty,
            after_quantity=after_qty,
            cost_price=cost,
            amount=round(qty * cost, 2),
            related_party=order.supplier_name or '',
            operator_id=user_id,
            operator_name=user_name,
            remark=order.remark or ''
        )
        db.session.add(log)
        
        # 更新商品库存
        product = ProductInfo.query.get(item.product_id)
        if product:
            product.current_stock = (product.current_stock or 0) + float(item.quantity or 0)
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，已入库'})

@app.route('/api/inventory/out', methods=['GET'])
@jwt_required()
def get_inventory_out_list():
    """获取出库单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    out_type = request.args.get('out_type', type=int)
    status = request.args.get('status', type=int)
    
    query = InventoryOut.query
    
    if keyword:
        query = query.filter(InventoryOut.out_no.contains(keyword))
    
    if out_type is not None:
        query = query.filter_by(out_type=out_type)
    
    if status is not None:
        query = query.filter_by(status=status)
    
    total = query.count()
    orders = query.order_by(InventoryOut.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/inventory/out', methods=['POST'])
@jwt_required()
def create_inventory_out():
    """创建出库单"""
    data = request.get_json()
    user_id = get_jwt_identity()
    
    last_order = InventoryOut.query.order_by(InventoryOut.id.desc()).first()
    out_no = generate_code('OUT', last_order.id if last_order else 0)
    
    order = InventoryOut(
        out_no=out_no,
        out_type=data.get('out_type', 1),
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        warehouse_id=data.get('warehouse_id', 1),
        warehouse_name=data.get('warehouse_name', '主仓库'),
        remark=data.get('remark'),
        related_order_id=data.get('related_order_id'),
        related_order_no=data.get('related_order_no'),
        created_by=user_id
    )
    
    db.session.add(order)
    db.session.flush()
    
    items = data.get('items', [])
    total_quantity = 0
    total_amount = 0
    
    for item_data in items:
        item = InventoryOutItem(
            out_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit_name=item_data.get('unit_name'),
            quantity=item_data.get('quantity', 0),
            unit_price=item_data.get('unit_price', 0),
            total_price=float(item_data.get('quantity', 0)) * float(item_data.get('unit_price', 0)),
            cost_price=item_data.get('cost_price', 0),
            batch_no=item_data.get('batch_no'),
            serial_no=item_data.get('serial_no'),
            remark=item_data.get('remark')
        )
        db.session.add(item)
        total_quantity += float(item_data.get('quantity', 0))
        total_amount += float(item.quantity) * float(item.unit_price or 0)
    
    order.total_quantity = total_quantity
    order.total_amount = total_amount
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'out_no': out_no}})

@app.route('/api/inventory/out/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_inventory_out(id):
    """审核出库单"""
    order = InventoryOut.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '出库单不存在'}), 404
    
    if order.status != 0:
        return jsonify({'code': 400, 'message': '出库单状态不正确'}), 400
    
    data = request.get_json()
    user_id = get_jwt_identity()
    
    # 检查库存是否充足
    items = InventoryOutItem.query.filter_by(out_id=id).all()
    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.warehouse_id
        ).first()
        
        if not stock or float(stock.available_quantity or 0) < float(item.quantity or 0):
            return jsonify({'code': 400, 'message': f'商品【{item.product_name}】库存不足'}), 400
    
    # 更新出库单状态
    order.status = 2
    order.auditor_id = user_id
    order.auditor_name = data.get('auditor_name', '')
    order.audit_time = datetime.now()
    
    # 出库类型映射
    out_type_map = {1: '销售出库', 2: '维修领料', 3: '调拨出库', 4: '拆卸出库', 5: '其他出库'}
    order_type_text = out_type_map.get(order.out_type, '出库')
    user_name = get_current_user_name()
    
    # 更新库存
    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.warehouse_id
        ).first()
        
        before_qty = 0
        if stock:
            before_qty = float(stock.quantity or 0)
            stock.quantity = before_qty - float(item.quantity or 0)
            stock.available_quantity = float(stock.available_quantity or 0) - float(item.quantity or 0)
        
        after_qty = float(stock.quantity or 0) if stock else 0
        cost = float(item.cost_price or (stock.cost_price if stock else 0) or 0)
        qty = float(item.quantity or 0)

        # 写入 InventoryLog
        log = InventoryLog(
            product_id=item.product_id,
            product_code=item.product_code,
            product_name=item.product_name,
            warehouse_id=order.warehouse_id,
            warehouse_name=order.warehouse_name,
            change_type='out',
            order_type=order_type_text,
            order_id=order.id,
            order_no=order.out_no,
            quantity_change=-qty,
            before_quantity=before_qty,
            after_quantity=after_qty,
            cost_price=cost,
            amount=round(qty * cost, 2),
            related_party=order.customer_name or '',
            operator_id=user_id,
            operator_name=user_name,
            remark=order.remark or ''
        )
        db.session.add(log)
        
        # 更新商品库存
        product = ProductInfo.query.get(item.product_id)
        if product:
            product.current_stock = (product.current_stock or 0) - float(item.quantity or 0)
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，已出库'})

# 库存盘点
@app.route('/api/inventory/check', methods=['GET'])
@jwt_required()
def get_inventory_check_list():
    """获取盘点单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    
    query = InventoryCheck.query
    
    if status is not None and status != '':
        query = query.filter_by(status=status)
    
    if keyword:
        query = query.filter(
            db.or_(
                InventoryCheck.check_no.like(f'%{keyword}%'),
                InventoryCheck.warehouse_name.like(f'%{keyword}%'),
                InventoryCheck.shelf_name.like(f'%{keyword}%')
            )
        )
    
    total = query.count()
    orders = query.order_by(InventoryCheck.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    result_list = []
    for o in orders.items:
        item = to_dict(o)
        # 统计盘盈/盘亏数
        check_items = InventoryCheckItem.query.filter_by(check_id=o.id).all()
        profit_count = len([i for i in check_items if i.diff_quantity and float(i.diff_quantity) > 0])
        loss_count = len([i for i in check_items if i.diff_quantity and float(i.diff_quantity) < 0])
        item['total_items'] = len(check_items)
        item['profit_items'] = profit_count
        item['loss_items'] = loss_count
        result_list.append(item)
    
    return jsonify({
        'code': 200,
        'data': {
            'list': result_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/inventory/check', methods=['POST'])
@jwt_required()
def create_inventory_check():
    """创建盘点单（支持按仓库/货架筛选）"""
    data = request.get_json()
    user_id = get_jwt_identity()
    
    last_order = InventoryCheck.query.order_by(InventoryCheck.id.desc()).first()
    check_no = generate_code('CHK', last_order.id if last_order else 0)
    
    warehouse_id = data.get('warehouse_id', 1)
    warehouse_name = data.get('warehouse_name', '主仓库')
    shelf_id = data.get('shelf_id')
    shelf_name = data.get('shelf_name', '')
    
    order = InventoryCheck(
        check_no=check_no,
        warehouse_id=warehouse_id,
        warehouse_name=warehouse_name,
        shelf_id=shelf_id,
        shelf_name=shelf_name,
        check_date=datetime.now().date(),
        remark=data.get('remark'),
        created_by=user_id
    )
    
    db.session.add(order)
    db.session.flush()
    
    # 获取当前库存并添加到盘点明细
    stocks_query = InventoryStock.query.filter_by(warehouse_id=warehouse_id)
    if shelf_id:
        stocks_query = stocks_query.filter_by(shelf_id=shelf_id)
    stocks = stocks_query.all()
    
    for stock in stocks:
        item = InventoryCheckItem(
            check_id=order.id,
            product_id=stock.product_id,
            product_code=stock.product_code,
            product_name=stock.product_name,
            system_quantity=stock.quantity,
            actual_quantity=stock.quantity,
            diff_quantity=0,
            cost_price=stock.cost_price,
            diff_amount=0
        )
        db.session.add(item)
    
    order.total_quantity = len(stocks)
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'check_no': check_no}})

@app.route('/api/inventory/check/<int:id>/diff-report', methods=['GET'])
@jwt_required()
def get_check_diff_report(id):
    """获取盘点差异报表"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404
    
    items = InventoryCheckItem.query.filter_by(check_id=id).all()
    
    profit_items = []  # 盘盈
    loss_items = []    # 盘亏
    normal_items = []  # 正常
    
    total_profit_qty = 0
    total_profit_amount = 0
    total_loss_qty = 0
    total_loss_amount = 0
    
    for i in items:
        diff = float(i.diff_quantity or 0)
        diff_amt = float(i.diff_amount or 0)
        item_data = {
            'product_code': i.product_code,
            'product_name': i.product_name,
            'specification': i.specification,
            'unit_name': i.unit_name,
            'system_quantity': float(i.system_quantity or 0),
            'actual_quantity': float(i.actual_quantity or 0),
            'diff_quantity': diff,
            'cost_price': float(i.cost_price or 0),
            'diff_amount': diff_amt,
            'remark': i.remark
        }
        
        if diff > 0:
            profit_items.append(item_data)
            total_profit_qty += diff
            total_profit_amount += diff_amt
        elif diff < 0:
            loss_items.append(item_data)
            total_loss_qty += abs(diff)
            total_loss_amount += abs(diff_amt)
        else:
            normal_items.append(item_data)
    
    return jsonify({
        'code': 200,
        'data': {
            'order_info': {
                'check_no': order.check_no,
                'warehouse_name': order.warehouse_name,
                'shelf_name': order.shelf_name,
                'check_date': str(order.check_date) if order.check_date else '',
                'status': order.status
            },
            'summary': {
                'total_items': len(items),
                'profit_count': len(profit_items),
                'loss_count': len(loss_items),
                'normal_count': len(normal_items),
                'total_profit_qty': total_profit_qty,
                'total_profit_amount': total_profit_amount,
                'total_loss_qty': total_loss_qty,
                'total_loss_amount': total_loss_amount
            },
            'profit_items': profit_items,
            'loss_items': loss_items,
            'normal_items': normal_items
        }
    })

@app.route('/api/inventory/check/diff-report/export', methods=['GET'])
@jwt_required()
def export_check_diff_report():
    """导出盘点差异报表为Excel"""
    check_id = request.args.get('check_id', type=int)
    if not check_id:
        return jsonify({'code': 400, 'message': '缺少check_id参数'}), 400
    
    order = InventoryCheck.query.get(check_id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404
    
    items = InventoryCheckItem.query.filter_by(check_id=check_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '盘点差异报表'
    
    # 标题行
    ws.merge_cells('A1:I1')
    ws['A1'] = f'盘点差异报表 - {order.check_no}'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    
    ws.merge_cells('A2:I2')
    ws['A2'] = f'仓库: {order.warehouse_name}  货架: {order.shelf_name or "全部"}  日期: {order.check_date}'
    
    # 表头
    headers = ['商品编码', '商品名称', '规格', '单位', '账面数量', '实盘数量', '差异数', '成本价', '差异金额', '类型']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
        ws.cell(row=4, column=col).font = openpyxl.styles.Font(bold=True)
    
    # 数据行
    for row_idx, item in enumerate(items, 5):
        diff = float(item.diff_quantity or 0)
        diff_type = '盘盈' if diff > 0 else ('盘亏' if diff < 0 else '正常')
        ws.cell(row=row_idx, column=1, value=item.product_code)
        ws.cell(row=row_idx, column=2, value=item.product_name)
        ws.cell(row=row_idx, column=3, value=item.specification or '')
        ws.cell(row=row_idx, column=4, value=item.unit_name or '')
        ws.cell(row=row_idx, column=5, value=float(item.system_quantity or 0))
        ws.cell(row=row_idx, column=6, value=float(item.actual_quantity or 0))
        ws.cell(row=row_idx, column=7, value=diff)
        ws.cell(row=row_idx, column=8, value=float(item.cost_price or 0))
        ws.cell(row=row_idx, column=9, value=float(item.diff_amount or 0))
        ws.cell(row=row_idx, column=10, value=diff_type)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'盘点差异报表_{order.check_no}.xlsx'
    )

# ============================================
# API路由 - 财务管理
# ============================================

@app.route('/api/finance/accounts', methods=['GET'])
@jwt_required()
def get_finance_accounts():
    """获取财务账户列表"""
    accounts = FinanceAccount.query.filter_by(status=1).all()
    return jsonify({
        'code': 200,
        'data': [{
            'id': a.id,
            'account_name': a.account_name,
            'account_type': a.account_type,
            'account_no': a.account_no,
            'balance': float(a.balance) if a.balance else 0.00,
            'remark': a.remark
        } for a in accounts]
    })

@app.route('/api/finance/records', methods=['GET'])
@jwt_required()
def get_finance_records():
    """获取财务流水"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    account_id = request.args.get('account_id', type=int)
    record_type = request.args.get('record_type', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    keyword = request.args.get('keyword', '')
    
    query = FinanceRecord.query
    
    if account_id:
        query = query.filter_by(account_id=account_id)
    
    if record_type:
        query = query.filter_by(record_type=record_type)
    
    if start_date:
        query = query.filter(FinanceRecord.created_at >= start_date)
    
    if end_date:
        query = query.filter(FinanceRecord.created_at <= end_date + ' 23:59:59')
    
    if keyword:
        keyword_like = f'%{keyword}%'
        query = query.filter(
            db.or_(
                FinanceRecord.account_name.like(keyword_like),
                FinanceRecord.remark.like(keyword_like),
                FinanceRecord.related_no.like(keyword_like)
            )
        )
    
    total = query.count()
    records = query.order_by(FinanceRecord.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )
    
    # 构建返回数据，补充关联信息
    result_list = []
    for r in records.items:
        item = to_dict(r)
        # 补充操作人名称
        if r.created_by:
            creator = SysUser.query.get(r.created_by)
            item['created_by_name'] = creator.real_name if creator else ''
        else:
            item['created_by_name'] = ''
        
        # 补充客户/供应商名称
        item['customer_name'] = ''
        item['supplier_name'] = ''
        if r.related_type == 'sale' and r.related_id:
            # 从应收记录获取客户名称
            receivable = FinanceReceivable.query.filter_by(related_id=r.related_id, related_type='sale').first()
            if receivable and receivable.customer_name:
                item['customer_name'] = receivable.customer_name
            elif r.related_id:
                # 从备注中提取或关联查询
                sale_order = db.session.query(SaleOrder).get(r.related_id) if hasattr(db.Model, 'SaleOrder') else None
        elif r.related_type == 'purchase' and r.related_id:
            payable = FinancePayable.query.filter_by(related_id=r.related_id, related_type='purchase').first()
            if payable and payable.supplier_name:
                item['supplier_name'] = payable.supplier_name
        elif r.related_type == 'work_order' and r.related_id:
            wo = WorkOrder.query.get(r.related_id)
            if wo and wo.customer_id:
                customer = BaseCustomer.query.get(wo.customer_id)
                if customer:
                    item['customer_name'] = customer.customer_name
        
        result_list.append(item)
    
    return jsonify({
        'code': 200,
        'data': {
            'list': result_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

# ============================================
# API路由 - 仪表盘统计
# ============================================

@app.route('/api/dashboard/stats', methods=['GET'])
@jwt_required()
def get_dashboard_stats():
    """获取仪表盘统计数据"""
    today = datetime.now().date()
    
    # 基础统计
    customer_count = BaseCustomer.query.filter_by(status=1).count()
    supplier_count = BaseSupplier.query.filter_by(status=1).count()
    product_count = ProductInfo.query.filter_by(status=1).count()
    
    # 工单统计
    wo_total = WorkOrder.query.count()
    wo_pending = WorkOrder.query.filter(WorkOrder.status.in_([0, 1, 2, 3, 4, 5, 6, 7])).count()
    wo_completed = WorkOrder.query.filter_by(status=9).count()
    wo_today = WorkOrder.query.filter(db.func.date(WorkOrder.created_at) == today).count()
    
    # 今日收入
    today_records = FinanceRecord.query.filter(
        FinanceRecord.record_type == 1,
        db.func.date(FinanceRecord.created_at) == today
    ).all()
    today_income = sum(float(r.amount or 0) for r in today_records)
    
    # 待处理工单按状态分组
    wo_by_status = db.session.query(
        WorkOrder.status,
        db.func.count(WorkOrder.id)
    ).filter(WorkOrder.status < 9).group_by(WorkOrder.status).all()
    
    status_stats = {s: c for s, c in wo_by_status}
    
    return jsonify({
        'code': 200,
        'data': {
            'customer_count': customer_count,
            'supplier_count': supplier_count,
            'product_count': product_count,
            'wo_total': wo_total,
            'wo_pending': wo_pending,
            'wo_completed': wo_completed,
            'wo_today': wo_today,
            'today_income': today_income,
            'wo_status_stats': {
                'pending': status_stats.get(0, 0),
                'accepted': status_stats.get(1, 0),
                'assigning': status_stats.get(2, 0),
                'processing': status_stats.get(3, 0),
                'waiting_parts': status_stats.get(4, 0),
                'waiting_audit': status_stats.get(5, 0),
                'parts_inbound': status_stats.get(6, 0),
                'repairing': status_stats.get(7, 0),
                'waiting_settle': status_stats.get(8, 0)
            }
        }
    })

@app.route('/api/dashboard/workorder-trend', methods=['GET'])
@jwt_required()
def get_workorder_trend():
    """获取工单趋势数据（最近7天）"""
    today = datetime.now().date()
    result = []
    
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        count = WorkOrder.query.filter(db.func.date(WorkOrder.created_at) == date).count()
        completed = WorkOrder.query.filter(
            db.func.date(WorkOrder.actual_time) == date,
            WorkOrder.status == 9
        ).count()
        result.append({
            'date': date.strftime('%m-%d'),
            'count': count,
            'completed': completed
        })
    
    return jsonify({'code': 200, 'data': result})

@app.route('/api/dashboard/income-trend', methods=['GET'])
@jwt_required()
def get_income_trend():
    """获取收入趋势数据（最近7天）"""
    today = datetime.now().date()
    result = []
    
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        records = FinanceRecord.query.filter(
            FinanceRecord.record_type == 1,
            db.func.date(FinanceRecord.created_at) == date
        ).all()
        income = sum(float(r.amount or 0) for r in records)
        result.append({
            'date': date.strftime('%m-%d'),
            'income': income
        })
    
    return jsonify({'code': 200, 'data': result})

# ============================================
# API路由 - 系统设置
# ============================================

# --- 办公室管理 ---
@app.route('/api/offices', methods=['GET'])
@jwt_required()
def get_offices():
    """获取办公室列表"""
    offices = Office.query.order_by(Office.sort_order.asc(), Office.id.asc()).all()
    return jsonify({
        'code': 200,
        'data': [o.to_dict() for o in offices]
    })

@app.route('/api/offices', methods=['POST'])
@jwt_required()
def create_office():
    """创建办公室"""
    data = request.get_json()
    if not data.get('name'):
        return jsonify({'code': 400, 'message': '办公室名称不能为空'}), 400

    # 检查编码是否重复
    if data.get('code'):
        existing = Office.query.filter_by(code=data['code']).first()
        if existing:
            return jsonify({'code': 400, 'message': '办公室编码已存在'}), 400

    office = Office(
        name=data['name'],
        code=data.get('code', ''),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    db.session.add(office)
    db.session.commit()
    return jsonify({'code': 200, 'message': '创建成功', 'data': office.to_dict()})

@app.route('/api/offices/<int:office_id>', methods=['PUT'])
@jwt_required()
def update_office(office_id):
    """更新办公室"""
    office = Office.query.get(office_id)
    if not office:
        return jsonify({'code': 404, 'message': '办公室不存在'}), 404

    data = request.get_json()
    if 'name' in data:
        office.name = data['name']
    if 'code' in data:
        # 检查编码是否重复
        existing = Office.query.filter(Office.code == data['code'], Office.id != office_id).first()
        if existing:
            return jsonify({'code': 400, 'message': '办公室编码已存在'}), 400
        office.code = data['code']
    if 'sort_order' in data:
        office.sort_order = data['sort_order']
    if 'status' in data:
        office.status = data['status']

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功', 'data': office.to_dict()})

@app.route('/api/offices/<int:office_id>', methods=['DELETE'])
@jwt_required()
def delete_office(office_id):
    """删除办公室"""
    office = Office.query.get(office_id)
    if not office:
        return jsonify({'code': 404, 'message': '办公室不存在'}), 404

    # 检查是否有关联资产
    from models import Asset
    has_assets = Asset.query.filter_by(office_id=office_id).first()
    if has_assets:
        return jsonify({'code': 400, 'message': '该办公室有关联资产，无法删除'}), 400

    db.session.delete(office)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

@app.route('/api/settings/wo-types', methods=['GET'])
@jwt_required()
def get_wo_types():
    """获取工单类型列表"""
    types = WoType.query.filter_by(status=1).order_by(WoType.sort_order).all()
    return jsonify({
        'code': 200,
        'data': [to_dict(t) for t in types]
    })

@app.route('/api/settings/projects', methods=['GET'])
@jwt_required()
def get_projects():
    """获取维修项目列表"""
    projects = Project.query.filter_by(status=1).all()
    return jsonify({
        'code': 200,
        'data': [to_dict(p) for p in projects]
    })

# 打印模板管理
@app.route('/api/settings/print-templates', methods=['GET'])
@jwt_required()
def get_print_templates():
    """获取打印模板列表"""
    template_type = request.args.get('template_type', '')
    query = PrintTemplate.query.filter_by(status=1)
    if template_type:
        query = query.filter_by(template_type=template_type)
    templates = query.order_by(PrintTemplate.is_default.desc(), PrintTemplate.id).all()
    return jsonify({'code': 200, 'data': [to_dict(t) for t in templates]})

@app.route('/api/settings/print-templates', methods=['POST'])
@jwt_required()
def create_print_template():
    """新增打印模板"""
    data = request.get_json()
    template = PrintTemplate(
        template_name=data.get('template_name'),
        template_type=data.get('template_type'),
        description=data.get('description', ''),
        header_content=data.get('header_content', ''),
        body_content=data.get('body_content', ''),
        footer_content=data.get('footer_content', ''),
        style_content=data.get('style_content', ''),
        paper_size=data.get('paper_size', 'A4'),
        paper_width=data.get('paper_width', 210),
        paper_height=data.get('paper_height', 297),
        is_default=data.get('is_default', 0),
        status=1
    )
    if template.is_default:
        PrintTemplate.query.filter_by(template_type=template.template_type).update({'is_default': 0})
    db.session.add(template)
    db.session.commit()
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': template.id}})

@app.route('/api/settings/print-templates/<int:id>', methods=['PUT'])
@jwt_required()
def update_print_template(id):
    """更新打印模板"""
    template = PrintTemplate.query.get(id)
    if not template:
        return jsonify({'code': 404, 'message': '模板不存在'}), 404
    data = request.get_json()
    for field in ['template_name', 'template_type', 'description', 'header_content', 'body_content', 'footer_content', 'style_content', 'paper_size', 'paper_width', 'paper_height']:
        if field in data:
            setattr(template, field, data[field])
    # 处理默认值：显式传入 is_default 时进行切换
    if 'is_default' in data:
        if data['is_default']:
            # 设为默认：先清除同类型所有模板的默认值，再设置当前模板
            PrintTemplate.query.filter_by(template_type=template.template_type).update({'is_default': 0})
            template.is_default = 1
        else:
            # 取消默认：仅清除当前模板的默认值
            template.is_default = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/settings/print-templates/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_print_template(id):
    """删除打印模板"""
    template = PrintTemplate.query.get(id)
    if not template:
        return jsonify({'code': 404, 'message': '模板不存在'}), 404
    db.session.delete(template)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

@app.route('/api/settings/print-templates/init-defaults', methods=['POST'])
@jwt_required()
def init_default_print_templates():
    """初始化默认打印模板"""
    # 先清空旧模板数据
    PrintTemplate.query.delete()
    db.session.flush()

    templates = [
        # === 工单管理 (work_order) ===
        {'template_name': '工单-标准A4', 'template_type': 'work_order', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>维修工单</h1><p>工单号：{{wo_no}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户姓名：{{customer_name}}</td><td>联系电话：{{customer_phone}}</td><td>设备品牌：{{device_brand}}</td><td>设备型号：{{device_model}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">故障描述</th></tr><tr><td colspan="4">{{fault_desc}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">维修项目</th></tr><tr><td colspan="4">{{repair_items}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">费用明细</th></tr><tr><td colspan="4">{{fee_detail}}</td></tr><tr><td colspan="2">合计金额：</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:40px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>工程师签字：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '工单-简洁A5', 'template_type': 'work_order', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<div style="text-align:center"><h2>维修工单</h2></div>',
         'body_content': '<p>工单号：{{wo_no}} | {{created_at}}</p><p>客户：{{customer_name}} | 电话：{{customer_phone}}</p><p>设备：{{device_brand}} {{device_model}}</p><hr><p>故障：{{fault_desc}}</p><p>维修：{{repair_items}}</p><hr><p style="font-weight:bold">合计：¥{{total_amount}}</p>',
         'footer_content': '<p>客户签字：______ 工程师：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '工单-详细带条码', 'template_type': 'work_order', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>维修工单（详细）</h1><p>工单号：{{wo_no}}</p><div style="margin:10px 0">[条码区域: {{wo_no}}]</div><p>日期：{{created_at}} | 状态：{{status_name}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户姓名：{{customer_name}}</td><td>联系电话：{{customer_phone}}</td><td>设备品牌：{{device_brand}}</td><td>设备型号：{{device_model}}</td></tr><tr><td>设备SN：{{device_sn}}</td><td>设备类型：{{device_type}}</td><td>工单类型：{{wo_type_name}}</td><td>优先级：{{priority_name}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">故障描述</th></tr><tr><td colspan="4">{{fault_desc}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">维修项目及配件</th></tr><tr><td colspan="4">{{repair_items}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">费用明细</th></tr><tr><td colspan="4">{{fee_detail}}</td></tr><tr><td colspan="2">合计金额：</td><td colspan="2" style="font-weight:bold;font-size:16px">¥{{total_amount}}</td></tr><tr><td colspan="4">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:40px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>工程师签字：______________</div><div>日期：______________</div></div><p style="text-align:center;color:#999;font-size:10px;margin-top:20px">此单一式两联，客户联和存根联</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 接件管理 (receive) ===
        {'template_name': '接件单-标准A4', 'template_type': 'receive', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>设备接件单</h1><p>接件单号：{{receive_no}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户姓名：{{customer_name}}</td><td>联系电话：{{customer_phone}}</td><td colspan="2">地址：{{address}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">设备信息</th></tr><tr><td>设备品牌：{{device_brand}}</td><td>设备型号：{{device_model}}</td><td>设备SN：{{device_sn}}</td><td>设备类型：{{device_type}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">故障描述</th></tr><tr><td colspan="4">{{fault_desc}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">外观检查</th></tr><tr><td colspan="4">{{appearance}}</td></tr><tr><td>附件：{{accessories}}</td><td colspan="3">预计费用：¥{{estimate_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:40px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>接待人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '接件单-简洁版', 'template_type': 'receive', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">接件单</h2>',
         'body_content': '<p>单号：{{receive_no}} | {{created_at}}</p><p>客户：{{customer_name}} | {{customer_phone}}</p><p>设备：{{device_brand}} {{device_model}}</p><p>故障：{{fault_desc}}</p><p>预计费用：¥{{estimate_amount}}</p>',
         'footer_content': '<p>客户签字：______ 接待人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '接件单-带照片', 'template_type': 'receive', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>设备接件单（含照片）</h1><p>接件单号：{{receive_no}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="2" style="background:#f0f0f0">客户信息</th><th colspan="2" style="background:#f0f0f0">设备信息</th></tr><tr><td>客户：{{customer_name}}</td><td>电话：{{customer_phone}}</td><td>品牌：{{device_brand}}</td><td>型号：{{device_model}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">故障描述</th></tr><tr><td colspan="4">{{fault_desc}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">设备照片</th></tr><tr><td colspan="4" style="text-align:center;height:150px">[设备照片区域]</td></tr><tr><th colspan="4" style="background:#f0f0f0">外观及附件</th></tr><tr><td colspan="4">{{appearance}} | 附件：{{accessories}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px"><p>客户签字：______________ 接待人：______________</p></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 派单管理 (dispatch) ===
        {'template_name': '派工单-标准A4', 'template_type': 'dispatch', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>派工单</h1><p>工单号：{{wo_no}} | 派单时间：{{dispatch_time}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0" width="25%">客户信息</th><td>{{customer_name}} | {{customer_phone}}</td></tr><tr><th style="background:#f0f0f0">设备信息</th><td>{{device_brand}} {{device_model}} | SN:{{device_sn}}</td></tr><tr><th style="background:#f0f0f0">故障描述</th><td>{{fault_desc}}</td></tr><tr><th style="background:#f0f0f0">指派工程师</th><td>{{engineer_name}} | {{engineer_phone}}</td></tr><tr><th style="background:#f0f0f0">优先级</th><td>{{priority}}</td></tr><tr><th style="background:#f0f0f0">要求完成时间</th><td>{{deadline}}</td></tr><tr><th style="background:#f0f0f0">备注</th><td>{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:40px;display:flex;justify-content:space-between"><div>派单人：______________</div><div>工程师签字：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '派工单-简洁版', 'template_type': 'dispatch', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">派工单</h2>',
         'body_content': '<p>工单号：{{wo_no}}</p><p>客户：{{customer_name}} | {{customer_phone}}</p><p>设备：{{device_brand}} {{device_model}}</p><p>工程师：{{engineer_name}} | {{engineer_phone}}</p><p>故障：{{fault_desc}}</p>',
         'footer_content': '<p>派单人：______ 工程师：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '派工单-详细版', 'template_type': 'dispatch', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>派工单（详细）</h1><p>工单号：{{wo_no}} | 派单时间：{{dispatch_time}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">工单信息</th></tr><tr><td>工单号：{{wo_no}}</td><td>工单类型：{{wo_type}}</td><td>优先级：{{priority}}</td><td>截止时间：{{deadline}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>姓名：{{customer_name}}</td><td>电话：{{customer_phone}}</td><td>地址：{{address}}</td><td></td></tr><tr><th colspan="4" style="background:#f0f0f0">设备信息</th></tr><tr><td>品牌：{{device_brand}}</td><td>型号：{{device_model}}</td><td>SN：{{device_sn}}</td><td>类型：{{device_type}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">故障描述</th></tr><tr><td colspan="4">{{fault_desc}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">派工信息</th></tr><tr><td>工程师：{{engineer_name}}</td><td>电话：{{engineer_phone}}</td><td>派单人：{{dispatcher_name}}</td><td>派单方式：{{dispatch_type}}</td></tr><tr><th colspan="4" style="background:#f0f0f0">备注</th></tr><tr><td colspan="4">{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:40px;display:flex;justify-content:space-between"><div>派单人：______________</div><div>工程师签字：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 报价管理 (quote) ===
        {'template_name': '报价单-标准A4', 'template_type': 'quote', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>维修报价单</h1><p>报价单号：{{quote_no}} | 日期：{{created_at}}</p><p>有效期：{{valid_days}}天</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户姓名：{{customer_name}}</td><td>联系电话：{{customer_phone}}</td><td colspan="2">设备：{{device_brand}} {{device_model}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">项目</th><th style="background:#f0f0f0">金额</th><th style="background:#f0f0f0">备注</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px"><p>报价说明：{{remark}}</p><p>客户签字：______________ 日期：______________</p></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '报价单-简洁版', 'template_type': 'quote', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">报价单</h2>',
         'body_content': '<p>单号：{{quote_no}} | {{created_at}}</p><p>客户：{{customer_name}} | {{customer_phone}}</p><p>设备：{{device_brand}} {{device_model}}</p><hr><p>{{items}}</p><p style="font-weight:bold">合计：¥{{total_amount}}</p>',
         'footer_content': '<p>签字：______ 日期：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '报价单-详细带条款', 'template_type': 'quote', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>维修报价单（正式）</h1><p>报价单号：{{quote_no}} | 报价日期：{{created_at}} | 有效期：{{valid_days}}天</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户：{{customer_name}}</td><td>电话：{{customer_phone}}</td><td>设备：{{device_brand}} {{device_model}}</td><td>SN：{{device_sn}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">维修项目</th><th style="background:#f0f0f0">单价</th><th style="background:#f0f0f0">小计</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold;font-size:16px">¥{{total_amount}}</td></tr></table><h3>服务条款</h3><ol><li>维修后保修{{warranty_days}}天</li><li>报价有效期内价格不变</li><li>如需额外配件将另行通知</li></ol>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>报价人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 销售管理 (sale) ===
        {'template_name': '销售单-标准A4', 'template_type': 'sale', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>销售单</h1><p>销售单号：{{sale_no}} | 日期：{{sale_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户名称：{{customer_name}}</td><td>联系电话：{{customer_phone}}</td><td colspan="2">地址：{{address}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>经办人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '销售单-简洁版', 'template_type': 'sale', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">销售单</h2>',
         'body_content': '<p>单号：{{sale_no}} | {{sale_date}}</p><p>客户：{{customer_name}}</p><hr><p>{{items}}</p><p style="font-weight:bold">合计：¥{{total_amount}}</p>',
         'footer_content': '<p>签字：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '销售单-含税版', 'template_type': 'sale', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>销售单（含税）</h1><p>销售单号：{{sale_no}} | 日期：{{sale_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户：{{customer_name}}</td><td>电话：{{customer_phone}}</td><td>地址：{{address}}</td><td></td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品</th><th style="background:#f0f0f0">数量×单价</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right">小计</td><td colspan="2">¥{{subtotal}}</td></tr><tr><td colspan="2" style="text-align:right">税额</td><td colspan="2">¥{{tax_amount}}</td></tr><tr><td colspan="2" style="text-align:right;font-weight:bold">合计（含税）</td><td colspan="2" style="font-weight:bold;font-size:16px">¥{{total_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>经办人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 采购管理 (purchase) ===
        {'template_name': '采购单-标准A4', 'template_type': 'purchase', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>采购单</h1><p>采购单号：{{purchase_no}} | 日期：{{purchase_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">供应商信息</th></tr><tr><td>供应商：{{supplier_name}}</td><td>联系人：{{contact_name}}</td><td>电话：{{phone}}</td><td></td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>采购人：______________</div><div>审批人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '采购单-简洁版', 'template_type': 'purchase', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">采购单</h2>',
         'body_content': '<p>单号：{{purchase_no}} | {{purchase_date}}</p><p>供应商：{{supplier_name}}</p><hr><p>{{items}}</p><p style="font-weight:bold">合计：¥{{total_amount}}</p>',
         'footer_content': '<p>采购人：______ 审批人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '采购单-详细版', 'template_type': 'purchase', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>采购单（详细）</h1><p>采购单号：{{purchase_no}} | 日期：{{purchase_date}} | 交货日期：{{delivery_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">供应商信息</th></tr><tr><td>供应商：{{supplier_name}}</td><td>联系人：{{contact_name}}</td><td>电话：{{phone}}</td><td>地址：{{address}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品</th><th style="background:#f0f0f0">规格</th><th style="background:#f0f0f0">数量×单价</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>采购人：______________</div><div>审批人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 采购退货 (return_purchase) ===
        {'template_name': '采购退货单-标准A4', 'template_type': 'return_purchase', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>采购退货单</h1><p>退货单号：{{return_no}} | 日期：{{return_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">供应商信息</th></tr><tr><td>供应商：{{supplier_name}}</td><td>联系人：{{contact_name}}</td><td>电话：{{phone}}</td><td></td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">退货数量</th><th style="background:#f0f0f0">退货金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">退货原因：{{reason}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>经办人：______________</div><div>审批人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '采购退货单-简洁版', 'template_type': 'return_purchase', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">采购退货单</h2>',
         'body_content': '<p>单号：{{return_no}} | {{return_date}}</p><p>供应商：{{supplier_name}}</p><hr><p>{{items}}</p><p>合计：¥{{total_amount}}</p><p>原因：{{reason}}</p>',
         'footer_content': '<p>经办人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '采购退货单-详细版', 'template_type': 'return_purchase', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>采购退货单（详细）</h1><p>退货单号：{{return_no}} | 日期：{{return_date}} | 原采购单号：{{original_no}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">供应商信息</th></tr><tr><td>供应商：{{supplier_name}}</td><td>联系人：{{contact_name}}</td><td>电话：{{phone}}</td><td>地址：{{address}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品</th><th style="background:#f0f0f0">规格</th><th style="background:#f0f0f0">退货数量×单价</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">退货原因：{{reason}}</td></tr><tr><td colspan="4">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>经办人：______________</div><div>审批人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 销售退货 (return_sale) ===
        {'template_name': '销售退货单-标准A4', 'template_type': 'return_sale', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>销售退货单</h1><p>退货单号：{{return_no}} | 日期：{{return_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户：{{customer_name}}</td><td>电话：{{customer_phone}}</td><td colspan="2">地址：{{address}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">退货数量</th><th style="background:#f0f0f0">退货金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">退货原因：{{reason}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>经办人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '销售退货单-简洁版', 'template_type': 'return_sale', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">销售退货单</h2>',
         'body_content': '<p>单号：{{return_no}} | {{return_date}}</p><p>客户：{{customer_name}}</p><hr><p>{{items}}</p><p>合计：¥{{total_amount}}</p><p>原因：{{reason}}</p>',
         'footer_content': '<p>签字：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '销售退货单-详细版', 'template_type': 'return_sale', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>销售退货单（详细）</h1><p>退货单号：{{return_no}} | 日期：{{return_date}} | 原销售单号：{{original_no}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">客户信息</th></tr><tr><td>客户：{{customer_name}}</td><td>电话：{{customer_phone}}</td><td>地址：{{address}}</td><td></td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品</th><th style="background:#f0f0f0">规格</th><th style="background:#f0f0f0">退货数量×单价</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">退货原因：{{reason}}</td></tr><tr><td colspan="4">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>客户签字：______________</div><div>经办人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 入库管理 (inventory_in) ===
        {'template_name': '入库单-标准A4', 'template_type': 'inventory_in', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>入库单</h1><p>入库单号：{{in_no}} | 日期：{{in_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">入库信息</th></tr><tr><td>仓库：{{warehouse_name}}</td><td>入库类型：{{in_type}}</td><td>经办人：{{operator}}</td><td>关联单号：{{related_no}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>审核人：______________</div><div>入库人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '入库单-简洁版', 'template_type': 'inventory_in', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">入库单</h2>',
         'body_content': '<p>单号：{{in_no}} | {{in_date}}</p><p>仓库：{{warehouse_name}} | 类型：{{in_type}}</p><hr><p>{{items}}</p><p>合计：¥{{total_amount}}</p>',
         'footer_content': '<p>制单人：______ 入库人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '入库单-详细版', 'template_type': 'inventory_in', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>入库单（详细）</h1><p>入库单号：{{in_no}} | 日期：{{in_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">入库信息</th></tr><tr><td>仓库：{{warehouse_name}}</td><td>入库类型：{{in_type}}</td><td>经办人：{{operator}}</td><td>关联单号：{{related_no}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品编码</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">规格</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">单价</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="6" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="7">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>审核人：______________</div><div>入库人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 出库管理 (inventory_out) ===
        {'template_name': '出库单-标准A4', 'template_type': 'inventory_out', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>出库单</h1><p>出库单号：{{out_no}} | 日期：{{out_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">出库信息</th></tr><tr><td>仓库：{{warehouse_name}}</td><td>出库类型：{{out_type}}</td><td>经办人：{{operator}}</td><td>关联单号：{{related_no}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>审核人：______________</div><div>领用人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '出库单-简洁版', 'template_type': 'inventory_out', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">出库单</h2>',
         'body_content': '<p>单号：{{out_no}} | {{out_date}}</p><p>仓库：{{warehouse_name}} | 类型：{{out_type}}</p><hr><p>{{items}}</p><p>合计：¥{{total_amount}}</p>',
         'footer_content': '<p>制单人：______ 领用人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '出库单-详细版', 'template_type': 'inventory_out', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>出库单（详细）</h1><p>出库单号：{{out_no}} | 日期：{{out_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">出库信息</th></tr><tr><td>仓库：{{warehouse_name}}</td><td>出库类型：{{out_type}}</td><td>经办人：{{operator}}</td><td>关联单号：{{related_no}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品编码</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">规格</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">单价</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="6" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="7">备注：{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>审核人：______________</div><div>领用人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 盘点管理 (inventory_check) ===
        {'template_name': '盘点单-标准A4', 'template_type': 'inventory_check', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>库存盘点单</h1><p>盘点单号：{{check_no}} | 日期：{{check_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">盘点信息</th></tr><tr><td>仓库：{{warehouse_name}}</td><td>货架：{{shelf_name}}</td><td>盘点人：{{operator}}</td><td>状态：{{status}}</td></tr><tr><th style="background:#f0f0f0">商品编码</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">账面数量</th><th style="background:#f0f0f0">实盘数量</th><th style="background:#f0f0f0">差异</th></tr>{{items}}<tr><td colspan="4" style="text-align:right;font-weight:bold">盘盈/盘亏合计</td><td style="font-weight:bold">¥{{diff_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>盘点人：______________</div><div>审核人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '盘点单-简洁版', 'template_type': 'inventory_check', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">盘点单</h2>',
         'body_content': '<p>单号：{{check_no}} | {{check_date}}</p><p>仓库：{{warehouse_name}}</p><hr><p>{{items}}</p><p>差异金额：¥{{diff_amount}}</p>',
         'footer_content': '<p>盘点人：______ 审核人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '盘点差异报表', 'template_type': 'inventory_check', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>盘点差异报表</h1><p>盘点单号：{{check_no}} | 日期：{{check_date}} | 仓库：{{warehouse_name}}</p></div>',
         'body_content': '<h3>盘盈明细</h3><table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th>商品编码</th><th>商品名称</th><th>账面</th><th>实盘</th><th>差异</th><th>差异金额</th></tr>{{profit_items}}</table><h3>盘亏明细</h3><table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th>商品编码</th><th>商品名称</th><th>账面</th><th>实盘</th><th>差异</th><th>差异金额</th></tr>{{loss_items}}</table><p style="font-weight:bold">盘盈合计：¥{{profit_amount}} | 盘亏合计：¥{{loss_amount}} | 净差异：¥{{net_amount}}</p>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>盘点人：______________</div><div>审核人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 调拨管理 (transfer) ===
        {'template_name': '调拨单-标准A4', 'template_type': 'transfer', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>库存调拨单</h1><p>调拨单号：{{transfer_no}} | 日期：{{transfer_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">调拨信息</th></tr><tr><td>调出仓库：{{from_warehouse}}</td><td>调入仓库：{{to_warehouse}}</td><td>经办人：{{operator}}</td><td>状态：{{status}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="2" style="text-align:right;font-weight:bold">合计</td><td colspan="2" style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="4">调拨原因：{{reason}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>调出方签字：______________</div><div>调入方签字：______________</div><div>审批人：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '调拨单-简洁版', 'template_type': 'transfer', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">调拨单</h2>',
         'body_content': '<p>单号：{{transfer_no}} | {{transfer_date}}</p><p>{{from_warehouse}} → {{to_warehouse}}</p><hr><p>{{items}}</p><p>合计：¥{{total_amount}}</p>',
         'footer_content': '<p>调出方：______ 调入方：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '调拨单-详细版', 'template_type': 'transfer', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>库存调拨单（详细）</h1><p>调拨单号：{{transfer_no}} | 日期：{{transfer_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="4" style="background:#f0f0f0">调拨信息</th></tr><tr><td>调出仓库：{{from_warehouse}}</td><td>调入仓库：{{to_warehouse}}</td><td>经办人：{{operator}}</td><td>状态：{{status}}</td></tr><tr><th style="background:#f0f0f0">序号</th><th style="background:#f0f0f0">商品编码</th><th style="background:#f0f0f0">商品名称</th><th style="background:#f0f0f0">规格</th><th style="background:#f0f0f0">数量</th><th style="background:#f0f0f0">单价</th><th style="background:#f0f0f0">金额</th></tr>{{items}}<tr><td colspan="6" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td></tr><tr><td colspan="7">调拨原因：{{reason}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>调出方签字：______________</div><div>调入方签字：______________</div><div>审批人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 应收管理 (receivable) ===
        {'template_name': '应收对账单-标准A4', 'template_type': 'receivable', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>应收对账单</h1><p>客户：{{customer_name}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0">单号</th><th style="background:#f0f0f0">日期</th><th style="background:#f0f0f0">摘要</th><th style="background:#f0f0f0">应收金额</th><th style="background:#f0f0f0">已收金额</th><th style="background:#f0f0f0">未收金额</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td><td style="font-weight:bold">¥{{received_amount}}</td><td style="font-weight:bold">¥{{unpaid_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>客户确认：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '应收对账单-简洁版', 'template_type': 'receivable', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">应收对账单</h2>',
         'body_content': '<p>客户：{{customer_name}} | {{created_at}}</p><hr><p>{{items}}</p><p>应收合计：¥{{total_amount}} | 已收：¥{{received_amount}} | 未收：¥{{unpaid_amount}}</p>',
         'footer_content': '<p>确认签字：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '应收对账单-含逾期', 'template_type': 'receivable', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>应收对账单（含逾期）</h1><p>客户：{{customer_name}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0">单号</th><th style="background:#f0f0f0">日期</th><th style="background:#f0f0f0">摘要</th><th style="background:#f0f0f0">应收金额</th><th style="background:#f0f0f0">已收</th><th style="background:#f0f0f0">未收</th><th style="background:#f0f0f0">逾期天数</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td><td style="font-weight:bold">¥{{received_amount}}</td><td style="font-weight:bold;color:red">¥{{unpaid_amount}}</td><td></td></tr></table><p style="color:red;font-weight:bold">逾期总金额：¥{{overdue_amount}}</p>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>客户确认：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 应付管理 (payable) ===
        {'template_name': '应付对账单-标准A4', 'template_type': 'payable', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>应付对账单</h1><p>供应商：{{supplier_name}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0">单号</th><th style="background:#f0f0f0">日期</th><th style="background:#f0f0f0">摘要</th><th style="background:#f0f0f0">应付金额</th><th style="background:#f0f0f0">已付金额</th><th style="background:#f0f0f0">未付金额</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td><td style="font-weight:bold">¥{{paid_amount}}</td><td style="font-weight:bold">¥{{unpaid_amount}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>供应商确认：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '应付对账单-简洁版', 'template_type': 'payable', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">应付对账单</h2>',
         'body_content': '<p>供应商：{{supplier_name}} | {{created_at}}</p><hr><p>{{items}}</p><p>应付合计：¥{{total_amount}} | 已付：¥{{paid_amount}} | 未付：¥{{unpaid_amount}}</p>',
         'footer_content': '<p>确认签字：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '应付对账单-含逾期', 'template_type': 'payable', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>应付对账单（含逾期）</h1><p>供应商：{{supplier_name}} | 日期：{{created_at}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0">单号</th><th style="background:#f0f0f0">日期</th><th style="background:#f0f0f0">摘要</th><th style="background:#f0f0f0">应付金额</th><th style="background:#f0f0f0">已付</th><th style="background:#f0f0f0">未付</th><th style="background:#f0f0f0">逾期天数</th></tr>{{items}}<tr><td colspan="3" style="text-align:right;font-weight:bold">合计</td><td style="font-weight:bold">¥{{total_amount}}</td><td style="font-weight:bold">¥{{paid_amount}}</td><td style="font-weight:bold;color:red">¥{{unpaid_amount}}</td><td></td></tr></table><p style="color:red;font-weight:bold">逾期总金额：¥{{overdue_amount}}</p>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>制单人：______________</div><div>供应商确认：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 收款管理 (receipt) ===
        {'template_name': '收款凭证-标准A4', 'template_type': 'receipt', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>收款凭证</h1><p>凭证号：{{receipt_no}} | 日期：{{receipt_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0">客户名称</th><td>{{customer_name}}</td><th style="background:#f0f0f0">收款金额</th><td style="font-weight:bold;font-size:16px">¥{{amount}}</td></tr><tr><th style="background:#f0f0f0">收款方式</th><td>{{payment_method}}</td><th style="background:#f0f0f0">收款账户</th><td>{{account_name}}</td></tr><tr><th style="background:#f0f0f0">关联单号</th><td colspan="3">{{related_no}}</td></tr><tr><th style="background:#f0f0f0">经办人</th><td>{{operator}}</td><th style="background:#f0f0f0">备注</th><td>{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:50px;display:flex;justify-content:space-between"><div>收款人签字：______________</div><div>客户确认签字：______________</div><div>日期：______________</div></div><p style="text-align:center;color:#999;font-size:10px">此凭证一式两联</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '收款凭证-简洁版', 'template_type': 'receipt', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">收款凭证</h2>',
         'body_content': '<p>凭证号：{{receipt_no}} | {{receipt_date}}</p><p>客户：{{customer_name}}</p><p style="font-size:18px;font-weight:bold">收款金额：¥{{amount}}</p><p>方式：{{payment_method}} | 账户：{{account_name}}</p>',
         'footer_content': '<p>收款人：______ 客户：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '收款凭证-正式版', 'template_type': 'receipt', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:3px double #333;padding-bottom:15px"><h1 style="letter-spacing:10px">收  款  凭  证</h1><p>凭证号：{{receipt_no}} | 日期：{{receipt_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="8"><tr><td style="width:25%;background:#f5f5f5">今收到</td><td colspan="3">{{customer_name}}</td></tr><tr><td style="background:#f5f5f0">收款金额</td><td colspan="3" style="font-weight:bold;font-size:18px">人民币（大写）：{{amount_cn}}  ¥{{amount}}</td></tr><tr><td style="background:#f5f5f0">收款方式</td><td>{{payment_method}}</td><td style="background:#f5f5f0">收款账户</td><td>{{account_name}}</td></tr><tr><td style="background:#f5f5f0">关联单号</td><td>{{related_no}}</td><td style="background:#f5f5f0">经办人</td><td>{{operator}}</td></tr><tr><td style="background:#f5f5f0">备注</td><td colspan="3">{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:60px;display:flex;justify-content:space-between"><div>收款单位（章）：______________</div><div>客户确认（章）：______________</div></div><p style="text-align:center;color:#999;font-size:10px;margin-top:10px">此凭证一式两联，收款联和客户联</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:14px;padding:30px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:8px}'},

        # === 付款管理 (payment) ===
        {'template_name': '付款凭证-标准A4', 'template_type': 'payment', 'paper_size': 'A5', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>付款凭证</h1><p>凭证号：{{payment_no}} | 日期：{{payment_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th style="background:#f0f0f0">供应商名称</th><td>{{supplier_name}}</td><th style="background:#f0f0f0">付款金额</th><td style="font-weight:bold;font-size:16px">¥{{amount}}</td></tr><tr><th style="background:#f0f0f0">付款方式</th><td>{{payment_method}}</td><th style="background:#f0f0f0">付款账户</th><td>{{account_name}}</td></tr><tr><th style="background:#f0f0f0">关联单号</th><td colspan="3">{{related_no}}</td></tr><tr><th style="background:#f0f0f0">经办人</th><td>{{operator}}</td><th style="background:#f0f0f0">备注</th><td>{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:50px;display:flex;justify-content:space-between"><div>付款人签字：______________</div><div>审批人签字：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '付款凭证-简洁版', 'template_type': 'payment', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">付款凭证</h2>',
         'body_content': '<p>凭证号：{{payment_no}} | {{payment_date}}</p><p>供应商：{{supplier_name}}</p><p style="font-size:18px;font-weight:bold">付款金额：¥{{amount}}</p><p>方式：{{payment_method}} | 账户：{{account_name}}</p>',
         'footer_content': '<p>付款人：______ 审批人：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '付款凭证-正式版', 'template_type': 'payment', 'paper_size': 'A5',
         'header_content': '<div style="text-align:center;border-bottom:3px double #333;padding-bottom:15px"><h1 style="letter-spacing:10px">付  款  凭  证</h1><p>凭证号：{{payment_no}} | 日期：{{payment_date}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="8"><tr><td style="width:25%;background:#f5f5f5">今付给</td><td colspan="3">{{supplier_name}}</td></tr><tr><td style="background:#f5f5f0">付款金额</td><td colspan="3" style="font-weight:bold;font-size:18px">人民币（大写）：{{amount_cn}}  ¥{{amount}}</td></tr><tr><td style="background:#f5f5f0">付款方式</td><td>{{payment_method}}</td><td style="background:#f5f5f0">付款账户</td><td>{{account_name}}</td></tr><tr><td style="background:#f5f5f0">关联单号</td><td>{{related_no}}</td><td style="background:#f5f5f0">经办人</td><td>{{operator}}</td></tr><tr><td style="background:#f5f5f0">备注</td><td colspan="3">{{remark}}</td></tr></table>',
         'footer_content': '<div style="margin-top:60px;display:flex;justify-content:space-between"><div>付款单位（章）：______________</div><div>审批人（章）：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:14px;padding:30px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:8px}'},

        # === 工资发放 (salary) ===
        {'template_name': '工资条-标准A4', 'template_type': 'salary', 'paper_size': 'A4', 'is_default': 1,
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>工资条</h1><p>月份：{{salary_month}} | 部门：{{department}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><th colspan="2" style="background:#f0f0f0">姓名：{{employee_name}}</th><th style="background:#f0f0f0">岗位：{{position}}</th><th style="background:#f0f0f0">工号：{{employee_no}}</th></tr><tr><th style="background:#f0f0f0">基本工资</th><td>{{base_salary}}</td><th style="background:#f0f0f0">绩效工资</th><td>{{performance_salary}}</td></tr><tr><th style="background:#f0f0f0">提成</th><td>{{commission}}</td><th style="background:#f0f0f0">补贴</th><td>{{allowance}}</td></tr><tr><th style="background:#f0f0f0">加班费</th><td>{{overtime_pay}}</td><th style="background:#f0f0f0">奖金</th><td>{{bonus}}</td></tr><tr><th style="background:#f0f0f0">应发合计</th><td colspan="3" style="font-weight:bold">¥{{gross_salary}}</td></tr><tr><th style="background:#f0f0f0">社保</th><td>{{social_insurance}}</td><th style="background:#f0f0f0">公积金</th><td>{{housing_fund}}</td></tr><tr><th style="background:#f0f0f0">个税</th><td>{{tax}}</td><th style="background:#f0f0f0">其他扣款</th><td>{{other_deductions}}</td></tr><tr><th style="background:#f0f0f0">实发工资</th><td colspan="3" style="font-weight:bold;font-size:16px">¥{{net_salary}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>员工签字：______________</div><div>发放人：______________</div><div>日期：______________</div></div>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},
        {'template_name': '工资条-简洁版', 'template_type': 'salary', 'paper_size': 'A5', 'paper_width': 148, 'paper_height': 210,
         'header_content': '<h2 style="text-align:center">工资条</h2>',
         'body_content': '<p>{{salary_month}} | {{employee_name}} | {{department}}</p><hr><p>基本工资：{{base_salary}} | 绩效：{{performance_salary}}</p><p>提成：{{commission}} | 补贴：{{allowance}}</p><p>应发：¥{{gross_salary}}</p><p>社保：{{social_insurance}} | 个税：{{tax}}</p><p style="font-weight:bold;font-size:16px">实发：¥{{net_salary}}</p>',
         'footer_content': '<p>员工签字：______</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:11px;padding:15px}'},
        {'template_name': '工资条-保密版', 'template_type': 'salary', 'paper_size': 'A4',
         'header_content': '<div style="text-align:center;border-bottom:2px solid #333;padding-bottom:10px"><h1>工资条（保密）</h1><p style="color:red">⚠ 本文件包含机密信息，请妥善保管</p><p>月份：{{salary_month}}</p></div>',
         'body_content': '<table width="100%" border="1" cellspacing="0" cellpadding="5"><tr><td style="background:#f0f0f0">姓名</td><td>{{employee_name}}</td><td style="background:#f0f0f0">部门</td><td>{{department}}</td><td style="background:#f0f0f0">岗位</td><td>{{position}}</td></tr><tr><td style="background:#f0f0f0">基本工资</td><td>{{base_salary}}</td><td style="background:#f0f0f0">绩效</td><td>{{performance_salary}}</td><td style="background:#f0f0f0">提成</td><td>{{commission}}</td></tr><tr><td style="background:#f0f0f0">补贴</td><td>{{allowance}}</td><td style="background:#f0f0f0">加班费</td><td>{{overtime_pay}}</td><td style="background:#f0f0f0">奖金</td><td>{{bonus}}</td></tr><tr style="background:#e8f5e9"><td style="font-weight:bold">应发合计</td><td colspan="5" style="font-weight:bold">¥{{gross_salary}}</td></tr><tr><td style="background:#f0f0f0">社保</td><td>{{social_insurance}}</td><td style="background:#f0f0f0">公积金</td><td>{{housing_fund}}</td><td style="background:#f0f0f0">个税</td><td>{{tax}}</td></tr><tr><td style="background:#f0f0f0">其他扣款</td><td colspan="5">{{other_deductions}}</td></tr><tr style="background:#fff3e0"><td style="font-weight:bold;font-size:14px">实发工资</td><td colspan="5" style="font-weight:bold;font-size:16px">¥{{net_salary}}</td></tr></table>',
         'footer_content': '<div style="margin-top:30px;display:flex;justify-content:space-between"><div>员工签字：______________</div><div>发放人：______________</div><div>日期：______________</div></div><p style="text-align:center;color:#999;font-size:10px">本工资条为机密文件，严禁外传</p>',
         'style_content': 'body{font-family:SimSun,serif;font-size:12px;padding:20px}table{border-collapse:collapse;width:100%}th,td{border:1px solid #333;padding:5px}th{background:#f5f5f5}'},

        # === 设备标签 (device_label) ===
        # 模板1 - 设备标签-标准版 (60x40mm)，每页21个（3列x7行）
        {'template_name': '设备标签-标准版', 'template_type': 'device_label', 'paper_size': 'A4', 'is_default': 1,
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{device_brand}} {{device_model}}</div><div class="info">SN:{{device_sn}}</div><div class="date">{{receive_date}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:70mm;height:40mm;border:1px solid #ccc;box-sizing:border-box;padding:2mm;page-break-inside:avoid}.no{font-size:10px;font-weight:bold;border-bottom:1px dashed #999;padding-bottom:1mm;margin-bottom:1mm}.info{font-size:8px;line-height:1.3;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.date{font-size:7px;color:#666;margin-top:auto}'},        # 模板2 - 设备标签-小巧版 (40x30mm)，每页45个（5列x9行）
        {'template_name': '设备标签-小巧版', 'template_type': 'device_label', 'paper_size': 'A4',
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{device_model}}</div><div class="date">{{receive_date}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:42mm;height:30mm;border:1px solid #ccc;box-sizing:border-box;padding:1.5mm;page-break-inside:avoid}.no{font-size:8px;font-weight:bold;border-bottom:1px dashed #999;padding-bottom:0.5mm;margin-bottom:0.5mm}.info{font-size:7px;line-height:1.2;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.date{font-size:6px;color:#666;margin-top:auto}'},        # 模板3 - 设备标签-最小版 (30x20mm)，每页98个（7列x14行）
        {'template_name': '设备标签-最小版', 'template_type': 'device_label', 'paper_size': 'A4',
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="date">{{receive_date}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:30mm;height:20mm;border:1px solid #ccc;box-sizing:border-box;padding:1mm;page-break-inside:avoid}.no{font-size:7px;font-weight:bold}.date{font-size:6px;color:#666}'},
        # === 客户自带标签 (customer_label) ===
        # 模板1 - 客户标签-标准版 (60x40mm)，每页21个（3列x7行）
        {'template_name': '客户标签-标准版', 'template_type': 'customer_label', 'paper_size': 'A4', 'is_default': 1,
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="badge">客户自带</div><div class="no">{{receive_no}}</div><div class="info">{{customer_name}} {{customer_phone}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:70mm;height:40mm;border:1px solid #409EFF;box-sizing:border-box;padding:2mm;page-break-inside:avoid}.badge{font-size:7px;background:#409EFF;color:#fff;padding:0.3mm 1.5mm;border-radius:1px;display:inline-block;margin-bottom:0.5mm}.no{font-size:10px;font-weight:bold;border-bottom:1px dashed #409EFF;padding-bottom:1mm;margin-bottom:1mm}.info{font-size:8px;line-height:1.3;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.date{font-size:7px;color:#666;margin-top:auto}'},        # 模板2 - 客户标签-小巧版 (40x30mm)，每页45个（5列x9行）
        {'template_name': '客户标签-小巧版', 'template_type': 'customer_label', 'paper_size': 'A4',
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
<div class="label"><div class="no">{{receive_no}}</div><div class="info">{{customer_name}}</div><div class="info">{{item_name}}</div><div class="date">{{receive_date}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:42mm;height:30mm;border:1px solid #409EFF;box-sizing:border-box;padding:1.5mm;page-break-inside:avoid}.no{font-size:8px;font-weight:bold;border-bottom:1px dashed #409EFF;padding-bottom:0.5mm;margin-bottom:0.5mm}.info{font-size:7px;line-height:1.2;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.date{font-size:6px;color:#666;margin-top:auto}'},        # 模板3 - 客户标签-最小版 (30x20mm)，每页98个（7列x14行）
        {'template_name': '客户标签-最小版', 'template_type': 'customer_label', 'paper_size': 'A4',
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
<div class="label"><div class="badge">自带</div><div class="no">{{receive_no}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:30mm;height:20mm;border:1px solid #409EFF;box-sizing:border-box;padding:1mm;page-break-inside:avoid}.badge{font-size:6px;background:#409EFF;color:#fff;padding:0.2mm 1mm;border-radius:1px;display:inline-block}.no{font-size:7px;font-weight:bold;margin-top:0.5mm}'},
        # === 商品条码标签 (product_barcode) ===
        # 模板1 - 商品条码-标准版 (60x40mm)，每页21个（3列x7行）
        {'template_name': '商品条码-标准版', 'template_type': 'product_barcode', 'paper_size': 'A4', 'is_default': 1,
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||||||||||||</div><div class="code">{{barcode}}</div><div class="info">规格:{{specification}}</div><div class="price">¥{{sale_price}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:70mm;height:40mm;border:1px solid #ccc;box-sizing:border-box;padding:2mm;page-break-inside:avoid}.name{font-size:9px;font-weight:bold;margin-bottom:1mm;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.barcode{font-size:14px;letter-spacing:1px;color:#000;text-align:center}.code{font-size:7px;text-align:center;margin-bottom:1mm;font-family:"Courier New",monospace}.info{font-size:7px;color:#666}.price{font-size:10px;color:#E6A23C;font-weight:bold;margin-top:1mm}'},        # 模板2 - 商品条码-小巧版 (40x30mm)，每页45个（5列x9行）
        {'template_name': '商品条码-小巧版', 'template_type': 'product_barcode', 'paper_size': 'A4',
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="name">{{product_name}}</div><div class="barcode">||||||||||||||</div><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:42mm;height:30mm;border:1px solid #ccc;box-sizing:border-box;padding:1.5mm;page-break-inside:avoid}.name{font-size:7px;font-weight:bold;margin-bottom:0.5mm;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.barcode{font-size:12px;letter-spacing:0.5px;color:#000;text-align:center}.code{font-size:6px;text-align:center;margin-bottom:0.5mm;font-family:"Courier New",monospace}.price{font-size:9px;color:#E6A23C;font-weight:bold}'},        # 模板3 - 商品条码-最小版 (30x20mm)，每页98个（7列x14行）
        {'template_name': '商品条码-最小版', 'template_type': 'product_barcode', 'paper_size': 'A4',
         'header_content': '',
         'body_content': '''<div class="label-grid">
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
<div class="label"><div class="code">{{barcode}}</div><div class="price">¥{{sale_price}}</div></div>
</div>''',
         'footer_content': '',
         'style_content': 'body{font-family:"Microsoft YaHei",sans-serif;margin:0;padding:0}.label-grid{display:flex;flex-wrap:wrap;width:210mm}.label{width:30mm;height:20mm;border:1px solid #ccc;box-sizing:border-box;padding:1mm;page-break-inside:avoid}.code{font-size:7px;text-align:center;font-family:"Courier New",monospace}.price{font-size:8px;color:#E6A23C;font-weight:bold;text-align:center;margin-top:0.5mm}'},
    ]

    for t in templates:
        db.session.add(PrintTemplate(**t))

    db.session.commit()
    return jsonify({'code': 200, 'message': f'已初始化{len(templates)}个默认打印模板'})

@app.route('/api/users', methods=['GET'])
@jwt_required()
def get_users():
    """获取用户列表"""
    users = SysUser.query.filter_by(is_deleted=0).all()
    return jsonify({
        'code': 200,
        'data': [{
            'id': u.id,
            'username': u.username,
            'real_name': u.real_name,
            'phone': u.phone,
            'email': u.email,
            'role_id': u.role_id,
            'department': u.department,
            'position': u.position,
            'base_salary': float(u.base_salary) if u.base_salary else 0,
            'status': u.status,
            'last_login_time': u.last_login_time.strftime('%Y-%m-%d %H:%M:%S') if u.last_login_time else None
        } for u in users]
    })

# ============================================
# 健康检查
# ============================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({'code': 200, 'message': '服务正常运行', 'timestamp': datetime.now().isoformat()})

# ============================================
# 新增数据模型 - 报价管理 / 自有设备
# ============================================

class QuoteOrder(db.Model):
    """报价单"""
    __tablename__ = 'quote_order'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    quote_no = db.Column(db.String(50), unique=True, nullable=False)
    customer_id = db.Column(db.BigInteger)
    customer_name = db.Column(db.String(100))
    customer_phone = db.Column(db.String(20))
    contact_name = db.Column(db.String(50))
    address = db.Column(db.String(255))
    quote_date = db.Column(db.Date, default=datetime.now().date)
    valid_until = db.Column(db.Date)
    total_amount = db.Column(db.Numeric(15, 2), default=0.00)
    remark = db.Column(db.Text)
    status = db.Column(db.Integer, default=0)  # 0待确认 1已确认 2已失效 3已转工单 4已转接件 5已转销售
    related_type = db.Column(db.String(50))
    related_id = db.Column(db.BigInteger)
    created_by = db.Column(db.BigInteger)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class QuoteOrderItem(db.Model):
    """报价单明细"""
    __tablename__ = 'quote_order_item'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    quote_id = db.Column(db.BigInteger, nullable=False)
    product_name = db.Column(db.String(200))
    specification = db.Column(db.String(100))
    brand = db.Column(db.String(50))
    unit = db.Column(db.String(20))
    quantity = db.Column(db.Numeric(10, 2), default=1)
    unit_price = db.Column(db.Numeric(15, 2), default=0.00)
    subtotal = db.Column(db.Numeric(15, 2), default=0.00)
    remark = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.now)

class OwnDevice(db.Model):
    """自有设备"""
    __tablename__ = 'own_device'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    asset_no = db.Column(db.String(50), unique=True, nullable=False)
    device_type = db.Column(db.String(50))
    device_model = db.Column(db.String(100))
    serial_number = db.Column(db.String(100))
    cpu = db.Column(db.String(100))
    memory = db.Column(db.String(50))
    hard_disk = db.Column(db.String(100))
    system = db.Column(db.String(50))
    system_version = db.Column(db.String(50))
    accessories = db.Column(db.Text)
    account = db.Column(db.String(100))
    password = db.Column(db.String(255))
    password_remark = db.Column(db.String(255))
    purchase_date = db.Column(db.Date)
    warranty_expire = db.Column(db.Date)
    location = db.Column(db.String(100))
    user_id = db.Column(db.BigInteger)
    cost = db.Column(db.Numeric(15, 2), default=0.00)
    depreciation = db.Column(db.Numeric(15, 2), default=0.00)
    status = db.Column(db.Integer, default=0)  # 0正常 1维修中 2报废 3外借
    remark = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class ReceiveOrderLog(db.Model):
    """接件单操作日志"""
    __tablename__ = 'receive_order_log'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    receive_order_id = db.Column(db.BigInteger, nullable=False)
    action = db.Column(db.String(50))
    old_status = db.Column(db.Integer)
    new_status = db.Column(db.Integer)
    content = db.Column(db.Text)
    operator_id = db.Column(db.BigInteger)
    operator_name = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.now)


# --- 采购单管理模块 ---

class PurchaseOrder(db.Model):
    """采购单"""
    __tablename__ = 'purchase_order'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    order_no = db.Column(db.String(50), unique=True, nullable=False)
    supplier_id = db.Column(db.BigInteger)
    supplier_name = db.Column(db.String(100))
    order_date = db.Column(db.Date)
    delivery_date = db.Column(db.Date)
    total_amount = db.Column(db.Numeric(15, 2), default=0)
    total_quantity = db.Column(db.Integer, default=0)
    status = db.Column(db.Integer, default=0)  # 0待审核 1已审核 2已完成 3已取消
    has_invoice = db.Column(db.Integer, default=0)  # 0未收发票 1已收发票
    remark = db.Column(db.Text)
    created_by = db.Column(db.BigInteger)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class PurchaseOrderItem(db.Model):
    """采购单明细"""
    __tablename__ = 'purchase_order_item'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    order_id = db.Column(db.BigInteger, nullable=False)
    product_id = db.Column(db.BigInteger)
    product_name = db.Column(db.String(200))
    specification = db.Column(db.String(100))
    unit = db.Column(db.String(20))
    quantity = db.Column(db.Integer, default=0)
    price = db.Column(db.Numeric(15, 4), default=0)
    amount = db.Column(db.Numeric(15, 2), default=0)
    received_qty = db.Column(db.Integer, default=0)
    remark = db.Column(db.Text)


# --- 销售单管理模块 ---

class SalesOrder(db.Model):
    """销售单"""
    __tablename__ = 'sales_order'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    order_no = db.Column(db.String(50), unique=True, nullable=False)
    customer_id = db.Column(db.BigInteger)
    customer_name = db.Column(db.String(100))
    customer_phone = db.Column(db.String(20))  # 客户电话
    customer_address = db.Column(db.String(255))  # 客户地址
    contact_name = db.Column(db.String(50))  # 联系人
    order_date = db.Column(db.Date)
    total_amount = db.Column(db.Numeric(15, 2), default=0)
    total_quantity = db.Column(db.Integer, default=0)
    discount_amount = db.Column(db.Numeric(15, 2), default=0)  # 折扣金额
    freight_amount = db.Column(db.Numeric(15, 2), default=0)  # 运费
    actual_amount = db.Column(db.Numeric(15, 2), default=0)  # 实际应收金额
    paid_amount = db.Column(db.Numeric(15, 2), default=0)  # 已收金额
    payment_method = db.Column(db.String(20))  # 付款方式：现金/转账/支付宝/微信/赊账
    delivery_method = db.Column(db.String(20))  # 交货方式：自提/送货/快递
    salesperson_id = db.Column(db.BigInteger)  # 销售人员ID
    salesperson_name = db.Column(db.String(50))  # 销售人员姓名
    status = db.Column(db.Integer, default=0)  # 0待审核 1已审核 2已出库 3已完成 4已取消
    remark = db.Column(db.Text)
    created_by = db.Column(db.BigInteger)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    has_invoice = db.Column(db.Integer, default=0)  # 是否开具发票：0否 1是
    has_receipt = db.Column(db.Integer, default=0)  # 是否开具收据：0否 1是


class SalesOrderItem(db.Model):
    """销售单明细"""
    __tablename__ = 'sales_order_item'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    order_id = db.Column(db.BigInteger, nullable=False)
    product_id = db.Column(db.BigInteger)
    product_name = db.Column(db.String(200))
    specification = db.Column(db.String(100))
    unit = db.Column(db.String(20))
    quantity = db.Column(db.Integer, default=0)
    price = db.Column(db.Numeric(15, 4), default=0)
    amount = db.Column(db.Numeric(15, 2), default=0)
    delivered_qty = db.Column(db.Integer, default=0)
    remark = db.Column(db.Text)


class SalesInvoice(db.Model):
    """销售发票"""
    __tablename__ = 'sales_invoice'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    order_id = db.Column(db.BigInteger, nullable=False)  # 关联销售单ID
    order_no = db.Column(db.String(50))  # 关联销售单号
    invoice_type = db.Column(db.String(30))  # 发票类型：普通发票/增值税专用发票/电子发票
    invoice_status = db.Column(db.String(20), default='未开票')  # 开票状态：未开票/已开票/作废
    invoice_no = db.Column(db.String(50))  # 发票编号
    invoice_date = db.Column(db.Date)  # 开票日期
    # 客户开票信息
    buyer_name = db.Column(db.String(200))  # 客户开票抬头（公司名称/个人）
    buyer_tax_no = db.Column(db.String(50))  # 统一社会信用代码/税号
    buyer_address = db.Column(db.String(255))  # 开票地址
    buyer_phone = db.Column(db.String(50))  # 联系电话
    buyer_bank = db.Column(db.String(100))  # 开户行
    buyer_bank_account = db.Column(db.String(50))  # 银行账号
    # 开票项目明细（JSON格式存储）
    items = db.Column(db.Text)  # 发票明细JSON：[{name,spec,qty,price,amount,tax_rate,tax,total}]
    total_amount = db.Column(db.Numeric(15, 2), default=0)  # 金额合计（不含税）
    total_tax = db.Column(db.Numeric(15, 2), default=0)  # 税额合计
    total_with_tax = db.Column(db.Numeric(15, 2), default=0)  # 价税合计
    tax_rate = db.Column(db.Numeric(5, 2), default=0)  # 默认税率%
    remark = db.Column(db.Text)  # 开票备注
    attachment = db.Column(db.String(500))  # 发票附件路径
    created_by = db.Column(db.BigInteger)  # 开票人ID
    created_by_name = db.Column(db.String(50))  # 开票人姓名
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class SalesReceipt(db.Model):
    """销售收据"""
    __tablename__ = 'sales_receipt'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    order_id = db.Column(db.BigInteger, nullable=False)  # 关联销售单ID
    order_no = db.Column(db.String(50))  # 关联销售单号
    receipt_no = db.Column(db.String(50))  # 收据编号（自动生成）
    receipt_date = db.Column(db.Date)  # 收款日期
    # 客户信息
    customer_name = db.Column(db.String(100))  # 客户名称
    customer_phone = db.Column(db.String(20))  # 联系方式
    # 金额信息
    total_amount = db.Column(db.Numeric(15, 2), default=0)  # 应收金额
    paid_amount = db.Column(db.Numeric(15, 2), default=0)  # 实收金额
    payment_method = db.Column(db.String(20))  # 收款方式
    # 商品明细（JSON格式）
    items = db.Column(db.Text)  # 收据商品明细JSON
    remark = db.Column(db.Text)  # 备注/收款说明
    payee = db.Column(db.String(50))  # 收款人
    status = db.Column(db.Integer, default=1)  # 状态：1有效 0作废
    created_by = db.Column(db.BigInteger)  # 开具人ID
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class AssetType(db.Model):
    """资产类型"""
    __tablename__ = 'asset_type'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    type_code = db.Column(db.String(50), unique=True, nullable=False)  # 类型编码: network/computer/printer等
    type_name = db.Column(db.String(50), nullable=False)  # 类型名称: 网络类设备/电脑办公类等
    icon = db.Column(db.String(50))  # 图标
    sort_order = db.Column(db.Integer, default=0)  # 排序
    status = db.Column(db.Integer, default=1)  # 0停用 1启用
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class Asset(db.Model):
    """资产台账"""
    __tablename__ = 'asset'
    id = db.Column(db.BigInteger, primary_key=True, autoincrement=True)
    asset_no = db.Column(db.String(50), unique=True)  # 资产编号
    
    # 归属信息（客户优先维度）
    customer_id = db.Column(db.BigInteger, nullable=False)  # 归属客户ID
    customer_name = db.Column(db.String(100), nullable=False)  # 归属客户名称
    office_id = db.Column(db.BigInteger)  # 所属办公室ID
    office_name = db.Column(db.String(50))  # 所属办公室名称
    location = db.Column(db.String(100))  # 存放位置
    
    # 基础信息
    asset_type_id = db.Column(db.BigInteger, nullable=False)  # 资产类型ID
    asset_type_name = db.Column(db.String(50))  # 资产类型名称
    asset_name = db.Column(db.String(100), nullable=False)  # 资产名称
    device_no = db.Column(db.String(50))  # 设备编号
    sn_code = db.Column(db.String(100))  # SN序列号
    
    # 时间信息
    register_date = db.Column(db.Date)  # 登记日期
    purchase_date = db.Column(db.Date)  # 采购日期
    warranty_expire_date = db.Column(db.Date)  # 质保到期日
    
    # 状态信息
    warranty_status = db.Column(db.Integer, default=1)  # 0过保 1在保
    asset_status = db.Column(db.Integer, default=1)  # 1正常使用 2维修中 3闲置 4报废 5停用
    
    # 责任人信息
    responsible_person = db.Column(db.String(50))  # 使用责任人
    contact_phone = db.Column(db.String(20))  # 联系电话
    
    # 通用技术字段（所有类型都显示）
    ip_address = db.Column(db.String(50))  # IP地址
    login_password = db.Column(db.String(100))  # 设备登录密码
    remark = db.Column(db.Text)  # 备注信息
    
    # 专属字段（JSON存储）
    asset_data = db.Column(db.JSON)  # 设备类型专属字段
    
    # 关联信息
    sales_order_id = db.Column(db.BigInteger)  # 关联销售单ID
    sales_order_no = db.Column(db.String(50))  # 关联销售单号
    
    # 审计字段
    created_by = db.Column(db.BigInteger)
    created_by_name = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


# 采购单状态映射
PO_STATUS_MAP = {
    0: '待审核',
    1: '已审核',
    2: '已完成',
    3: '已取消'
}

# 销售单状态映射
SO_STATUS_MAP = {
    0: '待审核',
    1: '已审核',
    2: '已出库',
    3: '已完成',
    4: '已取消'
}

# 接件单状态映射
RO_STATUS_MAP = {
    0: '已登记',
    1: '检测中',
    2: '待报价',
    3: '待客户确认',
    4: '待领料/采购',
    5: '维修中',
    6: '待测试',
    7: '待取件',
    8: '待结算',
    9: '已完成',
    10: '送修外店',
    11: '外店已报价',
    12: '待外店维修',
    13: '外店维修中',
    14: '外店取回待测试',
    15: '已取消'
}

# 报价单状态映射
QUOTE_STATUS_MAP = {
    0: '待确认',
    1: '已确认',
    2: '已失效',
    3: '已转工单',
    4: '已转接件',
    5: '已转销售'
}

# ============================================
# API路由 - 接件管理（完善）
# ============================================

@app.route('/api/receiveorders/<int:id>', methods=['PUT'])
@jwt_required()
def update_receiveorder(id):
    """更新接件单（含设备、配件、员工信息）"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    data = request.get_json()

    # 1. 更新主表字段（仅更新模型中存在的字段）
    for field in ['customer_id', 'customer_name', 'customer_phone',
                  'receive_type', 'external_shop_id', 'external_shop_name', 'remark']:
        if field in data:
            setattr(order, field, data[field])
    
    # 更新接待员工和维修工程师信息
    if 'reception_user_id' in data:
        reception_user_id = data.get('reception_user_id')
        order.receiver_id = reception_user_id
        if reception_user_id:
            reception_user = SysUser.query.get(reception_user_id)
            if reception_user:
                order.receiver_name = reception_user.real_name or reception_user.username
    
    if 'engineer_user_id' in data:
        engineer_user_id = data.get('engineer_user_id')
        order.engineer_id = engineer_user_id
        if engineer_user_id:
            engineer_user = SysUser.query.get(engineer_user_id)
            if engineer_user:
                order.engineer_name = engineer_user.real_name or engineer_user.username

    # 2. 更新设备信息
    if 'devices' in data and isinstance(data['devices'], list):
        # 删除旧设备
        ReceiveOrderDevice.query.filter_by(receive_order_id=id).delete()
        # 添加新设备
        for dev in data['devices']:
            device = ReceiveOrderDevice(
                receive_order_id=id,
                device_type=dev.get('device_type', ''),
                device_brand=dev.get('brand') or dev.get('device_brand', ''),
                device_model=dev.get('model') or dev.get('device_model', ''),
                device_sn=dev.get('serial_number') or dev.get('device_sn', ''),
                device_name=dev.get('device_name', ''),
                fault_desc=dev.get('fault_desc', ''),
                appearance_desc=dev.get('appearance_desc', ''),
                cpu=dev.get('cpu', ''),
                memory=dev.get('memory', ''),
                disk=dev.get('hard_disk') or dev.get('disk', ''),
                os=dev.get('os', ''),
                os_version=dev.get('os_version', ''),
                toner_model=dev.get('toner_model', ''),
                drum_model=dev.get('cartridge_model') or dev.get('drum_model', ''),
                monitor_brand=dev.get('monitor_brand', ''),
                camera_count=dev.get('camera_count', 0),
                port_count=dev.get('port_count', 0),
                ip_address=dev.get('ip_address', ''),
                firmware_version=dev.get('firmware_version', ''),
                accessories=dev.get('accessories', '')
            )
            db.session.add(device)
        db.session.flush()

    # 3. 更新配件信息
    if 'parts' in data and isinstance(data['parts'], list):
        # 删除旧配件
        ReceiveOrderPart.query.filter_by(receive_order_id=id).delete()
        # 添加新配件
        for part in data['parts']:
            p = ReceiveOrderPart(
                receive_order_id=id,
                product_name=part.get('name') or part.get('product_name', ''),
                specification=part.get('specification', ''),
                unit_name=part.get('unit', ''),
                quantity=part.get('quantity', 1),
                unit_price=part.get('unit_price', 0),
                total_price=(part.get('quantity', 1) or 1) * (part.get('unit_price', 0) or 0),
                status=part.get('status', '完好'),
                remark=part.get('remark', ''),
                source='自带'
            )
            db.session.add(p)

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功', 'data': to_dict(order)})

@app.route('/api/receiveorders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_receiveorder(id):
    """删除接件单 - 只有状态为已登记(0)或已取消(15)的接件单才能删除"""
    try:
        order = ReceiveOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '接件单不存在'}), 404

        if order.status not in (0, 15):
            return jsonify({'code': 400, 'message': f'当前状态为【{RO_STATUS_MAP.get(order.status, "未知")}】，只有已登记或已取消的接件单才能删除'}), 400

        # 删除关联数据
        ReceiveOrderDevice.query.filter_by(receive_order_id=id).delete()
        ReceiveOrderPart.query.filter_by(receive_order_id=id).delete()
        ReceiveOrderLog.query.filter_by(receive_order_id=id).delete()

        db.session.delete(order)
        db.session.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除接件单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'删除接件单失败: {str(e)}'}), 500


@app.route('/api/receiveorders/<int:id>/detect', methods=['POST'])
@jwt_required()
def detect_receiveorder(id):
    """工程师检测"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：已登记(0)才能检测
    if order.status not in [0]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行检测操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    can_repair = data.get('can_repair', True)

    # 保存检测信息
    order.detect_result = data.get('detect_result', '')
    order.detect_fault_reason = data.get('detect_fault_reason', '')
    order.detect_repair_plan = data.get('detect_repair_plan', '')
    detect_parts = data.get('detect_parts')
    if detect_parts is not None:
        order.detect_parts = json.dumps(detect_parts, ensure_ascii=False) if isinstance(detect_parts, (list, dict)) else str(detect_parts)

    # 设置工程师
    if data.get('engineer_id'):
        order.engineer_id = data['engineer_id']
        order.engineer_name = data.get('engineer_name', '')

    if can_repair:
        order.status = 2  # 待报价
        content = f'检测完成，可维修。故障原因：{order.detect_fault_reason}，维修方案：{order.detect_repair_plan}'
    else:
        order.status = 9  # 送修外店
        content = f'检测完成，本店无法维修。故障原因：{order.detect_fault_reason}'

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='工程师检测',
        old_status=old_status,
        new_status=order.status,
        content=content,
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '检测完成',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/quote', methods=['POST'])
@jwt_required()
def quote_receiveorder(id):
    """生成报价"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待报价(2)才能报价
    if order.status not in [2]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行报价操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    labor_cost = float(data.get('labor_cost', 0))
    material_cost = float(data.get('material_cost', 0))
    other_cost = float(data.get('other_cost', 0))

    order.quote_labor_cost = labor_cost
    order.quote_material_cost = material_cost
    order.quote_other_cost = other_cost
    order.quote_total = labor_cost + material_cost + other_cost
    order.total_amount = order.quote_total
    order.status = 3  # 待客户确认

    # 创建报价单关联
    items = data.get('items', [])
    if items:
        last_quote = QuoteOrder.query.order_by(QuoteOrder.id.desc()).first()
        quote_no = generate_code('QO', last_quote.id if last_quote else 0)
        quote_order = QuoteOrder(
            quote_no=quote_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            customer_phone=order.customer_phone,
            total_amount=order.quote_total,
            status=0,
            related_type='receive_order',
            related_id=id,
            created_by=user_id
        )
        db.session.add(quote_order)
        db.session.flush()

        for item in items:
            quote_item = QuoteOrderItem(
                quote_id=quote_order.id,
                product_name=item.get('product_name', ''),
                specification=item.get('specification', ''),
                brand=item.get('brand', ''),
                unit=item.get('unit', ''),
                quantity=item.get('quantity', 1),
                unit_price=item.get('unit_price', 0),
                subtotal=float(item.get('quantity', 1)) * float(item.get('unit_price', 0)),
                remark=item.get('remark', '')
            )
            db.session.add(quote_item)

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='生成报价',
        old_status=old_status,
        new_status=3,
        content=f'报价：人工费 {labor_cost}，材料费 {material_cost}，其他费用 {other_cost}，合计 {order.quote_total}',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '报价成功',
        'data': {
            'status': order.status,
            'status_text': RO_STATUS_MAP.get(order.status, '未知'),
            'quote_total': float(order.quote_total)
        }
    })


@app.route('/api/receiveorders/<int:id>/confirm', methods=['POST'])
@jwt_required()
def confirm_receiveorder(id):
    """客户确认/拒绝报价"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待客户确认(3)才能确认/拒绝
    if order.status not in [3]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行确认操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    confirmed = data.get('confirmed')

    if confirmed == 1:
        # 客户确认
        order.quote_confirmed = 1
        order.quote_confirm_time = datetime.now()
        order.status = 4  # 待领料/采购
        content = '客户已确认报价'
    elif confirmed == 2:
        # 客户拒绝
        order.quote_confirmed = 2
        order.quote_confirm_time = datetime.now()
        order.status = 14  # 已取消
        reject_reason = data.get('reject_reason', '')
        content = f'客户拒绝报价，原因：{reject_reason}'
    else:
        return jsonify({'code': 400, 'message': '参数错误，confirmed应为1(确认)或2(拒绝)'}), 400

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='客户确认',
        old_status=old_status,
        new_status=order.status,
        content=content,
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '操作成功',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/allocate', methods=['POST'])
@jwt_required()
def allocate_receiveorder(id):
    """领料/采购"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待领料/采购(4)才能领料
    if order.status not in [4]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行领料操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    items = data.get('items', [])
    if not items:
        return jsonify({'code': 400, 'message': '领料明细不能为空'}), 400

    all_in_stock = True  # 是否全部有库存
    out_items = []  # 有库存的配件
    purchase_items = []  # 需要采购的配件

    for item in items:
        product_id = item.get('product_id')
        quantity = float(item.get('quantity', 1))

        # 查询商品信息
        product = ProductInfo.query.get(product_id) if product_id else None

        # 查询库存
        stock = InventoryStock.query.filter_by(product_id=product_id).first()
        available_qty = float(stock.available_quantity or 0) if stock else 0

        part_info = {
            'product_id': product_id,
            'product_name': product.product_name if product else item.get('product_name', ''),
            'product_code': product.product_code if product else item.get('product_code', ''),
            'specification': product.specification if product else item.get('specification', ''),
            'unit_name': product.unit_name if product else item.get('unit_name', ''),
            'quantity': quantity,
            'unit_price': float(item.get('unit_price', 0)),
            'cost_price': float(product.cost_price or 0) if product else 0,
            'remark': item.get('remark', '')
        }
        part_info['total_price'] = part_info['quantity'] * part_info['unit_price']

        if available_qty >= quantity:
            # 有库存，记录为出库
            part_info['source'] = 1
            part_info['status'] = 1
            out_items.append(part_info)
        else:
            # 无库存，记录为采购
            part_info['source'] = 2
            part_info['status'] = 2
            purchase_items.append(part_info)
            all_in_stock = False

    # 创建出库单（有库存的配件）
    if out_items:
        last_out = InventoryOut.query.order_by(InventoryOut.id.desc()).first()
        out_no = generate_code('OUT', last_out.id if last_out else 0)
        out_order = InventoryOut(
            out_no=out_no,
            out_type=2,  # 维修领料
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            warehouse_id=1,
            warehouse_name='主仓库',
            status=1,  # 已审核（直接出库）
            auditor_id=user_id,
            auditor_name=user_name,
            audit_time=datetime.now(),
            remark=f'接件单{order.receive_no}维修领料',
            related_order_id=id,
            related_order_no=order.receive_no,
            created_by=user_id
        )
        db.session.add(out_order)
        db.session.flush()

        for item in out_items:
            out_item = InventoryOutItem(
                out_id=out_order.id,
                product_id=item['product_id'],
                product_code=item['product_code'],
                product_name=item['product_name'],
                specification=item['specification'],
                unit_name=item['unit_name'],
                quantity=item['quantity'],
                unit_price=item['unit_price'],
                total_price=item['total_price'],
                cost_price=item['cost_price'],
                remark=item['remark']
            )
            db.session.add(out_item)
            item['inventory_out_item_id'] = out_item.id

            # 扣减库存
            stock = InventoryStock.query.filter_by(product_id=item['product_id']).first()
            if stock:
                stock.quantity = float(stock.quantity or 0) - item['quantity']
                stock.available_quantity = float(stock.available_quantity or 0) - item['quantity']

    # 创建采购预订单（无库存的配件）
    if purchase_items:
        last_po = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
        po_no = generate_code('PO', last_po.id if last_po else 0)
        po_order = PurchaseOrder(
            order_no=po_no,
            supplier_id=data.get('supplier_id'),
            supplier_name=data.get('supplier_name', ''),
            order_date=datetime.now().date(),
            status=0,  # 待审核
            remark=f'接件单{order.receive_no}维修配件采购',
            created_by=user_id
        )
        db.session.add(po_order)
        db.session.flush()

        for item in purchase_items:
            po_item = PurchaseOrderItem(
                order_id=po_order.id,
                product_id=item['product_id'],
                product_name=item['product_name'],
                specification=item['specification'],
                unit=item['unit_name'],
                quantity=int(item['quantity']),
                price=item['cost_price'],
                amount=item['quantity'] * item['cost_price'],
                remark=item['remark']
            )
            db.session.add(po_item)
            item['purchase_order_item_id'] = po_item.id

    # 保存配件明细到接件单
    all_items = out_items + purchase_items
    for item in all_items:
        part = ReceiveOrderPart(
            receive_order_id=id,
            product_id=item['product_id'],
            product_name=item['product_name'],
            product_code=item['product_code'],
            specification=item['specification'],
            unit_name=item['unit_name'],
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            total_price=item['total_price'],
            cost_price=item['cost_price'],
            source=item['source'],
            inventory_out_item_id=item.get('inventory_out_item_id'),
            purchase_order_item_id=item.get('purchase_order_item_id'),
            status=item['status'],
            remark=item['remark']
        )
        db.session.add(part)

    # 更新状态（领料后总是进入维修中，采购可后续处理）
    old_status = order.status
    order.status = 5  # 进入维修中
    if all_in_stock:
        content = f'领料完成（全部有库存），共{len(all_items)}项配件，状态变更为维修中'
    else:
        content = f'领料完成，{len(out_items)}项已出库，{len(purchase_items)}项已创建采购单，状态变更为维修中'

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='领料/采购',
        old_status=old_status,
        new_status=order.status,
        content=content,
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '领料操作完成',
        'data': {
            'status': order.status,
            'status_text': RO_STATUS_MAP.get(order.status, '未知'),
            'out_count': len(out_items),
            'purchase_count': len(purchase_items)
        }
    })


@app.route('/api/receiveorders/<int:id>/finish', methods=['POST'])
@jwt_required()
def finish_receiveorder(id):
    """提交完工"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：维修中(5)才能提交完工
    if order.status not in [5]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许提交完工'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    order.finish_report = data.get('finish_report', '')
    finish_photos = data.get('finish_photos')
    if finish_photos is not None:
        order.finish_photos = json.dumps(finish_photos, ensure_ascii=False) if isinstance(finish_photos, (list, dict)) else str(finish_photos)
    order.status = 6  # 待测试

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='提交完工',
        old_status=old_status,
        new_status=6,
        content=f'工程师提交完工报告',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '完工提交成功',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/test', methods=['POST'])
@jwt_required()
def test_receiveorder(id):
    """测试设备"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待测试(6)才能测试
    if order.status not in [6]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行测试操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    test_result = data.get('test_result')
    test_remark = data.get('test_remark', '')

    order.test_result = test_result
    order.test_remark = test_remark

    if test_result == 1:
        # 测试通过
        order.status = 7  # 待取件
        content = '设备测试通过'
    elif test_result == 2:
        # 测试未通过
        order.status = 5  # 返回维修中
        content = f'设备测试未通过，原因：{test_remark}，返回维修中'
    else:
        return jsonify({'code': 400, 'message': '参数错误，test_result应为1(通过)或2(未通过)'}), 400

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='设备测试',
        old_status=old_status,
        new_status=order.status,
        content=content,
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '测试操作完成',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/notify', methods=['POST'])
@jwt_required()
def notify_receiveorder(id):
    """通知取件"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待取件(7)才能通知
    if order.status not in [7]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行通知操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    order.notify_time = datetime.now()
    order.notify_method = data.get('notify_method', '电话')
    # 状态保持7(待取件)

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='通知取件',
        old_status=order.status,
        new_status=order.status,
        content=f'已通过【{order.notify_method}】通知客户取件',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '通知成功',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/settle', methods=['POST'])
@jwt_required()
def settle_receiveorder(id):
    """完工结算 - 计算费用、创建销售单和财务/应收记录，状态从7(待取件)变为8(待结算)
    支持现金结算（直接入账）和签单结算（生成应收）"""
    try:
        order = ReceiveOrder.query.get(id)
        if not order:
            return jsonify({'code': 404, 'message': '接件单不存在'}), 404

        # 允许状态7(待取件)进行结算
        if order.status not in [7]:
            return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许结算，需为待取件'}), 400

        data = request.get_json()
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        
        # 结算方式：cash=现金结算, credit=签单结算
        settle_type = data.get('settle_type', 'cash')

        # 获取工单已领备件
        parts = ReceiveOrderPart.query.filter_by(receive_order_id=id).all()
        
        # 自动计算已用备件和未用备件
        used_parts = []
        unused_parts = []
        parts_cost_total = 0.0
        
        for part in parts:
            used_qty = float(data.get(f'used_qty_{part.id}', part.used_quantity or part.quantity or 0))
            unused_qty = float(part.quantity or 0) - used_qty
            
            part.used_quantity = used_qty
            
            if used_qty > 0:
                part.status = 1
                used_parts.append({
                    'product_name': part.product_name,
                    'specification': part.specification,
                    'quantity': used_qty,
                    'unit_price': float(part.unit_price or 0),
                    'total': used_qty * float(part.unit_price or 0)
                })
                parts_cost_total += used_qty * float(part.unit_price or 0)
            
            if unused_qty > 0:
                part.status = 2
                if part.product_id:
                    try:
                        product = Product.query.get(part.product_id)
                        if product:
                            product.stock = float(product.stock or 0) + unused_qty
                    except Exception as e:
                        logger.warning(f'退回库存失败: {str(e)}')
                unused_parts.append({
                    'product_name': part.product_name,
                    'specification': part.specification,
                    'quantity': unused_qty
                })

        # 更新费用信息
        labor_hours = data.get('labor_hours')
        labor_unit_price = data.get('labor_unit_price')
        labor_cost = data.get('labor_cost')
        other_cost = data.get('other_cost', 0)
        account_id = data.get('account_id')

        if labor_hours is not None:
            order.labor_hours = labor_hours
        if labor_unit_price is not None:
            order.labor_unit_price = labor_unit_price
        if labor_cost is not None:
            order.labor_cost = labor_cost
        elif labor_hours and labor_unit_price:
            order.labor_cost = float(labor_hours) * float(labor_unit_price)
        
        if data.get('parts_cost') is not None:
            order.parts_cost = data.get('parts_cost')
        else:
            order.parts_cost = parts_cost_total

        # 计算总费用
        total = (float(order.labor_cost or 0) + float(order.parts_cost or 0) +
                 float(order.material_cost or 0) + float(order.transport_cost or 0) +
                 float(other_cost or 0) + float(order.service_fee or 0))
        order.total_cost = total

        # 创建销售单关联
        last_so = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
        so_no = generate_code('XS', last_so.id if last_so else 0)
        sales_order = SalesOrder(
            order_no=so_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            customer_phone=order.customer_phone,
            customer_address=order.customer_address,
            order_date=datetime.now().date(),
            total_amount=total,
            total_quantity=1,
            actual_amount=total,
            paid_amount=0,
            salesperson_id=user_id,
            salesperson_name=user_name,
            status=1,
            remark=f'接件单{order.receive_no}结算生成',
            created_by=user_id
        )
        db.session.add(sales_order)
        db.session.flush()
        order.related_sales_id = sales_order.id

        # 根据结算方式处理财务
        finance_record = None
        receivable_record = None
        
        if settle_type == 'cash' and account_id:
            account = FinanceAccount.query.get(account_id)
            if account:
                balance_before = float(account.balance or 0)
                account.balance = balance_before + total

                finance_record = FinanceRecord(
                    account_id=account.id,
                    account_name=account.account_name,
                    record_type=1,
                    amount=total,
                    balance_before=balance_before,
                    balance_after=balance_before + total,
                    related_type='receive_order',
                    related_id=id,
                    related_no=order.receive_no,
                    remark=f'接件单现金结算：{order.receive_no}',
                    created_by=user_id
                )
                db.session.add(finance_record)
                db.session.flush()
                order.related_finance_id = finance_record.id
        elif settle_type == 'credit':
            today = datetime.now().strftime('%Y%m%d')
            prefix_ys = 'YS' + today
            last_receivable = FinanceReceivable.query.filter(FinanceReceivable.receivable_no.like(prefix_ys + '%')).order_by(FinanceReceivable.receivable_no.desc()).first()
            if last_receivable and last_receivable.receivable_no and len(last_receivable.receivable_no) > len(prefix_ys):
                seq_ys = int(last_receivable.receivable_no[len(prefix_ys):]) + 1
            else:
                seq_ys = 1
            receivable_no = prefix_ys + str(seq_ys).zfill(4)
            
            receivable_record = FinanceReceivable(
                receivable_no=receivable_no,
                related_type='receive_order',
                related_id=id,
                related_no=order.receive_no,
                customer_id=order.customer_id,
                customer_name=order.customer_name,
                total_amount=total,
                received_amount=0,
                remaining_amount=total,
                status=0,
                remark=f'接件单签单结算：{order.receive_no}',
                created_at=datetime.now(),
                updated_at=datetime.now()
            )
            db.session.add(receivable_record)
            db.session.flush()

        # 更新状态：7(待取件) -> 8(待结算)
        old_status = order.status
        order.status = 8
        order.settlement_status = 1
        order.settlement_account_id = account_id if settle_type == 'cash' else None
        order.settlement_time = datetime.now()

        # 记录日志
        settle_type_text = '现金结算' if settle_type == 'cash' else '签单结算'
        log_content = f'完工{settle_type_text}，总费用：{total}元'
        if used_parts:
            used_str = '、'.join([f'{p["product_name"]}x{p["quantity"]}' for p in used_parts])
            log_content += f'。已用配件：{used_str}'
        if unused_parts:
            unused_str = '、'.join([f'{p["product_name"]}x{p["quantity"]}' for p in unused_parts])
            log_content += f'。已退回配件：{unused_str}'
        
        log = ReceiveOrderLog(
            receive_order_id=id,
            action='完工结算',
            old_status=old_status,
            new_status=8,
            content=log_content,
            operator_id=user_id,
            operator_name=user_name
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({'code': 200, 'message': '结算成功', 'data': {
            'status': 8,
            'status_text': '待结算',
            'total_cost': total
        }})
    except Exception as e:
        db.session.rollback()
        logger.error(f'接件单结算失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'结算失败: {str(e)}'}), 500


@app.route('/api/receiveorders/<int:id>/complete', methods=['POST'])
@jwt_required()
def complete_receiveorder(id):
    """取件完成"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待结算(8)才能完成
    if order.status not in [8]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行完成操作，需先结算'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    order.status = 9  # 已完成
    order.complete_time = datetime.now()

    # 自动生成财务记录
    account_id = data.get('account_id')
    account = None
    if account_id:
        account = FinanceAccount.query.get(account_id)

    if account:
        balance_before = float(account.balance or 0)
        amount = float(order.quote_total or order.total_amount or 0)
        account.balance = balance_before + amount

        record = FinanceRecord(
            account_id=account.id,
            account_name=account.account_name,
            record_type=1,  # 收入
            amount=amount,
            balance_before=balance_before,
            balance_after=balance_before + amount,
            related_type='receive_order',
            related_id=id,
            related_no=order.receive_no,
            remark=f'接件单结算：{order.receive_no}',
            created_by=user_id
        )
        db.session.add(record)
        db.session.flush()
        order.finance_record_id = record.id
        order.paid_amount = amount

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='取件完成',
        old_status=old_status,
        new_status=8,
        content=f'客户取件完成，接件单结算金额：{order.paid_amount or order.quote_total or order.total_amount}',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '取件完成',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/external-send', methods=['POST'])
@jwt_required()
def external_send_receiveorder(id):
    """送修外店"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：检测中(1)或检测后无法维修(9)才能送修外店
    if order.status not in [1, 9]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行送修外店操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    order.receive_type = 2  # 外送修
    order.external_shop_id = data.get('external_shop_id')
    order.external_shop_name = data.get('external_shop_name')
    order.external_repair_reason = data.get('external_repair_reason', '')
    order.external_send_date = datetime.now().date()
    order.status = 9  # 送修外店

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='送修外店',
        old_status=old_status,
        new_status=9,
        content=f'送修外店：{order.external_shop_name}，原因：{order.external_repair_reason}',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '已送修外店',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/external-quote', methods=['POST'])
@jwt_required()
def external_quote_receiveorder(id):
    """外店报价"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：送修外店(9)才能外店报价
    if order.status not in [9]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行外店报价操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    order.external_quote = float(data.get('external_quote', 0))
    order.status = 10  # 外店已报价

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='外店报价',
        old_status=old_status,
        new_status=10,
        content=f'外店报价：{order.external_quote}元',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '外店报价已录入',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/customer-quote', methods=['POST'])
@jwt_required()
def customer_quote_receiveorder(id):
    """给客户报价（外店流程）"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：外店已报价(10)才能给客户报价
    if order.status not in [10]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行客户报价操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    labor_cost = float(data.get('labor_cost', 0))
    material_cost = float(data.get('material_cost', 0))
    other_cost = float(data.get('other_cost', 0))

    order.quote_labor_cost = labor_cost
    order.quote_material_cost = material_cost
    order.quote_other_cost = other_cost
    order.quote_total = labor_cost + material_cost + other_cost
    order.total_amount = order.quote_total
    order.status = 3  # 待客户确认

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='给客户报价',
        old_status=old_status,
        new_status=3,
        content=f'给客户报价（外店流程）：人工费 {labor_cost}，材料费 {material_cost}，其他费用 {other_cost}，合计 {order.quote_total}',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '客户报价已生成',
        'data': {
            'status': order.status,
            'status_text': RO_STATUS_MAP.get(order.status, '未知'),
            'quote_total': float(order.quote_total)
        }
    })


@app.route('/api/receiveorders/<int:id>/external-confirm', methods=['POST'])
@jwt_required()
def external_confirm_receiveorder(id):
    """确认送修（外店流程）"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：待外店维修(11)才能确认
    if order.status not in [11]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行确认送修操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    order.status = 12  # 外店维修中

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='确认送修',
        old_status=old_status,
        new_status=12,
        content=f'确认送修外店：{order.external_shop_name}，设备已送出',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '已确认送修',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/external-return', methods=['POST'])
@jwt_required()
def external_return_receiveorder(id):
    """取回设备（外店流程）"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：外店维修中(12)才能取回
    if order.status not in [12]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行取回操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    return_date = data.get('external_return_date')
    if return_date:
        order.external_return_date = datetime.strptime(return_date, '%Y-%m-%d').date() if isinstance(return_date, str) else return_date
    else:
        order.external_return_date = datetime.now().date()

    order.external_round = (order.external_round or 1) + 1
    order.status = 13  # 外店取回待测试

    # 记录到外店往返记录
    history = []
    if order.external_history:
        try:
            history = json.loads(order.external_history)
        except (json.JSONDecodeError, TypeError):
            history = []
    history.append({
        'round': order.external_round,
        'shop_name': order.external_shop_name,
        'send_date': str(order.external_send_date) if order.external_send_date else '',
        'return_date': str(order.external_return_date),
        'quote': float(order.external_quote or 0),
        'operator': user_name,
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })
    order.external_history = json.dumps(history, ensure_ascii=False)

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='取回设备',
        old_status=old_status,
        new_status=13,
        content=f'从外店取回设备，第{order.external_round}次往返，待测试',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '设备已取回',
        'data': {
            'status': order.status,
            'status_text': RO_STATUS_MAP.get(order.status, '未知'),
            'external_round': order.external_round
        }
    })


@app.route('/api/receiveorders/<int:id>/external-retest', methods=['POST'])
@jwt_required()
def external_retest_receiveorder(id):
    """外店取回后测试"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 验证当前状态：外店取回待测试(13)才能测试
    if order.status not in [13]:
        return jsonify({'code': 400, 'message': f'当前状态【{RO_STATUS_MAP.get(order.status, "未知")}】不允许执行测试操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    test_result = data.get('test_result')
    test_remark = data.get('test_remark', '')

    order.test_result = test_result
    order.test_remark = test_remark

    if test_result == 1:
        # 测试通过
        order.status = 7  # 待取件
        content = '外店取回后测试通过'
    elif test_result == 2:
        # 测试未通过，再次送修
        order.status = 9  # 送修外店
        content = f'外店取回后测试未通过，原因：{test_remark}，需再次送修外店'
    else:
        return jsonify({'code': 400, 'message': '参数错误，test_result应为1(通过)或2(未通过)'}), 400

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='外店取回后测试',
        old_status=old_status,
        new_status=order.status,
        content=content,
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '测试操作完成',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/cancel', methods=['POST'])
@jwt_required()
def cancel_receiveorder(id):
    """取消接件单"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    # 已完成的接件单不能取消
    if order.status in [8]:
        return jsonify({'code': 400, 'message': '已完成的接件单不能取消'}), 400

    # 已取消的不能重复取消
    if order.status == 14:
        return jsonify({'code': 400, 'message': '接件单已取消，请勿重复操作'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    old_status = order.status
    order.status = 14  # 已取消

    cancel_reason = data.get('cancel_reason', '') if data else ''

    # 记录日志
    log = ReceiveOrderLog(
        receive_order_id=id,
        action='取消',
        old_status=old_status,
        new_status=14,
        content=f'接件单已取消，原因：{cancel_reason}',
        operator_id=user_id,
        operator_name=user_name
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '接件单已取消',
        'data': {'status': order.status, 'status_text': RO_STATUS_MAP.get(order.status, '未知')}
    })


@app.route('/api/receiveorders/<int:id>/logs', methods=['GET'])
@jwt_required()
def get_receiveorder_logs(id):
    """获取接件单操作日志"""
    order = ReceiveOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '接件单不存在'}), 404

    logs = ReceiveOrderLog.query.filter_by(receive_order_id=id).order_by(
        ReceiveOrderLog.created_at.desc()
    ).limit(50).all()

    return jsonify({
        'code': 200,
        'data': [to_dict(l) for l in logs]
    })

# ============================================
# API路由 - 报价管理（全新模块）
# ============================================

@app.route('/api/quotes', methods=['GET'])
@jwt_required()
def get_quotes():
    """获取报价单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)

    query = QuoteOrder.query

    if keyword:
        query = query.filter(
            db.or_(
                QuoteOrder.quote_no.contains(keyword),
                QuoteOrder.customer_name.contains(keyword),
                QuoteOrder.customer_phone.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    orders = query.order_by(QuoteOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': o.id,
                'quote_no': o.quote_no,
                'customer_name': o.customer_name,
                'customer_phone': o.customer_phone,
                'total_amount': float(o.total_amount) if o.total_amount else 0.00,
                'status': o.status,
                'status_text': QUOTE_STATUS_MAP.get(o.status, '未知'),
                'quote_date': o.quote_date.strftime('%Y-%m-%d') if o.quote_date else None,
                'valid_until': o.valid_until.strftime('%Y-%m-%d') if o.valid_until else None,
                'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else None
            } for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/quotes/<int:id>', methods=['GET'])
@jwt_required()
def get_quote(id):
    """获取报价单详情"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    items = QuoteOrderItem.query.filter_by(quote_id=id).all()

    result = to_dict(order)
    result['status_text'] = QUOTE_STATUS_MAP.get(order.status, '未知')
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@app.route('/api/quotes', methods=['POST'])
@jwt_required()
def create_quote():
    """创建报价单"""
    data = request.get_json()
    user_id = get_jwt_identity()

    last_order = QuoteOrder.query.order_by(QuoteOrder.id.desc()).first()
    quote_no = generate_code('QT', last_order.id if last_order else 0)

    order = QuoteOrder(
        quote_no=quote_no,
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        customer_phone=data.get('customer_phone'),
        contact_name=data.get('contact_name'),
        address=data.get('address'),
        quote_date=data.get('quote_date'),
        valid_until=data.get('valid_until'),
        total_amount=0,
        remark=data.get('remark'),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    # 添加明细
    items = data.get('items', [])
    total_amount = 0
    for item_data in items:
        qty = float(item_data.get('quantity', 1))
        price = float(item_data.get('unit_price', 0))
        subtotal = qty * price
        total_amount += subtotal

        item = QuoteOrderItem(
            quote_id=order.id,
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            brand=item_data.get('brand'),
            unit=item_data.get('unit'),
            quantity=qty,
            unit_price=price,
            subtotal=subtotal,
            remark=item_data.get('remark')
        )
        db.session.add(item)

    order.total_amount = total_amount
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'quote_no': quote_no}})


@app.route('/api/quotes/<int:id>', methods=['PUT'])
@jwt_required()
def update_quote(id):
    """更新报价单"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待确认的报价单可以修改'}), 400

    data = request.get_json()

    for field in ['customer_id', 'customer_name', 'customer_phone',
                  'contact_name', 'address', 'quote_date', 'valid_until', 'remark']:
        if field in data:
            setattr(order, field, data[field])

    # 更新明细
    if 'items' in data:
        # 删除旧明细
        QuoteOrderItem.query.filter_by(quote_id=id).delete()
        total_amount = 0
        for item_data in data['items']:
            qty = float(item_data.get('quantity', 1))
            price = float(item_data.get('unit_price', 0))
            subtotal = qty * price
            total_amount += subtotal
            item = QuoteOrderItem(
                quote_id=id,
                product_name=item_data.get('product_name'),
                specification=item_data.get('specification'),
                brand=item_data.get('brand'),
                unit=item_data.get('unit'),
                quantity=qty,
                unit_price=price,
                subtotal=subtotal,
                remark=item_data.get('remark')
            )
            db.session.add(item)
        order.total_amount = total_amount

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/quotes/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_quote(id):
    """删除报价单"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    if order.status in [1, 3, 4, 5]:
        return jsonify({'code': 400, 'message': '该报价单已确认或已转换，不能删除'}), 400

    order.status = 2  # 标记为已失效
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/quotes/<int:id>/confirm', methods=['POST'])
@jwt_required()
def confirm_quote(id):
    """确认报价单"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待确认的报价单可以确认'}), 400

    order.status = 1
    db.session.commit()

    return jsonify({'code': 200, 'message': '报价单已确认'})


@app.route('/api/quotes/<int:id>/to-workorder', methods=['POST'])
@jwt_required()
def quote_to_workorder(id):
    """报价单转工单"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    if order.status != 1:
        return jsonify({'code': 400, 'message': '只有已确认的报价单可以转工单'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    # 生成工单号
    last_wo = WorkOrder.query.order_by(WorkOrder.id.desc()).first()
    wo_no = generate_code('WO', last_wo.id if last_wo else 0)

    wo = WorkOrder(
        wo_no=wo_no,
        wo_type=data.get('wo_type', '维修'),
        customer_id=order.customer_id,
        customer_name=order.customer_name,
        customer_phone=order.customer_phone,
        fault_desc=order.remark,
        labor_cost=data.get('labor_cost', 0),
        parts_cost=float(order.total_amount or 0) - float(data.get('labor_cost', 0)),
        total_cost=order.total_amount,
        status=0,
        created_by=user_id
    )

    db.session.add(wo)

    order.status = 3
    order.related_type = 'work_order'
    order.related_id = wo.id
    db.session.flush()

    db.session.commit()

    return jsonify({'code': 200, 'message': '已转工单', 'data': {'wo_id': wo.id, 'wo_no': wo_no}})


@app.route('/api/quotes/<int:id>/to-receive', methods=['POST'])
@jwt_required()
def quote_to_receive(id):
    """报价单转接件单"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    if order.status != 1:
        return jsonify({'code': 400, 'message': '只有已确认的报价单可以转接件单'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    last_ro = ReceiveOrder.query.order_by(ReceiveOrder.id.desc()).first()
    receive_no = generate_code('RO', last_ro.id if last_ro else 0)

    ro = ReceiveOrder(
        receive_no=receive_no,
        customer_id=order.customer_id,
        customer_name=order.customer_name,
        customer_phone=order.customer_phone,
        total_amount=order.total_amount,
        remark=order.remark,
        created_by=user_id
    )

    db.session.add(ro)

    order.status = 4
    order.related_type = 'receive_order'
    order.related_id = ro.id
    db.session.flush()

    db.session.commit()

    return jsonify({'code': 200, 'message': '已转接件单', 'data': {'receive_order_id': ro.id, 'receive_no': receive_no}})


@app.route('/api/quotes/<int:id>/to-sales', methods=['POST'])
@jwt_required()
def quote_to_sales(id):
    """报价单转销售单"""
    order = QuoteOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '报价单不存在'}), 404

    if order.status != 1:
        return jsonify({'code': 400, 'message': '只有已确认的报价单可以转销售单'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    # 创建出库单（销售出库）
    last_out = InventoryOut.query.order_by(InventoryOut.id.desc()).first()
    out_no = generate_code('OUT', last_out.id if last_out else 0)

    out_order = InventoryOut(
        out_no=out_no,
        out_type=1,  # 销售出库
        customer_id=order.customer_id,
        customer_name=order.customer_name,
        warehouse_id=data.get('warehouse_id', 1),
        warehouse_name=data.get('warehouse_name', '主仓库'),
        total_amount=order.total_amount,
        remark=f'报价单 {order.quote_no} 转销售',
        created_by=user_id
    )

    db.session.add(out_order)

    # 添加出库明细
    items = QuoteOrderItem.query.filter_by(quote_id=id).all()
    for item in items:
        out_item = InventoryOutItem(
            out_id=out_order.id,
            product_name=item.product_name,
            specification=item.specification,
            unit_name=item.unit,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total_price=item.subtotal,
            remark=item.remark
        )
        db.session.add(out_item)

    order.status = 5
    order.related_type = 'inventory_out'
    order.related_id = out_order.id
    db.session.flush()

    db.session.commit()

    return jsonify({'code': 200, 'message': '已转销售单', 'data': {'out_id': out_order.id, 'out_no': out_no}})

# ============================================
# API路由 - 设备管理（完善）
# ============================================

@app.route('/api/devices', methods=['GET'])
@jwt_required()
def get_devices():
    """获取客户设备列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    customer_id = request.args.get('customer_id', type=int)

    query = DeviceArchive.query.filter_by(status=1)

    if keyword:
        query = query.filter(
            db.or_(
                DeviceArchive.device_code.contains(keyword),
                DeviceArchive.device_type.contains(keyword),
                DeviceArchive.device_brand.contains(keyword),
                DeviceArchive.device_model.contains(keyword),
                DeviceArchive.device_sn.contains(keyword),
                DeviceArchive.device_name.contains(keyword)
            )
        )

    if customer_id:
        query = query.filter_by(customer_id=customer_id)

    total = query.count()
    devices = query.order_by(DeviceArchive.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    # 设备类型映射
    device_type_map = {
        'desktop': '台式机', 'laptop': '笔记本', 'server': '服务器',
        'printer': '打印机', 'network': '网络设备',
        'office_supplies': '办公文具', 'monitor': '监控设备', 'other': '其他'
    }

    # 获取所有相关的客户信息
    customer_ids = list(set(d.customer_id for d in devices.items if d.customer_id))
    customers = {}
    if customer_ids:
        customer_list = BaseCustomer.query.filter(BaseCustomer.id.in_(customer_ids)).all()
        customers = {c.id: c for c in customer_list}

    device_list = []
    for d in devices.items:
        item = to_dict(d)
        item['device_type_label'] = device_type_map.get(d.device_type, d.device_type or '未知')
        item['brand'] = d.device_brand
        item['model'] = d.device_model
        item['serial_number'] = d.device_sn
        # 关联客户信息
        customer = customers.get(d.customer_id)
        if customer:
            item['customer_name'] = customer.customer_name
            item['contact_name'] = customer.contact_name
            item['phone'] = customer.phone
        else:
            item['customer_name'] = ''
            item['contact_name'] = ''
            item['phone'] = ''
        device_list.append(item)

    return jsonify({
        'code': 200,
        'data': {
            'list': device_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/devices/<int:id>', methods=['GET'])
@jwt_required()
def get_device(id):
    """获取设备详情"""
    device = DeviceArchive.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '设备不存在'}), 404

    # 设备类型映射
    device_type_map = {
        'desktop': '台式机', 'laptop': '笔记本', 'server': '服务器',
        'printer': '打印机', 'network': '网络设备',
        'office_supplies': '办公文具', 'monitor': '监控设备', 'other': '其他'
    }

    data = to_dict(device)
    # 添加前端需要的字段映射
    data['brand'] = device.device_brand
    data['model'] = device.device_model
    data['serial_number'] = device.device_sn
    data['device_type_label'] = device_type_map.get(device.device_type, device.device_type or '未知')
    # 关联客户信息
    if device.customer_id:
        customer = BaseCustomer.query.get(device.customer_id)
        if customer:
            data['customer_name'] = customer.customer_name
            data['contact_name'] = customer.contact_name
            data['phone'] = customer.phone

    return jsonify({'code': 200, 'data': data})


@app.route('/api/devices', methods=['POST'])
@jwt_required()
def create_device():
    """创建客户设备"""
    data = request.get_json()

    last_device = DeviceArchive.query.order_by(DeviceArchive.id.desc()).first()
    device_code = generate_code('DEV', last_device.id if last_device else 0)

    # 处理日期字段：空字符串转为None
    def parse_date(val):
        if not val or str(val).strip() == '':
            return None
        return val

    device = DeviceArchive(
        device_code=device_code,
        customer_id=data.get('customer_id'),
        device_type=data.get('device_type'),
        device_name=data.get('device_name'),
        device_brand=data.get('brand') or data.get('device_brand'),
        device_model=data.get('model') or data.get('device_model'),
        device_sn=data.get('serial_number') or data.get('device_sn'),
        device_imei=data.get('device_imei'),
        device_password=data.get('device_password'),
        ip_address=data.get('ip_address'),
        port=data.get('port'),
        quantity=data.get('quantity', 1),
        cpu=data.get('cpu'),
        memory=data.get('memory'),
        disk=data.get('disk'),
        os=data.get('os'),
        os_version=data.get('os_version'),
        accessories=data.get('accessories'),
        account=data.get('account'),
        password=data.get('password'),
        password_remark=data.get('password_remark'),
        purchase_date=parse_date(data.get('purchase_date')),
        warranty_expire=parse_date(data.get('warranty_expire')),
        remark=data.get('remark')
    )

    db.session.add(device)

    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': device.id, 'device_code': device_code}})


@app.route('/api/devices/<int:id>', methods=['PUT'])
@jwt_required()
def update_device(id):
    """更新客户设备"""
    device = DeviceArchive.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '设备不存在'}), 404

    data = request.get_json()

    for field in ['customer_id', 'device_type', 'device_name',
                  'device_brand', 'brand', 'device_model', 'model',
                  'device_sn', 'serial_number', 'device_imei',
                  'device_password', 'ip_address', 'port', 'quantity',
                  'cpu', 'memory', 'disk', 'os', 'os_version', 'accessories',
                  'account', 'password', 'password_remark',
                  'purchase_date', 'warranty_expire', 'remark']:
        if field in data:
            val = data[field]
            # 字段名映射
            if field == 'brand':
                setattr(device, 'device_brand', val)
            elif field == 'model':
                setattr(device, 'device_model', val)
            elif field == 'serial_number':
                setattr(device, 'device_sn', val)
            else:
                setattr(device, field, val)

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/devices/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_device(id):
    """删除客户设备"""
    device = DeviceArchive.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '设备不存在'}), 404

    device.status = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/devices/own', methods=['GET'])
@jwt_required()
def get_own_devices():
    """获取自有设备列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)

    query = OwnDevice.query

    if keyword:
        query = query.filter(
            db.or_(
                OwnDevice.asset_no.contains(keyword),
                OwnDevice.device_type.contains(keyword),
                OwnDevice.device_model.contains(keyword),
                OwnDevice.serial_number.contains(keyword),
                OwnDevice.location.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    devices = query.order_by(OwnDevice.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(d, exclude=['password']) for d in devices.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/devices/own/<int:id>', methods=['GET'])
@jwt_required()
def get_own_device(id):
    """获取自有设备详情"""
    device = OwnDevice.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '自有设备不存在'}), 404
    return jsonify({'code': 200, 'data': to_dict(device, exclude=['password'])})


@app.route('/api/devices/own', methods=['POST'])
@jwt_required()
def create_own_device():
    """创建自有设备"""
    data = request.get_json()

    last_device = OwnDevice.query.order_by(OwnDevice.id.desc()).first()
    asset_no = generate_code('ASSET', last_device.id if last_device else 0)

    device = OwnDevice(
        asset_no=asset_no,
        device_type=data.get('device_type'),
        device_model=data.get('device_model'),
        serial_number=data.get('serial_number'),
        cpu=data.get('cpu'),
        memory=data.get('memory'),
        hard_disk=data.get('hard_disk'),
        system=data.get('system'),
        system_version=data.get('system_version'),
        accessories=data.get('accessories'),
        account=data.get('account'),
        password=generate_password_hash(data.get('password', '')) if data.get('password') else '',
        password_remark=data.get('password_remark'),
        purchase_date=data.get('purchase_date'),
        warranty_expire=data.get('warranty_expire'),
        location=data.get('location'),
        user_id=data.get('user_id'),
        cost=data.get('cost', 0),
        depreciation=data.get('depreciation', 0),
        remark=data.get('remark')
    )

    db.session.add(device)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': device.id, 'asset_no': asset_no}})


@app.route('/api/devices/own/<int:id>', methods=['PUT'])
@jwt_required()
def update_own_device(id):
    """更新自有设备"""
    device = OwnDevice.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '自有设备不存在'}), 404

    data = request.get_json()

    for field in ['device_type', 'device_model', 'serial_number', 'cpu', 'memory',
                  'hard_disk', 'system', 'system_version', 'accessories', 'account',
                  'password_remark', 'purchase_date', 'warranty_expire', 'location',
                  'user_id', 'cost', 'depreciation', 'status', 'remark']:
        if field in data:
            setattr(device, field, data[field])

    if data.get('password'):
        device.password = generate_password_hash(data['password'])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/devices/own/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_own_device(id):
    """删除自有设备"""
    device = OwnDevice.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '自有设备不存在'}), 404

    db.session.delete(device)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/devices/<int:id>/repair-history', methods=['GET'])
@jwt_required()
def get_device_repair_history(id):
    """获取设备维修历史"""
    device = DeviceArchive.query.get(id)
    if not device:
        return jsonify({'code': 404, 'message': '设备不存在'}), 404

    # 通过接件单设备明细查找关联工单
    rod_ids = db.session.query(ReceiveOrderDevice.id).filter(
        ReceiveOrderDevice.device_archive_id == id
    ).subquery()

    # 查找关联的工单
    work_orders = WorkOrder.query.filter(
        WorkOrder.receive_order_id.in_(
            db.session.query(ReceiveOrderDevice.receive_order_id).filter(
                ReceiveOrderDevice.device_archive_id == id
            )
        )
    ).order_by(WorkOrder.created_at.desc()).limit(20).all()

    # 也通过设备SN直接匹配工单
    if device.device_sn:
        sn_orders = WorkOrder.query.filter(
            WorkOrder.device_sn == device.device_sn
        ).order_by(WorkOrder.created_at.desc()).limit(20).all()
        # 合并去重
        existing_ids = {wo.id for wo in work_orders}
        for wo in sn_orders:
            if wo.id not in existing_ids:
                work_orders.append(wo)

    return jsonify({
        'code': 200,
        'data': [{
            'id': wo.id,
            'wo_no': wo.wo_no,
            'fault_desc': wo.fault_desc,
            'status': wo.status,
            'status_text': WO_STATUS_MAP.get(wo.status, '未知'),
            'total_cost': float(wo.total_cost) if wo.total_cost else 0.00,
            'created_at': wo.created_at.strftime('%Y-%m-%d %H:%M:%S') if wo.created_at else None,
            'actual_time': wo.actual_time.strftime('%Y-%m-%d %H:%M:%S') if wo.actual_time else None
        } for wo in work_orders]
    })

# ============================================
# API路由 - 发票管理（完善）
# ============================================

@app.route('/api/invoices', methods=['GET'])
@jwt_required()
def get_invoices():
    """获取发票列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    invoice_type = request.args.get('invoice_type', type=int)
    status = request.args.get('status', type=int)

    query = FinanceInvoice.query

    if keyword:
        query = query.filter(
            db.or_(
                FinanceInvoice.invoice_no.contains(keyword),
                FinanceInvoice.customer_name.contains(keyword)
            )
        )

    if invoice_type is not None:
        query = query.filter_by(invoice_type=invoice_type)

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    invoices = query.order_by(FinanceInvoice.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(i) for i in invoices.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/invoices/<int:id>', methods=['GET'])
@jwt_required()
def get_invoice(id):
    """获取发票详情"""
    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    return jsonify({'code': 200, 'data': to_dict(invoice)})


@app.route('/api/invoices', methods=['POST'])
@jwt_required()
def create_invoice():
    """创建发票"""
    data = request.get_json()
    user_id = get_jwt_identity()

    last_invoice = FinanceInvoice.query.order_by(FinanceInvoice.id.desc()).first()
    invoice_no = generate_code('INV', last_invoice.id if last_invoice else 0)

    invoice = FinanceInvoice(
        invoice_no=invoice_no,
        invoice_type=data.get('invoice_type', 1),
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        amount=data.get('amount', 0),
        tax_amount=data.get('tax_amount', 0),
        total_amount=data.get('total_amount', 0),
        related_type=data.get('related_type'),
        related_id=data.get('related_id'),
        related_no=data.get('related_no'),
        remark=data.get('remark'),
        created_by=user_id
    )

    db.session.add(invoice)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': invoice.id, 'invoice_no': invoice_no}})


@app.route('/api/invoices/<int:id>', methods=['PUT'])
@jwt_required()
def update_invoice(id):
    """更新发票"""
    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '已作废的发票不能修改'}), 400

    data = request.get_json()

    for field in ['invoice_type', 'customer_id', 'customer_name',
                  'amount', 'tax_amount', 'total_amount',
                  'related_type', 'related_id', 'related_no', 'remark']:
        if field in data:
            setattr(invoice, field, data[field])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/invoices/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_invoice(id):
    """作废发票"""
    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '发票已作废'}), 400

    invoice.status = 2
    db.session.commit()
    return jsonify({'code': 200, 'message': '发票已作废'})


@app.route('/api/invoices/<int:id>/cancel', methods=['POST'])
@jwt_required()
def cancel_invoice(id):
    """作废发票"""
    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '发票已作废'}), 400

    invoice.status = 2
    db.session.commit()
    return jsonify({'code': 200, 'message': '发票已作废'})


@app.route('/api/invoices/<int:id>/red-flush', methods=['POST'])
@jwt_required()
def red_flush_invoice(id):
    """红冲发票"""
    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '已作废的发票不能红冲'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    # 创建红冲发票（负数金额）
    last_invoice = FinanceInvoice.query.order_by(FinanceInvoice.id.desc()).first()
    red_no = generate_code('RINV', last_invoice.id if last_invoice else 0)

    red_invoice = FinanceInvoice(
        invoice_no=red_no,
        invoice_type=invoice.invoice_type,
        customer_id=invoice.customer_id,
        customer_name=invoice.customer_name,
        amount=-(float(invoice.amount or 0)),
        tax_amount=-(float(invoice.tax_amount or 0)),
        total_amount=-(float(invoice.total_amount or 0)),
        related_type='red_flush',
        related_id=id,
        related_no=invoice.invoice_no,
        remark=f'红冲发票：{invoice.invoice_no}。{data.get("remark", "")}',
        status=1,
        created_by=user_id
    )

    db.session.add(red_invoice)

    # 原发票标记为已红冲
    invoice.status = 2
    invoice.remark = (invoice.remark or '') + f' [已红冲，红冲发票号：{red_no}]'

    db.session.commit()

    return jsonify({'code': 200, 'message': '红冲成功', 'data': {'red_invoice_id': red_invoice.id, 'red_no': red_no}})

# ============================================
# API路由 - 库存盘点（完善）
# ============================================

@app.route('/api/inventory/check/<int:id>', methods=['GET'])
@jwt_required()
def get_inventory_check(id):
    """获取盘点单详情"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    items = InventoryCheckItem.query.filter_by(check_id=id).all()

    result = to_dict(order)
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@app.route('/api/inventory/check/<int:id>', methods=['PUT'])
@jwt_required()
def update_inventory_check(id):
    """更新盘点单（录入实际数量）"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status != 0 and order.status != 1:
        return jsonify({'code': 400, 'message': '盘点单状态不允许修改'}), 400

    data = request.get_json()

    if 'remark' in data:
        order.remark = data['remark']

    # 更新盘点状态为盘点中
    if order.status == 0:
        order.status = 1

    # 更新明细的实际数量
    if 'items' in data:
        total_diff = 0
        total_diff_amount = 0
        for item_data in data['items']:
            item_id = item_data.get('id')
            if item_id:
                item = InventoryCheckItem.query.get(item_id)
                if item and item.check_id == id:
                    actual_qty = float(item_data.get('actual_quantity', item.system_quantity or 0))
                    item.actual_quantity = actual_qty
                    item.diff_quantity = actual_qty - float(item.system_quantity or 0)
                    item.diff_amount = float(item.diff_quantity or 0) * float(item.cost_price or 0)
                    total_diff += abs(float(item.diff_quantity or 0))
                    total_diff_amount += abs(float(item.diff_amount or 0))

        order.diff_quantity = total_diff
        order.diff_amount = total_diff_amount

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/inventory/check/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_inventory_check(id):
    """审核盘点单（自动更新库存）"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status != 1:
        return jsonify({'code': 400, 'message': '只有盘点中的盘点单可以审核'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 更新库存
    items = InventoryCheckItem.query.filter_by(check_id=id).all()
    for item in items:
        if item.diff_quantity and float(item.diff_quantity) != 0:
            stock = InventoryStock.query.filter_by(
                product_id=item.product_id,
                warehouse_id=order.warehouse_id
            ).first()

            if stock:
                diff = float(item.diff_quantity)
                before_qty = float(stock.quantity or 0)
                stock.quantity = before_qty + diff
                stock.available_quantity = float(stock.available_quantity or 0) + diff
                after_qty = stock.quantity

                # 写入 InventoryLog
                log = InventoryLog(
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    warehouse_id=order.warehouse_id,
                    warehouse_name=order.warehouse_name,
                    change_type='adjust',
                    order_type='盘点调整',
                    order_id=order.id,
                    order_no=order.check_no,
                    quantity_change=diff,
                    before_quantity=before_qty,
                    after_quantity=after_qty,
                    cost_price=float(item.cost_price or 0),
                    amount=round(abs(diff) * float(item.cost_price or 0), 2),
                    related_party='',
                    operator_id=user_id,
                    operator_name=user_name,
                    remark=f'盘点调整：系统{before_qty}，实际{float(item.actual_quantity or 0)}，差异{diff}'
                )
                db.session.add(log)

                # 更新商品库存
                product = ProductInfo.query.get(item.product_id)
                if product:
                    product.current_stock = (product.current_stock or 0) + diff

    order.status = 2  # 已完成
    db.session.commit()

    return jsonify({'code': 200, 'message': '审核成功，库存已更新'})


@app.route('/api/inventory/check/<int:id>/items', methods=['GET'])
@jwt_required()
def get_inventory_check_items(id):
    """获取盘点明细"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    items = InventoryCheckItem.query.filter_by(check_id=id).all()

    return jsonify({
        'code': 200,
        'data': [{
            'id': i.id,
            'product_id': i.product_id,
            'product_code': i.product_code,
            'product_name': i.product_name,
            'specification': i.specification,
            'unit_name': i.unit_name,
            'system_quantity': float(i.system_quantity) if i.system_quantity else 0,
            'actual_quantity': float(i.actual_quantity) if i.actual_quantity else 0,
            'diff_quantity': float(i.diff_quantity) if i.diff_quantity else 0,
            'cost_price': float(i.cost_price) if i.cost_price else 0,
            'diff_amount': float(i.diff_amount) if i.diff_amount else 0,
            'remark': i.remark
        } for i in items]
    })


@app.route('/api/inventory/check/<int:id>/items', methods=['PUT'])
@jwt_required()
def update_inventory_check_items(id):
    """保存盘点明细（批量更新实际数量）"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '已完成的盘点单不能修改'}), 400

    data = request.get_json()
    items_data = data.get('items', [])

    total_diff = 0
    total_diff_amount = 0

    for item_data in items_data:
        item_id = item_data.get('id')
        if item_id:
            item = InventoryCheckItem.query.get(item_id)
            if item and item.check_id == id:
                actual_qty = float(item_data.get('actual_quantity', item.system_quantity or 0))
                item.actual_quantity = actual_qty
                item.diff_quantity = actual_qty - float(item.system_quantity or 0)
                item.diff_amount = float(item.diff_quantity or 0) * float(item.cost_price or 0)
                total_diff += abs(float(item.diff_quantity or 0))
                total_diff_amount += abs(float(item.diff_amount or 0))

    # 更新盘点单状态和汇总
    if order.status == 0:
        order.status = 1  # 盘点中
    order.diff_quantity = total_diff
    order.diff_amount = total_diff_amount

    db.session.commit()
    return jsonify({'code': 200, 'message': '保存成功'})


@app.route('/api/inventory/check/<int:id>/complete', methods=['POST'])
@jwt_required()
def complete_inventory_check(id):
    """完成盘点（审核并更新库存）"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '盘点单已完成'}), 400

    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 更新库存
    items = InventoryCheckItem.query.filter_by(check_id=id).all()
    for item in items:
        if item.diff_quantity and float(item.diff_quantity) != 0:
            stock = InventoryStock.query.filter_by(
                product_id=item.product_id,
                warehouse_id=order.warehouse_id
            ).first()

            if stock:
                diff = float(item.diff_quantity)
                before_qty = float(stock.quantity or 0)
                stock.quantity = before_qty + diff
                stock.available_quantity = float(stock.available_quantity or 0) + diff
                after_qty = stock.quantity

                # 写入库存日志
                log = InventoryLog(
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    warehouse_id=order.warehouse_id,
                    warehouse_name=order.warehouse_name,
                    change_type='adjust',
                    order_type='盘点调整',
                    order_id=order.id,
                    order_no=order.check_no,
                    quantity_change=diff,
                    before_quantity=before_qty,
                    after_quantity=after_qty,
                    cost_price=float(item.cost_price or 0),
                    amount=round(abs(diff) * float(item.cost_price or 0), 2),
                    related_party='',
                    operator_id=user_id,
                    operator_name=user_name,
                    remark=f'盘点调整：系统{before_qty}，实际{float(item.actual_quantity or 0)}，差异{diff}'
                )
                db.session.add(log)

                # 更新商品库存
                product = ProductInfo.query.get(item.product_id)
                if product:
                    product.current_stock = (product.current_stock or 0) + diff

    order.status = 2  # 已完成
    db.session.commit()

    return jsonify({'code': 200, 'message': '盘点完成，库存已更新'})


@app.route('/api/inventory/check/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_inventory_check(id):
    """取消/删除盘点单"""
    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '已完成的盘点单不能取消'}), 400

    # 删除盘点明细
    InventoryCheckItem.query.filter_by(check_id=id).delete()

    # 删除盘点单
    db.session.delete(order)
    db.session.commit()

    return jsonify({'code': 200, 'message': '取消成功'})


# ============================================
# API路由 - 采购单管理
# ============================================

@app.route('/api/purchase/orders', methods=['GET'])
@jwt_required()
def get_purchase_orders():
    """获取采购单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    supplier_id = request.args.get('supplier_id', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = PurchaseOrder.query

    if keyword:
        query = query.filter(
            db.or_(
                PurchaseOrder.order_no.contains(keyword),
                PurchaseOrder.supplier_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)

    if start_date:
        query = query.filter(PurchaseOrder.order_date >= start_date)

    if end_date:
        query = query.filter(PurchaseOrder.order_date <= end_date)

    total = query.count()
    orders = query.order_by(PurchaseOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    # 批量查询操作人名称
    user_ids = [o.created_by for o in orders.items if o.created_by]
    users_map = {}
    if user_ids:
        users = SysUser.query.filter(SysUser.id.in_(user_ids)).all()
        users_map = {u.id: (u.real_name or u.username) for u in users}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': o.id,
                'purchase_no': o.order_no,
                'order_no': o.order_no,
                'supplier_id': o.supplier_id,
                'supplier_name': o.supplier_name,
                'purchase_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else None,
                'order_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else None,
                'delivery_date': o.delivery_date.strftime('%Y-%m-%d') if o.delivery_date else None,
                'total_amount': float(o.total_amount) if o.total_amount else 0,
                'paid_amount': 0,
                'total_quantity': o.total_quantity,
                'status': o.status,
                'status_text': PO_STATUS_MAP.get(o.status, '未知'),
                'has_invoice': o.has_invoice,
                'operator_name': users_map.get(o.created_by, ''),
                'remark': o.remark,
                'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else None
            } for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/purchase/orders/<int:id>', methods=['GET'])
@jwt_required()
def get_purchase_order(id):
    """获取采购单详情"""
    order = PurchaseOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '采购单不存在'}), 404

    items = PurchaseOrderItem.query.filter_by(order_id=id).all()

    result = to_dict(order)
    result['status_text'] = PO_STATUS_MAP.get(order.status, '未知')
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@app.route('/api/purchase/orders', methods=['POST'])
@jwt_required()
def create_purchase_order():
    """创建采购单"""
    data = request.get_json()
    user_id = get_jwt_identity()

    # 生成采购单号
    last_order = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
    order_no = generate_code('PO', last_order.id if last_order else 0)

    order = PurchaseOrder(
        order_no=order_no,
        supplier_id=data.get('supplier_id'),
        supplier_name=data.get('supplier_name'),
        order_date=data.get('order_date'),
        delivery_date=data.get('delivery_date'),
        remark=data.get('remark'),
        has_invoice=data.get('has_invoice', 0),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    # 添加明细
    items = data.get('items', [])
    total_amount = 0
    total_quantity = 0

    for item_data in items:
        qty = int(item_data.get('quantity', 0))
        price = float(item_data.get('price') or item_data.get('unit_price') or 0)
        amount = qty * price
        total_amount += amount
        total_quantity += qty

        item = PurchaseOrderItem(
            order_id=order.id,
            product_id=item_data.get('product_id'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit=item_data.get('unit'),
            quantity=qty,
            price=price,
            amount=amount,
            remark=item_data.get('remark')
        )
        db.session.add(item)

    order.total_amount = total_amount
    order.total_quantity = total_quantity
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'order_no': order_no}})


@app.route('/api/purchase/orders/<int:id>', methods=['PUT'])
@jwt_required()
def update_purchase_order(id):
    """更新采购单"""
    order = PurchaseOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '采购单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核的采购单可以修改'}), 400

    data = request.get_json()

    for field in ['supplier_id', 'supplier_name', 'order_date', 'delivery_date', 'remark']:
        if field in data:
            setattr(order, field, data[field])

    if 'has_invoice' in data:
        order.has_invoice = data['has_invoice']

    # 更新明细
    if 'items' in data:
        # 删除旧明细
        PurchaseOrderItem.query.filter_by(order_id=id).delete()
        total_amount = 0
        total_quantity = 0

        for item_data in data['items']:
            qty = int(item_data.get('quantity', 0))
            price = float(item_data.get('price') or item_data.get('unit_price') or 0)
            amount = qty * price
            total_amount += amount
            total_quantity += qty

            item = PurchaseOrderItem(
                order_id=id,
                product_id=item_data.get('product_id'),
                product_name=item_data.get('product_name'),
                specification=item_data.get('specification'),
                unit=item_data.get('unit'),
                quantity=qty,
                price=price,
                amount=amount,
                remark=item_data.get('remark')
            )
            db.session.add(item)

        order.total_amount = total_amount
        order.total_quantity = total_quantity

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/purchase/orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_purchase_order(id):
    """删除采购单"""
    order = PurchaseOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '采购单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '已完成的采购单不能删除'}), 400

    # 软删除
    order.status = 3  # 已取消
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/purchase/orders/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_purchase_order(id):
    """审核采购单 - 自动入库、更新库存、生成应付账款和财务流水"""
    order = PurchaseOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '采购单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核的采购单可以审核'}), 400

    try:
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        today = datetime.now().strftime('%Y%m%d')

        # 1. 更新采购单状态
        order.status = 1  # 已审核

        # 2. 检查是否已有入库单（避免重复创建）
        existing_in = InventoryIn.query.filter_by(related_order_id=id).first()
        if existing_in:
            db.session.commit()
            return jsonify({'code': 200, 'message': '审核成功（入库单已存在）'})

        # 3. 获取采购单明细
        items = PurchaseOrderItem.query.filter_by(order_id=id).all()
        if not items:
            db.session.commit()
            return jsonify({'code': 200, 'message': '审核成功（无明细商品）'})

        # 4. 生成入库单号 RK + YYYYMMDD + 4位序号
        prefix_rk = 'RK' + today
        last_in = InventoryIn.query.filter(InventoryIn.in_no.like(prefix_rk + '%')).order_by(InventoryIn.in_no.desc()).first()
        if last_in and last_in.in_no and len(last_in.in_no) > len(prefix_rk):
            seq = int(last_in.in_no[len(prefix_rk):]) + 1
        else:
            seq = 1
        in_no = prefix_rk + str(seq).zfill(4)

        # 5. 创建入库单
        inventory_in = InventoryIn(
            in_no=in_no,
            in_type=1,  # 采购入库
            supplier_id=order.supplier_id,
            supplier_name=order.supplier_name,
            warehouse_id=1,
            warehouse_name='主仓库',
            total_quantity=sum(item.quantity or 0 for item in items),
            total_amount=float(order.total_amount or 0),
            status=2,  # 直接入库
            auditor_id=user_id,
            auditor_name=user_name,
            audit_time=datetime.now(),
            related_order_id=id,
            related_order_no=order.order_no,
            created_by=user_id
        )
        db.session.add(inventory_in)
        db.session.flush()  # 获取 in_id

        # 6. 创建入库单明细 + 更新库存 + 写入库存日志 + 更新商品现存量
        for item in items:
            # 入库明细
            in_item = InventoryInItem(
                in_id=inventory_in.id,
                product_id=item.product_id,
                product_code=None,
                product_name=item.product_name,
                specification=item.specification,
                unit_name=item.unit,
                quantity=item.quantity,
                unit_price=item.price,
                total_price=item.amount,
                cost_price=item.price
            )
            db.session.add(in_item)

            # 更新库存
            stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
            before_qty = stock.quantity if stock else 0
            if stock:
                stock.quantity = (stock.quantity or 0) + (item.quantity or 0)
                stock.available_quantity = (stock.available_quantity or 0) + (item.quantity or 0)
                if item.price:
                    stock.cost_price = item.price
            else:
                # 自动创建库存记录
                product = ProductInfo.query.get(item.product_id) if item.product_id else None
                stock = InventoryStock(
                    product_id=item.product_id,
                    product_code=product.product_code if product else None,
                    product_name=item.product_name,
                    warehouse_id=1,
                    warehouse_name='主仓库',
                    quantity=item.quantity or 0,
                    available_quantity=item.quantity or 0,
                    cost_price=item.price or 0
                )
                db.session.add(stock)
                before_qty = 0

            after_qty = (stock.quantity or 0)

            # 写入库存日志
            inv_log = InventoryLog(
                product_id=item.product_id,
                product_code=stock.product_code if stock else None,
                product_name=item.product_name,
                warehouse_id=1,
                warehouse_name='主仓库',
                change_type='in',
                order_type='采购入库',
                order_id=inventory_in.id,
                order_no=in_no,
                quantity_change=item.quantity or 0,
                before_quantity=before_qty,
                after_quantity=after_qty,
                cost_price=item.price,
                amount=item.amount,
                related_party=order.supplier_name,
                operator_id=user_id,
                operator_name=user_name,
                remark=f'采购单{order.order_no}审核自动入库'
            )
            db.session.add(inv_log)

            # 更新商品现存量
            if item.product_id:
                product_info = ProductInfo.query.get(item.product_id)
                if product_info:
                    product_info.current_stock = (product_info.current_stock or 0) + (item.quantity or 0)

        # 7. 生成应付编号 YF + YYYYMMDD + 4位序号
        prefix_yf = 'YF' + today
        last_payable = FinancePayable.query.filter(FinancePayable.payable_no.like(prefix_yf + '%')).order_by(FinancePayable.payable_no.desc()).first()
        if last_payable and last_payable.payable_no and len(last_payable.payable_no) > len(prefix_yf):
            seq_yf = int(last_payable.payable_no[len(prefix_yf):]) + 1
        else:
            seq_yf = 1
        payable_no = prefix_yf + str(seq_yf).zfill(4)

        # 8. 生成应付账款
        total_amount = float(order.total_amount or 0)
        payable = FinancePayable(
            payable_no=payable_no,
            related_type='purchase',
            related_id=id,
            related_no=order.order_no,
            supplier_id=order.supplier_id,
            supplier_name=order.supplier_name,
            total_amount=total_amount,
            paid_amount=0,
            remaining_amount=total_amount,
            status=0,  # 待付款
            remark=f'采购单{order.order_no}审核自动生成'
        )
        db.session.add(payable)

        # 9. 生成财务流水（支出）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=2,  # 支出
            amount=total_amount,
            balance_before=0,
            balance_after=0,
            related_type='purchase',
            related_id=id,
            related_no=order.order_no,
            remark=f'采购单{order.order_no}审核，应付金额{total_amount}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        db.session.commit()
        return jsonify({'code': 200, 'message': '审核成功，已自动入库并生成应付账款'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'审核失败：{str(e)}'}), 500


# ============================================
# API路由 - 销售单管理
# ============================================

@app.route('/api/sales/orders', methods=['GET'])
@jwt_required()
def get_sales_orders():
    """获取销售单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    customer_id = request.args.get('customer_id', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    has_invoice = request.args.get('has_invoice', type=int)
    has_receipt = request.args.get('has_receipt', type=int)

    query = SalesOrder.query

    if keyword:
        query = query.filter(
            db.or_(
                SalesOrder.order_no.contains(keyword),
                SalesOrder.customer_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if customer_id:
        query = query.filter_by(customer_id=customer_id)

    if start_date:
        query = query.filter(SalesOrder.order_date >= start_date)

    if end_date:
        query = query.filter(SalesOrder.order_date <= end_date)

    if has_invoice is not None:
        query = query.filter_by(has_invoice=has_invoice)

    if has_receipt is not None:
        query = query.filter_by(has_receipt=has_receipt)

    total = query.count()
    orders = query.order_by(SalesOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    # 批量查询操作人名称
    user_ids = [o.created_by for o in orders.items if o.created_by]
    users_map = {}
    if user_ids:
        users = SysUser.query.filter(SysUser.id.in_(user_ids)).all()
        users_map = {u.id: (u.real_name or u.username) for u in users}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': o.id,
                'sales_no': o.order_no,
                'order_no': o.order_no,
                'customer_id': o.customer_id,
                'customer_name': o.customer_name,
                'sales_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else None,
                'order_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else None,
                'total_amount': float(o.total_amount) if o.total_amount else 0,
                'received_amount': 0,
                'total_quantity': o.total_quantity,
                'status': o.status,
                'status_text': SO_STATUS_MAP.get(o.status, '未知'),
                'remark': o.remark,
                'has_invoice': o.has_invoice or 0,
                'has_receipt': o.has_receipt or 0,
                'operator_name': users_map.get(o.created_by, ''),
                'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else None
            } for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/sales/orders/<int:id>', methods=['GET'])
@jwt_required()
def get_sales_order(id):
    """获取销售单详情"""
    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    items = SalesOrderItem.query.filter_by(order_id=id).all()

    result = to_dict(order)
    result['status_text'] = SO_STATUS_MAP.get(order.status, '未知')
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@app.route('/api/sales/orders', methods=['POST'])
@jwt_required()
def create_sales_order():
    """创建销售单"""
    data = request.get_json()
    user_id = get_jwt_identity()

    # 生成销售单号
    last_order = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
    order_no = generate_code('SO', last_order.id if last_order else 0)

    order = SalesOrder(
        order_no=order_no,
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        order_date=data.get('order_date'),
        remark=data.get('remark'),
        has_invoice=data.get('has_invoice', 0),
        has_receipt=data.get('has_receipt', 0),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    # 添加明细
    items = data.get('items', [])
    total_amount = 0
    total_quantity = 0

    for item_data in items:
        qty = int(item_data.get('quantity', 0))
        price = float(item_data.get('price', 0))
        amount = qty * price
        total_amount += amount
        total_quantity += qty

        item = SalesOrderItem(
            order_id=order.id,
            product_id=item_data.get('product_id'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit=item_data.get('unit'),
            quantity=qty,
            price=price,
            amount=amount,
            remark=item_data.get('remark')
        )
        db.session.add(item)

    order.total_amount = total_amount
    order.total_quantity = total_quantity
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'order_no': order_no}})


@app.route('/api/sales/orders/<int:id>', methods=['PUT'])
@jwt_required()
def update_sales_order(id):
    """更新销售单"""
    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    if order.status not in [0, 1]:
        return jsonify({'code': 400, 'message': '该状态的销售单不能修改'}), 400

    data = request.get_json()

    for field in ['customer_id', 'customer_name', 'order_date', 'remark', 'has_invoice', 'has_receipt']:
        if field in data:
            setattr(order, field, data[field])

    # 更新明细
    if 'items' in data:
        # 删除旧明细
        SalesOrderItem.query.filter_by(order_id=id).delete()
        total_amount = 0
        total_quantity = 0

        for item_data in data['items']:
            qty = int(item_data.get('quantity', 0))
            price = float(item_data.get('price', 0))
            amount = qty * price
            total_amount += amount
            total_quantity += qty

            item = SalesOrderItem(
                order_id=id,
                product_id=item_data.get('product_id'),
                product_name=item_data.get('product_name'),
                specification=item_data.get('specification'),
                unit=item_data.get('unit'),
                quantity=qty,
                price=price,
                amount=amount,
                remark=item_data.get('remark')
            )
            db.session.add(item)

        order.total_amount = total_amount
        order.total_quantity = total_quantity

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/sales/orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_sales_order(id):
    """删除销售单"""
    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    if order.status == 3:
        return jsonify({'code': 400, 'message': '已完成的销售单不能删除'}), 400

    # 软删除
    order.status = 4  # 已取消
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/sales/orders/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_sales_order(id):
    """审核销售单 - 自动出库、扣减库存、生成应收账款和财务流水"""
    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核的销售单可以审核'}), 400

    try:
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        today = datetime.now().strftime('%Y%m%d')

        # 1. 更新销售单状态为已审核
        order.status = 1

        # 2. 检查是否已有出库单（避免重复创建）
        existing_out = InventoryOut.query.filter_by(related_order_id=id).first()
        if existing_out:
            order.status = 2  # 已出库
            db.session.commit()
            return jsonify({'code': 200, 'message': '审核成功（出库单已存在）'})

        # 3. 获取销售单明细
        items = SalesOrderItem.query.filter_by(order_id=id).all()
        if not items:
            order.status = 2
            db.session.commit()
            return jsonify({'code': 200, 'message': '审核成功（无明细商品）'})

        # 4. 检查库存是否充足
        for item in items:
            stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
            available = (stock.available_quantity or 0) if stock else 0
            if available < (item.quantity or 0):
                return jsonify({'code': 400, 'message': f'商品"{item.product_name}"库存不足，可用库存{available}，需要{item.quantity}'}), 400

        # 5. 生成出库单号 CK + YYYYMMDD + 4位序号
        prefix_ck = 'CK' + today
        last_out = InventoryOut.query.filter(InventoryOut.out_no.like(prefix_ck + '%')).order_by(InventoryOut.out_no.desc()).first()
        if last_out and last_out.out_no and len(last_out.out_no) > len(prefix_ck):
            seq = int(last_out.out_no[len(prefix_ck):]) + 1
        else:
            seq = 1
        out_no = prefix_ck + str(seq).zfill(4)

        # 6. 创建出库单
        inventory_out = InventoryOut(
            out_no=out_no,
            out_type=1,  # 销售出库
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            warehouse_id=1,
            warehouse_name='主仓库',
            total_quantity=sum(item.quantity or 0 for item in items),
            total_amount=float(order.total_amount or 0),
            status=2,  # 直接出库
            auditor_id=user_id,
            auditor_name=user_name,
            audit_time=datetime.now(),
            related_order_id=id,
            related_order_no=order.order_no,
            created_by=user_id
        )
        db.session.add(inventory_out)
        db.session.flush()  # 获取 out_id

        # 7. 创建出库单明细 + 扣减库存 + 写入库存日志 + 更新商品现存量
        for item in items:
            # 出库明细
            out_item = InventoryOutItem(
                out_id=inventory_out.id,
                product_id=item.product_id,
                product_code=None,
                product_name=item.product_name,
                specification=item.specification,
                unit_name=item.unit,
                quantity=item.quantity,
                unit_price=item.price,
                total_price=item.amount,
                cost_price=item.price
            )
            db.session.add(out_item)

            # 扣减库存
            stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
            before_qty = stock.quantity if stock else 0
            if stock:
                stock.quantity = (stock.quantity or 0) - (item.quantity or 0)
                stock.available_quantity = (stock.available_quantity or 0) - (item.quantity or 0)
            after_qty = (stock.quantity or 0) if stock else 0

            # 写入库存日志
            inv_log = InventoryLog(
                product_id=item.product_id,
                product_code=stock.product_code if stock else None,
                product_name=item.product_name,
                warehouse_id=1,
                warehouse_name='主仓库',
                change_type='out',
                order_type='销售出库',
                order_id=inventory_out.id,
                order_no=out_no,
                quantity_change=-(item.quantity or 0),
                before_quantity=before_qty,
                after_quantity=after_qty,
                cost_price=item.price,
                amount=item.amount,
                related_party=order.customer_name,
                operator_id=user_id,
                operator_name=user_name,
                remark=f'销售单{order.order_no}审核自动出库'
            )
            db.session.add(inv_log)

            # 更新商品现存量
            if item.product_id:
                product_info = ProductInfo.query.get(item.product_id)
                if product_info:
                    product_info.current_stock = max(0, (product_info.current_stock or 0) - (item.quantity or 0))

        # 8. 生成应收编号 YS + YYYYMMDD + 4位序号
        prefix_ys = 'YS' + today
        last_receivable = FinanceReceivable.query.filter(FinanceReceivable.receivable_no.like(prefix_ys + '%')).order_by(FinanceReceivable.receivable_no.desc()).first()
        if last_receivable and last_receivable.receivable_no and len(last_receivable.receivable_no) > len(prefix_ys):
            seq_ys = int(last_receivable.receivable_no[len(prefix_ys):]) + 1
        else:
            seq_ys = 1
        receivable_no = prefix_ys + str(seq_ys).zfill(4)

        # 9. 生成应收账款
        total_amount = float(order.actual_amount or order.total_amount or 0)
        receivable = FinanceReceivable(
            receivable_no=receivable_no,
            related_type='sale',
            related_id=id,
            related_no=order.order_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            total_amount=total_amount,
            received_amount=float(order.paid_amount or 0),
            remaining_amount=total_amount - float(order.paid_amount or 0),
            status=1 if (order.paid_amount and float(order.paid_amount) > 0 and float(order.paid_amount) < total_amount) else (2 if (order.paid_amount and float(order.paid_amount) >= total_amount) else 0),
            remark=f'销售单{order.order_no}审核自动生成'
        )
        db.session.add(receivable)

        # 10. 生成财务流水（收入）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=1,  # 收入
            amount=total_amount,
            balance_before=0,
            balance_after=0,
            related_type='sale',
            related_id=id,
            related_no=order.order_no,
            remark=f'销售单{order.order_no}审核，应收金额{total_amount}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        # 11. 更新销售单状态为已出库
        order.status = 2

        db.session.commit()
        return jsonify({'code': 200, 'message': '审核成功，已自动出库并生成应收账款'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'审核失败：{str(e)}'}), 500


@app.route('/api/sales/orders/<int:id>/print', methods=['GET'])
@jwt_required()
def print_sales_order(id):
    """打印销售单/收据"""
    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404
    
    # 获取明细
    items = SalesOrderItem.query.filter_by(order_id=id).all()
    
    # 获取客户信息
    customer = BaseCustomer.query.get(order.customer_id) if order.customer_id else None
    
    # 构建打印数据
    print_data = {
        'order_no': order.order_no,
        'order_date': order.order_date.strftime('%Y-%m-%d') if order.order_date else '',
        'customer_name': order.customer_name or '',
        'customer_phone': order.customer_phone or (customer.phone if customer else ''),
        'customer_address': order.customer_address or (customer.address if customer else ''),
        'contact_name': order.contact_name or (customer.contact_name if customer else ''),
        'salesperson_name': order.salesperson_name or '',
        'payment_method': order.payment_method or '',
        'delivery_method': order.delivery_method or '',
        'items': [{
            'product_name': item.product_name,
            'specification': item.specification or '',
            'unit': item.unit or '',
            'quantity': int(item.quantity),
            'price': float(item.price),
            'amount': float(item.amount)
        } for item in items],
        'total_quantity': int(order.total_quantity),
        'total_amount': float(order.total_amount),
        'discount_amount': float(order.discount_amount) if order.discount_amount else 0,
        'freight_amount': float(order.freight_amount) if order.freight_amount else 0,
        'actual_amount': float(order.actual_amount) if order.actual_amount else float(order.total_amount),
        'paid_amount': float(order.paid_amount) if order.paid_amount else 0,
        'remark': order.remark or '',
        'print_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    return jsonify({
        'code': 200,
        'data': print_data
    })


# ============================================
# API路由 - 销售发票管理
# ============================================

@app.route('/api/sales/invoices', methods=['GET'])
@jwt_required()
def get_sales_invoices():
    """获取发票列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    order_id = request.args.get('order_id', type=int)
    invoice_status = request.args.get('invoice_status', '')
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesInvoice.query

    if order_id:
        query = query.filter_by(order_id=order_id)

    if invoice_status:
        query = query.filter_by(invoice_status=invoice_status)

    if keyword:
        query = query.filter(
            db.or_(
                SalesInvoice.order_no.contains(keyword),
                SalesInvoice.buyer_name.contains(keyword),
                SalesInvoice.invoice_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= start_date)

    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= end_date)

    total = query.count()
    invoices = query.order_by(SalesInvoice.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': inv.id,
                'order_id': inv.order_id,
                'order_no': inv.order_no,
                'invoice_type': inv.invoice_type,
                'invoice_status': inv.invoice_status,
                'invoice_no': inv.invoice_no,
                'invoice_date': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else None,
                'buyer_name': inv.buyer_name,
                'total_amount': float(inv.total_amount) if inv.total_amount else 0,
                'total_tax': float(inv.total_tax) if inv.total_tax else 0,
                'total_with_tax': float(inv.total_with_tax) if inv.total_with_tax else 0,
                'tax_rate': float(inv.tax_rate) if inv.tax_rate else 0,
                'created_by_name': inv.created_by_name,
                'created_at': inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else None
            } for inv in invoices.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/sales/invoices', methods=['POST'])
@jwt_required()
def create_or_update_sales_invoice():
    """创建或更新发票"""
    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    order_id = data.get('order_id')
    if not order_id:
        return jsonify({'code': 400, 'message': '缺少order_id参数'}), 400

    order = SalesOrder.query.get(order_id)
    if not order:
        return jsonify({'code': 404, 'message': '关联的销售单不存在'}), 404

    # 检查是否已存在该订单的发票
    existing = SalesInvoice.query.filter_by(order_id=order_id).first()

    if existing:
        # 更新已有发票
        for field in ['invoice_type', 'invoice_no', 'invoice_date', 'buyer_name', 'buyer_tax_no',
                       'buyer_address', 'buyer_phone', 'buyer_bank', 'buyer_bank_account',
                       'items', 'total_amount', 'total_tax', 'total_with_tax', 'tax_rate',
                       'remark', 'attachment', 'invoice_status']:
            if field in data:
                setattr(existing, field, data[field])

        existing.updated_at = datetime.now()

        # 发票关联核销：更新发票状态为已开票时，关联应收账款
        if data.get('invoice_status') == '已开票':
            receivable = FinanceReceivable.query.filter_by(
                related_type='sale',
                related_id=order_id
            ).first()
            if receivable:
                receivable.invoice_id = existing.id
                invoice_amount = float(data.get('total_with_tax') or data.get('total_amount') or 0)
                receivable_amount = float(receivable.total_amount or 0)
                if invoice_amount >= receivable_amount:
                    receivable.status = 2  # 已结清
                receivable.updated_at = datetime.now()

        db.session.commit()
        return jsonify({'code': 200, 'message': '更新成功', 'data': {'id': existing.id}})
    else:
        # 创建新发票
        items_json = data.get('items', '[]')
        if isinstance(items_json, list):
            import json
            items_json = json.dumps(items_json, ensure_ascii=False)

        invoice = SalesInvoice(
            order_id=order_id,
            order_no=order.order_no,
            invoice_type=data.get('invoice_type', '普通发票'),
            invoice_status=data.get('invoice_status', '未开票'),
            invoice_no=data.get('invoice_no', ''),
            invoice_date=data.get('invoice_date'),
            buyer_name=data.get('buyer_name', ''),
            buyer_tax_no=data.get('buyer_tax_no', ''),
            buyer_address=data.get('buyer_address', ''),
            buyer_phone=data.get('buyer_phone', ''),
            buyer_bank=data.get('buyer_bank', ''),
            buyer_bank_account=data.get('buyer_bank_account', ''),
            items=items_json,
            total_amount=data.get('total_amount', 0),
            total_tax=data.get('total_tax', 0),
            total_with_tax=data.get('total_with_tax', 0),
            tax_rate=data.get('tax_rate', 0),
            remark=data.get('remark', ''),
            attachment=data.get('attachment', ''),
            created_by=user_id,
            created_by_name=user_name
        )
        db.session.add(invoice)

        # 更新销售单的has_invoice标记
        order.has_invoice = 1

        # 发票关联核销：查找关联的应收账款并更新
        if data.get('invoice_status') == '已开票':
            receivable = FinanceReceivable.query.filter_by(
                related_type='sale',
                related_id=order_id
            ).first()
            if receivable:
                receivable.invoice_id = invoice.id
                invoice_amount = float(data.get('total_with_tax') or data.get('total_amount') or 0)
                receivable_amount = float(receivable.total_amount or 0)
                if invoice_amount >= receivable_amount:
                    receivable.status = 2  # 已结清
                receivable.updated_at = datetime.now()

        db.session.commit()
        return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': invoice.id}})


@app.route('/api/sales/invoices/<int:id>', methods=['GET'])
@jwt_required()
def get_sales_invoice(id):
    """获取发票详情"""
    invoice = SalesInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    result = to_dict(invoice)
    # 解析items JSON
    if result.get('items'):
        try:
            import json
            result['items'] = json.loads(result['items'])
        except (json.JSONDecodeError, TypeError):
            pass

    return jsonify({'code': 200, 'data': result})


@app.route('/api/sales/invoices/<int:id>/status', methods=['PUT'])
@jwt_required()
def update_sales_invoice_status(id):
    """更新发票状态"""
    invoice = SalesInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    data = request.get_json()
    new_status = data.get('invoice_status')
    if not new_status or new_status not in ['未开票', '已开票', '作废']:
        return jsonify({'code': 400, 'message': '无效的开票状态，允许值：未开票/已开票/作废'}), 400

    invoice.invoice_status = new_status
    invoice.updated_at = datetime.now()
    db.session.commit()

    return jsonify({'code': 200, 'message': '状态更新成功', 'data': {'invoice_status': new_status}})


@app.route('/api/sales/invoices/export', methods=['GET'])
@jwt_required()
def export_sales_invoices():
    """导出发票列表"""
    keyword = request.args.get('keyword', '')
    invoice_status = request.args.get('invoice_status', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesInvoice.query
    if keyword:
        query = query.filter(
            db.or_(
                SalesInvoice.order_no.contains(keyword),
                SalesInvoice.buyer_name.contains(keyword),
                SalesInvoice.invoice_no.contains(keyword)
            )
        )
    if invoice_status:
        query = query.filter_by(invoice_status=invoice_status)
    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= start_date)
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= end_date)

    invoices = query.order_by(SalesInvoice.created_at.desc()).all()

    data = []
    for inv in invoices:
        data.append({
            '销售单号': inv.order_no or '',
            '发票类型': inv.invoice_type or '',
            '发票编号': inv.invoice_no or '',
            '开票日期': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else '',
            '开票状态': inv.invoice_status or '',
            '客户抬头': inv.buyer_name or '',
            '税号': inv.buyer_tax_no or '',
            '金额合计': float(inv.total_amount) if inv.total_amount else 0,
            '税额合计': float(inv.total_tax) if inv.total_tax else 0,
            '价税合计': float(inv.total_with_tax) if inv.total_with_tax else 0,
            '开票人': inv.created_by_name or '',
            '创建时间': inv.created_at.strftime('%Y-%m-%d %H:%M') if inv.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '发票列表'
    if data:
        ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'发票列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# API路由 - 销售收据管理
# ============================================

@app.route('/api/sales/receipts', methods=['GET'])
@jwt_required()
def get_sales_receipts():
    """获取收据列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    order_id = request.args.get('order_id', type=int)
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesReceipt.query.filter_by(status=1)

    if order_id:
        query = query.filter_by(order_id=order_id)

    if keyword:
        query = query.filter(
            db.or_(
                SalesReceipt.order_no.contains(keyword),
                SalesReceipt.customer_name.contains(keyword),
                SalesReceipt.receipt_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(SalesReceipt.receipt_date >= start_date)

    if end_date:
        query = query.filter(SalesReceipt.receipt_date <= end_date)

    total = query.count()
    receipts = query.order_by(SalesReceipt.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': r.id,
                'order_id': r.order_id,
                'order_no': r.order_no,
                'receipt_no': r.receipt_no,
                'receipt_date': r.receipt_date.strftime('%Y-%m-%d') if r.receipt_date else None,
                'customer_name': r.customer_name,
                'customer_phone': r.customer_phone,
                'total_amount': float(r.total_amount) if r.total_amount else 0,
                'paid_amount': float(r.paid_amount) if r.paid_amount else 0,
                'payment_method': r.payment_method,
                'payee': r.payee,
                'status': r.status,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None
            } for r in receipts.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/sales/receipts', methods=['POST'])
@jwt_required()
def create_sales_receipt():
    """开具收据"""
    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    order_id = data.get('order_id')
    if not order_id:
        return jsonify({'code': 400, 'message': '缺少order_id参数'}), 400

    order = SalesOrder.query.get(order_id)
    if not order:
        return jsonify({'code': 404, 'message': '关联的销售单不存在'}), 404

    # 自动生成收据编号：SK + 年月日 + 4位序号
    today_str = datetime.now().strftime('%Y%m%d')
    prefix = f'SK{today_str}'
    last_receipt = SalesReceipt.query.filter(SalesReceipt.receipt_no.like(f'{prefix}%')).order_by(SalesReceipt.id.desc()).first()
    if last_receipt and last_receipt.receipt_no:
        last_seq = int(last_receipt.receipt_no[-4:]) if len(last_receipt.receipt_no) >= 4 else 0
    else:
        last_seq = 0
    receipt_no = f'{prefix}{str(last_seq + 1).zfill(4)}'

    items_json = data.get('items', '[]')
    if isinstance(items_json, list):
        import json
        items_json = json.dumps(items_json, ensure_ascii=False)

    receipt = SalesReceipt(
        order_id=order_id,
        order_no=order.order_no,
        receipt_no=receipt_no,
        receipt_date=data.get('receipt_date') or datetime.now().strftime('%Y-%m-%d'),
        customer_name=data.get('customer_name', order.customer_name),
        customer_phone=data.get('customer_phone', order.customer_phone),
        total_amount=data.get('total_amount', 0),
        paid_amount=data.get('paid_amount', 0),
        payment_method=data.get('payment_method', ''),
        items=items_json,
        remark=data.get('remark', ''),
        payee=data.get('payee', user_name),
        status=1,
        created_by=user_id
    )
    db.session.add(receipt)

    # 更新销售单的has_receipt标记
    order.has_receipt = 1

    db.session.commit()
    return jsonify({'code': 200, 'message': '收据开具成功', 'data': {'id': receipt.id, 'receipt_no': receipt_no}})


@app.route('/api/sales/receipts/<int:id>', methods=['GET'])
@jwt_required()
def get_sales_receipt(id):
    """获取收据详情"""
    receipt = SalesReceipt.query.get(id)
    if not receipt:
        return jsonify({'code': 404, 'message': '收据不存在'}), 404

    result = to_dict(receipt)
    # 解析items JSON
    if result.get('items'):
        try:
            import json
            result['items'] = json.loads(result['items'])
        except (json.JSONDecodeError, TypeError):
            pass

    return jsonify({'code': 200, 'data': result})


@app.route('/api/sales/receipts/<int:id>/print', methods=['GET'])
@jwt_required()
def print_sales_receipt(id):
    """获取收据打印数据"""
    receipt = SalesReceipt.query.get(id)
    if not receipt:
        return jsonify({'code': 404, 'message': '收据不存在'}), 404

    if receipt.status == 0:
        return jsonify({'code': 400, 'message': '该收据已作废，无法打印'}), 400

    # 解析items
    items = []
    if receipt.items:
        try:
            import json
            items = json.loads(receipt.items)
        except (json.JSONDecodeError, TypeError):
            pass

    print_data = {
        'receipt_no': receipt.receipt_no,
        'receipt_date': receipt.receipt_date.strftime('%Y-%m-%d') if receipt.receipt_date else '',
        'customer_name': receipt.customer_name or '',
        'customer_phone': receipt.customer_phone or '',
        'items': items,
        'total_amount': float(receipt.total_amount) if receipt.total_amount else 0,
        'paid_amount': float(receipt.paid_amount) if receipt.paid_amount else 0,
        'payment_method': receipt.payment_method or '',
        'remark': receipt.remark or '',
        'payee': receipt.payee or '',
        'print_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    return jsonify({'code': 200, 'data': print_data})


@app.route('/api/sales/receipts/<int:id>/void', methods=['PUT'])
@jwt_required()
def void_sales_receipt(id):
    """作废收据"""
    receipt = SalesReceipt.query.get(id)
    if not receipt:
        return jsonify({'code': 404, 'message': '收据不存在'}), 404

    if receipt.status == 0:
        return jsonify({'code': 400, 'message': '该收据已作废'}), 400

    receipt.status = 0
    receipt.updated_at = datetime.now()
    db.session.commit()

    return jsonify({'code': 200, 'message': '收据已作废'})


@app.route('/api/sales/receipts/export', methods=['GET'])
@jwt_required()
def export_sales_receipts():
    """导出收据列表"""
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesReceipt.query.filter_by(status=1)
    if keyword:
        query = query.filter(
            db.or_(
                SalesReceipt.order_no.contains(keyword),
                SalesReceipt.customer_name.contains(keyword),
                SalesReceipt.receipt_no.contains(keyword)
            )
        )
    if start_date:
        query = query.filter(SalesReceipt.receipt_date >= start_date)
    if end_date:
        query = query.filter(SalesReceipt.receipt_date <= end_date)

    receipts = query.order_by(SalesReceipt.created_at.desc()).all()

    data = []
    for r in receipts:
        data.append({
            '收据编号': r.receipt_no or '',
            '销售单号': r.order_no or '',
            '收款日期': r.receipt_date.strftime('%Y-%m-%d') if r.receipt_date else '',
            '客户名称': r.customer_name or '',
            '联系方式': r.customer_phone or '',
            '应收金额': float(r.total_amount) if r.total_amount else 0,
            '实收金额': float(r.paid_amount) if r.paid_amount else 0,
            '收款方式': r.payment_method or '',
            '收款人': r.payee or '',
            '备注': r.remark or '',
            '创建时间': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '收据列表'
    if data:
        ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'收据列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# API路由 - 销售数据修复
# ============================================

@app.route('/api/sales/fix-prices', methods=['POST'])
@jwt_required()
def fix_sales_prices():
    """修复销售单明细中价格为0的数据，从商品表重新读取销售价"""
    fixed_count = 0
    fixed_order_ids = set()
    
    # 查找所有价格为0的明细
    items = SalesOrderItem.query.filter(
        (SalesOrderItem.price == 0) | (SalesOrderItem.price == None),
        SalesOrderItem.product_id != None
    ).all()
    
    for item in items:
        product = ProductInfo.query.get(item.product_id)
        if product and product.sale_price:
            item.price = float(product.sale_price)
            item.amount = int(item.quantity or 0) * float(product.sale_price)
            fixed_count += 1
            fixed_order_ids.add(item.order_id)
    
    # 重新计算受影响订单的总金额
    for order_id in fixed_order_ids:
        order_items = SalesOrderItem.query.filter_by(order_id=order_id).all()
        total = sum(float(i.amount or 0) for i in order_items)
        total_qty = sum(int(i.quantity or 0) for i in order_items)
        order = SalesOrder.query.get(order_id)
        if order:
            order.total_amount = total
            order.actual_amount = total - float(order.discount_amount or 0) + float(order.freight_amount or 0)
            order.total_quantity = total_qty
    
    db.session.commit()
    return jsonify({'code': 200, 'message': f'修复完成，共修复 {fixed_count} 条明细，更新 {len(fixed_order_ids)} 个销售单'})


# ============================================
# API路由 - 仓库管理
# ============================================

@app.route('/api/warehouses', methods=['GET'])
@jwt_required()
def get_warehouses():
    """获取仓库列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)

    query = Warehouse.query

    if keyword:
        query = query.filter(
            db.or_(
                Warehouse.warehouse_name.contains(keyword),
                Warehouse.warehouse_code.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    warehouses = query.order_by(Warehouse.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(w) for w in warehouses.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/warehouses', methods=['POST'])
@jwt_required()
def create_warehouse():
    """新增仓库"""
    data = request.get_json()

    if not data.get('warehouse_name'):
        return jsonify({'code': 400, 'message': '仓库名称不能为空'}), 400

    # 检查编码唯一性
    if data.get('warehouse_code'):
        exists = Warehouse.query.filter_by(warehouse_code=data['warehouse_code']).first()
        if exists:
            return jsonify({'code': 400, 'message': '仓库编码已存在'}), 400

    warehouse = Warehouse(
        warehouse_code=data.get('warehouse_code'),
        warehouse_name=data['warehouse_name'],
        address=data.get('address', ''),
        contact_person=data.get('contact_person', ''),
        contact_phone=data.get('contact_phone', ''),
        remark=data.get('remark', ''),
        status=data.get('status', 1)
    )

    db.session.add(warehouse)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': warehouse.id}})


@app.route('/api/warehouses/<int:id>', methods=['PUT'])
@jwt_required()
def update_warehouse(id):
    """编辑仓库"""
    warehouse = Warehouse.query.get(id)
    if not warehouse:
        return jsonify({'code': 404, 'message': '仓库不存在'}), 404

    data = request.get_json()

    if data.get('warehouse_code') and data['warehouse_code'] != warehouse.warehouse_code:
        exists = Warehouse.query.filter_by(warehouse_code=data['warehouse_code']).first()
        if exists:
            return jsonify({'code': 400, 'message': '仓库编码已存在'}), 400

    for field in ['warehouse_code', 'warehouse_name', 'address', 'contact_person', 'contact_phone', 'remark', 'status']:
        if field in data:
            setattr(warehouse, field, data[field])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/warehouses/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_warehouse(id):
    """删除仓库（有库存关联的不允许删除）"""
    warehouse = Warehouse.query.get(id)
    if not warehouse:
        return jsonify({'code': 404, 'message': '仓库不存在'}), 404

    # 检查是否有库存关联
    stock_count = InventoryStock.query.filter_by(warehouse_id=id).count()
    if stock_count > 0:
        return jsonify({'code': 400, 'message': f'该仓库下有 {stock_count} 条库存记录，不允许删除'}), 400

    db.session.delete(warehouse)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/warehouses/<int:id>/status', methods=['PUT'])
@jwt_required()
def toggle_warehouse_status(id):
    """启用/禁用仓库"""
    warehouse = Warehouse.query.get(id)
    if not warehouse:
        return jsonify({'code': 404, 'message': '仓库不存在'}), 404

    data = request.get_json()
    warehouse.status = data.get('status', 0)
    db.session.commit()

    status_text = '启用' if warehouse.status == 1 else '禁用'
    return jsonify({'code': 200, 'message': f'{status_text}成功'})


# ============================================
# API路由 - 库存变动明细
# ============================================

@app.route('/api/inventory/logs', methods=['GET'])
@jwt_required()
def get_inventory_logs():
    """库存变动明细列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    product_name = request.args.get('product_name', '')
    change_type = request.args.get('change_type', '')
    order_type = request.args.get('order_type', '')
    related_party = request.args.get('related_party', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = InventoryLog.query

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    if product_name:
        query = query.filter(InventoryLog.product_name.contains(product_name))

    if change_type:
        query = query.filter_by(change_type=change_type)

    if order_type:
        query = query.filter_by(order_type=order_type)

    if related_party:
        query = query.filter(InventoryLog.related_party.contains(related_party))

    if start_date:
        query = query.filter(InventoryLog.created_at >= start_date)

    if end_date:
        query = query.filter(InventoryLog.created_at <= end_date + ' 23:59:59')

    total = query.count()
    logs = query.order_by(InventoryLog.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(l) for l in logs.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/inventory/logs/export', methods=['GET'])
@jwt_required()
def export_inventory_logs():
    """导出库存变动明细"""
    warehouse_id = request.args.get('warehouse_id', type=int)
    product_name = request.args.get('product_name', '')
    change_type = request.args.get('change_type', '')
    order_type = request.args.get('order_type', '')
    related_party = request.args.get('related_party', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = InventoryLog.query

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    if product_name:
        query = query.filter(InventoryLog.product_name.contains(product_name))

    if change_type:
        query = query.filter_by(change_type=change_type)

    if order_type:
        query = query.filter_by(order_type=order_type)

    if related_party:
        query = query.filter(InventoryLog.related_party.contains(related_party))

    if start_date:
        query = query.filter(InventoryLog.created_at >= start_date)

    if end_date:
        query = query.filter(InventoryLog.created_at <= end_date + ' 23:59:59')

    logs = query.order_by(InventoryLog.created_at.desc()).all()

    # 变动类型映射
    change_type_map = {
        'in': '入库',
        'out': '出库',
        'adjust': '调整',
        'transfer': '调拨'
    }

    data = []
    for log in logs:
        data.append({
            '商品名称': log.product_name or '',
            '仓库': log.warehouse_name or '',
            '变动类型': change_type_map.get(log.change_type, log.change_type or ''),
            '变动数量': float(log.quantity_change) if log.quantity_change else 0,
            '变动前库存': float(log.before_quantity) if log.before_quantity else 0,
            '变动后库存': float(log.after_quantity) if log.after_quantity else 0,
            '关联单号': log.order_no or '',
            '操作人': log.operator_name or '',
            '创建时间': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '库存变动明细'
    # 始终写入表头
    headers = ['商品名称', '仓库', '变动类型', '变动数量', '变动前库存', '变动后库存', '关联单号', '操作人', '创建时间']
    ws.append(headers)
    for row in data:
        ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'库存变动明细_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# API路由 - 调拨管理
# ============================================

@app.route('/api/transfer/orders', methods=['GET'])
@jwt_required()
def get_transfer_orders():
    """调拨单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    transfer_type = request.args.get('transfer_type', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = TransferOrder.query

    if keyword:
        query = query.filter(TransferOrder.transfer_no.contains(keyword))

    if status is not None:
        query = query.filter_by(status=status)

    if transfer_type is not None:
        query = query.filter_by(transfer_type=transfer_type)

    if start_date:
        query = query.filter(TransferOrder.created_at >= start_date)

    if end_date:
        query = query.filter(TransferOrder.created_at <= end_date + ' 23:59:59')

    total = query.count()
    orders = query.order_by(TransferOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/transfer/orders', methods=['POST'])
@jwt_required()
def create_transfer_order():
    """创建调拨单（同价/变价）"""
    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 自动生成调拨单号：DB + 年月日 + 4位序号
    date_str = datetime.now().strftime("%Y%m%d")
    last_order = TransferOrder.query.filter(
        TransferOrder.transfer_no.like(f'DB{date_str}%')
    ).order_by(TransferOrder.id.desc()).first()

    if last_order:
        last_seq = int(last_order.transfer_no[-4:]) if last_order.transfer_no else 0
        seq = str(last_seq + 1).zfill(4)
    else:
        seq = '0001'
    transfer_no = f'DB{date_str}{seq}'

    transfer_type = data.get('transfer_type', 1)  # 1同价调拨 2变价调拨

    order = TransferOrder(
        transfer_no=transfer_no,
        from_warehouse_id=data.get('from_warehouse_id'),
        from_warehouse_name=data.get('from_warehouse_name', ''),
        to_warehouse_id=data.get('to_warehouse_id'),
        to_warehouse_name=data.get('to_warehouse_name', ''),
        total_quantity=0,
        status=0,
        transfer_type=transfer_type,
        from_cost_price=data.get('from_cost_price', 0),
        to_cost_price=data.get('to_cost_price', 0),
        operator_id=user_id,
        operator_name=user_name,
        remark=data.get('remark', ''),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    items = data.get('items', [])
    total_quantity = 0

    for item_data in items:
        item = TransferOrderItem(
            transfer_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification', ''),
            unit_name=item_data.get('unit_name', ''),
            quantity=item_data.get('quantity', 0),
            from_cost_price=item_data.get('from_cost_price', 0),
            to_cost_price=item_data.get('to_cost_price', 0),
            remark=item_data.get('remark', '')
        )
        db.session.add(item)
        total_quantity += float(item_data.get('quantity', 0))

    order.total_quantity = total_quantity
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'transfer_no': transfer_no}})


@app.route('/api/transfer/orders/<int:id>', methods=['GET'])
@jwt_required()
def get_transfer_order(id):
    """调拨单详情"""
    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    items = TransferOrderItem.query.filter_by(transfer_id=id).all()

    result = to_dict(order)
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@app.route('/api/transfer/orders/<int:id>', methods=['PUT'])
@jwt_required()
def update_transfer_order(id):
    """编辑调拨单（待审核状态）"""
    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调拨单可以编辑'}), 400

    data = request.get_json()

    for field in ['from_warehouse_id', 'from_warehouse_name', 'to_warehouse_id', 'to_warehouse_name',
                   'transfer_type', 'from_cost_price', 'to_cost_price', 'remark']:
        if field in data:
            setattr(order, field, data[field])

    # 更新明细
    if 'items' in data:
        # 删除旧明细
        TransferOrderItem.query.filter_by(transfer_id=id).delete()

        total_quantity = 0
        for item_data in data['items']:
            item = TransferOrderItem(
                transfer_id=id,
                product_id=item_data.get('product_id'),
                product_code=item_data.get('product_code'),
                product_name=item_data.get('product_name'),
                specification=item_data.get('specification', ''),
                unit_name=item_data.get('unit_name', ''),
                quantity=item_data.get('quantity', 0),
                from_cost_price=item_data.get('from_cost_price', 0),
                to_cost_price=item_data.get('to_cost_price', 0),
                remark=item_data.get('remark', '')
            )
            db.session.add(item)
            total_quantity += float(item_data.get('quantity', 0))

        order.total_quantity = total_quantity

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/transfer/orders/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_transfer_order(id):
    """审核调拨单"""
    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调拨单可以审核'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    items = TransferOrderItem.query.filter_by(transfer_id=id).all()

    # 检查调出仓库库存是否充足
    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.from_warehouse_id
        ).first()

        if not stock or float(stock.available_quantity or 0) < float(item.quantity or 0):
            return jsonify({'code': 400, 'message': f'商品【{item.product_name}】在调出仓库库存不足'}), 400

    # 更新调拨单状态
    order.status = 1
    order.auditor_id = user_id
    order.auditor_name = user_name
    order.audit_time = datetime.now()

    # 处理库存变动
    for item in items:
        qty = float(item.quantity or 0)
        from_cost = float(item.from_cost_price or 0)
        to_cost = float(item.to_cost_price or 0)

        # 调出仓库减少库存
        from_stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.from_warehouse_id
        ).first()

        if from_stock:
            before_from_qty = float(from_stock.quantity or 0)
            from_stock.quantity = before_from_qty - qty
            from_stock.available_quantity = float(from_stock.available_quantity or 0) - qty

            # 写入调出仓库的 InventoryLog
            log_out = InventoryLog(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=order.from_warehouse_id,
                warehouse_name=order.from_warehouse_name,
                change_type='transfer',
                order_type='调拨出库',
                order_id=order.id,
                order_no=order.transfer_no,
                quantity_change=-qty,
                before_quantity=before_from_qty,
                after_quantity=before_from_qty - qty,
                cost_price=from_cost,
                amount=qty * from_cost,
                related_party=order.to_warehouse_name or '',
                operator_id=user_id,
                operator_name=user_name,
                remark=f'调拨至{order.to_warehouse_name}'
            )
            db.session.add(log_out)

        # 调入仓库增加库存
        to_stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.to_warehouse_id
        ).first()

        if to_stock:
            before_to_qty = float(to_stock.quantity or 0)
            to_stock.quantity = before_to_qty + qty
            to_stock.available_quantity = float(to_stock.available_quantity or 0) + qty
            # 变价调拨时更新调入仓库成本价
            if order.transfer_type == 2 and to_cost > 0:
                to_stock.cost_price = to_cost
        else:
            before_to_qty = 0
            to_stock = InventoryStock(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=order.to_warehouse_id,
                warehouse_name=order.to_warehouse_name,
                quantity=qty,
                frozen_quantity=0,
                available_quantity=qty,
                cost_price=to_cost if (order.transfer_type == 2 and to_cost > 0) else from_cost
            )
            db.session.add(to_stock)

        # 写入调入仓库的 InventoryLog
        log_in = InventoryLog(
            product_id=item.product_id,
            product_code=item.product_code,
            product_name=item.product_name,
            warehouse_id=order.to_warehouse_id,
            warehouse_name=order.to_warehouse_name,
            change_type='transfer',
            order_type='调拨入库',
            order_id=order.id,
            order_no=order.transfer_no,
            quantity_change=qty,
            before_quantity=before_to_qty,
            after_quantity=before_to_qty + qty,
            cost_price=to_cost if (order.transfer_type == 2 and to_cost > 0) else from_cost,
            amount=qty * (to_cost if (order.transfer_type == 2 and to_cost > 0) else from_cost),
            related_party=order.from_warehouse_name or '',
            operator_id=user_id,
            operator_name=user_name,
            remark=f'从{order.from_warehouse_name}调入'
        )
        db.session.add(log_in)

        # 变价调拨时生成 FinanceRecord 财务流水（成本差异）
        if order.transfer_type == 2 and from_cost != to_cost:
            diff_amount = qty * (to_cost - from_cost)
            finance = FinanceRecord(
                account_id=1,
                account_name='默认账户',
                record_type=2 if diff_amount > 0 else 1,
                amount=abs(diff_amount),
                balance_before=0,
                balance_after=0,
                related_type='transfer',
                related_id=order.id,
                related_no=order.transfer_no,
                remark=f'变价调拨成本差异：{item.product_name} x {qty}，原成本{from_cost}，新成本{to_cost}',
                created_by=user_id
            )
            db.session.add(finance)

    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，调拨已完成'})


@app.route('/api/transfer/orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_transfer_order(id):
    """删除调拨单（待审核状态）"""
    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调拨单可以删除'}), 400

    # 删除明细
    TransferOrderItem.query.filter_by(transfer_id=id).delete()
    db.session.delete(order)
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/transfer/orders/export', methods=['GET'])
@jwt_required()
def export_transfer_orders():
    """导出调拨单"""
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = TransferOrder.query

    if keyword:
        query = query.filter(TransferOrder.transfer_no.contains(keyword))

    if status is not None:
        query = query.filter_by(status=status)

    if start_date:
        query = query.filter(TransferOrder.created_at >= start_date)

    if end_date:
        query = query.filter(TransferOrder.created_at <= end_date + ' 23:59:59')

    orders = query.order_by(TransferOrder.created_at.desc()).all()

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '调拨单列表'

    headers = ['调拨单号', '调出仓库', '调入仓库', '调拨类型', '总数量', '状态', '经手人', '备注', '创建时间']
    ws.append(headers)

    status_map = {0: '待审核', 1: '已审核', 2: '已完成'}
    type_map = {1: '同价调拨', 2: '变价调拨'}

    for o in orders:
        ws.append([
            o.transfer_no,
            o.from_warehouse_name or '',
            o.to_warehouse_name or '',
            type_map.get(o.transfer_type, ''),
            float(o.total_quantity or 0),
            status_map.get(o.status, ''),
            o.operator_name or '',
            o.remark or '',
            o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
        ])

    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name=f'调拨单列表_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# API路由 - 成本调价管理
# ============================================

@app.route('/api/cost-adjusts', methods=['GET'])
@jwt_required()
def get_cost_adjusts():
    """成本调价单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = CostAdjust.query

    if keyword:
        query = query.filter(
            db.or_(
                CostAdjust.adjust_no.contains(keyword),
                CostAdjust.product_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    if start_date:
        query = query.filter(CostAdjust.created_at >= start_date)

    if end_date:
        query = query.filter(CostAdjust.created_at <= end_date + ' 23:59:59')

    total = query.count()
    adjusts = query.order_by(CostAdjust.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(a) for a in adjusts.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/cost-adjusts', methods=['POST'])
@jwt_required()
def create_cost_adjust():
    """创建成本调价单"""
    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 自动生成调价单号：CB + 年月日 + 4位序号
    date_str = datetime.now().strftime("%Y%m%d")
    last_adjust = CostAdjust.query.filter(
        CostAdjust.adjust_no.like(f'CB{date_str}%')
    ).order_by(CostAdjust.id.desc()).first()

    if last_adjust:
        last_seq = int(last_adjust.adjust_no[-4:]) if last_adjust.adjust_no else 0
        seq = str(last_seq + 1).zfill(4)
    else:
        seq = '0001'
    adjust_no = f'CB{date_str}{seq}'

    old_cost = float(data.get('old_cost_price', 0))
    new_cost = float(data.get('new_cost_price', 0))
    qty = float(data.get('adjust_quantity', 0))
    adjust_amount = round((new_cost - old_cost) * qty, 2)

    adjust = CostAdjust(
        adjust_no=adjust_no,
        warehouse_id=data.get('warehouse_id'),
        warehouse_name=data.get('warehouse_name', ''),
        product_id=data.get('product_id'),
        product_code=data.get('product_code', ''),
        product_name=data.get('product_name', ''),
        old_cost_price=old_cost,
        new_cost_price=new_cost,
        adjust_quantity=qty,
        adjust_amount=adjust_amount,
        status=0,
        remark=data.get('remark', ''),
        created_by=user_id,
        created_by_name=user_name
    )

    db.session.add(adjust)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': adjust.id, 'adjust_no': adjust_no}})


@app.route('/api/cost-adjusts/<int:id>', methods=['GET'])
@jwt_required()
def get_cost_adjust(id):
    """调价单详情"""
    adjust = CostAdjust.query.get(id)
    if not adjust:
        return jsonify({'code': 404, 'message': '调价单不存在'}), 404

    return jsonify({'code': 200, 'data': to_dict(adjust)})


@app.route('/api/cost-adjusts/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_cost_adjust(id):
    """审核成本调价单"""
    adjust = CostAdjust.query.get(id)
    if not adjust:
        return jsonify({'code': 404, 'message': '调价单不存在'}), 404

    if adjust.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调价单可以审核'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 更新 InventoryStock 中该商品在该仓库的 cost_price
    stock = InventoryStock.query.filter_by(
        product_id=adjust.product_id,
        warehouse_id=adjust.warehouse_id
    ).first()

    old_cost = float(adjust.old_cost_price or 0)
    new_cost = float(adjust.new_cost_price or 0)
    qty = float(adjust.adjust_quantity or 0)

    if stock:
        stock.cost_price = new_cost
        before_qty = float(stock.quantity or 0)
    else:
        before_qty = 0

    # 生成 InventoryLog 记录
    log = InventoryLog(
        product_id=adjust.product_id,
        product_code=adjust.product_code,
        product_name=adjust.product_name,
        warehouse_id=adjust.warehouse_id,
        warehouse_name=adjust.warehouse_name,
        change_type='adjust',
        order_type='成本调价',
        order_id=adjust.id,
        order_no=adjust.adjust_no,
        quantity_change=0,  # 成本调价不改变数量
        before_quantity=before_qty,
        after_quantity=before_qty,
        cost_price=new_cost,
        amount=adjust.adjust_amount,
        related_party='',
        operator_id=user_id,
        operator_name=user_name,
        remark=f'成本调价：{old_cost} -> {new_cost}，数量{qty}'
    )
    db.session.add(log)

    # 生成 FinanceRecord 财务流水（成本差异）
    if adjust.adjust_amount and float(adjust.adjust_amount) != 0:
        diff_amount = float(adjust.adjust_amount)
        finance = FinanceRecord(
            account_id=1,
            account_name='默认账户',
            record_type=2 if diff_amount > 0 else 1,
            amount=abs(diff_amount),
            balance_before=0,
            balance_after=0,
            related_type='cost_adjust',
            related_id=adjust.id,
            related_no=adjust.adjust_no,
            remark=f'成本调价差异：{adjust.product_name} x {qty}，原成本{old_cost}，新成本{new_cost}',
            created_by=user_id
        )
        db.session.add(finance)

    # 更新调价单状态
    adjust.status = 1
    adjust.audited_by = user_id
    adjust.audited_by_name = user_name
    adjust.audited_at = datetime.now()

    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，成本已更新'})


@app.route('/api/cost-adjusts/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_cost_adjust(id):
    """删除调价单（待审核状态）"""
    adjust = CostAdjust.query.get(id)
    if not adjust:
        return jsonify({'code': 404, 'message': '调价单不存在'}), 404

    if adjust.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调价单可以删除'}), 400

    db.session.delete(adjust)
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


@app.route('/api/cost-adjusts/export', methods=['GET'])
@jwt_required()
def export_cost_adjusts():
    """导出成本调价单"""
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = CostAdjust.query

    if keyword:
        query = query.filter(
            db.or_(
                CostAdjust.adjust_no.contains(keyword),
                CostAdjust.product_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if start_date:
        query = query.filter(CostAdjust.created_at >= start_date)

    if end_date:
        query = query.filter(CostAdjust.created_at <= end_date + ' 23:59:59')

    adjusts = query.order_by(CostAdjust.created_at.desc()).all()

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成本调价单列表'

    headers = ['调价单号', '仓库', '商品编码', '商品名称', '原成本价', '新成本价', '调整数量', '调整金额', '状态', '创建人', '创建时间']
    ws.append(headers)

    status_map = {0: '待审核', 1: '已审核', 2: '已取消'}

    for a in adjusts:
        ws.append([
            a.adjust_no,
            a.warehouse_name or '',
            a.product_code or '',
            a.product_name or '',
            float(a.old_cost_price or 0),
            float(a.new_cost_price or 0),
            float(a.adjust_quantity or 0),
            float(a.adjust_amount or 0),
            status_map.get(a.status, ''),
            a.created_by_name or '',
            a.created_at.strftime('%Y-%m-%d %H:%M:%S') if a.created_at else ''
        ])

    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name=f'成本调价单列表_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# API路由 - 库存查询增强（导出）
# ============================================

@app.route('/api/inventory/stock/export', methods=['GET'])
@jwt_required()
def export_inventory_stock():
    """导出库存查询结果"""
    keyword = request.args.get('keyword', '')
    warehouse_id = request.args.get('warehouse_id', type=int)

    query = InventoryStock.query

    if keyword:
        query = query.filter(
            db.or_(
                InventoryStock.product_name.contains(keyword),
                InventoryStock.product_code.contains(keyword)
            )
        )

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    stocks = query.order_by(InventoryStock.updated_at.desc()).all()

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '库存查询'

    headers = ['商品编码', '商品名称', '仓库', '库存数量', '冻结数量', '可用数量', '成本价', '库存金额', '采购价', '安全库存', '库存状态']
    ws.append(headers)

    for s in stocks:
        product = ProductInfo.query.get(s.product_id)
        purchase_price = float(product.purchase_price) if product and product.purchase_price else 0
        safe_stock = product.min_stock if product else 0
        qty = float(s.quantity or 0)

        if safe_stock and safe_stock > 0:
            if qty <= 0:
                stock_status = '缺货'
            elif qty < safe_stock:
                stock_status = '不足'
            elif qty > safe_stock * 3:
                stock_status = '过剩'
            else:
                stock_status = '正常'
        else:
            stock_status = '正常'

        ws.append([
            s.product_code or '',
            s.product_name or '',
            s.warehouse_name or '',
            qty,
            float(s.frozen_quantity or 0),
            float(s.available_quantity or 0),
            float(s.cost_price or 0),
            round(qty * float(s.cost_price or 0), 2),
            purchase_price,
            safe_stock,
            stock_status
        ])

    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name=f'库存查询_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# API路由 - 财务账户管理（完善）
# ============================================

@app.route('/api/finance/accounts', methods=['POST'])
@jwt_required()
def create_finance_account():
    """创建财务账户"""
    data = request.get_json()

    account = FinanceAccount(
        account_name=data.get('account_name'),
        account_type=data.get('account_type', 1),
        account_no=data.get('account_no'),
        balance=data.get('balance', 0),
        remark=data.get('remark')
    )

    db.session.add(account)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': account.id}})


@app.route('/api/finance/accounts/<int:id>', methods=['PUT'])
@jwt_required()
def update_finance_account(id):
    """更新财务账户"""
    account = FinanceAccount.query.get(id)
    if not account:
        return jsonify({'code': 404, 'message': '账户不存在'}), 404

    data = request.get_json()

    for field in ['account_name', 'account_type', 'account_no', 'remark']:
        if field in data:
            setattr(account, field, data[field])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@app.route('/api/finance/accounts/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_finance_account(id):
    """删除财务账户"""
    account = FinanceAccount.query.get(id)
    if not account:
        return jsonify({'code': 404, 'message': '账户不存在'}), 404

    # 软删除
    account.status = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


# ============================================
# API路由 - 账户管理增强
# ============================================

@app.route('/api/finance/accounts/transfer', methods=['POST'])
@jwt_required()
def transfer_finance_account():
    """账户转账"""
    try:
        current_user_id = get_jwt_identity()
        current_user_name = get_current_user_name()
        data = request.get_json()
        if not data:
            return jsonify({'code': 400, 'message': '请求数据不能为空'}), 400

        from_account_id = data.get('from_account_id')
        to_account_id = data.get('to_account_id')
        amount = data.get('amount')
        remark = data.get('remark', '')

        if not from_account_id or not to_account_id:
            return jsonify({'code': 400, 'message': '转出账户和转入账户不能为空'}), 400
        if from_account_id == to_account_id:
            return jsonify({'code': 400, 'message': '转出账户和转入账户不能相同'}), 400
        if not amount or float(amount) <= 0:
            return jsonify({'code': 400, 'message': '转账金额必须大于0'}), 400

        amount_val = float(amount)

        from_account = FinanceAccount.query.get(from_account_id)
        to_account = FinanceAccount.query.get(to_account_id)

        if not from_account:
            return jsonify({'code': 404, 'message': '转出账户不存在'}), 404
        if not to_account:
            return jsonify({'code': 404, 'message': '转入账户不存在'}), 404
        if from_account.status != 1:
            return jsonify({'code': 400, 'message': '转出账户已停用'}), 400
        if to_account.status != 1:
            return jsonify({'code': 400, 'message': '转入账户已停用'}), 400
        if float(from_account.balance) < amount_val:
            return jsonify({'code': 400, 'message': '转出账户余额不足'}), 400

        related_no = 'TR' + str(int(datetime.now().timestamp()))

        # 扣减转出账户余额
        from_balance_before = float(from_account.balance)
        from_account.balance = from_balance_before - amount_val

        # 增加转入账户余额
        to_balance_before = float(to_account.balance)
        to_account.balance = to_balance_before + amount_val

        # 生成转出流水记录（支出）
        from_record = FinanceRecord(
            account_id=from_account.id,
            account_name=from_account.account_name,
            record_type=2,
            amount=amount_val,
            balance_before=from_balance_before,
            balance_after=from_balance_before - amount_val,
            related_type='transfer',
            related_id=to_account.id,
            related_no=related_no,
            remark=f'转账至 {to_account.account_name}' + (f'：{remark}' if remark else ''),
            created_at=datetime.now(),
            created_by=current_user_id
        )

        # 生成转入流水记录（收入）
        to_record = FinanceRecord(
            account_id=to_account.id,
            account_name=to_account.account_name,
            record_type=1,
            amount=amount_val,
            balance_before=to_balance_before,
            balance_after=to_balance_before + amount_val,
            related_type='transfer',
            related_id=from_account.id,
            related_no=related_no,
            remark=f'从 {from_account.account_name} 转入' + (f'：{remark}' if remark else ''),
            created_at=datetime.now(),
            created_by=current_user_id
        )

        db.session.add(from_record)
        db.session.add(to_record)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '转账成功',
            'data': {
                'related_no': related_no,
                'from_account': {
                    'id': from_account.id,
                    'account_name': from_account.account_name,
                    'balance_after': float(from_account.balance)
                },
                'to_account': {
                    'id': to_account.id,
                    'account_name': to_account.account_name,
                    'balance_after': float(to_account.balance)
                }
            }
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'账户转账失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'转账失败: {str(e)}'}), 500


@app.route('/api/finance/accounts/<int:id>/adjust', methods=['POST'])
@jwt_required()
def adjust_finance_account(id):
    """余额调整（盘盈/盘亏）"""
    try:
        current_user_id = get_jwt_identity()
        current_user_name = get_current_user_name()
        data = request.get_json()
        if not data:
            return jsonify({'code': 400, 'message': '请求数据不能为空'}), 400

        adjust_amount = data.get('adjust_amount')
        remark = data.get('remark', '')

        if adjust_amount is None:
            return jsonify({'code': 400, 'message': '调整金额不能为空'}), 400

        adjust_amount_val = float(adjust_amount)
        if adjust_amount_val == 0:
            return jsonify({'code': 400, 'message': '调整金额不能为0'}), 400

        account = FinanceAccount.query.get(id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404
        if account.status != 1:
            return jsonify({'code': 400, 'message': '账户已停用，无法调整'}), 400

        balance_before = float(account.balance)
        balance_after = balance_before + adjust_amount_val

        # 余额不能为负数
        if balance_after < 0:
            return jsonify({'code': 400, 'message': '调整后余额不能为负数'}), 400

        account.balance = balance_after

        # 确定流水类型：正数为收入（盘盈），负数为支出（盘亏）
        record_type = 1 if adjust_amount_val > 0 else 2
        adjust_type_text = '盘盈' if adjust_amount_val > 0 else '盘亏'

        record = FinanceRecord(
            account_id=account.id,
            account_name=account.account_name,
            record_type=record_type,
            amount=abs(adjust_amount_val),
            balance_before=balance_before,
            balance_after=balance_after,
            related_type='adjust',
            related_id=id,
            related_no='AD' + str(int(datetime.now().timestamp())),
            remark=f'{adjust_type_text}调整' + (f'：{remark}' if remark else ''),
            created_at=datetime.now(),
            created_by=current_user_id
        )

        db.session.add(record)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '余额调整成功',
            'data': {
                'id': account.id,
                'account_name': account.account_name,
                'balance_before': balance_before,
                'balance_after': balance_after,
                'adjust_amount': adjust_amount_val,
                'adjust_type': adjust_type_text
            }
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'余额调整失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'余额调整失败: {str(e)}'}), 500


@app.route('/api/finance/records/enhanced', methods=['GET'])
@jwt_required()
def get_finance_records_enhanced():
    """增强版流水查询"""
    try:
        account_id = request.args.get('account_id', type=int)
        record_type = request.args.get('record_type', type=int)
        related_type = request.args.get('related_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        keyword = request.args.get('keyword')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        if page < 1:
            page = 1
        if page_size < 1 or page_size > 100:
            page_size = 20

        query = FinanceRecord.query

        if account_id:
            query = query.filter(FinanceRecord.account_id == account_id)
        if record_type:
            query = query.filter(FinanceRecord.record_type == record_type)
        if related_type:
            query = query.filter(FinanceRecord.related_type == related_type)
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(FinanceRecord.created_at >= start_dt)
            except ValueError:
                return jsonify({'code': 400, 'message': '开始日期格式不正确，应为YYYY-MM-DD'}), 400
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                # 包含当天全天，设置到23:59:59
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
                query = query.filter(FinanceRecord.created_at <= end_dt)
            except ValueError:
                return jsonify({'code': 400, 'message': '结束日期格式不正确，应为YYYY-MM-DD'}), 400
        if keyword:
            keyword_like = f'%{keyword}%'
            query = query.filter(
                db.or_(
                    FinanceRecord.account_name.like(keyword_like),
                    FinanceRecord.remark.like(keyword_like),
                    FinanceRecord.related_no.like(keyword_like)
                )
            )

        query = query.order_by(FinanceRecord.created_at.desc())

        total = query.count()
        records = query.offset((page - 1) * page_size).limit(page_size).all()

        account_type_map = {1: '现金', 2: '银行', 3: '支付宝', 4: '微信'}
        record_type_map = {1: '收入', 2: '支出'}

        items = []
        for record in records:
            # 查找账户类型
            account = FinanceAccount.query.get(record.account_id)
            account_type_text = ''
            if account:
                account_type_text = account_type_map.get(account.account_type, '未知')

            items.append({
                'id': record.id,
                'account_id': record.account_id,
                'account_name': record.account_name,
                'account_type_text': account_type_text,
                'record_type': record.record_type,
                'record_type_text': record_type_map.get(record.record_type, '未知'),
                'amount': float(record.amount) if record.amount else 0,
                'balance_before': float(record.balance_before) if record.balance_before else 0,
                'balance_after': float(record.balance_after) if record.balance_after else 0,
                'related_type': record.related_type,
                'related_id': record.related_id,
                'related_no': record.related_no,
                'remark': record.remark,
                'created_at': record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else None,
                'created_by': record.created_by
            })

        return jsonify({
            'code': 200,
            'message': '查询成功',
            'data': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'items': items
            }
        })
    except Exception as e:
        logger.error(f'增强版流水查询失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'查询失败: {str(e)}'}), 500


# ============================================
# API路由 - 导入导出
# ============================================

# 导入必要的库
try:
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_SUPPORT = True
    logger.info('openpyxl已加载，Excel导入导出功能可用')
except ImportError:
    EXCEL_SUPPORT = False
    openpyxl = None
    logger.warning('openpyxl未安装，Excel导入导出功能不可用')


def export_to_excel(data, columns, filename):
    """导出数据到Excel（使用openpyxl，不依赖pandas）"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导出功能不可用，请安装openpyxl'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '数据'

    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(vertical='center')
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 自动调整列宽
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(2, len(data) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def read_excel_data(file):
    """读取Excel数据（使用openpyxl，不依赖pandas）"""
    if not EXCEL_SUPPORT:
        return None, 'openpyxl未安装，Excel导入功能不可用'

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            return None, 'Excel文件为空或没有数据行'

        headers = [str(h) if h else '' for h in all_rows[0]]
        data = []
        for row in all_rows[1:]:
            if any(cell is not None for cell in row):
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                data.append(row_dict)
        return data, None
    except Exception as e:
        return None, f'读取Excel失败：{str(e)}'


# 客户导入导出
@app.route('/api/customers/export', methods=['GET'])
@jwt_required()
def export_customers():
    """导出客户"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导出功能不可用，请安装pandas'}), 500

    customers = BaseCustomer.query.filter_by(status=1).all()

    data = []
    for c in customers:
        data.append({
            '客户编码': c.customer_code,
            '客户名称': c.customer_name,
            '客户类型': '个人' if c.customer_type == 1 else '企业',
            '联系人': c.contact_name,
            '电话': c.phone,
            '电话2': c.phone2,
            '邮箱': c.email,
            '地址': c.address,
            '折扣率': float(c.discount_rate) if c.discount_rate else 100.00,
            '信用额度': float(c.credit_limit) if c.credit_limit else 0.00,
            '税号': c.tax_number,
            '开户行': c.bank_name,
            '银行账号': c.bank_account,
            '备注': c.remark
        })

    columns = ['客户编码', '客户名称', '客户类型', '联系人', '电话', '电话2', '邮箱', '地址',
               '折扣率', '信用额度', '税号', '开户行', '银行账号', '备注']

    return export_to_excel(data, columns, f'客户列表_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/api/customers/import', methods=['POST'])
@jwt_required()
def import_customers():
    """导入客户"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导入功能不可用，请安装pandas'}), 500

    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    # Excel 文件白名单校验
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in {'xlsx', 'xls', 'csv'}:
        return jsonify({'code': 400, 'message': f'仅支持 xlsx/xls/csv 格式，当前: {ext}'}), 400
    if (file.mimetype or '').lower() not in {
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',  # 浏览器有时给这个
        'text/csv',
    }:
        return jsonify({'code': 400, 'message': '文件 MIME 类型不符'}), 400

    try:
        data, error = read_excel_data(file)
        if error:
            return jsonify({'code': 400, 'message': error}), 400

        success_count = 0
        error_count = 0
        errors = []

        for idx, row in enumerate(data):
            try:
                # 验证必填字段
                customer_name = row.get('客户名称')
                if not customer_name or str(customer_name).strip() == '':
                    error_count += 1
                    errors.append(f'第{idx + 2}行：客户名称不能为空')
                    continue

                # 生成客户编码
                last_customer = BaseCustomer.query.order_by(BaseCustomer.id.desc()).first()
                customer_code = generate_code('C', last_customer.id if last_customer else 0)

                def get_val(row, key):
                    v = row.get(key)
                    if v is None:
                        return None
                    s = str(v).strip()
                    return s if s else None

                def get_float(row, key, default=0.0):
                    v = row.get(key)
                    if v is None:
                        return default
                    try:
                        return float(v)
                    except:
                        return default

                customer = BaseCustomer(
                    customer_code=customer_code,
                    customer_name=str(customer_name).strip(),
                    customer_type=1 if get_val(row, '客户类型') == '个人' else 2,
                    pinyin_code=generate_pinyin_code(str(customer_name)),
                    contact_name=get_val(row, '联系人'),
                    phone=get_val(row, '电话'),
                    phone2=get_val(row, '电话2'),
                    email=get_val(row, '邮箱'),
                    address=get_val(row, '地址'),
                    discount_rate=get_float(row, '折扣率', 100.00),
                    credit_limit=get_float(row, '信用额度', 0.00),
                    tax_number=get_val(row, '税号'),
                    bank_name=get_val(row, '开户行'),
                    bank_account=get_val(row, '银行账号'),
                    remark=get_val(row, '备注')
                )

                db.session.add(customer)
                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f'第{idx + 2}行：{str(e)}')

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'导入完成，成功{success_count}条，失败{error_count}条',
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # 只返回前10个错误
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导入失败：{str(e)}'}), 500


# 供应商导入导出
@app.route('/api/suppliers/export', methods=['GET'])
@jwt_required()
def export_suppliers():
    """导出供应商"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导出功能不可用，请安装pandas'}), 500

    suppliers = BaseSupplier.query.filter_by(status=1).all()

    data = []
    for s in suppliers:
        data.append({
            '供应商编码': s.supplier_code,
            '供应商名称': s.supplier_name,
            '联系人': s.contact_name,
            '电话': s.phone,
            '邮箱': s.email,
            '地址': s.address,
            '税号': s.tax_number,
            '开户行': s.bank_name,
            '银行账号': s.bank_account,
            '是否维修合作方': '是' if s.is_repair_partner else '否',
            '备注': s.remark
        })

    columns = ['供应商编码', '供应商名称', '联系人', '电话', '邮箱', '地址',
               '税号', '开户行', '银行账号', '是否维修合作方', '备注']

    return export_to_excel(data, columns, f'供应商列表_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/api/suppliers/import', methods=['POST'])
@jwt_required()
def import_suppliers():
    """导入供应商"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导入功能不可用，请安装pandas'}), 500

    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    # Excel 文件白名单
    _ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if _ext not in {'xlsx', 'xls', 'csv'}:
        return jsonify({'code': 400, 'message': f'仅支持 xlsx/xls/csv 格式，当前: {_ext}'}), 400
    if (file.mimetype or '').lower() not in {
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',
        'text/csv',
    }:
        return jsonify({'code': 400, 'message': '文件 MIME 类型不符'}), 400

    try:
        data, error = read_excel_data(file)
        if error:
            return jsonify({'code': 400, 'message': error}), 400

        success_count = 0
        error_count = 0
        errors = []

        for idx, row in enumerate(data):
            try:
                supplier_name = row.get('供应商名称')
                if not supplier_name or str(supplier_name).strip() == '':
                    error_count += 1
                    errors.append(f'第{idx + 2}行：供应商名称不能为空')
                    continue

                last_supplier = BaseSupplier.query.order_by(BaseSupplier.id.desc()).first()
                supplier_code = generate_code('S', last_supplier.id if last_supplier else 0)

                def get_val(row, key):
                    v = row.get(key)
                    if v is None:
                        return None
                    s = str(v).strip()
                    return s if s else None

                is_repair = 1 if get_val(row, '是否维修合作方') == '是' else 0

                supplier = BaseSupplier(
                    supplier_code=supplier_code,
                    supplier_name=str(supplier_name).strip(),
                    pinyin_code=generate_pinyin_code(str(supplier_name)),
                    contact_name=get_val(row, '联系人'),
                    phone=get_val(row, '电话'),
                    email=get_val(row, '邮箱'),
                    address=get_val(row, '地址'),
                    tax_number=get_val(row, '税号'),
                    bank_name=get_val(row, '开户行'),
                    bank_account=get_val(row, '银行账号'),
                    is_repair_partner=is_repair,
                    remark=get_val(row, '备注')
                )

                db.session.add(supplier)
                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f'第{idx + 2}行：{str(e)}')

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'导入完成，成功{success_count}条，失败{error_count}条',
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导入失败：{str(e)}'}), 500


# 商品导入导出
@app.route('/api/products/export', methods=['GET'])
@jwt_required()
def export_products():
    """导出商品"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导出功能不可用，请安装pandas'}), 500

    products = ProductInfo.query.filter_by(status=1).all()

    data = []
    for p in products:
        data.append({
            '商品编码': p.product_code,
            '条形码': p.barcode,
            '商品名称': p.product_name,
            '分类': p.category_name,
            '品牌': p.brand,
            '规格': p.specification,
            '单位': p.unit_name,
            '采购价': float(p.purchase_price) if p.purchase_price else 0.00,
            '销售价': float(p.sale_price) if p.sale_price else 0.00,
            '会员价': float(p.member_price) if p.member_price else 0.00,
            '批发价': float(p.wholesale_price) if p.wholesale_price else 0.00,
            '成本价': float(p.cost_price) if p.cost_price else 0.00,
            '库存下限': p.min_stock,
            '库存上限': p.max_stock,
            '当前库存': p.current_stock,
            '是否序列号管理': '是' if p.is_serial else '否',
            '是否批次管理': '是' if p.is_batch else '否',
            '备注': p.remark
        })

    columns = ['商品编码', '条形码', '商品名称', '分类', '品牌', '规格', '单位',
               '采购价', '销售价', '会员价', '批发价', '成本价',
               '库存下限', '库存上限', '当前库存',
               '是否序列号管理', '是否批次管理', '备注']

    return export_to_excel(data, columns, f'商品列表_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/api/products/import', methods=['POST'])
@jwt_required()
def import_products():
    """导入商品"""
    if not EXCEL_SUPPORT:
        return jsonify({'code': 500, 'message': 'Excel导入功能不可用，请安装openpyxl'}), 500

    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    # Excel 文件白名单
    _ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if _ext not in {'xlsx', 'xls', 'csv'}:
        return jsonify({'code': 400, 'message': f'仅支持 xlsx/xls/csv 格式，当前: {_ext}'}), 400
    if (file.mimetype or '').lower() not in {
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/octet-stream',
        'text/csv',
    }:
        return jsonify({'code': 400, 'message': '文件 MIME 类型不符'}), 400

    try:
        data, error = read_excel_data(file)
        if error:
            return jsonify({'code': 400, 'message': error}), 400

        success_count = 0
        error_count = 0
        errors = []

        for idx, row in enumerate(data):
            try:
                product_name = row.get('商品名称')
                if not product_name or str(product_name).strip() == '':
                    error_count += 1
                    errors.append(f'第{idx + 2}行：商品名称不能为空')
                    continue

                last_product = ProductInfo.query.order_by(ProductInfo.id.desc()).first()
                product_code = generate_code('P', last_product.id if last_product else 0)

                def get_val(row, key):
                    v = row.get(key)
                    if v is None:
                        return None
                    s = str(v).strip()
                    return s if s else None

                def get_float(row, key, default=0.0):
                    v = row.get(key)
                    if v is None:
                        return default
                    try:
                        return float(v)
                    except:
                        return default

                def get_int(row, key, default=0):
                    v = row.get(key)
                    if v is None:
                        return default
                    try:
                        return int(float(v))
                    except:
                        return default

                is_serial = 1 if get_val(row, '是否序列号管理') == '是' else 0
                is_batch = 1 if get_val(row, '是否批次管理') == '是' else 0

                product = ProductInfo(
                    product_code=product_code,
                    barcode=get_val(row, '条形码'),
                    product_name=str(product_name).strip(),
                    pinyin_code=generate_pinyin_code(str(product_name)),
                    category_name=get_val(row, '分类'),
                    brand=get_val(row, '品牌'),
                    specification=get_val(row, '规格'),
                    unit_name=get_val(row, '单位'),
                    purchase_price=get_float(row, '采购价'),
                    sale_price=get_float(row, '销售价'),
                    member_price=get_float(row, '会员价'),
                    wholesale_price=get_float(row, '批发价'),
                    cost_price=get_float(row, '成本价'),
                    min_stock=get_int(row, '库存下限'),
                    max_stock=get_int(row, '库存上限'),
                    is_serial=is_serial,
                    is_batch=is_batch,
                    remark=get_val(row, '备注')
                )

                db.session.add(product)
                db.session.flush()

                # 创建库存记录
                stock = InventoryStock(
                    product_id=product.id,
                    product_code=product.product_code,
                    product_name=product.product_name,
                    quantity=0,
                    frozen_quantity=0,
                    available_quantity=0,
                    cost_price=product.cost_price
                )
                db.session.add(stock)

                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f'第{idx + 2}行：{str(e)}')

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'导入完成，成功{success_count}条，失败{error_count}条',
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导入失败：{str(e)}'}), 500


# ============================================
# 接件配件API
# ============================================

@app.route('/api/receiveorders/<int:receive_id>/accessories', methods=['GET'])
@jwt_required()
def get_accessories(receive_id):
    """获取接件单配件列表"""
    accessories = DeviceAccessory.query.filter_by(receive_order_id=receive_id).all()
    return jsonify({'code': 200, 'data': [to_dict(a) for a in accessories]})

@app.route('/api/receiveorders/<int:receive_id>/accessories', methods=['POST'])
@jwt_required()
def add_accessory(receive_id):
    """添加配件"""
    data = request.get_json()
    accessory = DeviceAccessory(
        receive_order_id=receive_id,
        accessory_name=data.get('accessory_name'),
        quantity=data.get('quantity', 1),
        status=data.get('status', '完好'),
        remark=data.get('remark')
    )
    db.session.add(accessory)
    db.session.commit()
    return jsonify({'code': 200, 'message': '添加成功', 'data': to_dict(accessory)})

@app.route('/api/receiveorders/accessories/<int:accessory_id>', methods=['PUT'])
@jwt_required()
def update_accessory(accessory_id):
    """更新配件"""
    accessory = DeviceAccessory.query.get_or_404(accessory_id)
    data = request.get_json()
    if 'accessory_name' in data: accessory.accessory_name = data['accessory_name']
    if 'quantity' in data: accessory.quantity = data['quantity']
    if 'status' in data: accessory.status = data['status']
    if 'remark' in data: accessory.remark = data['remark']
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/receiveorders/accessories/<int:accessory_id>', methods=['DELETE'])
@jwt_required()
def delete_accessory(accessory_id):
    """删除配件"""
    accessory = DeviceAccessory.query.get_or_404(accessory_id)
    db.session.delete(accessory)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# 接件照片API
# ============================================

@app.route('/api/receiveorders/<int:receive_id>/photos', methods=['GET'])
@jwt_required()
def get_photos(receive_id):
    """获取接件单照片列表"""
    photos = DevicePhoto.query.filter_by(receive_order_id=receive_id).all()
    return jsonify({'code': 200, 'data': [to_dict(p) for p in photos]})

@app.route('/api/receiveorders/<int:receive_id>/photos', methods=['POST'])
@jwt_required()
def upload_photo(receive_id):
    """上传接件照片（带白名单校验）"""
    photo_type = request.form.get('photo_type', '整体照')
    remark = request.form.get('remark', '')
    file = request.files.get('file')
    if not file:
        return jsonify({'code': 400, 'message': '请选择文件'}), 400

    # 文件白名单校验：扩展名 + MIME
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in {'png', 'jpg', 'jpeg', 'gif'}:
        return jsonify({'code': 400, 'message': f'不支持的图片格式: {ext}'}), 400
    if (file.mimetype or '').lower() not in {'image/png', 'image/jpeg', 'image/gif'}:
        return jsonify({'code': 400, 'message': '文件 MIME 类型不符'}), 400

    # 用 UUID 重命名，避免路径穿越和文件名冲突
    import uuid as _uuid
    safe_name = f'{_uuid.uuid4().hex}.{ext}'
    upload_dir = os.path.join('uploads', 'receive_photos', str(receive_id))
    os.makedirs(upload_dir, exist_ok=True)
    filepath = os.path.join(upload_dir, safe_name)
    file.save(filepath)
    photo_url = f"/uploads/receive_photos/{receive_id}/{safe_name}"
    photo = DevicePhoto(receive_order_id=receive_id, photo_type=photo_type, photo_url=photo_url, remark=remark)
    db.session.add(photo)
    db.session.commit()
    return jsonify({'code': 200, 'message': '上传成功', 'data': to_dict(photo)})

# ============================================
# 接件签字API
# ============================================

@app.route('/api/receiveorders/<int:receive_id>/sign', methods=['POST'])
@jwt_required()
def customer_sign(receive_id):
    """客户签字"""
    order = ReceiveOrder.query.get_or_404(receive_id)
    data = request.get_json()
    order.customer_sign = data.get('sign_url')
    order.sign_time = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '签字成功'})

# ============================================
# 派单API
# ============================================

@app.route('/api/dispatch/manual', methods=['POST'])
@jwt_required()
def manual_dispatch():
    """手动派单"""
    data = request.get_json()
    wo_id = data.get('wo_id')
    staff_id = data.get('staff_id')
    remark = data.get('remark', '')
    
    wo = WorkOrder.query.get_or_404(wo_id)
    staff = SysUser.query.get_or_404(staff_id)
    
    # 创建派单记录
    record = DispatchRecord(
        wo_id=wo_id,
        dispatch_type='manual',
        dispatcher_id=get_current_user_id(),
        dispatcher_name=get_current_user_name(),
        staff_id=staff_id,
        staff_name=staff.real_name or staff.username,
        staff_phone=staff.phone,
        remark=remark
    )
    db.session.add(record)
    
    # 更新工单
    wo.assigned_user_id = staff_id
    wo.assigned_user_name = staff.real_name or staff.username
    wo.assigned_time = datetime.now()
    wo.status = 1  # 已派单
    wo.status_name = '已派单'
    
    # 记录日志
    log = DispatchLog(wo_id=wo_id, action='派单', content=f"派单给{staff.real_name or staff.username}", operator_id=get_current_user_id(), operator_name=get_current_user_name())
    db.session.add(log)
    
    db.session.commit()
    return jsonify({'code': 200, 'message': '派单成功'})

@app.route('/api/dispatch/<int:record_id>/accept', methods=['POST'])
@jwt_required()
def accept_dispatch(record_id):
    """技术员接单"""
    record = DispatchRecord.query.get_or_404(record_id)
    record.accept_status = 1
    record.accept_time = datetime.now()
    
    wo = WorkOrder.query.get(record.wo_id)
    wo.status = 3  # 处理中
    wo.status_name = '处理中'
    
    log = DispatchLog(wo_id=record.wo_id, action='接单', content=f"{record.staff_name}已接单", operator_id=record.staff_id, operator_name=record.staff_name)
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '接单成功'})

@app.route('/api/dispatch/<int:record_id>/reject', methods=['POST'])
@jwt_required()
def reject_dispatch(record_id):
    """技术员拒单"""
    data = request.get_json()
    record = DispatchRecord.query.get_or_404(record_id)
    record.accept_status = 2
    record.reject_reason = data.get('reason', '')
    
    log = DispatchLog(wo_id=record.wo_id, action='拒单', content=f"{record.staff_name}拒单: {data.get('reason', '')}", operator_id=record.staff_id, operator_name=record.staff_name)
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '拒单成功'})

@app.route('/api/dispatch/<int:record_id>/arrive', methods=['POST'])
@jwt_required()
def arrive_dispatch(record_id):
    """确认到达"""
    record = DispatchRecord.query.get_or_404(record_id)
    record.arrive_time = datetime.now()
    log = DispatchLog(wo_id=record.wo_id, action='到达', content=f"{record.staff_name}已到达客户现场", operator_id=record.staff_id, operator_name=record.staff_name)
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '已确认到达'})

@app.route('/api/dispatch/<int:record_id>/finish', methods=['POST'])
@jwt_required()
def finish_dispatch(record_id):
    """确认完成"""
    record = DispatchRecord.query.get_or_404(record_id)
    record.finish_time = datetime.now()
    wo = WorkOrder.query.get(record.wo_id)
    wo.status = 5  # 已完成
    wo.status_name = '已完成'
    wo.actual_time = datetime.now()
    log = DispatchLog(wo_id=record.wo_id, action='完成', content=f"{record.staff_name}已完成工单", operator_id=record.staff_id, operator_name=record.staff_name)
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '工单已完成'})

@app.route('/api/dispatch/<int:record_id>/redirect', methods=['POST'])
@jwt_required()
def redirect_dispatch(record_id):
    """改派"""
    data = request.get_json()
    record = DispatchRecord.query.get_or_404(record_id)
    record.accept_status = 2
    record.reject_reason = '改派'
    
    new_staff_id = data.get('staff_id')
    new_staff = SysUser.query.get_or_404(new_staff_id)
    
    new_record = DispatchRecord(
        wo_id=record.wo_id,
        dispatch_type='manual',
        dispatcher_id=get_current_user_id(),
        dispatcher_name=get_current_user_name(),
        staff_id=new_staff_id,
        staff_name=new_staff.real_name or new_staff.username,
        staff_phone=new_staff.phone,
        remark=data.get('remark', '改派')
    )
    db.session.add(new_record)
    
    wo = WorkOrder.query.get(record.wo_id)
    wo.assigned_user_id = new_staff_id
    wo.assigned_user_name = new_staff.real_name or new_staff.username
    
    log = DispatchLog(wo_id=record.wo_id, action='改派', content=f"从{record.staff_name}改派给{new_staff.real_name or new_staff.username}", operator_id=get_current_user_id(), operator_name=get_current_user_name())
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '改派成功'})

@app.route('/api/dispatch/records', methods=['GET'])
@jwt_required()
def get_all_dispatch_records():
    """获取所有派单记录列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', '')
    
    query = DispatchRecord.query
    if keyword:
        query = query.join(WorkOrder, DispatchRecord.wo_id == WorkOrder.id).filter(
            db.or_(
                WorkOrder.wo_no.contains(keyword),
                WorkOrder.customer_name.contains(keyword),
                DispatchRecord.staff_name.contains(keyword)
            )
        )
    if status:
        query = query.filter(DispatchRecord.accept_status == int(status))
    
    total = query.count()
    records = query.order_by(DispatchRecord.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    # 补充工单信息
    result = []
    for r in records:
        d = to_dict(r)
        wo = WorkOrder.query.get(r.wo_id)
        if wo:
            d['wo_type'] = wo.wo_type
            d['customer_name'] = wo.customer_name
            d['customer_phone'] = wo.customer_phone
        result.append(d)
    
    return jsonify({
        'code': 200,
        'data': {
            'list': result,
            'total': total
        }
    })

@app.route('/api/dispatch/records/<int:wo_id>', methods=['GET'])
@jwt_required()
def get_dispatch_records(wo_id):
    """获取工单派单记录"""
    records = DispatchRecord.query.filter_by(wo_id=wo_id).order_by(DispatchRecord.created_at.desc()).all()
    return jsonify({'code': 200, 'data': [to_dict(r) for r in records]})

@app.route('/api/dispatch/logs/<int:wo_id>', methods=['GET'])
@jwt_required()
def get_dispatch_logs(wo_id):
    """获取工单派单日志"""
    logs = DispatchLog.query.filter_by(wo_id=wo_id).order_by(DispatchLog.created_at.desc()).all()
    return jsonify({'code': 200, 'data': [to_dict(l) for l in logs]})

@app.route('/api/dispatch/staff-list', methods=['GET'])
@jwt_required()
def get_staff_list():
    """获取技术员列表（用于派单选择）"""
    staff_list = SysUser.query.filter(SysUser.status == 1).all()
    result = []
    for s in staff_list:
        # 查询技术员状态
        ss = StaffStatus.query.filter_by(staff_id=s.id).first()
        # 查询今日派单数
        today = datetime.now().date()
        today_count = DispatchRecord.query.filter(
            DispatchRecord.staff_id == s.id,
            DispatchRecord.dispatch_time >= datetime.combine(today, datetime.min.time()),
            DispatchRecord.accept_status.in_([0, 1])
        ).count()
        result.append({
            'id': s.id,
            'name': s.real_name or s.username,
            'phone': s.phone,
            'online_status': ss.online_status if ss else 0,
            'today_count': today_count,
            'max_daily': ss.max_daily if ss else 10,
            'skills': ss.skills if ss else '',
            'rating': float(ss.rating) if ss else 5.0
        })
    return jsonify({'code': 200, 'data': result})

@app.route('/api/dispatch/pending', methods=['GET'])
@jwt_required()
def get_pending_dispatch():
    """获取待派单工单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    keyword = request.args.get('keyword', '')

    query = WorkOrder.query.filter(WorkOrder.status == 0)
    if keyword:
        query = query.filter(
            db.or_(
                WorkOrder.wo_no.contains(keyword),
                WorkOrder.customer_name.contains(keyword),
                WorkOrder.customer_phone.contains(keyword)
            )
        )

    total = query.count()
    orders = query.order_by(WorkOrder.priority.desc(), WorkOrder.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    # 批量获取关联信息
    order_ids = [o.id for o in orders]

    # 查询员工名称
    user_ids = set()
    for o in orders:
        if o.reception_user_id: user_ids.add(o.reception_user_id)
        if o.engineer_user_id: user_ids.add(o.engineer_user_id)
    users_map = {}
    if user_ids:
        for u in SysUser.query.filter(SysUser.id.in_(user_ids)).all():
            users_map[u.id] = u.real_name or u.username

    result = []
    for o in orders:
        d = to_dict(o)
        # 关联接待员工名称
        d['reception_user_name'] = users_map.get(o.reception_user_id, '')
        # 关联维修工程师名称
        d['engineer_user_name'] = users_map.get(o.engineer_user_id, '')
        # 设备信息直接从工单获取
        d['device_brand'] = o.device_brand or ''
        d['device_model'] = o.device_model or ''
        d['device_sn'] = o.device_sn or ''
        d['device_type'] = o.device_type or ''
        result.append(d)

    return jsonify({
        'code': 200,
        'data': {
            'list': result,
            'total': total
        }
    })

@app.route('/api/dispatch/stats', methods=['GET'])
@jwt_required()
def get_dispatch_stats():
    """获取派单统计"""
    today = datetime.now().date()
    today_start = datetime.combine(today, datetime.min.time())
    
    total_dispatch = DispatchRecord.query.filter(DispatchRecord.dispatch_time >= today_start).count()
    accepted = DispatchRecord.query.filter(DispatchRecord.dispatch_time >= today_start, DispatchRecord.accept_status == 1).count()
    rejected = DispatchRecord.query.filter(DispatchRecord.dispatch_time >= today_start, DispatchRecord.accept_status == 2).count()
    pending = DispatchRecord.query.filter(DispatchRecord.dispatch_time >= today_start, DispatchRecord.accept_status == 0).count()
    
    return jsonify({
        'code': 200,
        'data': {
            'total_dispatch': total_dispatch,
            'accepted': accepted,
            'rejected': rejected,
            'pending': pending,
            'accept_rate': round(accepted / total_dispatch * 100, 1) if total_dispatch > 0 else 0
        }
    })

# ============================================
# 批量操作API（直接内嵌，避免文件路径问题）
# ============================================

def check_business_record(ids):
    """检查商品是否有业务记录"""
    has_record_ids = []
    in_records = InventoryInItem.query.filter(
        InventoryInItem.product_id.in_(ids)
    ).with_entities(InventoryInItem.product_id).distinct().all()
    has_record_ids.extend([r[0] for r in in_records])
    out_records = InventoryOutItem.query.filter(
        InventoryOutItem.product_id.in_(ids)
    ).with_entities(InventoryOutItem.product_id).distinct().all()
    has_record_ids.extend([r[0] for r in out_records])
    wo_records = WorkOrderPart.query.filter(
        WorkOrderPart.product_id.in_(ids)
    ).with_entities(WorkOrderPart.product_id).distinct().all()
    has_record_ids.extend([r[0] for r in wo_records])
    po_records = PurchaseOrderItem.query.filter(
        PurchaseOrderItem.product_id.in_(ids)
    ).with_entities(PurchaseOrderItem.product_id).distinct().all()
    has_record_ids.extend([r[0] for r in po_records])
    so_records = SalesOrderItem.query.filter(
        SalesOrderItem.product_id.in_(ids)
    ).with_entities(SalesOrderItem.product_id).distinct().all()
    has_record_ids.extend([r[0] for r in so_records])
    return list(set(has_record_ids))

@app.route('/api/products/batch-update-category', methods=['POST'])
@jwt_required()
def batch_update_category():
    """批量修改分类"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        category_id = data.get('category_id')
        category_name = data.get('category_name', '')
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要操作的商品'}), 400
        if not category_id:
            return jsonify({'code': 400, 'message': '请选择新分类'}), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        for product in products:
            product.category_id = category_id
            product.category_name = category_name
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功修改{len(products)}个商品的分类'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量修改分类失败，请稍后重试'}), 500

@app.route('/api/products/batch-update-price', methods=['POST'])
@jwt_required()
def batch_update_price():
    """批量修改价格"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要操作的商品'}), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        update_fields = {}
        for field in ['purchase_price', 'sale_price', 'cost_price', 'member_price', 'wholesale_price', 'customer_price']:
            if field in data and data[field] is not None:
                update_fields[field] = data[field]
        if not update_fields:
            return jsonify({'code': 400, 'message': '请至少填写一个价格字段'}), 400
        for product in products:
            for field, value in update_fields.items():
                setattr(product, field, value)
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功修改{len(products)}个商品的价格'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量修改价格失败，请稍后重试'}), 500

@app.route('/api/products/batch-update-stock-warning', methods=['POST'])
@jwt_required()
def batch_update_stock_warning():
    """批量设置库存预警"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        min_stock = data.get('min_stock')
        max_stock = data.get('max_stock')
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要操作的商品'}), 400
        if min_stock is None:
            return jsonify({'code': 400, 'message': '请设置最低库存'}), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        for product in products:
            product.min_stock = min_stock
            if max_stock is not None:
                product.max_stock = max_stock
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功设置{len(products)}个商品的库存预警'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量设置库存预警失败，请稍后重试'}), 500

@app.route('/api/products/batch-update-sort', methods=['POST'])
@jwt_required()
def batch_update_sort():
    """批量修改排序"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        sort_order = data.get('sort_order')
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要操作的商品'}), 400
        if sort_order is None:
            return jsonify({'code': 400, 'message': '请设置排序值'}), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        for product in products:
            product.sort_order = sort_order
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功修改{len(products)}个商品的排序'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量修改排序失败，请稍后重试'}), 500

@app.route('/api/products/batch-disable', methods=['POST'])
@jwt_required()
def batch_disable():
    """批量禁用"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要操作的商品'}), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        for product in products:
            product.status = 0
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功禁用{len(products)}个商品'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量禁用失败，请稍后重试'}), 500

@app.route('/api/products/batch-enable', methods=['POST'])
@jwt_required()
def batch_enable():
    """批量启用"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要操作的商品'}), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        for product in products:
            product.status = 1
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功启用{len(products)}个商品'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量启用失败，请稍后重试'}), 500

@app.route('/api/products/batch-delete', methods=['POST'])
@jwt_required()
def batch_delete_products():
    """批量删除商品"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要删除的商品'}), 400
        has_record_ids = check_business_record(ids)
        if has_record_ids:
            record_products = ProductInfo.query.filter(
                ProductInfo.id.in_(has_record_ids)
            ).with_entities(ProductInfo.product_name).all()
            names = [p[0] for p in record_products[:3]]
            return jsonify({
                'code': 400,
                'message': f'以下商品已存在业务记录，不允许删除：{"、".join(names)}{"等" if len(has_record_ids) > 3 else ""}'
            }), 400
        products = ProductInfo.query.filter(ProductInfo.id.in_(ids)).all()
        count = len(products)
        for product in products:
            db.session.delete(product)
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功删除{count}个商品'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量删除失败，请稍后重试'}), 500

@app.route('/api/customers/batch-delete', methods=['POST'])
@jwt_required()
def batch_delete_customers():
    """批量删除客户"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要删除的客户'}), 400
        customers = BaseCustomer.query.filter(BaseCustomer.id.in_(ids)).all()
        count = len(customers)
        for customer in customers:
            db.session.delete(customer)
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功删除{count}个客户'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量删除客户失败，请稍后重试'}), 500

@app.route('/api/suppliers/batch-delete', methods=['POST'])
@jwt_required()
def batch_delete_suppliers():
    """批量删除供应商"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要删除的供应商'}), 400
        suppliers = BaseSupplier.query.filter(BaseSupplier.id.in_(ids)).all()
        count = len(suppliers)
        for supplier in suppliers:
            db.session.delete(supplier)
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功删除{count}个供应商'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量删除供应商失败，请稍后重试'}), 500

@app.route('/api/workorders/batch-delete', methods=['POST'])
@jwt_required()
def batch_delete_workorders():
    """批量删除工单"""
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        if not ids:
            return jsonify({'code': 400, 'message': '请选择要删除的工单'}), 400
        # 检查是否有已完成的工单
        completed = WorkOrder.query.filter(WorkOrder.id.in_(ids), WorkOrder.status == 6).all()
        if completed:
            return jsonify({'code': 400, 'message': '选中的工单中包含已完成的工单，无法删除'}), 400
        workorders = WorkOrder.query.filter(WorkOrder.id.in_(ids)).all()
        count = len(workorders)
        for wo in workorders:
            db.session.delete(wo)
        db.session.commit()
        return jsonify({'code': 200, 'message': f'成功删除{count}个工单'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': '批量删除工单失败，请稍后重试'}), 500

# ============================================
# 启动应用
# ============================================

# 注意：原代码此处有 startup 块，会调用 db.create_all() 和 init_data()。
# 这两个调用：
#   1. init_data 在文件下方才定义（L15285），NameError 永远触发
#   2. 与阶段 3 启用的 flask-migrate 流程冲突
# 已删除。建表改用 `flask db upgrade`；种子数据改由 setup 端点触发（阶段 0/3）。

# ============================================
# API路由 - 预订单管理
# ============================================

@app.route('/api/pre-orders', methods=['GET'])
@jwt_required()
def get_preorders():
    """获取预订单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    keyword = request.args.get('keyword', '')
    pre_type = request.args.get('pre_type', type=int)
    status = request.args.get('status', type=int)

    query = PreOrder.query
    if keyword:
        query = query.filter(db.or_(
            PreOrder.pre_no.contains(keyword),
            PreOrder.customer_name.contains(keyword),
            PreOrder.supplier_name.contains(keyword)
        ))
    if pre_type is not None:
        query = query.filter(PreOrder.pre_type == pre_type)
    if status is not None:
        query = query.filter(PreOrder.status == status)

    total = query.count()
    orders = query.order_by(PreOrder.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders],
            'total': total
        }
    })

@app.route('/api/pre-orders/<int:id>', methods=['GET'])
@jwt_required()
def get_preorder(id):
    """获取预订单详情"""
    order = PreOrder.query.get_or_404(id)
    items = PreOrderItem.query.filter_by(pre_id=id).all()
    order_dict = to_dict(order)
    order_dict['items'] = [to_dict(i) for i in items]
    return jsonify({'code': 200, 'data': order_dict})

@app.route('/api/pre-orders', methods=['POST'])
@jwt_required()
def create_preorder():
    """创建预订单"""
    data = request.get_json()
    user_id = get_jwt_identity()

    last = PreOrder.query.order_by(PreOrder.id.desc()).first()
    pre_no = generate_code('PRE', last.id if last else 0)

    order = PreOrder(
        pre_no=pre_no,
        pre_type=data.get('pre_type', 1),
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        supplier_id=data.get('supplier_id'),
        supplier_name=data.get('supplier_name'),
        remark=data.get('remark'),
        created_by=user_id
    )
    db.session.add(order)
    db.session.flush()

    # 添加明细
    items = data.get('items', [])
    total_quantity = 0
    total_amount = 0
    for item_data in items:
        item = PreOrderItem(
            pre_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            quantity=item_data.get('quantity', 0),
            unit_price=item_data.get('unit_price', 0),
            total_price=item_data.get('quantity', 0) * item_data.get('unit_price', 0),
            remark=item_data.get('remark')
        )
        db.session.add(item)
        total_quantity += item_data.get('quantity', 0)
        total_amount += item.total_price

    order.total_quantity = total_quantity
    order.total_amount = total_amount
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'pre_no': pre_no}})

@app.route('/api/pre-orders/<int:id>', methods=['PUT'])
@jwt_required()
def update_preorder(id):
    """更新预订单"""
    order = PreOrder.query.get_or_404(id)
    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待处理的预订单可以修改'}), 400

    data = request.get_json()
    if 'customer_name' in data: order.customer_name = data['customer_name']
    if 'supplier_name' in data: order.supplier_name = data['supplier_name']
    if 'remark' in data: order.remark = data['remark']

    # 更新明细
    if 'items' in data:
        PreOrderItem.query.filter_by(pre_id=id).delete()
        total_quantity = 0
        total_amount = 0
        for item_data in data['items']:
            item = PreOrderItem(
                pre_id=id,
                product_id=item_data.get('product_id'),
                product_code=item_data.get('product_code'),
                product_name=item_data.get('product_name'),
                quantity=item_data.get('quantity', 0),
                unit_price=item_data.get('unit_price', 0),
                total_price=item_data.get('quantity', 0) * item_data.get('unit_price', 0),
                remark=item_data.get('remark')
            )
            db.session.add(item)
            total_quantity += item_data.get('quantity', 0)
            total_amount += item.total_price
        order.total_quantity = total_quantity
        order.total_amount = total_amount

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})

@app.route('/api/pre-orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_preorder(id):
    """删除预订单"""
    order = PreOrder.query.get_or_404(id)
    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待处理的预订单可以删除'}), 400
    order.status = 2
    db.session.commit()
    return jsonify({'code': 200, 'message': '已取消'})

@app.route('/api/pre-orders/<int:id>/convert', methods=['POST'])
@jwt_required()
def convert_preorder(id):
    """预订单转正式单据"""
    order = PreOrder.query.get_or_404(id)
    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待处理的预订单可以转单'}), 400

    data = request.get_json()

    if order.pre_type == 1:
        # 采购预定 -> 采购单
        from app import PurchaseOrder, PurchaseOrderItem
        last_po = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
        po_no = generate_code('PO', last_po.id if last_po else 0)

        po = PurchaseOrder(
            order_no=po_no,
            supplier_id=order.supplier_id,
            supplier_name=order.supplier_name,
            total_amount=order.total_amount,
            status=0,
            remark=f'由预订单{order.pre_no}转换',
            created_by=get_jwt_identity()
        )
        db.session.add(po)
        db.session.flush()

        items = PreOrderItem.query.filter_by(pre_id=id).all()
        for item in items:
            po_item = PurchaseOrderItem(
                order_id=po.id,
                product_id=item.product_id,
                product_name=item.product_name,
                quantity=int(item.quantity) if item.quantity else 0,
                price=float(item.unit_price) if item.unit_price else 0,
                amount=float(item.total_price) if item.total_price else 0
            )
            db.session.add(po_item)

        order.related_order_id = po.id
        order.related_order_no = po_no

    elif order.pre_type == 2:
        # 销售预定 -> 销售单
        from app import SalesOrder, SalesOrderItem
        last_so = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
        so_no = generate_code('SO', last_so.id if last_so else 0)

        so = SalesOrder(
            order_no=so_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            total_amount=order.total_amount,
            total_quantity=int(order.total_quantity),
            status=0,
            remark=f'由预订单{order.pre_no}转换',
            created_by=get_jwt_identity()
        )
        db.session.add(so)
        db.session.flush()

        items = PreOrderItem.query.filter_by(pre_id=id).all()
        for item in items:
            so_item = SalesOrderItem(
                order_id=so.id,
                product_id=item.product_id,
                product_name=item.product_name,
                specification='',
                unit='个',
                quantity=int(item.quantity),
                price=item.unit_price,
                amount=item.total_price
            )
            db.session.add(so_item)

        order.related_order_id = so.id
        order.related_order_no = so_no

    order.status = 1
    db.session.commit()

    return jsonify({'code': 200, 'message': '转单成功', 'data': {'related_order_no': order.related_order_no}})

# ============================================
# API路由 - 退货管理
# ============================================

@app.route('/api/return-orders', methods=['GET'])
@jwt_required()
def get_return_orders():
    """获取退货单列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    keyword = request.args.get('keyword', '')
    return_type = request.args.get('return_type', type=int)
    status = request.args.get('status', type=int)
    
    query = ReturnOrder.query
    if keyword:
        query = query.filter(db.or_(
            ReturnOrder.return_no.contains(keyword),
            ReturnOrder.customer_name.contains(keyword),
            ReturnOrder.supplier_name.contains(keyword)
        ))
    if return_type is not None:
        query = query.filter(ReturnOrder.return_type == return_type)
    if status is not None:
        query = query.filter(ReturnOrder.status == status)
    
    total = query.count()
    orders = query.order_by(ReturnOrder.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return jsonify({'code': 200, 'data': {'list': [to_dict(o) for o in orders], 'total': total}})

@app.route('/api/return-orders/<int:id>', methods=['GET'])
@jwt_required()
def get_return_order(id):
    """获取退货单详情"""
    order = ReturnOrder.query.get_or_404(id)
    items = ReturnOrderItem.query.filter_by(return_id=id).all()
    order_dict = to_dict(order)
    order_dict['items'] = [to_dict(i) for i in items]
    return jsonify({'code': 200, 'data': order_dict})

@app.route('/api/return-orders', methods=['POST'])
@jwt_required()
def create_return_order():
    """创建退货单"""
    data = request.get_json()
    user_id = get_jwt_identity()
    
    last = ReturnOrder.query.order_by(ReturnOrder.id.desc()).first()
    return_no = generate_code('R', last.id if last else 0)
    
    order = ReturnOrder(
        return_no=return_no,
        return_type=data.get('return_type', 1),
        related_order_id=data.get('related_order_id'),
        related_order_no=data.get('related_order_no'),
        supplier_id=data.get('supplier_id'),
        supplier_name=data.get('supplier_name'),
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        refund_amount=data.get('refund_amount', 0),
        reason=data.get('reason'),
        remark=data.get('remark'),
        created_by=user_id
    )
    db.session.add(order)
    db.session.flush()
    
    items = data.get('items', [])
    total_quantity = 0
    total_amount = 0
    for item_data in items:
        qty = item_data.get('quantity', 0)
        price = item_data.get('unit_price', 0)
        item = ReturnOrderItem(
            return_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            quantity=qty,
            unit_price=price,
            total_price=qty * price,
            remark=item_data.get('remark')
        )
        db.session.add(item)
        total_quantity += qty
        total_amount += qty * price
    
    order.total_quantity = total_quantity
    order.total_amount = total_amount
    if not order.refund_amount:
        order.refund_amount = total_amount
    db.session.commit()
    
    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'return_no': return_no}})

@app.route('/api/return-orders/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_return_order(id):
    """审核退货单 - 自动处理库存、冲减应收/应付账款并生成财务流水"""
    order = ReturnOrder.query.get_or_404(id)
    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核的退货单可以审核'}), 400

    try:
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        refund_amount = float(order.refund_amount or order.total_amount or 0)
        today = datetime.now().strftime('%Y%m%d')

        items = ReturnOrderItem.query.filter_by(return_id=id).all()

        if order.return_type == 1:
            # ========== 采购退货：自动出库扣减库存 ==========
            for item in items:
                # 检查库存是否充足
                stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
                available = (stock.available_quantity or 0) if stock else 0
                if available < (item.quantity or 0):
                    return jsonify({'code': 400, 'message': f'商品"{item.product_name}"库存不足，可用库存{available}，需要{item.quantity}'}), 400

            # 生成出库单号 CK + YYYYMMDD + 4位序号
            prefix_ck = 'CK' + today
            last_out = InventoryOut.query.filter(InventoryOut.out_no.like(prefix_ck + '%')).order_by(InventoryOut.out_no.desc()).first()
            if last_out and last_out.out_no and len(last_out.out_no) > len(prefix_ck):
                seq = int(last_out.out_no[len(prefix_ck):]) + 1
            else:
                seq = 1
            out_no = prefix_ck + str(seq).zfill(4)

            # 创建出库单
            inventory_out = InventoryOut(
                out_no=out_no,
                out_type=5,  # 其他出库（采购退货）
                warehouse_id=1,
                warehouse_name='主仓库',
                total_quantity=sum(i.quantity or 0 for i in items),
                total_amount=sum((i.quantity or 0) * (i.unit_price or 0) for i in items),
                status=2,  # 已出库
                related_order_id=id,
                related_order_no=order.return_no,
                remark=f'采购退货单{order.return_no}出库',
                created_by=user_id,
                created_at=datetime.now()
            )
            db.session.add(inventory_out)
            db.session.flush()

            # 处理出库明细并扣减库存
            for item in items:
                # 创建出库明细
                out_item = InventoryOutItem(
                    out_id=inventory_out.id,
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    quantity=item.quantity,
                    unit_price=item.unit_price,
                    total_price=item.total_price,
                    remark=f'采购退货'
                )
                db.session.add(out_item)

                # 扣减库存
                stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
                before_qty = stock.quantity if stock else 0
                if stock:
                    stock.quantity = (stock.quantity or 0) - (item.quantity or 0)
                    stock.available_quantity = (stock.available_quantity or 0) - (item.quantity or 0)
                after_qty = (stock.quantity or 0) if stock else 0

                # 写入库存日志
                inv_log = InventoryLog(
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    warehouse_id=1,
                    warehouse_name='主仓库',
                    change_type='out',
                    order_type='采购退货出库',
                    order_id=inventory_out.id,
                    order_no=out_no,
                    quantity_change=-(item.quantity or 0),
                    before_quantity=before_qty,
                    after_quantity=after_qty,
                    cost_price=item.unit_price,
                    amount=item.total_price,
                    related_party=order.supplier_name,
                    operator_id=user_id,
                    operator_name=user_name,
                    remark=f'采购退货单{order.return_no}出库'
                )
                db.session.add(inv_log)

                # 更新商品现存量
                if item.product_id:
                    product_info = ProductInfo.query.get(item.product_id)
                    if product_info:
                        product_info.current_stock = max(0, (product_info.current_stock or 0) - (item.quantity or 0))

            # 冲减应付账款
            payable = FinancePayable.query.filter_by(
                related_type='purchase',
                related_id=order.related_order_id
            ).first()
            if payable and payable.status != 3:
                payable.total_amount = max(0, float(payable.total_amount or 0) - refund_amount)
                payable.remaining_amount = max(0, float(payable.remaining_amount or 0) - refund_amount)
                if payable.remaining_amount <= 0:
                    payable.status = 2
                elif float(payable.paid_amount or 0) > 0:
                    payable.status = 1
                else:
                    payable.status = 0

            # 生成财务流水
            finance_record = FinanceRecord(
                account_id=None,
                account_name='',
                record_type=1,
                amount=refund_amount,
                balance_before=0,
                balance_after=0,
                related_type='return_purchase',
                related_id=id,
                related_no=order.return_no,
                remark=f'采购退货单{order.return_no}审核，冲减应付{refund_amount}元',
                created_at=datetime.now(),
                created_by=user_id
            )
            db.session.add(finance_record)

            order.status = 2  # 已完成（已出库）

        elif order.return_type == 2:
            # ========== 销售退货：自动入库增加库存 ==========
            # 生成入库单号 RK + YYYYMMDD + 4位序号
            prefix_rk = 'RK' + today
            last_in = InventoryIn.query.filter(InventoryIn.in_no.like(prefix_rk + '%')).order_by(InventoryIn.in_no.desc()).first()
            if last_in and last_in.in_no and len(last_in.in_no) > len(prefix_rk):
                seq = int(last_in.in_no[len(prefix_rk):]) + 1
            else:
                seq = 1
            in_no = prefix_rk + str(seq).zfill(4)

            # 创建入库单
            inventory_in = InventoryIn(
                in_no=in_no,
                in_type=2,  # 退货入库
                warehouse_id=1,
                warehouse_name='主仓库',
                total_quantity=sum(i.quantity or 0 for i in items),
                total_amount=sum((i.quantity or 0) * (i.unit_price or 0) for i in items),
                status=2,  # 已入库
                related_order_id=id,
                related_order_no=order.return_no,
                remark=f'销售退货单{order.return_no}入库',
                created_by=user_id,
                created_at=datetime.now()
            )
            db.session.add(inventory_in)
            db.session.flush()

            # 处理入库明细并增加库存
            for item in items:
                # 创建入库明细
                in_item = InventoryInItem(
                    in_id=inventory_in.id,
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    quantity=item.quantity,
                    unit_price=item.unit_price,
                    total_price=item.total_price,
                    remark=f'销售退货'
                )
                db.session.add(in_item)

                # 增加库存
                stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
                before_qty = (stock.quantity or 0) if stock else 0
                if stock:
                    stock.quantity = (stock.quantity or 0) + (item.quantity or 0)
                    stock.available_quantity = (stock.available_quantity or 0) + (item.quantity or 0)
                else:
                    stock = InventoryStock(
                        product_id=item.product_id,
                        product_code=item.product_code,
                        product_name=item.product_name,
                        warehouse_id=1,
                        warehouse_name='主仓库',
                        quantity=item.quantity or 0,
                        available_quantity=item.quantity or 0,
                        cost_price=item.unit_price or 0
                    )
                    db.session.add(stock)
                    before_qty = 0

                after_qty = (stock.quantity or 0)

                # 写入库存日志
                inv_log = InventoryLog(
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    warehouse_id=1,
                    warehouse_name='主仓库',
                    change_type='in',
                    order_type='销售退货入库',
                    order_id=inventory_in.id,
                    order_no=in_no,
                    quantity_change=item.quantity or 0,
                    before_quantity=before_qty,
                    after_quantity=after_qty,
                    cost_price=item.unit_price,
                    amount=item.total_price,
                    related_party=order.customer_name,
                    operator_id=user_id,
                    operator_name=user_name,
                    remark=f'销售退货单{order.return_no}入库'
                )
                db.session.add(inv_log)

                # 更新商品现存量
                if item.product_id:
                    product_info = ProductInfo.query.get(item.product_id)
                    if product_info:
                        product_info.current_stock = (product_info.current_stock or 0) + (item.quantity or 0)

            # 冲减应收账款
            receivable = FinanceReceivable.query.filter_by(
                related_type='sale',
                related_id=order.related_order_id
            ).first()
            if receivable and receivable.status != 3:
                receivable.total_amount = max(0, float(receivable.total_amount or 0) - refund_amount)
                receivable.remaining_amount = max(0, float(receivable.remaining_amount or 0) - refund_amount)
                if receivable.remaining_amount <= 0:
                    receivable.status = 2
                elif float(receivable.received_amount or 0) > 0:
                    receivable.status = 1
                else:
                    receivable.status = 0

            # 生成财务流水
            finance_record = FinanceRecord(
                account_id=None,
                account_name='',
                record_type=2,
                amount=refund_amount,
                balance_before=0,
                balance_after=0,
                related_type='return_sale',
                related_id=id,
                related_no=order.return_no,
                remark=f'销售退货单{order.return_no}审核，冲减应收{refund_amount}元',
                created_at=datetime.now(),
                created_by=user_id
            )
            db.session.add(finance_record)

            order.status = 2  # 已完成（已入库）

        db.session.commit()
        return jsonify({'code': 200, 'message': '审核成功'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'审核失败：{str(e)}'}), 500

@app.route('/api/return-orders/<int:id>/stock-in', methods=['POST'])
@jwt_required()
def return_order_stock_in(id):
    """退货入库 - 更新库存、写入库存日志、更新商品现存量"""
    order = ReturnOrder.query.get_or_404(id)
    if order.status != 1:
        return jsonify({'code': 400, 'message': '只有已审核的退货单可以入库'}), 400

    try:
        user_id = get_jwt_identity()
        user_name = get_current_user_name()
        order_type_text = '采购退货入库' if order.return_type == 1 else '销售退货入库'
        related_party = order.supplier_name if order.return_type == 1 else order.customer_name

        items = ReturnOrderItem.query.filter_by(return_id=id).all()
        for item in items:
            # 更新库存
            stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
            before_qty = (stock.quantity or 0) if stock else 0
            if stock:
                stock.quantity = (stock.quantity or 0) + item.quantity
                stock.available_quantity = (stock.available_quantity or 0) + item.quantity
            else:
                # 自动创建库存记录
                product = ProductInfo.query.get(item.product_id) if item.product_id else None
                stock = InventoryStock(
                    product_id=item.product_id,
                    product_code=item.product_code or (product.product_code if product else None),
                    product_name=item.product_name,
                    warehouse_id=1,
                    warehouse_name='主仓库',
                    quantity=item.quantity or 0,
                    available_quantity=item.quantity or 0,
                    cost_price=item.unit_price or 0
                )
                db.session.add(stock)
                before_qty = 0

            after_qty = (stock.quantity or 0)

            # 记录入库流水
            in_record = InventoryIn(
                in_type=2,  # 退货入库
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                quantity=item.quantity,
                unit_price=item.unit_price,
                total_price=item.total_price,
                related_type='return_order',
                related_id=id,
                remark=f'退货单{order.return_no}入库'
            )
            db.session.add(in_record)

            # 写入库存日志
            inv_log = InventoryLog(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=1,
                warehouse_name='主仓库',
                change_type='in',
                order_type=order_type_text,
                order_id=id,
                order_no=order.return_no,
                quantity_change=item.quantity or 0,
                before_quantity=before_qty,
                after_quantity=after_qty,
                cost_price=item.unit_price,
                amount=item.total_price,
                related_party=related_party,
                operator_id=user_id,
                operator_name=user_name,
                remark=f'退货单{order.return_no}入库'
            )
            db.session.add(inv_log)

            # 更新商品现存量
            if item.product_id:
                product_info = ProductInfo.query.get(item.product_id)
                if product_info:
                    product_info.current_stock = (product_info.current_stock or 0) + (item.quantity or 0)

        order.status = 2
        db.session.commit()
        return jsonify({'code': 200, 'message': '入库成功'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'入库失败：{str(e)}'}), 500

@app.route('/api/return-orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_return_order(id):
    """取消退货单"""
    order = ReturnOrder.query.get_or_404(id)
    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核的退货单可以取消'}), 400
    order.status = 3
    db.session.commit()
    return jsonify({'code': 200, 'message': '已取消'})

# ============================================
# API路由 - 应收账款管理
# ============================================

@app.route('/api/finance/receivables', methods=['GET'])
@jwt_required()
def get_receivables():
    """应收账款列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    customer_id = request.args.get('customer_id', type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = FinanceReceivable.query

    if customer_id:
        query = query.filter_by(customer_id=customer_id)

    if status is not None:
        query = query.filter_by(status=status)

    if keyword:
        query = query.filter(
            db.or_(
                FinanceReceivable.receivable_no.contains(keyword),
                FinanceReceivable.customer_name.contains(keyword),
                FinanceReceivable.related_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(FinanceReceivable.created_at >= start_date)

    if end_date:
        query = query.filter(FinanceReceivable.created_at <= end_date + ' 23:59:59')

    total = query.count()
    receivables = query.order_by(FinanceReceivable.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    status_map = {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': r.id,
                'receivable_no': r.receivable_no,
                'related_type': r.related_type,
                'related_id': r.related_id,
                'related_no': r.related_no,
                'customer_id': r.customer_id,
                'customer_name': r.customer_name,
                'total_amount': float(r.total_amount) if r.total_amount else 0,
                'received_amount': float(r.received_amount) if r.received_amount else 0,
                'remaining_amount': float(r.remaining_amount) if r.remaining_amount else 0,
                'status': r.status,
                'status_text': status_map.get(r.status, '未知'),
                'invoice_id': r.invoice_id,
                'remark': r.remark,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None,
                'updated_at': r.updated_at.strftime('%Y-%m-%d %H:%M:%S') if r.updated_at else None
            } for r in receivables.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/finance/receivables/<int:id>', methods=['GET'])
@jwt_required()
def get_receivable(id):
    """应收账款详情"""
    r = FinanceReceivable.query.get(id)
    if not r:
        return jsonify({'code': 404, 'message': '应收记录不存在'}), 404

    status_map = {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'id': r.id,
            'receivable_no': r.receivable_no,
            'related_type': r.related_type,
            'related_id': r.related_id,
            'related_no': r.related_no,
            'customer_id': r.customer_id,
            'customer_name': r.customer_name,
            'total_amount': float(r.total_amount) if r.total_amount else 0,
            'received_amount': float(r.received_amount) if r.received_amount else 0,
            'remaining_amount': float(r.remaining_amount) if r.remaining_amount else 0,
            'status': r.status,
            'status_text': status_map.get(r.status, '未知'),
            'invoice_id': r.invoice_id,
            'remark': r.remark,
            'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None,
            'updated_at': r.updated_at.strftime('%Y-%m-%d %H:%M:%S') if r.updated_at else None
        }
    })


@app.route('/api/finance/receivables/<int:id>/receive', methods=['POST'])
@jwt_required()
def receive_receivable(id):
    """收款核销"""
    receivable = FinanceReceivable.query.get(id)
    if not receivable:
        return jsonify({'code': 404, 'message': '应收记录不存在'}), 404

    if receivable.status == 2:
        return jsonify({'code': 400, 'message': '该应收已结清'}), 400

    if receivable.status == 3:
        return jsonify({'code': 400, 'message': '该应收已取消'}), 400

    data = request.get_json()
    received_amount = data.get('received_amount')
    if not received_amount or float(received_amount) <= 0:
        return jsonify({'code': 400, 'message': '收款金额必须大于0'}), 400

    received_amount = float(received_amount)
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    try:
        # 更新应收金额
        receivable.received_amount = float(receivable.received_amount or 0) + received_amount
        receivable.remaining_amount = float(receivable.total_amount or 0) - float(receivable.received_amount or 0)

        if receivable.remaining_amount <= 0:
            receivable.remaining_amount = 0
            receivable.status = 2  # 已结清
        else:
            receivable.status = 1  # 部分收款

        receivable.updated_at = datetime.now()

        # 生成财务流水（收入）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=1,  # 收入
            amount=received_amount,
            balance_before=0,
            balance_after=0,
            related_type='receivable',
            related_id=id,
            related_no=receivable.receivable_no,
            remark=f'应收{receivable.receivable_no}收款{received_amount}元，操作人：{user_name}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        db.session.commit()
        return jsonify({'code': 200, 'message': '收款成功', 'data': {
            'received_amount': float(receivable.received_amount),
            'remaining_amount': float(receivable.remaining_amount),
            'status': receivable.status
        }})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'收款失败：{str(e)}'}), 500


@app.route('/api/finance/receivables/export', methods=['GET'])
@jwt_required()
def export_receivables():
    """导出应收账款"""
    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)
        customer_id = request.args.get('customer_id', type=int)
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        query = FinanceReceivable.query

        if customer_id:
            query = query.filter_by(customer_id=customer_id)
        if status is not None:
            query = query.filter_by(status=status)
        if keyword:
            query = query.filter(
                db.or_(
                    FinanceReceivable.receivable_no.contains(keyword),
                    FinanceReceivable.customer_name.contains(keyword),
                    FinanceReceivable.related_no.contains(keyword)
                )
            )
        if start_date:
            query = query.filter(FinanceReceivable.created_at >= start_date)
        if end_date:
            query = query.filter(FinanceReceivable.created_at <= end_date + ' 23:59:59')

        receivables = query.order_by(FinanceReceivable.created_at.desc()).all()

        status_map = {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '应收账款'

        headers = ['应收编号', '关联类型', '关联单号', '客户名称', '应收总额', '已收金额', '待收金额', '状态', '备注', '创建时间']
        ws.append(headers)

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for r in receivables:
            ws.append([
                r.receivable_no,
                r.related_type,
                r.related_no,
                r.customer_name,
                float(r.total_amount) if r.total_amount else 0,
                float(r.received_amount) if r.received_amount else 0,
                float(r.remaining_amount) if r.remaining_amount else 0,
                status_map.get(r.status, '未知'),
                r.remark or '',
                r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                         download_name=f'应收账款_{datetime.now().strftime("%Y%m%d")}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导出失败：{str(e)}'}), 500


# ============================================
# API路由 - 应付账款管理
# ============================================

@app.route('/api/finance/payables', methods=['GET'])
@jwt_required()
def get_payables():
    """应付账款列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    supplier_id = request.args.get('supplier_id', type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = FinancePayable.query

    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)

    if status is not None:
        query = query.filter_by(status=status)

    if keyword:
        query = query.filter(
            db.or_(
                FinancePayable.payable_no.contains(keyword),
                FinancePayable.supplier_name.contains(keyword),
                FinancePayable.related_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(FinancePayable.created_at >= start_date)

    if end_date:
        query = query.filter(FinancePayable.created_at <= end_date + ' 23:59:59')

    total = query.count()
    payables = query.order_by(FinancePayable.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    status_map = {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': p.id,
                'payable_no': p.payable_no,
                'related_type': p.related_type,
                'related_id': p.related_id,
                'related_no': p.related_no,
                'supplier_id': p.supplier_id,
                'supplier_name': p.supplier_name,
                'total_amount': float(p.total_amount) if p.total_amount else 0,
                'paid_amount': float(p.paid_amount) if p.paid_amount else 0,
                'remaining_amount': float(p.remaining_amount) if p.remaining_amount else 0,
                'status': p.status,
                'status_text': status_map.get(p.status, '未知'),
                'invoice_id': p.invoice_id,
                'remark': p.remark,
                'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None,
                'updated_at': p.updated_at.strftime('%Y-%m-%d %H:%M:%S') if p.updated_at else None
            } for p in payables.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@app.route('/api/finance/payables/<int:id>', methods=['GET'])
@jwt_required()
def get_payable(id):
    """应付账款详情"""
    p = FinancePayable.query.get(id)
    if not p:
        return jsonify({'code': 404, 'message': '应付记录不存在'}), 404

    status_map = {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'id': p.id,
            'payable_no': p.payable_no,
            'related_type': p.related_type,
            'related_id': p.related_id,
            'related_no': p.related_no,
            'supplier_id': p.supplier_id,
            'supplier_name': p.supplier_name,
            'total_amount': float(p.total_amount) if p.total_amount else 0,
            'paid_amount': float(p.paid_amount) if p.paid_amount else 0,
            'remaining_amount': float(p.remaining_amount) if p.remaining_amount else 0,
            'status': p.status,
            'status_text': status_map.get(p.status, '未知'),
            'invoice_id': p.invoice_id,
            'remark': p.remark,
            'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None,
            'updated_at': p.updated_at.strftime('%Y-%m-%d %H:%M:%S') if p.updated_at else None
        }
    })


@app.route('/api/finance/payables/<int:id>/pay', methods=['POST'])
@jwt_required()
def pay_payable(id):
    """付款核销"""
    payable = FinancePayable.query.get(id)
    if not payable:
        return jsonify({'code': 404, 'message': '应付记录不存在'}), 404

    if payable.status == 2:
        return jsonify({'code': 400, 'message': '该应付已结清'}), 400

    if payable.status == 3:
        return jsonify({'code': 400, 'message': '该应付已取消'}), 400

    data = request.get_json()
    paid_amount = data.get('paid_amount')
    if not paid_amount or float(paid_amount) <= 0:
        return jsonify({'code': 400, 'message': '付款金额必须大于0'}), 400

    paid_amount = float(paid_amount)
    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    try:
        # 更新应付金额
        payable.paid_amount = float(payable.paid_amount or 0) + paid_amount
        payable.remaining_amount = float(payable.total_amount or 0) - float(payable.paid_amount or 0)

        if payable.remaining_amount <= 0:
            payable.remaining_amount = 0
            payable.status = 2  # 已结清
        else:
            payable.status = 1  # 部分付款

        payable.updated_at = datetime.now()

        # 生成财务流水（支出）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=2,  # 支出
            amount=paid_amount,
            balance_before=0,
            balance_after=0,
            related_type='payable',
            related_id=id,
            related_no=payable.payable_no,
            remark=f'应付{payable.payable_no}付款{paid_amount}元，操作人：{user_name}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        db.session.commit()
        return jsonify({'code': 200, 'message': '付款成功', 'data': {
            'paid_amount': float(payable.paid_amount),
            'remaining_amount': float(payable.remaining_amount),
            'status': payable.status
        }})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'付款失败：{str(e)}'}), 500


@app.route('/api/finance/payables/export', methods=['GET'])
@jwt_required()
def export_payables():
    """导出应付账款"""
    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)
        supplier_id = request.args.get('supplier_id', type=int)
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        query = FinancePayable.query

        if supplier_id:
            query = query.filter_by(supplier_id=supplier_id)
        if status is not None:
            query = query.filter_by(status=status)
        if keyword:
            query = query.filter(
                db.or_(
                    FinancePayable.payable_no.contains(keyword),
                    FinancePayable.supplier_name.contains(keyword),
                    FinancePayable.related_no.contains(keyword)
                )
            )
        if start_date:
            query = query.filter(FinancePayable.created_at >= start_date)
        if end_date:
            query = query.filter(FinancePayable.created_at <= end_date + ' 23:59:59')

        payables = query.order_by(FinancePayable.created_at.desc()).all()

        status_map = {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '应付账款'

        headers = ['应付编号', '关联类型', '关联单号', '供应商名称', '应付总额', '已付金额', '待付金额', '状态', '备注', '创建时间']
        ws.append(headers)

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for p in payables:
            ws.append([
                p.payable_no,
                p.related_type,
                p.related_no,
                p.supplier_name,
                float(p.total_amount) if p.total_amount else 0,
                float(p.paid_amount) if p.paid_amount else 0,
                float(p.remaining_amount) if p.remaining_amount else 0,
                status_map.get(p.status, '未知'),
                p.remark or '',
                p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else ''
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                         download_name=f'应付账款_{datetime.now().strftime("%Y%m%d")}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导出失败：{str(e)}'}), 500


# ============================================
# API路由 - 采购发票
# ============================================

@app.route('/api/purchase/invoices', methods=['GET'])
@jwt_required()
def get_purchase_invoices():
    """采购发票列表"""
    try:
        supplier_id = request.args.get('supplier_id', type=int)
        status = request.args.get('status', type=int)
        keyword = request.args.get('keyword', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        query = PurchaseInvoice.query.filter(PurchaseInvoice.status != 3)

        if supplier_id:
            query = query.filter(PurchaseInvoice.supplier_id == supplier_id)
        if status is not None:
            query = query.filter(PurchaseInvoice.status == status)
        if keyword:
            query = query.filter(db.or_(
                PurchaseInvoice.invoice_no.like(f'%{keyword}%'),
                PurchaseInvoice.supplier_name.like(f'%{keyword}%'),
                PurchaseInvoice.purchase_order_no.like(f'%{keyword}%')
            ))
        if start_date:
            query = query.filter(PurchaseInvoice.invoice_date >= start_date)
        if end_date:
            query = query.filter(PurchaseInvoice.invoice_date <= end_date)

        query = query.order_by(PurchaseInvoice.created_at.desc())
        pagination = query.paginate(page=page, per_page=page_size, error_out=False)

        invoice_type_map = {1: '普通发票', 2: '增值税专用发票'}
        status_map = {0: '待认证', 1: '已认证', 2: '已抵扣', 3: '已作废'}

        items = []
        for inv in pagination.items:
            items.append({
                'id': inv.id,
                'invoice_no': inv.invoice_no,
                'invoice_code': inv.invoice_code,
                'invoice_type': inv.invoice_type,
                'invoice_type_name': invoice_type_map.get(inv.invoice_type, ''),
                'purchase_order_id': inv.purchase_order_id,
                'purchase_order_no': inv.purchase_order_no,
                'supplier_id': inv.supplier_id,
                'supplier_name': inv.supplier_name,
                'amount': float(inv.amount) if inv.amount else 0,
                'tax_rate': float(inv.tax_rate) if inv.tax_rate else 0,
                'tax_amount': float(inv.tax_amount) if inv.tax_amount else 0,
                'total_amount': float(inv.total_amount) if inv.total_amount else 0,
                'invoice_date': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else None,
                'status': inv.status,
                'status_name': status_map.get(inv.status, ''),
                'remark': inv.remark,
                'created_at': inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else None
            })

        return jsonify({
            'code': 200,
            'message': '查询成功',
            'data': {
                'total': pagination.total,
                'page': page,
                'page_size': page_size,
                'items': items
            }
        })
    except Exception as e:
        logger.error(f'查询采购发票列表失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'查询失败：{str(e)}'}), 500


@app.route('/api/purchase/invoices/<int:id>', methods=['GET'])
@jwt_required()
def get_purchase_invoice_detail(id):
    """采购发票详情"""
    try:
        invoice = PurchaseInvoice.query.get(id)
        if not invoice:
            return jsonify({'code': 404, 'message': '发票不存在'}), 404

        invoice_type_map = {1: '普通发票', 2: '增值税专用发票'}
        status_map = {0: '待认证', 1: '已认证', 2: '已抵扣', 3: '已作废'}

        return jsonify({
            'code': 200,
            'message': '查询成功',
            'data': {
                'id': invoice.id,
                'invoice_no': invoice.invoice_no,
                'invoice_code': invoice.invoice_code,
                'invoice_type': invoice.invoice_type,
                'invoice_type_name': invoice_type_map.get(invoice.invoice_type, ''),
                'purchase_order_id': invoice.purchase_order_id,
                'purchase_order_no': invoice.purchase_order_no,
                'supplier_id': invoice.supplier_id,
                'supplier_name': invoice.supplier_name,
                'amount': float(invoice.amount) if invoice.amount else 0,
                'tax_rate': float(invoice.tax_rate) if invoice.tax_rate else 0,
                'tax_amount': float(invoice.tax_amount) if invoice.tax_amount else 0,
                'total_amount': float(invoice.total_amount) if invoice.total_amount else 0,
                'invoice_date': invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else None,
                'status': invoice.status,
                'status_name': status_map.get(invoice.status, ''),
                'remark': invoice.remark,
                'created_at': invoice.created_at.strftime('%Y-%m-%d %H:%M:%S') if invoice.created_at else None,
                'updated_at': invoice.updated_at.strftime('%Y-%m-%d %H:%M:%S') if invoice.updated_at else None,
                'created_by': invoice.created_by
            }
        })
    except Exception as e:
        logger.error(f'查询采购发票详情失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'查询失败：{str(e)}'}), 500


@app.route('/api/purchase/invoices', methods=['POST'])
@jwt_required()
def create_purchase_invoice():
    """创建采购发票"""
    try:
        data = request.get_json()
        current_user_id = get_jwt_identity()

        # 自动生成发票号码
        today_str = datetime.now().strftime('%Y%m%d')
        prefix = f'PI{today_str}'
        last_invoice = PurchaseInvoice.query.filter(
            PurchaseInvoice.invoice_no.like(f'{prefix}%')
        ).order_by(PurchaseInvoice.id.desc()).first()

        if last_invoice:
            last_seq = int(last_invoice.invoice_no[-4:])
            new_seq = last_seq + 1
        else:
            new_seq = 1

        invoice_no = f'{prefix}{new_seq:04d}'

        purchase_order_id = data.get('purchase_order_id')
        purchase_order_no = data.get('purchase_order_no', '')
        supplier_id = data.get('supplier_id')
        supplier_name = data.get('supplier_name', '')

        # 如果提供了采购单ID，自动查询采购单信息
        if purchase_order_id:
            purchase_order = PurchaseOrder.query.get(purchase_order_id)
            if purchase_order:
                purchase_order_no = purchase_order.order_no
                if not supplier_id:
                    supplier_id = purchase_order.supplier_id
                if not supplier_name:
                    supplier_name = purchase_order.supplier_name

        invoice = PurchaseInvoice(
            invoice_no=invoice_no,
            invoice_code=data.get('invoice_code', ''),
            invoice_type=data.get('invoice_type', 1),
            purchase_order_id=purchase_order_id,
            purchase_order_no=purchase_order_no,
            supplier_id=supplier_id,
            supplier_name=supplier_name,
            amount=data.get('amount', 0),
            tax_rate=data.get('tax_rate', 0),
            tax_amount=data.get('tax_amount', 0),
            total_amount=data.get('total_amount', 0),
            invoice_date=data.get('invoice_date'),
            remark=data.get('remark', ''),
            created_by=current_user_id
        )

        db.session.add(invoice)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': {'id': invoice.id, 'invoice_no': invoice.invoice_no}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'创建采购发票失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'创建失败：{str(e)}'}), 500


@app.route('/api/purchase/invoices/<int:id>', methods=['PUT'])
@jwt_required()
def update_purchase_invoice(id):
    """更新采购发票"""
    try:
        invoice = PurchaseInvoice.query.get(id)
        if not invoice:
            return jsonify({'code': 404, 'message': '发票不存在'}), 404

        data = request.get_json()

        if 'invoice_code' in data:
            invoice.invoice_code = data['invoice_code']
        if 'invoice_type' in data:
            invoice.invoice_type = data['invoice_type']
        if 'purchase_order_id' in data:
            invoice.purchase_order_id = data['purchase_order_id']
        if 'purchase_order_no' in data:
            invoice.purchase_order_no = data['purchase_order_no']
        if 'supplier_id' in data:
            invoice.supplier_id = data['supplier_id']
        if 'supplier_name' in data:
            invoice.supplier_name = data['supplier_name']
        if 'amount' in data:
            invoice.amount = data['amount']
        if 'tax_rate' in data:
            invoice.tax_rate = data['tax_rate']
        if 'tax_amount' in data:
            invoice.tax_amount = data['tax_amount']
        if 'total_amount' in data:
            invoice.total_amount = data['total_amount']
        if 'invoice_date' in data:
            invoice.invoice_date = data['invoice_date']
        if 'remark' in data:
            invoice.remark = data['remark']

        invoice.updated_at = datetime.now()

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '更新成功',
            'data': {'id': invoice.id}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新采购发票失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'更新失败：{str(e)}'}), 500


@app.route('/api/purchase/invoices/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_purchase_invoice(id):
    """删除采购发票（软删除，设为已作废）"""
    try:
        invoice = PurchaseInvoice.query.get(id)
        if not invoice:
            return jsonify({'code': 404, 'message': '发票不存在'}), 404

        invoice.status = 3
        invoice.updated_at = datetime.now()

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '删除成功',
            'data': {'id': invoice.id}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除采购发票失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'删除失败：{str(e)}'}), 500


@app.route('/api/purchase/invoices/<int:id>/certify', methods=['POST'])
@jwt_required()
def certify_purchase_invoice(id):
    """采购发票认证"""
    try:
        invoice = PurchaseInvoice.query.get(id)
        if not invoice:
            return jsonify({'code': 404, 'message': '发票不存在'}), 404

        if invoice.status != 0:
            return jsonify({'code': 400, 'message': '只有待认证状态的发票才能进行认证操作'}), 400

        invoice.status = 1
        invoice.updated_at = datetime.now()

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '认证成功',
            'data': {'id': invoice.id, 'status': invoice.status}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'采购发票认证失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'认证失败：{str(e)}'}), 500


@app.route('/api/purchase/invoices/<int:id>/deduct', methods=['POST'])
@jwt_required()
def deduct_purchase_invoice(id):
    """采购发票抵扣"""
    try:
        invoice = PurchaseInvoice.query.get(id)
        if not invoice:
            return jsonify({'code': 404, 'message': '发票不存在'}), 404

        if invoice.status != 1:
            return jsonify({'code': 400, 'message': '只有已认证状态的发票才能进行抵扣操作'}), 400

        invoice.status = 2
        invoice.updated_at = datetime.now()

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '抵扣成功',
            'data': {'id': invoice.id, 'status': invoice.status}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'采购发票抵扣失败：{str(e)}')
        return jsonify({'code': 500, 'message': f'抵扣失败：{str(e)}'}), 500



# ============================================
# API路由 - 应收应付增强功能
# ============================================

@app.route('/api/finance/receivables/batch-receive', methods=['POST'])
@jwt_required()
def batch_receive_receivables():
    """批量收款"""
    data = request.get_json()
    items = data.get('items', [])
    account_id = data.get('account_id')
    remark = data.get('remark', '')

    if not items:
        return jsonify({'code': 400, 'message': '请选择要收款的应收记录'}), 400

    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 如果提供了account_id，查询账户信息
    account = None
    if account_id:
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

    success_items = []
    failed_items = []

    try:
        for item in items:
            try:
                receivable_id = item.get('id')
                amount = item.get('amount')

                if not receivable_id or not amount or float(amount) <= 0:
                    failed_items.append({'id': receivable_id, 'reason': '金额无效'})
                    continue

                amount = float(amount)
                receivable = FinanceReceivable.query.get(receivable_id)
                if not receivable:
                    failed_items.append({'id': receivable_id, 'reason': '应收记录不存在'})
                    continue

                if receivable.status == 2:
                    failed_items.append({'id': receivable_id, 'reason': '该应收已结清'})
                    continue

                if receivable.status == 3:
                    failed_items.append({'id': receivable_id, 'reason': '该应收已取消'})
                    continue

                if amount > float(receivable.remaining_amount or 0):
                    failed_items.append({'id': receivable_id, 'reason': f'收款金额不能大于待收金额{receivable.remaining_amount}'})
                    continue

                # 更新应收金额
                receivable.received_amount = float(receivable.received_amount or 0) + amount
                receivable.remaining_amount = float(receivable.total_amount or 0) - float(receivable.received_amount or 0)

                if receivable.remaining_amount <= 0:
                    receivable.remaining_amount = 0
                    receivable.status = 2  # 已结清
                else:
                    receivable.status = 1  # 部分收款

                receivable.updated_at = datetime.now()

                # 如果关联了账户，更新账户余额
                balance_before = 0
                balance_after = 0
                if account:
                    balance_before = float(account.balance or 0)
                    account.balance = balance_before + amount
                    balance_after = float(account.balance)

                # 生成财务流水（收入）
                finance_record = FinanceRecord(
                    account_id=account_id,
                    account_name=account.account_name if account else '',
                    record_type=1,  # 收入
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=balance_after,
                    related_type='receivable',
                    related_id=receivable_id,
                    related_no=receivable.receivable_no,
                    remark=f'应收{receivable.receivable_no}收款{amount}元，操作人：{user_name}{("，备注：" + remark) if remark else ""}',
                    created_at=datetime.now(),
                    created_by=user_id
                )
                db.session.add(finance_record)

                success_items.append({
                    'id': receivable_id,
                    'receivable_no': receivable.receivable_no,
                    'amount': amount,
                    'remaining_amount': float(receivable.remaining_amount),
                    'status': receivable.status
                })

            except Exception as e:
                logger.error(f'批量收款-单条处理失败: {str(e)}')
                failed_items.append({'id': item.get('id'), 'reason': str(e)})

        db.session.commit()
        return jsonify({'code': 200, 'message': f'批量收款完成，成功{len(success_items)}条，失败{len(failed_items)}条', 'data': {
            'success_items': success_items,
            'failed_items': failed_items
        }})

    except Exception as e:
        db.session.rollback()
        logger.error(f'批量收款失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'批量收款失败：{str(e)}'}), 500


@app.route('/api/finance/payables/batch-pay', methods=['POST'])
@jwt_required()
def batch_pay_payables():
    """批量付款"""
    data = request.get_json()
    items = data.get('items', [])
    account_id = data.get('account_id')
    remark = data.get('remark', '')

    if not items:
        return jsonify({'code': 400, 'message': '请选择要付款的应付记录'}), 400

    user_id = get_jwt_identity()
    user_name = get_current_user_name()

    # 如果提供了account_id，查询账户信息
    account = None
    if account_id:
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

    success_items = []
    failed_items = []

    try:
        for item in items:
            try:
                payable_id = item.get('id')
                amount = item.get('amount')

                if not payable_id or not amount or float(amount) <= 0:
                    failed_items.append({'id': payable_id, 'reason': '金额无效'})
                    continue

                amount = float(amount)
                payable = FinancePayable.query.get(payable_id)
                if not payable:
                    failed_items.append({'id': payable_id, 'reason': '应付记录不存在'})
                    continue

                if payable.status == 2:
                    failed_items.append({'id': payable_id, 'reason': '该应付已结清'})
                    continue

                if payable.status == 3:
                    failed_items.append({'id': payable_id, 'reason': '该应付已取消'})
                    continue

                if amount > float(payable.remaining_amount or 0):
                    failed_items.append({'id': payable_id, 'reason': f'付款金额不能大于待付金额{payable.remaining_amount}'})
                    continue

                # 更新应付金额
                payable.paid_amount = float(payable.paid_amount or 0) + amount
                payable.remaining_amount = float(payable.total_amount or 0) - float(payable.paid_amount or 0)

                if payable.remaining_amount <= 0:
                    payable.remaining_amount = 0
                    payable.status = 2  # 已结清
                else:
                    payable.status = 1  # 部分付款

                payable.updated_at = datetime.now()

                # 如果关联了账户，更新账户余额
                balance_before = 0
                balance_after = 0
                if account:
                    balance_before = float(account.balance or 0)
                    account.balance = balance_before - amount
                    balance_after = float(account.balance)

                # 生成财务流水（支出）
                finance_record = FinanceRecord(
                    account_id=account_id,
                    account_name=account.account_name if account else '',
                    record_type=2,  # 支出
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=balance_after,
                    related_type='payable',
                    related_id=payable_id,
                    related_no=payable.payable_no,
                    remark=f'应付{payable.payable_no}付款{amount}元，操作人：{user_name}{("，备注：" + remark) if remark else ""}',
                    created_at=datetime.now(),
                    created_by=user_id
                )
                db.session.add(finance_record)

                success_items.append({
                    'id': payable_id,
                    'payable_no': payable.payable_no,
                    'amount': amount,
                    'remaining_amount': float(payable.remaining_amount),
                    'status': payable.status
                })

            except Exception as e:
                logger.error(f'批量付款-单条处理失败: {str(e)}')
                failed_items.append({'id': item.get('id'), 'reason': str(e)})

        db.session.commit()
        return jsonify({'code': 200, 'message': f'批量付款完成，成功{len(success_items)}条，失败{len(failed_items)}条', 'data': {
            'success_items': success_items,
            'failed_items': failed_items
        }})

    except Exception as e:
        db.session.rollback()
        logger.error(f'批量付款失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'批量付款失败：{str(e)}'}), 500


@app.route('/api/finance/receivables/<int:id>/print', methods=['GET'])
@jwt_required()
def print_receivable(id):
    """应收账款打印数据"""
    receivable = FinanceReceivable.query.get(id)
    if not receivable:
        return jsonify({'code': 404, 'message': '应收记录不存在'}), 404

    try:
        # 查询收款记录
        records = FinanceRecord.query.filter_by(
            related_type='receivable',
            related_id=id
        ).order_by(FinanceRecord.created_at.asc()).all()

        record_list = []
        for r in records:
            record_list.append({
                'id': r.id,
                'amount': float(r.amount),
                'account_name': r.account_name or '',
                'remark': r.remark or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
                'created_by': r.created_by
            })

        # 计算逾期天数
        overdue_days = (datetime.now().date() - receivable.due_date).days if receivable.due_date and receivable.due_date < datetime.now().date() else 0

        # 公司信息（默认值）
        company_info = {
            'company_name': '',
            'address': '',
            'phone': '',
            'tax_no': ''
        }

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'receivable': {
                'id': receivable.id,
                'receivable_no': receivable.receivable_no,
                'related_type': receivable.related_type,
                'related_no': receivable.related_no,
                'customer_id': receivable.customer_id,
                'customer_name': receivable.customer_name,
                'total_amount': float(receivable.total_amount or 0),
                'received_amount': float(receivable.received_amount or 0),
                'remaining_amount': float(receivable.remaining_amount or 0),
                'status': receivable.status,
                'status_text': {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}.get(receivable.status, ''),
                'due_date': receivable.due_date.strftime('%Y-%m-%d') if receivable.due_date else '',
                'remark': receivable.remark or '',
                'created_at': receivable.created_at.strftime('%Y-%m-%d %H:%M:%S') if receivable.created_at else '',
                'updated_at': receivable.updated_at.strftime('%Y-%m-%d %H:%M:%S') if receivable.updated_at else ''
            },
            'company_info': company_info,
            'records': record_list,
            'overdue_days': overdue_days
        }})

    except Exception as e:
        logger.error(f'获取应收打印数据失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取打印数据失败：{str(e)}'}), 500


@app.route('/api/finance/payables/<int:id>/print', methods=['GET'])
@jwt_required()
def print_payable(id):
    """应付账款打印数据"""
    payable = FinancePayable.query.get(id)
    if not payable:
        return jsonify({'code': 404, 'message': '应付记录不存在'}), 404

    try:
        # 查询付款记录
        records = FinanceRecord.query.filter_by(
            related_type='payable',
            related_id=id
        ).order_by(FinanceRecord.created_at.asc()).all()

        record_list = []
        for r in records:
            record_list.append({
                'id': r.id,
                'amount': float(r.amount),
                'account_name': r.account_name or '',
                'remark': r.remark or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
                'created_by': r.created_by
            })

        # 计算逾期天数
        overdue_days = (datetime.now().date() - payable.due_date).days if payable.due_date and payable.due_date < datetime.now().date() else 0

        # 公司信息（默认值）
        company_info = {
            'company_name': '',
            'address': '',
            'phone': '',
            'tax_no': ''
        }

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'payable': {
                'id': payable.id,
                'payable_no': payable.payable_no,
                'related_type': payable.related_type,
                'related_no': payable.related_no,
                'supplier_id': payable.supplier_id,
                'supplier_name': payable.supplier_name,
                'total_amount': float(payable.total_amount or 0),
                'paid_amount': float(payable.paid_amount or 0),
                'remaining_amount': float(payable.remaining_amount or 0),
                'status': payable.status,
                'status_text': {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}.get(payable.status, ''),
                'due_date': payable.due_date.strftime('%Y-%m-%d') if payable.due_date else '',
                'remark': payable.remark or '',
                'created_at': payable.created_at.strftime('%Y-%m-%d %H:%M:%S') if payable.created_at else '',
                'updated_at': payable.updated_at.strftime('%Y-%m-%d %H:%M:%S') if payable.updated_at else ''
            },
            'company_info': company_info,
            'records': record_list,
            'overdue_days': overdue_days
        }})

    except Exception as e:
        logger.error(f'获取应付打印数据失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取打印数据失败：{str(e)}'}), 500


@app.route('/api/finance/receivables/summary', methods=['GET'])
@jwt_required()
def receivables_summary():
    """应收统计汇总"""
    try:
        today = datetime.now().date()

        # 应收总额：所有待收款+部分收款的remaining_amount之和
        total_receivable = db.session.query(
            db.func.coalesce(db.func.sum(FinanceReceivable.remaining_amount), 0)
        ).filter(
            FinanceReceivable.status.in_([0, 1])
        ).scalar()
        total_receivable = float(total_receivable or 0)

        # 逾期总额：due_date < 今天 且 status != 2 且 status != 3 的remaining_amount之和
        total_overdue = db.session.query(
            db.func.coalesce(db.func.sum(FinanceReceivable.remaining_amount), 0)
        ).filter(
            FinanceReceivable.status.in_([0, 1]),
            FinanceReceivable.due_date != None,
            FinanceReceivable.due_date < today
        ).scalar()
        total_overdue = float(total_overdue or 0)

        # 逾期笔数
        overdue_count = FinanceReceivable.query.filter(
            FinanceReceivable.status.in_([0, 1]),
            FinanceReceivable.due_date != None,
            FinanceReceivable.due_date < today
        ).count()

        # 今日收款总额
        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today, datetime.max.time())
        total_received_today = db.session.query(
            db.func.coalesce(db.func.sum(FinanceRecord.amount), 0)
        ).filter(
            FinanceRecord.related_type == 'receivable',
            FinanceRecord.record_type == 1,
            FinanceRecord.created_at >= today_start,
            FinanceRecord.created_at <= today_end
        ).scalar()
        total_received_today = float(total_received_today or 0)

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'total_receivable': total_receivable,
            'total_overdue': total_overdue,
            'overdue_count': overdue_count,
            'total_received_today': total_received_today
        }})

    except Exception as e:
        logger.error(f'获取应收统计汇总失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取统计汇总失败：{str(e)}'}), 500


@app.route('/api/finance/payables/summary', methods=['GET'])
@jwt_required()
def payables_summary():
    """应付统计汇总"""
    try:
        today = datetime.now().date()

        # 应付总额：所有待付款+部分付款的remaining_amount之和
        total_payable = db.session.query(
            db.func.coalesce(db.func.sum(FinancePayable.remaining_amount), 0)
        ).filter(
            FinancePayable.status.in_([0, 1])
        ).scalar()
        total_payable = float(total_payable or 0)

        # 逾期总额：due_date < 今天 且 status != 2 且 status != 3 的remaining_amount之和
        total_overdue = db.session.query(
            db.func.coalesce(db.func.sum(FinancePayable.remaining_amount), 0)
        ).filter(
            FinancePayable.status.in_([0, 1]),
            FinancePayable.due_date != None,
            FinancePayable.due_date < today
        ).scalar()
        total_overdue = float(total_overdue or 0)

        # 逾期笔数
        overdue_count = FinancePayable.query.filter(
            FinancePayable.status.in_([0, 1]),
            FinancePayable.due_date != None,
            FinancePayable.due_date < today
        ).count()

        # 今日付款总额
        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today, datetime.max.time())
        total_paid_today = db.session.query(
            db.func.coalesce(db.func.sum(FinanceRecord.amount), 0)
        ).filter(
            FinanceRecord.related_type == 'payable',
            FinanceRecord.record_type == 2,
            FinanceRecord.created_at >= today_start,
            FinanceRecord.created_at <= today_end
        ).scalar()
        total_paid_today = float(total_paid_today or 0)

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'total_payable': total_payable,
            'total_overdue': total_overdue,
            'overdue_count': overdue_count,
            'total_paid_today': total_paid_today
        }})

    except Exception as e:
        logger.error(f'获取应付统计汇总失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取统计汇总失败：{str(e)}'}), 500


# ============================================
# API路由 - 数据导出
# ============================================

@app.route('/api/workorders/export', methods=['GET'])
@jwt_required()
def export_workorders():
    """导出工单Excel"""
    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)
        wo_type = request.args.get('wo_type', '')
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')

        query = WorkOrder.query
        if keyword:
            query = query.filter(db.or_(
                WorkOrder.wo_no.contains(keyword),
                WorkOrder.customer_name.contains(keyword),
                WorkOrder.customer_phone.contains(keyword),
                WorkOrder.customer_company.contains(keyword)
            ))
        if status is not None:
            query = query.filter(WorkOrder.status == status)
        if wo_type:
            query = query.filter(WorkOrder.wo_type == wo_type)
        if date_start:
            query = query.filter(WorkOrder.created_at >= date_start)
        if date_end:
            query = query.filter(WorkOrder.created_at <= date_end + ' 23:59:59')

        orders = query.order_by(WorkOrder.created_at.desc()).all()

        data = []
        for o in orders:
            data.append({
                '工单号': o.wo_no,
                '工单类型': WO_TYPE_MAP.get(o.wo_type, o.wo_type or ''),
                '客户名称': o.customer_name,
                '客户单位': o.customer_company or '',
                '客户电话': o.customer_phone,
                '设备类型': o.device_type or '',
                '设备品牌': o.device_brand or '',
                '设备型号': o.device_model or '',
                '故障描述': o.fault_desc or '',
                '人工费': float(o.labor_cost or 0),
                '配件费': float(o.parts_cost or 0),
                '材料费': float(o.material_cost or 0),
                '总费用': float(o.total_cost or 0),
                '状态': WO_STATUS_MAP.get(o.status, str(o.status)),
                '指派工程师': o.assigned_user_name or '',
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else ''
            })

        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '工单列表'
        # 定义表头
        headers = ['工单号', '工单类型', '客户名称', '客户单位', '客户电话', '设备类型', '设备品牌', '设备型号', '故障描述', '人工费', '配件费', '材料费', '总费用', '状态', '指派工程师', '创建时间']
        ws.append(headers)
        for row in data:
            ws.append(list(row.values()))
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name=f'工单列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f'导出工单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出工单失败: {str(e)}'}), 500

@app.route('/api/sales/orders/export', methods=['GET'])
@jwt_required()
def export_sales_orders():
    """导出销售单"""
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    
    query = SalesOrder.query
    if keyword:
        query = query.filter(db.or_(SalesOrder.order_no.contains(keyword), SalesOrder.customer_name.contains(keyword)))
    if status is not None:
        query = query.filter(SalesOrder.status == status)
    if date_start:
        query = query.filter(SalesOrder.created_at >= date_start)
    if date_end:
        query = query.filter(SalesOrder.created_at <= date_end + ' 23:59:59')
    
    orders = query.order_by(SalesOrder.created_at.desc()).all()
    
    data = []
    for o in orders:
        data.append({
            '销售单号': o.order_no,
            '客户名称': o.customer_name,
            '客户电话': getattr(o, 'customer_phone', '') or '',
            '销售日期': o.order_date.strftime('%Y-%m-%d') if o.order_date else '',
            '总数量': int(o.total_quantity or 0),
            '总金额': float(o.total_amount or 0),
            '已收金额': float(getattr(o, 'paid_amount', 0) or 0),
            '状态': {0: '待审核', 1: '已审核', 2: '已出库', 3: '已完成', 4: '已取消'}.get(o.status, str(o.status)),
            '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else ''
        })
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '销售单列表'
    # 定义表头
    headers = ['销售单号', '客户名称', '客户电话', '销售日期', '总数量', '总金额', '已收金额', '状态', '创建时间']
    ws.append(headers)
    for row in data:
        ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name=f'销售单列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/purchase/orders/export', methods=['GET'])
@jwt_required()
def export_purchase_orders():
    """导出采购单"""
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    
    query = PurchaseOrder.query
    if keyword:
        query = query.filter(db.or_(PurchaseOrder.order_no.contains(keyword), PurchaseOrder.supplier_name.contains(keyword)))
    if status is not None:
        query = query.filter(PurchaseOrder.status == status)
    if date_start:
        query = query.filter(PurchaseOrder.created_at >= date_start)
    if date_end:
        query = query.filter(PurchaseOrder.created_at <= date_end + ' 23:59:59')
    
    orders = query.order_by(PurchaseOrder.created_at.desc()).all()
    
    data = []
    for o in orders:
        data.append({
            '采购单号': o.order_no,
            '供应商': o.supplier_name,
            '总金额': float(o.total_amount or 0),
            '状态': {0: '待审核', 1: '已审核', 2: '已入库', 3: '已取消'}.get(o.status, str(o.status)),
            '备注': o.remark or '',
            '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else ''
        })
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '采购单列表'
    # 定义表头
    headers = ['采购单号', '供应商', '总金额', '状态', '备注', '创建时间']
    ws.append(headers)
    for row in data:
        ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name=f'采购单列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/receiveorders/export', methods=['GET'])
@jwt_required()
def export_receive_orders():
    """导出接件单"""
    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')

        query = ReceiveOrder.query
        if keyword:
            query = query.filter(db.or_(
                ReceiveOrder.receive_no.contains(keyword),
                ReceiveOrder.customer_name.contains(keyword),
                ReceiveOrder.customer_phone.contains(keyword)
            ))
        if status is not None:
            query = query.filter(ReceiveOrder.status == status)
        if date_start:
            query = query.filter(ReceiveOrder.created_at >= date_start)
        if date_end:
            query = query.filter(ReceiveOrder.created_at <= date_end + ' 23:59:59')

        orders = query.order_by(ReceiveOrder.created_at.desc()).all()

        # 接件状态映射
        ro_status_map = {
            0: '待检测',
            1: '检测中',
            2: '待报价',
            3: '待确认',
            4: '维修中',
            5: '外送修',
            6: '待测试',
            7: '待取件',
            8: '已完成',
            9: '已取消'
        }

        # 接件类型映射
        ro_type_map = {
            1: '本店修',
            2: '外送修'
        }

        data = []
        for o in orders:
            data.append({
                '接件单号': o.receive_no,
                '客户名称': o.customer_name or '',
                '客户电话': o.customer_phone or '',
                '接件类型': ro_type_map.get(o.receive_type, '本店修'),
                '状态': ro_status_map.get(o.status, str(o.status)),
                '总金额': float(o.total_amount or 0),
                '已付金额': float(o.paid_amount or 0),
                '接待人': o.receiver_name or '',
                '工程师': o.engineer_name or '',
                '外送供应商': o.external_shop_name or '',
                '报价人工费': float(o.quote_labor_cost or 0),
                '报价材料费': float(o.quote_material_cost or 0),
                '报价其他费': float(o.quote_other_cost or 0),
                '报价总计': float(o.quote_total or 0),
                '检测结果': o.detect_result or '',
                '故障原因': o.detect_fault_reason or '',
                '维修方案': o.detect_repair_plan or '',
                '完工报告': o.finish_report or '',
                '测试结果': '通过' if o.test_result == 1 else ('未通过' if o.test_result == 2 else '待测试'),
                '备注': o.remark or '',
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else ''
            })

        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '接件单列表'
        # 定义表头
        headers = ['接件单号', '客户名称', '客户电话', '接件类型', '状态', '总金额', '已付金额', '接待人', '工程师', '外送供应商', '报价人工费', '报价材料费', '报价其他费', '报价总计', '检测结果', '故障原因', '维修方案', '完工报告', '测试结果', '备注', '创建时间']
        ws.append(headers)
        for row in data:
            ws.append(list(row.values()))
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name=f'接件单列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.error(f'导出接件单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出接件单失败: {str(e)}'}), 500

# ============================================
# RBAC 权限校验装饰器和辅助函数
# ============================================

def check_user_permission(user, permission_code):
    """检查用户是否拥有指定权限"""
    if not user.role_id:
        return False
    role = SysRole.query.get(user.role_id)
    if not role or role.status != 1:
        return False
    if role.permissions and isinstance(role.permissions, list) and '*' in role.permissions:
        return True  # 超级管理员
    perms = role.permissions if role.permissions else []
    return permission_code in perms

def check_user_menu_permission(user, menu_path):
    """检查用户是否有某个菜单的权限"""
    if not user.role_id:
        return False
    role = SysRole.query.get(user.role_id)
    if not role or role.status != 1:
        return False
    if role.permissions and isinstance(role.permissions, list) and '*' in role.permissions:
        return True
    # 查找菜单权限
    perm = SysPermission.query.filter_by(path=menu_path, type=1, status=1).first()
    if perm and perm.code:
        return perm.code in (role.permissions or [])
    return True  # 没有注册的菜单默认放行

def permission_required(permission_code):
    """权限校验装饰器，用于接口级别权限控制"""
    def wrapper(fn):
        @wraps(fn)
        @jwt_required()
        def decorated(*args, **kwargs):
            user_id = get_jwt_identity()
            user = SysUser.query.get(user_id)
            if not user:
                return jsonify({'code': 401, 'message': '用户不存在'}), 401
            if not check_user_permission(user, permission_code):
                return jsonify({'code': 403, 'message': '无权限执行此操作'}), 403
            return fn(*args, **kwargs)
        return decorated
    return wrapper

def get_user_permission_codes(user):
    """获取用户的所有权限code列表"""
    if not user.role_id:
        return []
    role = SysRole.query.get(user.role_id)
    if not role or role.status != 1:
        return []
    perms = role.permissions if role.permissions else []
    if isinstance(perms, list) and '*' in perms:
        all_perms = SysPermission.query.filter_by(status=1).all()
        return [p.code for p in all_perms]
    return perms

def build_permission_tree(permissions):
    """将权限列表构建为树形结构"""
    perm_dict = {}
    for p in permissions:
        perm_dict[p.id] = {
            'id': p.id,
            'name': p.name,
            'code': p.code,
            'type': p.type,
            'parent_id': p.parent_id,
            'path': p.path,
            'icon': p.icon,
            'sort_order': p.sort_order,
            'status': p.status,
            'children': []
        }
    tree = []
    for p in permissions:
        if p.parent_id and p.parent_id in perm_dict:
            perm_dict[p.parent_id]['children'].append(perm_dict[p.id])
        else:
            tree.append(perm_dict[p.id])
    tree.sort(key=lambda x: x.get('sort_order', 0))
    return tree

# ============================================
# RBAC API路由 - 角色管理
# ============================================

@app.route('/api/settings/roles', methods=['GET'])
@jwt_required()
def get_roles():
    """获取角色列表（分页、搜索）"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')

    query = SysRole.query
    if keyword:
        query = query.filter(
            db.or_(
                SysRole.role_name.contains(keyword),
                SysRole.role_code.contains(keyword)
            )
        )
    query = query.order_by(SysRole.created_at.desc())
    total = query.count()
    roles = query.offset((page - 1) * page_size).limit(page_size).all()

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(r) for r in roles],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/settings/roles/all', methods=['GET'])
@jwt_required()
def get_all_roles():
    """获取所有启用角色（下拉选择用）"""
    roles = SysRole.query.filter_by(status=1).order_by(SysRole.created_at).all()
    return jsonify({
        'code': 200,
        'data': [{'id': r.id, 'role_name': r.role_name, 'role_code': r.role_code} for r in roles]
    })

@app.route('/api/settings/roles', methods=['POST'])
@jwt_required()
def create_role():
    """创建角色"""
    data = request.get_json()
    role_name = data.get('role_name')
    role_code = data.get('role_code')
    description = data.get('description', '')
    permissions = data.get('permissions', [])

    if not role_name or not role_code:
        return jsonify({'code': 400, 'message': '角色名称和角色编码不能为空'}), 400

    existing = SysRole.query.filter_by(role_code=role_code).first()
    if existing:
        return jsonify({'code': 400, 'message': '角色编码已存在'}), 400

    role = SysRole(
        role_name=role_name,
        role_code=role_code,
        description=description,
        permissions=permissions
    )
    db.session.add(role)
    db.session.commit()
    return jsonify({'code': 200, 'message': '创建成功', 'data': to_dict(role)})

@app.route('/api/settings/roles/<int:role_id>', methods=['PUT'])
@jwt_required()
def update_role(role_id):
    """更新角色"""
    role = SysRole.query.get(role_id)
    if not role:
        return jsonify({'code': 404, 'message': '角色不存在'}), 404

    data = request.get_json()
    if 'role_name' in data:
        role.role_name = data['role_name']
    if 'description' in data:
        role.description = data['description']
    if 'status' in data:
        role.status = data['status']
    if 'permissions' in data:
        role.permissions = data['permissions']
    if 'data_scope' in data:
        # 数据权限通过permissions JSON字段中的data_scope处理
        perms = role.permissions if role.permissions else []
        if isinstance(perms, list):
            # 将data_scope存入permissions的额外信息中
            # permissions保持为权限code列表，data_scope单独存储
            pass
        # 将data_scope作为角色额外属性存储在permissions中
        # 格式：permissions为列表时保持不变，data_scope通过单独字段传递
        # 由于不能修改模型，将data_scope嵌入permissions JSON
        current_perms = role.permissions
        if isinstance(current_perms, dict):
            current_perms['data_scope'] = data['data_scope']
        elif isinstance(current_perms, list):
            # 转为dict格式存储，包含codes和data_scope
            role.permissions = {'codes': current_perms, 'data_scope': data['data_scope']}

    role.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功', 'data': to_dict(role)})

@app.route('/api/settings/roles/<int:role_id>', methods=['DELETE'])
@jwt_required()
def delete_role(role_id):
    """删除角色（不能删除admin角色）"""
    role = SysRole.query.get(role_id)
    if not role:
        return jsonify({'code': 404, 'message': '角色不存在'}), 404

    if role.role_code == 'admin':
        return jsonify({'code': 400, 'message': '不能删除超级管理员角色'}), 400

    # 检查是否有用户使用该角色
    user_count = SysUser.query.filter_by(role_id=role_id).count()
    if user_count > 0:
        return jsonify({'code': 400, 'message': f'该角色下有{user_count}个用户，不能删除'}), 400

    db.session.delete(role)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

@app.route('/api/settings/roles/<int:role_id>/permissions', methods=['GET'])
@jwt_required()
def get_role_permissions(role_id):
    """获取角色权限（返回权限树）"""
    role = SysRole.query.get(role_id)
    if not role:
        return jsonify({'code': 404, 'message': '角色不存在'}), 404

    # 获取角色已有的权限code列表
    role_perms = role.permissions if role.permissions else []
    if isinstance(role_perms, dict):
        role_perms = role_perms.get('codes', [])

    # 获取所有权限，构建树形结构，并标记角色已有的权限
    all_perms = SysPermission.query.filter_by(status=1).order_by(SysPermission.sort_order).all()
    perm_tree = build_permission_tree(all_perms)

    # 标记已选中的权限
    def mark_selected(nodes):
        for node in nodes:
            node['checked'] = node.get('code', '') in role_perms
            if node.get('children'):
                mark_selected(node['children'])
    mark_selected(perm_tree)

    return jsonify({
        'code': 200,
        'data': {
            'role': to_dict(role),
            'permission_tree': perm_tree,
            'selected_codes': role_perms
        }
    })

@app.route('/api/settings/roles/<int:role_id>/permissions', methods=['POST'])
@jwt_required()
def set_role_permissions(role_id):
    """配置角色权限（接收permission_ids列表）"""
    role = SysRole.query.get(role_id)
    if not role:
        return jsonify({'code': 404, 'message': '角色不存在'}), 404

    data = request.get_json()
    permission_ids = data.get('permission_ids', [])

    # 根据permission_ids查询对应的权限code
    if permission_ids:
        perms = SysPermission.query.filter(SysPermission.id.in_(permission_ids)).all()
        perm_codes = [p.code for p in perms]
    else:
        perm_codes = []

    # 保留data_scope（如果之前有的话）
    existing_perms = role.permissions
    data_scope = None
    if isinstance(existing_perms, dict):
        data_scope = existing_perms.get('data_scope')

    if data_scope:
        role.permissions = {'codes': perm_codes, 'data_scope': data_scope}
    else:
        role.permissions = perm_codes

    role.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '权限配置成功'})

# ============================================
# RBAC API路由 - 权限管理
# ============================================

@app.route('/api/settings/permissions', methods=['GET'])
@jwt_required()
def get_permissions():
    """获取权限列表（树形结构）"""
    all_perms = SysPermission.query.order_by(SysPermission.sort_order).all()
    perm_tree = build_permission_tree(all_perms)
    return jsonify({'code': 200, 'data': perm_tree})

@app.route('/api/settings/permissions/menus', methods=['GET'])
@jwt_required()
def get_menu_permissions():
    """获取菜单权限树（type=1）"""
    menu_perms = SysPermission.query.filter_by(type=1, status=1).order_by(SysPermission.sort_order).all()
    menu_tree = build_permission_tree(menu_perms)
    return jsonify({'code': 200, 'data': menu_tree})

@app.route('/api/settings/permissions', methods=['POST'])
@jwt_required()
def create_permission():
    """创建权限"""
    data = request.get_json()
    name = data.get('name')
    code = data.get('code')
    perm_type = data.get('type', 1)
    parent_id = data.get('parent_id', 0)
    path = data.get('path', '')
    icon = data.get('icon', '')
    sort_order = data.get('sort_order', 0)

    if not name or not code:
        return jsonify({'code': 400, 'message': '权限名称和权限编码不能为空'}), 400

    existing = SysPermission.query.filter_by(code=code).first()
    if existing:
        return jsonify({'code': 400, 'message': '权限编码已存在'}), 400

    perm = SysPermission(
        name=name,
        code=code,
        type=perm_type,
        parent_id=parent_id,
        path=path,
        icon=icon,
        sort_order=sort_order
    )
    db.session.add(perm)
    db.session.commit()
    return jsonify({'code': 200, 'message': '创建成功', 'data': to_dict(perm)})

@app.route('/api/settings/permissions/<int:perm_id>', methods=['PUT'])
@jwt_required()
def update_permission(perm_id):
    """更新权限"""
    perm = SysPermission.query.get(perm_id)
    if not perm:
        return jsonify({'code': 404, 'message': '权限不存在'}), 404

    data = request.get_json()
    if 'name' in data:
        perm.name = data['name']
    if 'type' in data:
        perm.type = data['type']
    if 'parent_id' in data:
        perm.parent_id = data['parent_id']
    if 'path' in data:
        perm.path = data['path']
    if 'icon' in data:
        perm.icon = data['icon']
    if 'sort_order' in data:
        perm.sort_order = data['sort_order']
    if 'status' in data:
        perm.status = data['status']

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功', 'data': to_dict(perm)})

@app.route('/api/settings/permissions/<int:perm_id>', methods=['DELETE'])
@jwt_required()
def delete_permission(perm_id):
    """删除权限"""
    perm = SysPermission.query.get(perm_id)
    if not perm:
        return jsonify({'code': 404, 'message': '权限不存在'}), 404

    # 检查是否有子权限
    children = SysPermission.query.filter_by(parent_id=perm_id).count()
    if children > 0:
        return jsonify({'code': 400, 'message': '该权限下有子权限，不能删除'}), 400

    db.session.delete(perm)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# ============================================
# RBAC API路由 - 用户管理
# ============================================

@app.route('/api/settings/users', methods=['GET'])
@jwt_required()
def get_settings_users():
    """获取用户列表（分页、搜索keyword、status筛选）"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)

    query = SysUser.query.filter_by(is_deleted=0)
    if keyword:
        query = query.filter(
            db.or_(
                SysUser.username.contains(keyword),
                SysUser.real_name.contains(keyword),
                SysUser.phone.contains(keyword),
                SysUser.email.contains(keyword)
            )
        )
    if status is not None:
        query = query.filter_by(status=status)

    query = query.order_by(SysUser.created_at.desc())
    total = query.count()
    users = query.offset((page - 1) * page_size).limit(page_size).all()

    # 附加角色名称
    user_list = []
    for u in users:
        item = to_dict(u, exclude=['password'])
        if u.role_id:
            role = SysRole.query.get(u.role_id)
            item['role_name'] = role.role_name if role else ''
            item['role_code'] = role.role_code if role else ''
        else:
            item['role_name'] = ''
            item['role_code'] = ''
        user_list.append(item)

    return jsonify({
        'code': 200,
        'data': {
            'list': user_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/settings/users', methods=['POST'])
@jwt_required()
def create_user():
    """创建用户"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    real_name = data.get('real_name', '')
    phone = data.get('phone', '')
    email = data.get('email', '')
    role_id = data.get('role_id')
    department = data.get('department', '')
    position = data.get('position', '')

    if not username or not password:
        return jsonify({'code': 400, 'message': '用户名和密码不能为空'}), 400

    existing = SysUser.query.filter_by(username=username).first()
    if existing:
        return jsonify({'code': 400, 'message': '用户名已存在'}), 400

    user = SysUser(
        username=username,
        password=generate_password_hash(password),
        real_name=real_name,
        phone=phone,
        email=email,
        role_id=role_id,
        department=department,
        position=position,
        base_salary=data.get('base_salary', 0),
        status=1,
        created_by=get_current_user_id()
    )
    db.session.add(user)
    db.session.commit()
    return jsonify({'code': 200, 'message': '创建成功', 'data': to_dict(user, exclude=['password'])})

@app.route('/api/settings/users/<int:user_id>', methods=['PUT'])
@jwt_required()
def update_user(user_id):
    """更新用户"""
    user = SysUser.query.get(user_id)
    if not user or user.is_deleted:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404

    data = request.get_json()
    if 'real_name' in data:
        user.real_name = data['real_name']
    if 'phone' in data:
        user.phone = data['phone']
    if 'email' in data:
        user.email = data['email']
    if 'department' in data:
        user.department = data['department']
    if 'position' in data:
        user.position = data['position']
    if 'avatar' in data:
        user.avatar = data['avatar']
    if 'base_salary' in data:
        user.base_salary = data['base_salary']
    if 'role_id' in data:
        user.role_id = data['role_id']
    if 'status' in data:
        user.status = data['status']

    user.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功', 'data': to_dict(user, exclude=['password'])})

@app.route('/api/settings/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
def delete_user(user_id):
    """删除用户（不能删除自己、不能删除admin）"""
    current_user_id = get_current_user_id()

    if current_user_id == user_id:
        return jsonify({'code': 400, 'message': '不能删除自己'}), 400

    user = SysUser.query.get(user_id)
    if not user or user.is_deleted:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404

    if user.username == 'admin':
        return jsonify({'code': 400, 'message': '不能删除管理员账号'}), 400

    # 软删除
    user.is_deleted = 1
    user.status = 0
    user.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})

# 操作日志API
@app.route('/api/settings/logs', methods=['GET'])
@jwt_required()
def get_operation_logs():
    """获取操作日志列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    user_id = request.args.get('user_id', type=int)
    action = request.args.get('action', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = OperationLog.query
    
    if user_id:
        query = query.filter_by(user_id=user_id)
    if action:
        query = query.filter(OperationLog.action.contains(action))
    if start_date:
        query = query.filter(OperationLog.created_at >= start_date)
    if end_date:
        query = query.filter(OperationLog.created_at <= end_date + ' 23:59:59')
    
    total = query.count()
    logs = query.order_by(OperationLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(l) for l in logs],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })

@app.route('/api/settings/logs', methods=['DELETE'])
@jwt_required()
def clear_operation_logs():
    """清空操作日志"""
    OperationLog.query.delete()
    db.session.commit()
    return jsonify({'code': 200, 'message': '清空成功'})

@app.route('/api/settings/users/<int:user_id>/status', methods=['PUT'])
@jwt_required()
def toggle_user_status(user_id):
    """启用/禁用用户"""
    current_user_id = get_current_user_id()

    if current_user_id == user_id:
        return jsonify({'code': 400, 'message': '不能禁用自己'}), 400

    user = SysUser.query.get(user_id)
    if not user or user.is_deleted:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404

    if user.username == 'admin':
        return jsonify({'code': 400, 'message': '不能禁用管理员账号'}), 400

    data = request.get_json()
    user.status = data.get('status', 0)
    user.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '状态更新成功'})

@app.route('/api/settings/users/<int:user_id>/password', methods=['PUT'])
@jwt_required()
def reset_user_password(user_id):
    """重置密码"""
    user = SysUser.query.get(user_id)
    if not user or user.is_deleted:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404

    data = request.get_json()
    new_password = data.get('password')
    if not new_password:
        return jsonify({'code': 400, 'message': '新密码不能为空'}), 400

    if len(new_password) < 6:
        return jsonify({'code': 400, 'message': '密码长度不能少于6位'}), 400

    user.password = generate_password_hash(new_password)
    user.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '密码重置成功'})

@app.route('/api/settings/users/<int:user_id>/roles', methods=['PUT'])
@jwt_required()
def assign_user_role(user_id):
    """分配角色（接收role_id）"""
    user = SysUser.query.get(user_id)
    if not user or user.is_deleted:
        return jsonify({'code': 404, 'message': '用户不存在'}), 404

    data = request.get_json()
    role_id = data.get('role_id')

    if role_id:
        role = SysRole.query.get(role_id)
        if not role:
            return jsonify({'code': 400, 'message': '角色不存在'}), 400

    user.role_id = role_id
    user.updated_at = datetime.now()
    db.session.commit()
    return jsonify({'code': 200, 'message': '角色分配成功'})

# ============================================
# RBAC API路由 - 权限初始化
# ============================================

@app.route('/api/settings/init-permissions', methods=['POST'])
@jwt_required()
def init_permissions():
    """初始化系统权限数据（菜单+按钮权限种子数据）"""
    try:
        # 菜单权限定义（parent_id=0）
        menu_permissions = [
            {'name': '首页', 'code': 'dashboard', 'path': '/dashboard', 'icon': 'dashboard', 'sort_order': 1},
            {'name': '工单管理', 'code': 'workorder', 'path': '/workorder', 'icon': 'workorder', 'sort_order': 2},
            {'name': '接件管理', 'code': 'receive', 'path': '/receive', 'icon': 'receive', 'sort_order': 3},
            {'name': '派单管理', 'code': 'dispatch', 'path': '/dispatch', 'icon': 'dispatch', 'sort_order': 4},
            {'name': '报价管理', 'code': 'quote', 'path': '/quote', 'icon': 'quote', 'sort_order': 5},
            {'name': '设备管理', 'code': 'device', 'path': '/device', 'icon': 'device', 'sort_order': 6},
            {'name': '商品管理', 'code': 'product', 'path': '/product', 'icon': 'product', 'sort_order': 7},
            {'name': '销售管理', 'code': 'sales', 'path': '/sales', 'icon': 'sales', 'sort_order': 8},
            {'name': '采购管理', 'code': 'purchase', 'path': '/purchase', 'icon': 'purchase', 'sort_order': 9},
            {'name': '采购预订', 'code': 'preorder', 'path': '/preorder', 'icon': 'preorder', 'sort_order': 10},
            {'name': '销售预订', 'code': 'preorder-sale', 'path': '/preorder-sale', 'icon': 'preorder-sale', 'sort_order': 11},
            {'name': '采购退货', 'code': 'purchase-return', 'path': '/purchase-return', 'icon': 'purchase-return', 'sort_order': 12},
            {'name': '销售退货', 'code': 'sales-return', 'path': '/sales-return', 'icon': 'sales-return', 'sort_order': 13},
            {'name': '库存查询', 'code': 'inventory', 'path': '/inventory', 'icon': 'inventory', 'sort_order': 14},
            {'name': '入库管理', 'code': 'inventory-in', 'path': '/inventory-in', 'icon': 'inventory-in', 'sort_order': 15},
            {'name': '出库管理', 'code': 'inventory-out', 'path': '/inventory-out', 'icon': 'inventory-out', 'sort_order': 16},
            {'name': '库存盘点', 'code': 'inventory-check', 'path': '/inventory-check', 'icon': 'inventory-check', 'sort_order': 17},
            {'name': '客户管理', 'code': 'customer', 'path': '/customer', 'icon': 'customer', 'sort_order': 18},
            {'name': '供应商管理', 'code': 'supplier', 'path': '/supplier', 'icon': 'supplier', 'sort_order': 19},
            {'name': '应收管理', 'code': 'finance-receivable', 'path': '/finance-receivable', 'icon': 'finance-receivable', 'sort_order': 20},
            {'name': '应付管理', 'code': 'finance-payable', 'path': '/finance-payable', 'icon': 'finance-payable', 'sort_order': 21},
            {'name': '账户管理', 'code': 'finance-account', 'path': '/finance-account', 'icon': 'finance-account', 'sort_order': 22},
            {'name': '收款管理', 'code': 'finance-receipt', 'path': '/finance-receipt', 'icon': 'finance-receipt', 'sort_order': 23},
            {'name': '付款管理', 'code': 'finance-payment', 'path': '/finance-payment', 'icon': 'finance-payment', 'sort_order': 24},
            {'name': '发票管理', 'code': 'invoice', 'path': '/invoice', 'icon': 'invoice', 'sort_order': 25},
            {'name': '用户管理', 'code': 'settings-users', 'path': '/settings/users', 'icon': 'settings-users', 'sort_order': 26},
            {'name': '角色管理', 'code': 'settings-roles', 'path': '/settings/roles', 'icon': 'settings-roles', 'sort_order': 27},
            {'name': '操作日志', 'code': 'settings-log', 'path': '/settings/log', 'icon': 'settings-log', 'sort_order': 28},
            {'name': '商品分类', 'code': 'settings-category', 'path': '/settings/category', 'icon': 'settings-category', 'sort_order': 29},
            {'name': '计量单位', 'code': 'settings-unit', 'path': '/settings/unit', 'icon': 'settings-unit', 'sort_order': 30},
            {'name': '打印模版', 'code': 'settings-print', 'path': '/settings/print', 'icon': 'settings-print', 'sort_order': 31},
        ]

        # 按钮权限定义
        button_actions = ['view', 'add', 'edit', 'delete', 'export', 'print']
        button_names = {
            'view': '查看', 'add': '新增', 'edit': '编辑',
            'delete': '删除', 'export': '导出', 'print': '打印'
        }

        # 创建菜单权限
        menu_id_map = {}
        for menu in menu_permissions:
            existing = SysPermission.query.filter_by(code=menu['code'], type=1).first()
            if not existing:
                perm = SysPermission(
                    name=menu['name'],
                    code=menu['code'],
                    type=1,
                    parent_id=0,
                    path=menu['path'],
                    icon=menu['icon'],
                    sort_order=menu['sort_order'],
                    status=1
                )
                db.session.add(perm)
                db.session.flush()  # 获取id
                menu_id_map[menu['code']] = perm.id
            else:
                menu_id_map[menu['code']] = existing.id

        # 创建按钮权限（每个菜单下挂载）
        btn_count = 0
        for menu_code, parent_id in menu_id_map.items():
            for action in button_actions:
                btn_code = f'{menu_code}:{action}'
                existing = SysPermission.query.filter_by(code=btn_code, type=2).first()
                if not existing:
                    perm = SysPermission(
                        name=f'{menu_code} - {button_names[action]}',
                        code=btn_code,
                        type=2,
                        parent_id=parent_id,
                        sort_order=button_actions.index(action) + 1,
                        status=1
                    )
                    db.session.add(perm)
                    btn_count += 1

        db.session.commit()
        return jsonify({
            'code': 200,
            'message': f'权限初始化成功，共{len(menu_permissions)}个菜单权限，新增{btn_count}个按钮权限'
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'权限初始化失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'权限初始化失败: {str(e)}'}), 500

# ============================================
# RBAC - init_data 增强版（替换原init_data）
# ============================================

def init_data():
    """初始化基础数据"""
    with app.app_context():
        try:
            # 创建默认角色
            admin_role = SysRole.query.filter_by(role_code='admin').first()
            if not admin_role:
                admin_role = SysRole(
                    role_name='超级管理员',
                    role_code='admin',
                    description='系统超级管理员，拥有所有权限',
                    permissions=['*']
                )
                db.session.add(admin_role)
                db.session.commit()

            # 创建更多默认角色
            default_roles = [
                {
                    'role_name': '前台接待',
                    'role_code': 'receptionist',
                    'description': '前台接待人员，负责接件和客户管理',
                    'permissions': [
                        'dashboard:view',
                        'workorder:view', 'workorder:add',
                        'receive:view', 'receive:add', 'receive:edit',
                        'customer:view', 'customer:add', 'customer:edit',
                        'quote:view', 'quote:add', 'quote:edit',
                        'device:view', 'device:add',
                    ]
                },
                {
                    'role_name': '维修工程师',
                    'role_code': 'engineer',
                    'description': '维修工程师，负责工单处理和设备维修',
                    'permissions': [
                        'dashboard:view',
                        'workorder:view', 'workorder:add', 'workorder:edit',
                        'receive:view', 'receive:edit',
                        'device:view', 'device:add', 'device:edit',
                        'inventory:view',
                    ]
                },
                {
                    'role_name': '仓库管理员',
                    'role_code': 'warehouse',
                    'description': '仓库管理员，负责库存和采购管理',
                    'permissions': [
                        'dashboard:view',
                        'inventory:view', 'inventory:add', 'inventory:edit',
                        'inventory-in:view', 'inventory-in:add', 'inventory-in:edit',
                        'inventory-out:view', 'inventory-out:add', 'inventory-out:edit',
                        'inventory-check:view', 'inventory-check:add', 'inventory-check:edit',
                        'purchase:view', 'purchase:add', 'purchase:edit',
                        'product:view', 'product:add', 'product:edit',
                        'supplier:view', 'supplier:add', 'supplier:edit',
                        'preorder:view', 'preorder:add', 'preorder:edit',
                        'purchase-return:view', 'purchase-return:add', 'purchase-return:edit',
                    ]
                },
                {
                    'role_name': '财务人员',
                    'role_code': 'finance',
                    'description': '财务人员，负责所有财务模块',
                    'permissions': [
                        'dashboard:view',
                        'finance-receivable:view', 'finance-receivable:add', 'finance-receivable:edit',
                        'finance-payable:view', 'finance-payable:add', 'finance-payable:edit',
                        'finance-account:view', 'finance-account:add', 'finance-account:edit',
                        'finance-receipt:view', 'finance-receipt:add', 'finance-receipt:edit',
                        'finance-payment:view', 'finance-payment:add', 'finance-payment:edit',
                        'invoice:view', 'invoice:add', 'invoice:edit',
                        'customer:view',
                        'supplier:view',
                        'sales:view',
                        'workorder:view',
                    ]
                },
                {
                    'role_name': '业务员',
                    'role_code': 'salesman',
                    'description': '业务员，负责客户、报价和销售',
                    'permissions': [
                        'dashboard:view',
                        'customer:view', 'customer:add', 'customer:edit',
                        'quote:view', 'quote:add', 'quote:edit',
                        'sales:view', 'sales:add', 'sales:edit',
                        'workorder:view', 'workorder:add',
                        'preorder-sale:view', 'preorder-sale:add', 'preorder-sale:edit',
                        'sales-return:view', 'sales-return:add', 'sales-return:edit',
                    ]
                },
            ]

            for role_data in default_roles:
                existing = SysRole.query.filter_by(role_code=role_data['role_code']).first()
                if not existing:
                    role = SysRole(**role_data)
                    db.session.add(role)
            db.session.commit()
            logger.info('默认角色检查完成')

            # 注意：不再自动创建 admin 用户。请通过 setup 端点或管理后台创建。
            # POST /api/setup/init {username, password} （仅空库时可用）
            if SysUser.query.count() == 0:
                logger.warning(
                    '数据库无用户，请通过 setup 端点创建初始管理员：'
                    'POST /api/setup/init {username, password}'
                )

            # 创建默认财务账户
            if FinanceAccount.query.count() == 0:
                accounts = [
                    FinanceAccount(account_name='现金', account_type=1, balance=0),
                    FinanceAccount(account_name='银行账户', account_type=2, balance=0),
                    FinanceAccount(account_name='微信', account_type=4, balance=0),
                    FinanceAccount(account_name='支付宝', account_type=3, balance=0)
                ]
                db.session.add_all(accounts)
                db.session.commit()
                logger.info('默认财务账户已创建')

            # 创建默认工单类型（逐个检查，避免重复）
            default_types = [
                {'type_name': '网络维修维护', 'type_code': 'network', 'default_labor_cost': 100, 'estimated_days': 1, 'sort_order': 1},
                {'type_name': '设备维修维护', 'type_code': 'device', 'default_labor_cost': 80, 'estimated_days': 3, 'sort_order': 2},
                {'type_name': '设备配送安装', 'type_code': 'delivery', 'default_labor_cost': 50, 'estimated_days': 1, 'sort_order': 3},
                {'type_name': '监控维修安装', 'type_code': 'monitor', 'default_labor_cost': 200, 'estimated_days': 2, 'sort_order': 4},
                {'type_name': '产品代购', 'type_code': 'purchase', 'default_labor_cost': 30, 'estimated_days': 3, 'sort_order': 5},
                {'type_name': '现场勘察', 'type_code': 'survey', 'default_labor_cost': 50, 'estimated_days': 1, 'sort_order': 6},
                {'type_name': '设备安装(客户自有)', 'type_code': 'install', 'default_labor_cost': 100, 'estimated_days': 1, 'sort_order': 7}
            ]
            for t in default_types:
                existing = WoType.query.filter_by(type_code=t['type_code']).first()
                if not existing:
                    wt = WoType(**t)
                    db.session.add(wt)
            db.session.commit()
            logger.info('默认工单类型检查完成')

            # 创建默认资产类型
            default_asset_types = [
                {'type_code': 'network', 'type_name': '网络类设备', 'icon': 'Connection', 'sort_order': 1},
                {'type_code': 'computer', 'type_name': '电脑办公类设备', 'icon': 'Monitor', 'sort_order': 2},
                {'type_code': 'printer', 'type_name': '打印复印扫描类设备', 'icon': 'Printer', 'sort_order': 3},
                {'type_code': 'security', 'type_name': '监控安防类设备', 'icon': 'VideoCamera', 'sort_order': 4},
                {'type_code': 'server', 'type_name': '服务器机房设备', 'icon': 'Box', 'sort_order': 5},
                {'type_code': 'access', 'type_name': '门禁考勤类设备', 'icon': 'Lock', 'sort_order': 6},
                {'type_code': 'audio', 'type_name': '音响广播类设备', 'icon': 'Microphone', 'sort_order': 7},
                {'type_code': 'other', 'type_name': '其他设备', 'icon': 'More', 'sort_order': 99},
            ]
            for at in default_asset_types:
                existing = AssetType.query.filter_by(type_code=at['type_code']).first()
                if not existing:
                    asset_type = AssetType(**at)
                    db.session.add(asset_type)
            db.session.commit()
            logger.info('默认资产类型检查完成')

        except Exception as e:
            logger.error(f'初始化数据失败: {str(e)}')
            db.session.rollback()


# ============================================
# 业绩统计API
# ============================================

@app.route('/api/statistics/revenue', methods=['GET'])
@jwt_required()
def get_revenue_statistics():
    """营收统计API"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        group_by = request.args.get('group_by', 'day')  # day/week/month

        if not start_date or not end_date:
            return jsonify({'code': 400, 'message': '缺少必要参数: start_date, end_date'}), 400

        # 构建日期范围条件
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        end_dt_str = end_date + ' 23:59:59'

        # 1. 销售单统计 (status=1或2)
        sales_query = SalesOrder.query.filter(
            SalesOrder.created_at >= start_date,
            SalesOrder.created_at <= end_dt_str,
            SalesOrder.status.in_([1, 2])
        )
        total_sales = float(sum(s.actual_amount or 0 for s in sales_query.all()))
        sales_count = sales_query.count()

        # 2. 工单统计 (settlement_status=1)
        workorder_query = WorkOrder.query.filter(
            WorkOrder.created_at >= start_date,
            WorkOrder.created_at <= end_dt_str,
            WorkOrder.settlement_status == 1
        )
        total_workorders = float(sum(w.total_amount or 0 for w in workorder_query.all()))
        workorder_count = workorder_query.count()

        # 3. 回款统计 (status=2)
        received_query = FinanceReceivable.query.filter(
            FinanceReceivable.created_at >= start_date,
            FinanceReceivable.created_at <= end_dt_str,
            FinanceReceivable.status == 2
        )
        total_received = float(sum(r.total_amount or 0 for r in received_query.all()))

        # 4. 未回款统计 (status=0或1)
        unreceived_query = FinanceReceivable.query.filter(
            FinanceReceivable.created_at >= start_date,
            FinanceReceivable.created_at <= end_dt_str,
            FinanceReceivable.status.in_([0, 1])
        )
        total_unreceived = float(sum(
            ((r.total_amount or 0) - (r.received_amount or 0)) for r in unreceived_query.all()
        ))

        # 5. 趋势数据
        trend_data = []
        if group_by == 'day':
            from sqlalchemy import func
            # 销售按天统计
            sales_by_day = db.session.query(
                func.date(SalesOrder.created_at).label('date'),
                func.sum(SalesOrder.actual_amount).label('amount')
            ).filter(
                SalesOrder.created_at >= start_date,
                SalesOrder.created_at <= end_dt_str,
                SalesOrder.status.in_([1, 2])
            ).group_by(func.date(SalesOrder.created_at)).all()

            # 工单按天统计
            wo_by_day = db.session.query(
                func.date(WorkOrder.created_at).label('date'),
                func.sum(WorkOrder.total_cost).label('amount')
            ).filter(
                WorkOrder.created_at >= start_date,
                WorkOrder.created_at <= end_dt_str,
                WorkOrder.settlement_status == 1
            ).group_by(func.date(WorkOrder.created_at)).all()

            # 合并数据
            date_map = {}
            for d, amt in sales_by_day:
                date_str = d.strftime('%Y-%m-%d') if d else ''
                if date_str not in date_map:
                    date_map[date_str] = {'sales_amount': 0, 'workorder_amount': 0}
                date_map[date_str]['sales_amount'] = float(amt or 0)

            for d, amt in wo_by_day:
                date_str = d.strftime('%Y-%m-%d') if d else ''
                if date_str not in date_map:
                    date_map[date_str] = {'sales_amount': 0, 'workorder_amount': 0}
                date_map[date_str]['workorder_amount'] = float(amt or 0)

            for date_key in sorted(date_map.keys()):
                trend_data.append({
                    'date': date_key,
                    'sales_amount': date_map[date_key]['sales_amount'],
                    'workorder_amount': date_map[date_key]['workorder_amount'],
                    'total_revenue': date_map[date_key]['sales_amount'] + date_map[date_key]['workorder_amount']
                })

        summary = {
            'total_sales': total_sales,
            'total_workorders': total_workorders,
            'total_revenue': total_sales + total_workorders,
            'total_received': total_received,
            'total_unreceived': total_unreceived,
            'sales_count': sales_count,
            'workorder_count': workorder_count
        }

        return jsonify({
            'code': 200,
            'data': {
                'summary': summary,
                'trend': trend_data
            }
        })

    except Exception as e:
        logger.error(f'营收统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'统计失败: {str(e)}'}), 500


@app.route('/api/statistics/employee', methods=['GET'])
@jwt_required()
def get_employee_statistics():
    """员工业绩统计API"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        user_id = request.args.get('user_id')

        if not start_date or not end_date:
            return jsonify({'code': 400, 'message': '缺少必要参数: start_date, end_date'}), 400

        end_dt_str = end_date + ' 23:59:59'

        # 查询所有有效用户
        user_query = SysUser.query.filter(SysUser.status == 1)
        if user_id:
            user_query = user_query.filter(SysUser.id == user_id)
        users = user_query.all()

        details = []
        total_sales_count = 0
        total_sales_amount = 0
        total_workorder_count = 0
        total_workorder_amount = 0

        for user in users:
            # 统计该用户的销售单 (作为创建者或关联者)
            user_sales = SalesOrder.query.filter(
                SalesOrder.created_at >= start_date,
                SalesOrder.created_at <= end_dt_str,
                SalesOrder.status.in_([1, 2])
            ).filter(
                (SalesOrder.created_by == user.id) | (SalesOrder.salesman_id == user.id)
            ).all()

            sales_count = len(user_sales)
            sales_amount = float(sum(s.actual_amount or 0 for s in user_sales))

            # 统计该用户的工单 (按assigned_to)
            user_workorders = WorkOrder.query.filter(
                WorkOrder.created_at >= start_date,
                WorkOrder.created_at <= end_dt_str,
                WorkOrder.settlement_status == 1,
                WorkOrder.assigned_to == user.id
            ).all()

            workorder_count = len(user_workorders)
            workorder_amount = float(sum(w.total_amount or 0 for w in user_workorders))

            if sales_count > 0 or workorder_count > 0:
                details.append({
                    'user_id': user.id,
                    'user_name': user.real_name or user.username,
                    'department': user.department or '',
                    'sales_count': sales_count,
                    'sales_amount': sales_amount,
                    'workorder_count': workorder_count,
                    'workorder_amount': workorder_amount,
                    'total_amount': sales_amount + workorder_amount
                })

                total_sales_count += sales_count
                total_sales_amount += sales_amount
                total_workorder_count += workorder_count
                total_workorder_amount += workorder_amount

        # 按总业绩排序
        details.sort(key=lambda x: x['total_amount'], reverse=True)

        summary = {
            'total_employees': len(details),
            'total_sales_count': total_sales_count,
            'total_sales_amount': total_sales_amount,
            'total_workorder_count': total_workorder_count,
            'total_workorder_amount': total_workorder_amount
        }

        return jsonify({
            'code': 200,
            'data': {
                'summary': summary,
                'details': details
            }
        })

    except Exception as e:
        logger.error(f'员工业绩统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'统计失败: {str(e)}'}), 500


@app.route('/api/statistics/customer', methods=['GET'])
@jwt_required()
def get_customer_statistics():
    """客户业绩统计API"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        customer_id = request.args.get('customer_id')

        if not start_date or not end_date:
            return jsonify({'code': 400, 'message': '缺少必要参数: start_date, end_date'}), 400

        end_dt_str = end_date + ' 23:59:59'

        # 查询销售单
        sales_query = SalesOrder.query.filter(
            SalesOrder.created_at >= start_date,
            SalesOrder.created_at <= end_dt_str,
            SalesOrder.status.in_([1, 2])
        )
        if customer_id:
            sales_query = sales_query.filter(SalesOrder.customer_id == customer_id)
        sales_orders = sales_query.all()

        # 按客户分组统计
        customer_stats = {}
        for order in sales_orders:
            cid = order.customer_id
            cname = order.customer_name or '未知客户'
            if cid not in customer_stats:
                customer_stats[cid] = {
                    'customer_id': cid,
                    'customer_name': cname,
                    'order_count': 0,
                    'total_amount': 0,
                    'received_amount': 0
                }
            customer_stats[cid]['order_count'] += 1
            customer_stats[cid]['total_amount'] += float(order.actual_amount or 0)

        # 查询回款记录
        receivable_query = FinanceReceivable.query.filter(
            FinanceReceivable.created_at >= start_date,
            FinanceReceivable.created_at <= end_dt_str
        )
        if customer_id:
            receivable_query = receivable_query.filter(FinanceReceivable.customer_id == customer_id)
        receivables = receivable_query.all()

        for rec in receivables:
            cid = rec.customer_id
            if cid and cid in customer_stats:
                customer_stats[cid]['received_amount'] += float(rec.received_amount or 0)

        # 转换为列表并排序
        top_customers = list(customer_stats.values())
        top_customers.sort(key=lambda x: x['total_amount'], reverse=True)

        total_orders = sum(c['order_count'] for c in top_customers)
        total_amount = sum(c['total_amount'] for c in top_customers)
        total_received = sum(c['received_amount'] for c in top_customers)

        summary = {
            'total_customers': len(top_customers),
            'total_orders': total_orders,
            'total_amount': total_amount,
            'total_received': total_received
        }

        return jsonify({
            'code': 200,
            'data': {
                'summary': summary,
                'top_customers': top_customers[:20]  # 返回前20名
            }
        })

    except Exception as e:
        logger.error(f'客户业绩统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'统计失败: {str(e)}'}), 500


@app.route('/api/statistics/product', methods=['GET'])
@jwt_required()
def get_product_statistics():
    """产品业绩统计API"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            return jsonify({'code': 400, 'message': '缺少必要参数: start_date, end_date'}), 400

        end_dt_str = end_date + ' 23:59:59'

        # 查询销售单明细
        from sqlalchemy import func
        items_query = db.session.query(SalesOrderItem).join(
            SalesOrder, SalesOrderItem.order_id == SalesOrder.id
        ).filter(
            SalesOrder.created_at >= start_date,
            SalesOrder.created_at <= end_dt_str,
            SalesOrder.status.in_([1, 2])
        )

        items = items_query.all()

        # 按产品统计
        product_stats = {}
        category_stats = {}

        for item in items:
            pid = item.product_id
            if pid not in product_stats:
                product_stats[pid] = {
                    'product_id': pid,
                    'product_name': item.product_name or '未知产品',
                    'category_name': item.category_name or '未分类',
                    'total_quantity': 0,
                    'total_amount': 0
                }
            product_stats[pid]['total_quantity'] += float(item.quantity or 0)
            product_stats[pid]['total_amount'] += float(item.total_price or 0)

        # 按分类统计
        for pid, stat in product_stats.items():
            cat_name = stat['category_name']
            if cat_name not in category_stats:
                category_stats[cat_name] = {
                    'category_name': cat_name,
                    'product_count': 0,
                    'total_quantity': 0,
                    'total_amount': 0
                }
            category_stats[cat_name]['product_count'] += 1
            category_stats[cat_name]['total_quantity'] += stat['total_quantity']
            category_stats[cat_name]['total_amount'] += stat['total_amount']

        # 转换为列表并排序
        top_products = list(product_stats.values())
        top_products.sort(key=lambda x: x['total_amount'], reverse=True)

        category_list = list(category_stats.values())
        category_list.sort(key=lambda x: x['total_amount'], reverse=True)

        total_quantity = sum(p['total_quantity'] for p in top_products)
        total_amount = sum(p['total_amount'] for p in top_products)

        summary = {
            'total_products': len(top_products),
            'total_quantity': total_quantity,
            'total_amount': total_amount
        }

        return jsonify({
            'code': 200,
            'data': {
                'summary': summary,
                'top_products': top_products[:20],  # 返回前20名
                'category_stats': category_list
            }
        })

    except Exception as e:
        logger.error(f'产品业绩统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'统计失败: {str(e)}'}), 500


@app.route('/api/statistics/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard_statistics():
    """综合仪表盘数据API"""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        today_start = today + ' 00:00:00'
        today_end = today + ' 23:59:59'

        # 本月日期范围
        current_year = datetime.now().year
        current_month = datetime.now().month
        month_start = f'{current_year}-{current_month:02d}-01'
        month_end = today_end

        # 1. 今日数据
        today_sales = SalesOrder.query.filter(
            SalesOrder.created_at >= today_start,
            SalesOrder.created_at <= today_end,
            SalesOrder.status.in_([1, 2])
        ).all()
        today_sales_count = len(today_sales)
        today_sales_amount = float(sum(s.actual_amount or 0 for s in today_sales))

        today_workorders = WorkOrder.query.filter(
            WorkOrder.created_at >= today_start,
            WorkOrder.created_at <= today_end,
            WorkOrder.settlement_status == 1
        ).all()
        today_workorder_count = len(today_workorders)
        today_workorder_amount = float(sum(w.total_amount or 0 for w in today_workorders))

        today_data = {
            'sales_count': today_sales_count,
            'sales_amount': today_sales_amount,
            'workorder_count': today_workorder_count,
            'workorder_amount': today_workorder_amount
        }

        # 2. 本月数据
        month_sales = SalesOrder.query.filter(
            SalesOrder.created_at >= month_start,
            SalesOrder.created_at <= month_end,
            SalesOrder.status.in_([1, 2])
        ).all()
        month_sales_count = len(month_sales)
        month_sales_amount = float(sum(s.actual_amount or 0 for s in month_sales))

        month_workorders = WorkOrder.query.filter(
            WorkOrder.created_at >= month_start,
            WorkOrder.created_at <= month_end,
            WorkOrder.settlement_status == 1
        ).all()
        month_workorder_count = len(month_workorders)
        month_workorder_amount = float(sum(w.total_amount or 0 for w in month_workorders))

        # 本月应收和回款
        month_receivables = FinanceReceivable.query.filter(
            FinanceReceivable.created_at >= month_start,
            FinanceReceivable.created_at <= month_end
        ).all()
        month_receivable_amount = float(sum(r.total_amount or 0 for r in month_receivables))
        month_received_amount = float(sum(r.received_amount or 0 for r in month_receivables))

        month_data = {
            'sales_count': month_sales_count,
            'sales_amount': month_sales_amount,
            'workorder_count': month_workorder_count,
            'workorder_amount': month_workorder_amount,
            'receivable_amount': month_receivable_amount,
            'received_amount': month_received_amount
        }

        # 3. 待处理数据
        pending_orders = SalesOrder.query.filter(SalesOrder.status == 0).count()
        pending_workorders = WorkOrder.query.filter(WorkOrder.settlement_status == 0).count()
        pending_receivables = FinanceReceivable.query.filter(
            FinanceReceivable.status.in_([0, 1])
        ).count()

        pending_data = {
            'pending_orders': pending_orders,
            'pending_workorders': pending_workorders,
            'pending_receivables': pending_receivables
        }

        # 4. 预警信息
        alerts = []

        # 库存预警 (假设有库存模型)
        low_stock_count = 0
        try:
            low_stock_count = Inventory.query.filter(Inventory.quantity <= Inventory.min_quantity).count()
            if low_stock_count > 0:
                alerts.append({
                    'type': 'inventory',
                    'message': f'有{low_stock_count}个产品库存不足',
                    'count': low_stock_count
                })
        except:
            pass

        # 逾期未回款预警
        overdue_receivables = FinanceReceivable.query.filter(
            FinanceReceivable.status.in_([0, 1]),
            FinanceReceivable.due_date < today
        ).count()
        if overdue_receivables > 0:
            alerts.append({
                'type': 'receivable',
                'message': f'有{overdue_receivables}笔应收账款已逾期',
                'count': overdue_receivables
            })

        # 待审核预警
        if pending_orders > 0:
            alerts.append({
                'type': 'order',
                'message': f'有{pending_orders}个销售单待审核',
                'count': pending_orders
            })

        if pending_workorders > 0:
            alerts.append({
                'type': 'workorder',
                'message': f'有{pending_workorders}个工单待结算',
                'count': pending_workorders
            })

        return jsonify({
            'code': 200,
            'data': {
                'today': today_data,
                'month': month_data,
                'pending': pending_data,
                'alerts': alerts
            }
        })

    except Exception as e:
        logger.error(f'仪表盘数据统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'统计失败: {str(e)}'}), 500


# ============================================
# API路由 - 工资发放模块
# ============================================

@app.route('/api/salary', methods=['GET'])
@jwt_required()
def get_salary_list():
    """工资列表"""
    try:
        user_id = request.args.get('user_id', type=int)
        department = request.args.get('department', '')
        salary_month = request.args.get('salary_month', '')
        status = request.args.get('status', type=int)
        keyword = request.args.get('keyword', '')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        query = Salary.query

        if user_id:
            query = query.filter(Salary.user_id == user_id)
        if department:
            query = query.filter(Salary.department == department)
        if salary_month:
            query = query.filter(Salary.salary_month == salary_month)
        if status is not None:
            query = query.filter(Salary.status == status)
        if keyword:
            query = query.filter(
                db.or_(
                    Salary.salary_no.like(f'%{keyword}%'),
                    Salary.user_name.like(f'%{keyword}%'),
                    Salary.position.like(f'%{keyword}%')
                )
            )

        query = query.order_by(Salary.created_at.desc())
        pagination = query.paginate(page=page, per_page=page_size, error_out=False)

        items = []
        for s in pagination.items:
            items.append({
                'id': s.id,
                'salary_no': s.salary_no,
                'user_id': s.user_id,
                'user_name': s.user_name,
                'department': s.department,
                'position': s.position,
                'salary_month': s.salary_month,
                'base_salary': float(s.base_salary) if s.base_salary else 0,
                'performance_salary': float(s.performance_salary) if s.performance_salary else 0,
                'commission': float(s.commission) if s.commission else 0,
                'subsidy': float(s.subsidy) if s.subsidy else 0,
                'deduction': float(s.deduction) if s.deduction else 0,
                'should_pay': float(s.should_pay) if s.should_pay else 0,
                'tax': float(s.tax) if s.tax else 0,
                'actual_pay': float(s.actual_pay) if s.actual_pay else 0,
                'account_id': s.account_id,
                'account_name': s.account_name,
                'status': s.status,
                'remark': s.remark,
                'created_at': s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else None,
                'paid_at': s.paid_at.strftime('%Y-%m-%d %H:%M:%S') if s.paid_at else None
            })

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'items': items,
                'total': pagination.total,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        logger.error(f'获取工资列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取工资列表失败: {str(e)}'}), 500


@app.route('/api/salary/<int:id>', methods=['GET'])
@jwt_required()
def get_salary_detail(id):
    """工资详情"""
    try:
        s = Salary.query.get(id)
        if not s:
            return jsonify({'code': 404, 'message': '工资单不存在'}), 404

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'id': s.id,
                'salary_no': s.salary_no,
                'user_id': s.user_id,
                'user_name': s.user_name,
                'department': s.department,
                'position': s.position,
                'salary_month': s.salary_month,
                'base_salary': float(s.base_salary) if s.base_salary else 0,
                'performance_salary': float(s.performance_salary) if s.performance_salary else 0,
                'commission': float(s.commission) if s.commission else 0,
                'subsidy': float(s.subsidy) if s.subsidy else 0,
                'deduction': float(s.deduction) if s.deduction else 0,
                'should_pay': float(s.should_pay) if s.should_pay else 0,
                'tax': float(s.tax) if s.tax else 0,
                'actual_pay': float(s.actual_pay) if s.actual_pay else 0,
                'account_id': s.account_id,
                'account_name': s.account_name,
                'status': s.status,
                'remark': s.remark,
                'created_at': s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else None,
                'updated_at': s.updated_at.strftime('%Y-%m-%d %H:%M:%S') if s.updated_at else None,
                'paid_at': s.paid_at.strftime('%Y-%m-%d %H:%M:%S') if s.paid_at else None
            }
        })
    except Exception as e:
        logger.error(f'获取工资详情失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取工资详情失败: {str(e)}'}), 500


@app.route('/api/salary', methods=['POST'])
@jwt_required()
def create_salary():
    """创建工资单"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()

        # 必填字段校验
        if not data.get('user_id') or not data.get('user_name') or not data.get('salary_month'):
            return jsonify({'code': 400, 'message': '员工ID、员工姓名和工资月份为必填项'}), 400

        # 自动生成工资单号：SA + 年月 + 4位序号
        month_str = data.get('salary_month', datetime.now().strftime('%Y%m'))
        prefix = f'SA{month_str.replace("-", "")}'
        last_salary = Salary.query.filter(Salary.salary_no.like(f'{prefix}%')).order_by(Salary.id.desc()).first()
        if last_salary and last_salary.salary_no:
            last_seq = int(last_salary.salary_no[-4:]) if len(last_salary.salary_no) >= 4 else 0
        else:
            last_seq = 0
        salary_no = f'{prefix}{str(last_seq + 1).zfill(4)}'

        # 计算应发金额
        base_salary = float(data.get('base_salary', 0))
        performance_salary = float(data.get('performance_salary', 0))
        commission = float(data.get('commission', 0))
        subsidy = float(data.get('subsidy', 0))
        deduction = float(data.get('deduction', 0))
        should_pay = base_salary + performance_salary + commission + subsidy - deduction

        salary = Salary(
            salary_no=salary_no,
            user_id=data.get('user_id'),
            user_name=data.get('user_name'),
            department=data.get('department', ''),
            position=data.get('position', ''),
            salary_month=data.get('salary_month'),
            base_salary=base_salary,
            performance_salary=performance_salary,
            commission=commission,
            subsidy=subsidy,
            deduction=deduction,
            should_pay=should_pay,
            tax=float(data.get('tax', 0)),
            actual_pay=float(data.get('actual_pay', 0)),
            remark=data.get('remark', ''),
            status=0,  # 待发放
            created_by=current_user_id
        )

        db.session.add(salary)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': {'id': salary.id, 'salary_no': salary_no}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'创建工资单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'创建工资单失败: {str(e)}'}), 500


@app.route('/api/salary/<int:id>', methods=['PUT'])
@jwt_required()
def update_salary(id):
    """更新工资单"""
    try:
        salary = Salary.query.get(id)
        if not salary:
            return jsonify({'code': 404, 'message': '工资单不存在'}), 404

        if salary.status != 0:
            return jsonify({'code': 400, 'message': '只有待发放的工资单可以修改'}), 400

        data = request.get_json()

        # 更新字段
        if 'user_id' in data:
            salary.user_id = data['user_id']
        if 'user_name' in data:
            salary.user_name = data['user_name']
        if 'department' in data:
            salary.department = data['department']
        if 'position' in data:
            salary.position = data['position']
        if 'salary_month' in data:
            salary.salary_month = data['salary_month']
        if 'base_salary' in data:
            salary.base_salary = float(data['base_salary'])
        if 'performance_salary' in data:
            salary.performance_salary = float(data['performance_salary'])
        if 'commission' in data:
            salary.commission = float(data['commission'])
        if 'subsidy' in data:
            salary.subsidy = float(data['subsidy'])
        if 'deduction' in data:
            salary.deduction = float(data['deduction'])
        if 'tax' in data:
            salary.tax = float(data['tax'])
        if 'actual_pay' in data:
            salary.actual_pay = float(data['actual_pay'])
        if 'remark' in data:
            salary.remark = data['remark']

        # 重新计算应发金额
        salary.should_pay = float(salary.base_salary or 0) + float(salary.performance_salary or 0) + \
                           float(salary.commission or 0) + float(salary.subsidy or 0) - float(salary.deduction or 0)

        db.session.commit()

        return jsonify({'code': 200, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新工资单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'更新工资单失败: {str(e)}'}), 500


@app.route('/api/salary/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_salary(id):
    """删除工资单（软删除）"""
    try:
        salary = Salary.query.get(id)
        if not salary:
            return jsonify({'code': 404, 'message': '工资单不存在'}), 404

        salary.status = 2  # 已取消
        db.session.commit()

        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除工资单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'删除工资单失败: {str(e)}'}), 500


@app.route('/api/salary/<int:id>/pay', methods=['POST'])
@jwt_required()
def pay_salary(id):
    """发放工资"""
    try:
        current_user_id = get_jwt_identity()
        current_user_name = get_current_user_name()

        salary = Salary.query.get(id)
        if not salary:
            return jsonify({'code': 404, 'message': '工资单不存在'}), 404

        if salary.status != 0:
            return jsonify({'code': 400, 'message': '只有待发放的工资单可以发放'}), 400

        data = request.get_json()
        account_id = data.get('account_id')

        if not account_id:
            return jsonify({'code': 400, 'message': '请选择发放账户'}), 400

        # 校验账户
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

        if account.status != 1:
            return jsonify({'code': 400, 'message': '账户已禁用'}), 400

        # 校验余额
        actual_pay = float(salary.actual_pay or 0)
        if actual_pay > 0 and float(account.balance or 0) < actual_pay:
            return jsonify({'code': 400, 'message': '账户余额不足'}), 400

        # 扣减账户余额
        account.balance = float(account.balance or 0) - actual_pay

        # 生成支出流水
        record = FinanceRecord(
            account_id=account.id,
            account_name=account.account_name,
            record_type=2,  # 支出
            amount=actual_pay,
            balance=account.balance,
            related_type='salary',
            related_id=salary.id,
            related_no=salary.salary_no,
            remark=f'工资发放 - {salary.user_name}({salary.salary_month})',
            created_by=current_user_id
        )
        db.session.add(record)

        # 更新工资单状态
        salary.status = 1  # 已发放
        salary.account_id = account.id
        salary.account_name = account.account_name
        salary.paid_at = datetime.now()

        db.session.commit()

        return jsonify({'code': 200, 'message': '发放成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'发放工资失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'发放工资失败: {str(e)}'}), 500


@app.route('/api/salary/statistics', methods=['GET'])
@jwt_required()
def get_salary_statistics():
    """工资统计"""
    try:
        salary_month = request.args.get('salary_month', datetime.now().strftime('%Y-%m'))

        # 基础统计
        total_should_pay = db.session.query(db.func.sum(Salary.should_pay)).filter(
            Salary.salary_month == salary_month,
            Salary.status != 2
        ).scalar() or 0

        total_actual_pay = db.session.query(db.func.sum(Salary.actual_pay)).filter(
            Salary.salary_month == salary_month,
            Salary.status != 2
        ).scalar() or 0

        total_tax = db.session.query(db.func.sum(Salary.tax)).filter(
            Salary.salary_month == salary_month,
            Salary.status != 2
        ).scalar() or 0

        total_count = Salary.query.filter(
            Salary.salary_month == salary_month,
            Salary.status != 2
        ).count()

        pending_count = Salary.query.filter(
            Salary.salary_month == salary_month,
            Salary.status == 0
        ).count()

        paid_count = Salary.query.filter(
            Salary.salary_month == salary_month,
            Salary.status == 1
        ).count()

        # 部门统计
        dept_stats = db.session.query(
            Salary.department,
            db.func.count(Salary.id).label('count'),
            db.func.sum(Salary.should_pay).label('should_pay'),
            db.func.sum(Salary.actual_pay).label('actual_pay')
        ).filter(
            Salary.salary_month == salary_month,
            Salary.status != 2
        ).group_by(Salary.department).all()

        department_stats = []
        for dept in dept_stats:
            department_stats.append({
                'department': dept.department or '未分配',
                'count': dept.count,
                'should_pay': float(dept.should_pay or 0),
                'actual_pay': float(dept.actual_pay or 0)
            })

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'salary_month': salary_month,
                'total_should_pay': float(total_should_pay),
                'total_actual_pay': float(total_actual_pay),
                'total_tax': float(total_tax),
                'total_count': total_count,
                'pending_count': pending_count,
                'paid_count': paid_count,
                'department_stats': department_stats
            }
        })
    except Exception as e:
        logger.error(f'获取工资统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取工资统计失败: {str(e)}'}), 500


# ============================================
# API路由 - 费用管理模块
# ============================================

@app.route('/api/expense', methods=['GET'])
@jwt_required()
def get_expense_list():
    """费用列表"""
    try:
        expense_type = request.args.get('expense_type', type=int)
        record_type = request.args.get('record_type', type=int)
        status = request.args.get('status', type=int)
        partner_type = request.args.get('partner_type', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        keyword = request.args.get('keyword', '')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        query = Expense.query

        if expense_type:
            query = query.filter(Expense.expense_type == expense_type)
        if record_type:
            query = query.filter(Expense.record_type == record_type)
        if status is not None:
            query = query.filter(Expense.status == status)
        if partner_type:
            query = query.filter(Expense.partner_type == partner_type)
        if start_date:
            query = query.filter(Expense.expense_date >= start_date)
        if end_date:
            query = query.filter(Expense.expense_date <= end_date)
        if keyword:
            query = query.filter(
                db.or_(
                    Expense.expense_no.like(f'%{keyword}%'),
                    Expense.expense_name.like(f'%{keyword}%'),
                    Expense.partner_name.like(f'%{keyword}%')
                )
            )

        query = query.order_by(Expense.created_at.desc())
        pagination = query.paginate(page=page, per_page=page_size, error_out=False)

        # 费用类型映射
        type_map = {1: '日常费用', 2: '其他收入', 3: '运营支出', 4: '管理费用'}

        items = []
        for e in pagination.items:
            items.append({
                'id': e.id,
                'expense_no': e.expense_no,
                'expense_name': e.expense_name,
                'expense_type': e.expense_type,
                'expense_type_name': type_map.get(e.expense_type, '其他'),
                'amount': float(e.amount) if e.amount else 0,
                'record_type': e.record_type,
                'record_type_name': '收入' if e.record_type == 1 else '支出',
                'account_id': e.account_id,
                'account_name': e.account_name,
                'partner_type': e.partner_type,
                'partner_id': e.partner_id,
                'partner_name': e.partner_name,
                'expense_date': e.expense_date.strftime('%Y-%m-%d') if e.expense_date else None,
                'status': e.status,
                'attachment': e.attachment,
                'remark': e.remark,
                'created_at': e.created_at.strftime('%Y-%m-%d %H:%M:%S') if e.created_at else None
            })

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'items': items,
                'total': pagination.total,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        logger.error(f'获取费用列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取费用列表失败: {str(e)}'}), 500


@app.route('/api/expense/<int:id>', methods=['GET'])
@jwt_required()
def get_expense_detail(id):
    """费用详情"""
    try:
        e = Expense.query.get(id)
        if not e:
            return jsonify({'code': 404, 'message': '费用单不存在'}), 404

        type_map = {1: '日常费用', 2: '其他收入', 3: '运营支出', 4: '管理费用'}

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'id': e.id,
                'expense_no': e.expense_no,
                'expense_name': e.expense_name,
                'expense_type': e.expense_type,
                'expense_type_name': type_map.get(e.expense_type, '其他'),
                'amount': float(e.amount) if e.amount else 0,
                'record_type': e.record_type,
                'record_type_name': '收入' if e.record_type == 1 else '支出',
                'account_id': e.account_id,
                'account_name': e.account_name,
                'partner_type': e.partner_type,
                'partner_id': e.partner_id,
                'partner_name': e.partner_name,
                'expense_date': e.expense_date.strftime('%Y-%m-%d') if e.expense_date else None,
                'status': e.status,
                'attachment': e.attachment,
                'remark': e.remark,
                'created_at': e.created_at.strftime('%Y-%m-%d %H:%M:%S') if e.created_at else None,
                'updated_at': e.updated_at.strftime('%Y-%m-%d %H:%M:%S') if e.updated_at else None
            }
        })
    except Exception as e:
        logger.error(f'获取费用详情失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取费用详情失败: {str(e)}'}), 500


@app.route('/api/expense', methods=['POST'])
@jwt_required()
def create_expense():
    """创建费用"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()

        # 必填字段校验
        if not data.get('expense_name'):
            return jsonify({'code': 400, 'message': '费用名称为必填项'}), 400

        # 自动生成费用单号：EX + 年月日 + 4位序号
        today_str = datetime.now().strftime('%Y%m%d')
        prefix = f'EX{today_str}'
        last_expense = Expense.query.filter(Expense.expense_no.like(f'{prefix}%')).order_by(Expense.id.desc()).first()
        if last_expense and last_expense.expense_no:
            last_seq = int(last_expense.expense_no[-4:]) if len(last_expense.expense_no) >= 4 else 0
        else:
            last_seq = 0
        expense_no = f'{prefix}{str(last_seq + 1).zfill(4)}'

        # 处理费用日期
        expense_date = data.get('expense_date')
        if expense_date:
            expense_date = datetime.strptime(expense_date, '%Y-%m-%d').date()

        expense = Expense(
            expense_no=expense_no,
            expense_name=data.get('expense_name'),
            expense_type=data.get('expense_type', 1),
            amount=float(data.get('amount', 0)),
            record_type=data.get('record_type', 2),
            partner_type=data.get('partner_type', ''),
            partner_id=data.get('partner_id'),
            partner_name=data.get('partner_name', ''),
            expense_date=expense_date,
            remark=data.get('remark', ''),
            status=0,  # 待处理
            created_by=current_user_id
        )

        db.session.add(expense)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': {'id': expense.id, 'expense_no': expense_no}
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'创建费用单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'创建费用单失败: {str(e)}'}), 500


@app.route('/api/expense/<int:id>', methods=['PUT'])
@jwt_required()
def update_expense(id):
    """更新费用"""
    try:
        expense = Expense.query.get(id)
        if not expense:
            return jsonify({'code': 404, 'message': '费用单不存在'}), 404

        if expense.status != 0:
            return jsonify({'code': 400, 'message': '只有待处理的费用单可以修改'}), 400

        data = request.get_json()

        # 更新字段
        if 'expense_name' in data:
            expense.expense_name = data['expense_name']
        if 'expense_type' in data:
            expense.expense_type = data['expense_type']
        if 'amount' in data:
            expense.amount = float(data['amount'])
        if 'record_type' in data:
            expense.record_type = data['record_type']
        if 'partner_type' in data:
            expense.partner_type = data['partner_type']
        if 'partner_id' in data:
            expense.partner_id = data['partner_id']
        if 'partner_name' in data:
            expense.partner_name = data['partner_name']
        if 'expense_date' in data:
            expense.expense_date = datetime.strptime(data['expense_date'], '%Y-%m-%d').date()
        if 'remark' in data:
            expense.remark = data['remark']

        db.session.commit()

        return jsonify({'code': 200, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'更新费用单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'更新费用单失败: {str(e)}'}), 500


@app.route('/api/expense/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_expense(id):
    """删除费用（软删除）"""
    try:
        expense = Expense.query.get(id)
        if not expense:
            return jsonify({'code': 404, 'message': '费用单不存在'}), 404

        expense.status = 2  # 已取消
        db.session.commit()

        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'删除费用单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'删除费用单失败: {str(e)}'}), 500


@app.route('/api/expense/<int:id>/confirm', methods=['POST'])
@jwt_required()
def confirm_expense(id):
    """确认费用"""
    try:
        current_user_id = get_jwt_identity()

        expense = Expense.query.get(id)
        if not expense:
            return jsonify({'code': 404, 'message': '费用单不存在'}), 404

        if expense.status != 0:
            return jsonify({'code': 400, 'message': '只有待处理的费用单可以确认'}), 400

        data = request.get_json()
        account_id = data.get('account_id')

        if not account_id:
            return jsonify({'code': 400, 'message': '请选择收支账户'}), 400

        # 校验账户
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

        if account.status != 1:
            return jsonify({'code': 400, 'message': '账户已禁用'}), 400

        amount = float(expense.amount or 0)

        # 如果是支出，校验余额
        if expense.record_type == 2 and amount > 0:
            if float(account.balance or 0) < amount:
                return jsonify({'code': 400, 'message': '账户余额不足'}), 400
            # 扣减余额
            account.balance = float(account.balance or 0) - amount
        else:
            # 收入增加余额
            account.balance = float(account.balance or 0) + amount

        # 生成流水
        record = FinanceRecord(
            account_id=account.id,
            account_name=account.account_name,
            record_type=expense.record_type,  # 1收入 2支出
            amount=amount,
            balance=account.balance,
            related_type='expense',
            related_id=expense.id,
            related_no=expense.expense_no,
            remark=f'{expense.expense_name} - {expense.partner_name or ""}',
            created_by=current_user_id
        )
        db.session.add(record)

        # 更新费用单状态
        expense.status = 1  # 已确认
        expense.account_id = account.id
        expense.account_name = account.account_name

        db.session.commit()

        return jsonify({'code': 200, 'message': '确认成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f'确认费用失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'确认费用失败: {str(e)}'}), 500


@app.route('/api/expense/statistics', methods=['GET'])
@jwt_required()
def get_expense_statistics():
    """费用统计"""
    try:
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        query = Expense.query.filter(Expense.status == 1)  # 只统计已确认的

        if start_date:
            query = query.filter(Expense.expense_date >= start_date)
        if end_date:
            query = query.filter(Expense.expense_date <= end_date)

        # 总收入
        total_income = db.session.query(db.func.sum(Expense.amount)).filter(
            Expense.status == 1,
            Expense.record_type == 1
        )
        if start_date:
            total_income = total_income.filter(Expense.expense_date >= start_date)
        if end_date:
            total_income = total_income.filter(Expense.expense_date <= end_date)
        total_income = total_income.scalar() or 0

        # 总支出
        total_expense = db.session.query(db.func.sum(Expense.amount)).filter(
            Expense.status == 1,
            Expense.record_type == 2
        )
        if start_date:
            total_expense = total_expense.filter(Expense.expense_date >= start_date)
        if end_date:
            total_expense = total_expense.filter(Expense.expense_date <= end_date)
        total_expense = total_expense.scalar() or 0

        # 按类型统计
        type_stats_query = db.session.query(
            Expense.expense_type,
            Expense.record_type,
            db.func.sum(Expense.amount).label('total')
        ).filter(Expense.status == 1)

        if start_date:
            type_stats_query = type_stats_query.filter(Expense.expense_date >= start_date)
        if end_date:
            type_stats_query = type_stats_query.filter(Expense.expense_date <= end_date)

        type_stats_query = type_stats_query.group_by(Expense.expense_type, Expense.record_type).all()

        type_map = {1: '日常费用', 2: '其他收入', 3: '运营支出', 4: '管理费用'}

        # 整理类型统计
        type_data = {}
        for t in type_stats_query:
            type_id = t.expense_type
            if type_id not in type_data:
                type_data[type_id] = {'income': 0, 'expense': 0}
            if t.record_type == 1:
                type_data[type_id]['income'] = float(t.total or 0)
            else:
                type_data[type_id]['expense'] = float(t.total or 0)

        type_stats = []
        for type_id, amounts in type_data.items():
            type_stats.append({
                'expense_type': type_id,
                'type_name': type_map.get(type_id, '其他'),
                'income': amounts['income'],
                'expense': amounts['expense'],
                'net': amounts['income'] - amounts['expense']
            })

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'total_income': float(total_income),
                'total_expense': float(total_expense),
                'net_amount': float(total_income - total_expense),
                'type_stats': type_stats
            }
        })
    except Exception as e:
        logger.error(f'获取费用统计失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取费用统计失败: {str(e)}'}), 500


# ==================== 对账管理API ====================

# related_type 文字映射
RELATED_TYPE_MAP = {
    'sale': '销售单',
    'purchase': '采购单',
    'work_order': '工单',
    'return_sale': '销售退货',
    'return_purchase': '采购退货',
    'receivable': '应收收款',
    'payable': '应付付款',
    'transfer': '转账',
    'adjust': '调整',
    'salary': '工资发放',
    'expense': '费用'
}

# status 文字映射
STATUS_MAP = {
    0: '待处理',
    1: '部分处理',
    2: '已结清',
    3: '已取消'
}


@app.route('/api/reconciliation/customer', methods=['GET'])
@jwt_required()
def get_customer_reconciliation():
    """客户对账单"""
    try:
        customer_id = request.args.get('customer_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not customer_id:
            return jsonify({'code': 400, 'message': '客户ID不能为空'}), 400

        # 获取客户信息
        customer = BaseCustomer.query.get(customer_id)
        if not customer:
            return jsonify({'code': 404, 'message': '客户不存在'}), 404

        customer_info = {
            'id': customer.id,
            'name': customer.customer_name,
            'contact': customer.contact_name,
            'phone': customer.phone
        }

        # 构建应收账款查询
        receivable_query = FinanceReceivable.query.filter(
            FinanceReceivable.customer_id == customer_id
        )

        if start_date:
            receivable_query = receivable_query.filter(FinanceReceivable.created_at >= start_date)
        if end_date:
            receivable_query = receivable_query.filter(FinanceReceivable.created_at <= end_date + ' 23:59:59')

        receivables = receivable_query.order_by(FinanceReceivable.created_at.desc()).all()

        # 计算汇总数据
        total_receivable = sum(float(r.total_amount or 0) for r in receivables)
        total_received = sum(float(r.received_amount or 0) for r in receivables)
        total_remaining = sum(float(r.remaining_amount or 0) for r in receivables)
        receivable_count = len(receivables)
        received_count = len([r for r in receivables if r.status == 2])

        summary = {
            'total_receivable': total_receivable,
            'total_received': total_received,
            'total_remaining': total_remaining,
            'receivable_count': receivable_count,
            'received_count': received_count
        }

        # 构建明细列表
        details = []
        for r in receivables:
            details.append({
                'receivable_no': r.receivable_no,
                'related_no': r.related_no,
                'related_type': r.related_type,
                'related_type_text': RELATED_TYPE_MAP.get(r.related_type, r.related_type),
                'total_amount': float(r.total_amount or 0),
                'received_amount': float(r.received_amount or 0),
                'remaining_amount': float(r.remaining_amount or 0),
                'status': r.status,
                'status_text': STATUS_MAP.get(r.status, '未知'),
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            })

        # 获取相关收款流水
        record_query = FinanceRecord.query.filter(
            FinanceRecord.related_type == 'receivable'
        )

        # 获取该客户的所有应收单号
        receivable_nos = [r.receivable_no for r in receivables]
        if receivable_nos:
            record_query = record_query.filter(FinanceRecord.related_no.in_(receivable_nos))
        else:
            record_query = record_query.filter(FinanceRecord.related_no == '')

        if start_date:
            record_query = record_query.filter(FinanceRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.filter(FinanceRecord.created_at <= end_date + ' 23:59:59')

        records_data = record_query.order_by(FinanceRecord.created_at.desc()).all()

        records = []
        for r in records_data:
            records.append({
                'record_no': r.id,
                'amount': float(r.amount or 0),
                'account_name': r.account_name,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            })

        return jsonify({
            'code': 200,
            'data': {
                'customer_info': customer_info,
                'period': {
                    'start_date': start_date or '',
                    'end_date': end_date or ''
                },
                'summary': summary,
                'details': details,
                'records': records
            }
        })
    except Exception as e:
        logger.error(f'获取客户对账单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取客户对账单失败: {str(e)}'}), 500


@app.route('/api/reconciliation/customers', methods=['GET'])
@jwt_required()
def get_customers_reconciliation():
    """客户对账单列表（多客户汇总）"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        keyword = request.args.get('keyword', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))

        # 构建客户查询
        customer_query = BaseCustomer.query
        if keyword:
            customer_query = customer_query.filter(
                db.or_(
                    BaseCustomer.customer_name.like(f'%{keyword}%'),
                    BaseCustomer.contact_name.like(f'%{keyword}%'),
                    BaseCustomer.phone.like(f'%{keyword}%')
                )
            )

        total_count = customer_query.count()
        customers = customer_query.offset((page - 1) * page_size).limit(page_size).all()

        # 构建结果列表
        result = []
        for customer in customers:
            # 查询该客户的应收账款
            receivable_query = FinanceReceivable.query.filter(
                FinanceReceivable.customer_id == customer.id
            )

            if start_date:
                receivable_query = receivable_query.filter(FinanceReceivable.created_at >= start_date)
            if end_date:
                receivable_query = receivable_query.filter(FinanceReceivable.created_at <= end_date + ' 23:59:59')

            receivables = receivable_query.all()

            total_receivable = sum(float(r.total_amount or 0) for r in receivables)
            total_received = sum(float(r.received_amount or 0) for r in receivables)
            total_remaining = sum(float(r.remaining_amount or 0) for r in receivables)

            result.append({
                'customer_id': customer.id,
                'customer_name': customer.customer_name,
                'contact': customer.contact_name,
                'phone': customer.phone,
                'total_receivable': total_receivable,
                'total_received': total_received,
                'total_remaining': total_remaining,
                'receivable_count': len(receivables)
            })

        return jsonify({
            'code': 200,
            'data': {
                'items': result,
                'total': total_count,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        logger.error(f'获取客户对账单列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取客户对账单列表失败: {str(e)}'}), 500


@app.route('/api/reconciliation/supplier', methods=['GET'])
@jwt_required()
def get_supplier_reconciliation():
    """供应商对账单"""
    try:
        supplier_id = request.args.get('supplier_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not supplier_id:
            return jsonify({'code': 400, 'message': '供应商ID不能为空'}), 400

        # 获取供应商信息
        supplier = BaseSupplier.query.get(supplier_id)
        if not supplier:
            return jsonify({'code': 404, 'message': '供应商不存在'}), 404

        supplier_info = {
            'id': supplier.id,
            'name': supplier.supplier_name,
            'contact': supplier.contact_name,
            'phone': supplier.phone
        }

        # 构建应付账款查询
        payable_query = FinancePayable.query.filter(
            FinancePayable.supplier_id == supplier_id
        )

        if start_date:
            payable_query = payable_query.filter(FinancePayable.created_at >= start_date)
        if end_date:
            payable_query = payable_query.filter(FinancePayable.created_at <= end_date + ' 23:59:59')

        payables = payable_query.order_by(FinancePayable.created_at.desc()).all()

        # 计算汇总数据
        total_payable = sum(float(p.total_amount or 0) for p in payables)
        total_paid = sum(float(p.paid_amount or 0) for p in payables)
        total_remaining = sum(float(p.remaining_amount or 0) for p in payables)
        payable_count = len(payables)
        paid_count = len([p for p in payables if p.status == 2])

        summary = {
            'total_payable': total_payable,
            'total_paid': total_paid,
            'total_remaining': total_remaining,
            'payable_count': payable_count,
            'paid_count': paid_count
        }

        # 构建明细列表
        details = []
        for p in payables:
            details.append({
                'payable_no': p.payable_no,
                'related_no': p.related_no,
                'related_type': p.related_type,
                'related_type_text': RELATED_TYPE_MAP.get(p.related_type, p.related_type),
                'total_amount': float(p.total_amount or 0),
                'paid_amount': float(p.paid_amount or 0),
                'remaining_amount': float(p.remaining_amount or 0),
                'status': p.status,
                'status_text': STATUS_MAP.get(p.status, '未知'),
                'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else ''
            })

        # 获取相关付款流水
        record_query = FinanceRecord.query.filter(
            FinanceRecord.related_type == 'payable'
        )

        # 获取该供应商的所有应付单号
        payable_nos = [p.payable_no for p in payables]
        if payable_nos:
            record_query = record_query.filter(FinanceRecord.related_no.in_(payable_nos))
        else:
            record_query = record_query.filter(FinanceRecord.related_no == '')

        if start_date:
            record_query = record_query.filter(FinanceRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.filter(FinanceRecord.created_at <= end_date + ' 23:59:59')

        records_data = record_query.order_by(FinanceRecord.created_at.desc()).all()

        records = []
        for r in records_data:
            records.append({
                'record_no': r.id,
                'amount': float(r.amount or 0),
                'account_name': r.account_name,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            })

        return jsonify({
            'code': 200,
            'data': {
                'supplier_info': supplier_info,
                'period': {
                    'start_date': start_date or '',
                    'end_date': end_date or ''
                },
                'summary': summary,
                'details': details,
                'records': records
            }
        })
    except Exception as e:
        logger.error(f'获取供应商对账单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取供应商对账单失败: {str(e)}'}), 500


@app.route('/api/reconciliation/suppliers', methods=['GET'])
@jwt_required()
def get_suppliers_reconciliation():
    """供应商对账单列表（多供应商汇总）"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        keyword = request.args.get('keyword', '')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))

        # 构建供应商查询
        supplier_query = BaseSupplier.query
        if keyword:
            supplier_query = supplier_query.filter(
                db.or_(
                    BaseSupplier.supplier_name.like(f'%{keyword}%'),
                    BaseSupplier.contact_name.like(f'%{keyword}%'),
                    BaseSupplier.phone.like(f'%{keyword}%')
                )
            )

        total_count = supplier_query.count()
        suppliers = supplier_query.offset((page - 1) * page_size).limit(page_size).all()

        # 构建结果列表
        result = []
        for supplier in suppliers:
            # 查询该供应商的应付账款
            payable_query = FinancePayable.query.filter(
                FinancePayable.supplier_id == supplier.id
            )

            if start_date:
                payable_query = payable_query.filter(FinancePayable.created_at >= start_date)
            if end_date:
                payable_query = payable_query.filter(FinancePayable.created_at <= end_date + ' 23:59:59')

            payables = payable_query.all()

            total_payable = sum(float(p.total_amount or 0) for p in payables)
            total_paid = sum(float(p.paid_amount or 0) for p in payables)
            total_remaining = sum(float(p.remaining_amount or 0) for p in payables)

            result.append({
                'supplier_id': supplier.id,
                'supplier_name': supplier.supplier_name,
                'contact': supplier.contact_name,
                'phone': supplier.phone,
                'total_payable': total_payable,
                'total_paid': total_paid,
                'total_remaining': total_remaining,
                'payable_count': len(payables)
            })

        return jsonify({
            'code': 200,
            'data': {
                'items': result,
                'total': total_count,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        logger.error(f'获取供应商对账单列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取供应商对账单列表失败: {str(e)}'}), 500


@app.route('/api/reconciliation/account', methods=['GET'])
@jwt_required()
def get_account_reconciliation():
    """账户对账单"""
    try:
        account_id = request.args.get('account_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not account_id:
            return jsonify({'code': 400, 'message': '账户ID不能为空'}), 400

        # 获取账户信息
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

        account_info = {
            'id': account.id,
            'account_name': account.account_name,
            'account_type': account.account_type,
            'account_no': account.account_no
        }

        # 计算期初余额
        if start_date:
            # 获取start_date前的最后一条流水
            last_record_before = FinanceRecord.query.filter(
                FinanceRecord.account_id == account_id,
                FinanceRecord.created_at < start_date
            ).order_by(FinanceRecord.created_at.desc()).first()
            
            if last_record_before:
                opening_balance = float(last_record_before.balance_after or 0)
            else:
                opening_balance = float(account.balance or 0)
        else:
            opening_balance = float(account.balance or 0)

        # 计算期末余额
        if end_date:
            # 获取end_date当天的最后一条流水
            last_record_in_period = FinanceRecord.query.filter(
                FinanceRecord.account_id == account_id,
                FinanceRecord.created_at <= end_date + ' 23:59:59'
            ).order_by(FinanceRecord.created_at.desc()).first()
            
            if last_record_in_period:
                closing_balance = float(last_record_in_period.balance_after or 0)
            else:
                closing_balance = opening_balance
        else:
            closing_balance = float(account.balance or 0)

        # 构建流水查询
        record_query = FinanceRecord.query.filter(
            FinanceRecord.account_id == account_id
        )

        if start_date:
            record_query = record_query.filter(FinanceRecord.created_at >= start_date)
        if end_date:
            record_query = record_query.filter(FinanceRecord.created_at <= end_date + ' 23:59:59')

        records_data = record_query.order_by(FinanceRecord.created_at.desc()).all()

        # 计算本期收入和支出
        total_income = sum(float(r.amount or 0) for r in records_data if r.record_type == 1)
        total_expense = sum(float(r.amount or 0) for r in records_data if r.record_type == 2)
        net_amount = total_income - total_expense

        # 构建流水列表
        records = []
        for r in records_data:
            record_type_text = '收入' if r.record_type == 1 else '支出'
            records.append({
                'record_type': r.record_type,
                'record_type_text': record_type_text,
                'amount': float(r.amount or 0),
                'balance_before': float(r.balance_before or 0),
                'balance_after': float(r.balance_after or 0),
                'related_type': r.related_type,
                'related_type_text': RELATED_TYPE_MAP.get(r.related_type, r.related_type or ''),
                'related_no': r.related_no,
                'remark': getattr(r, 'remark', '') or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            })

        return jsonify({
            'code': 200,
            'data': {
                'account_info': account_info,
                'period': {
                    'start_date': start_date or '',
                    'end_date': end_date or ''
                },
                'opening_balance': opening_balance,
                'closing_balance': closing_balance,
                'total_income': total_income,
                'total_expense': total_expense,
                'net_amount': net_amount,
                'records': records
            }
        })
    except Exception as e:
        logger.error(f'获取账户对账单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取账户对账单失败: {str(e)}'}), 500


@app.route('/api/reconciliation/accounts', methods=['GET'])
@jwt_required()
def get_accounts_reconciliation():
    """账户对账单列表（多账户汇总）"""
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))

        # 查询所有账户
        account_query = FinanceAccount.query

        total_count = account_query.count()
        accounts = account_query.offset((page - 1) * page_size).limit(page_size).all()

        # 构建结果列表
        result = []
        for account in accounts:
            # 计算期初余额
            if start_date:
                last_record_before = FinanceRecord.query.filter(
                    FinanceRecord.account_id == account.id,
                    FinanceRecord.created_at < start_date
                ).order_by(FinanceRecord.created_at.desc()).first()
                
                if last_record_before:
                    opening_balance = float(last_record_before.balance_after or 0)
                else:
                    opening_balance = float(account.balance or 0)
            else:
                opening_balance = float(account.balance or 0)

            # 计算期末余额
            if end_date:
                last_record_in_period = FinanceRecord.query.filter(
                    FinanceRecord.account_id == account.id,
                    FinanceRecord.created_at <= end_date + ' 23:59:59'
                ).order_by(FinanceRecord.created_at.desc()).first()
                
                if last_record_in_period:
                    closing_balance = float(last_record_in_period.balance_after or 0)
                else:
                    closing_balance = opening_balance
            else:
                closing_balance = float(account.balance or 0)

            # 查询本期流水
            record_query = FinanceRecord.query.filter(
                FinanceRecord.account_id == account.id
            )

            if start_date:
                record_query = record_query.filter(FinanceRecord.created_at >= start_date)
            if end_date:
                record_query = record_query.filter(FinanceRecord.created_at <= end_date + ' 23:59:59')

            records_data = record_query.all()

            total_income = sum(float(r.amount or 0) for r in records_data if r.record_type == 1)
            total_expense = sum(float(r.amount or 0) for r in records_data if r.record_type == 2)
            net_amount = total_income - total_expense

            result.append({
                'account_id': account.id,
                'account_name': account.account_name,
                'account_type': account.account_type,
                'account_no': account.account_no,
                'opening_balance': opening_balance,
                'closing_balance': closing_balance,
                'total_income': total_income,
                'total_expense': total_expense,
                'net_amount': net_amount,
                'record_count': len(records_data)
            })

        return jsonify({
            'code': 200,
            'data': {
                'list': result,
                'total': total_count,
                'page': page,
                'page_size': page_size
            }
        })
    except Exception as e:
        logger.error(f'获取账户对账单列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取账户对账单列表失败: {str(e)}'}), 500


@app.route('/api/purchase/orders/<int:order_id>/invoice', methods=['POST'])
@jwt_required()
def create_purchase_order_invoice(order_id):
    """为采购单创建发票"""
    order = PurchaseOrder.query.get(order_id)
    if not order:
        return jsonify({'code': 404, 'message': '采购单不存在'}), 404

    data = request.get_json()

    # 生成发票号
    date_str = datetime.now().strftime('%Y%m%d')
    last_invoice = PurchaseInvoice.query.filter(
        PurchaseInvoice.invoice_no.like(f'PI{date_str}%')
    ).order_by(PurchaseInvoice.id.desc()).first()

    if last_invoice:
        last_no = int(last_invoice.invoice_no[-4:])
        new_no = f'PI{date_str}{last_no + 1:04d}'
    else:
        new_no = f'PI{date_str}0001'

    # 计算税额和合计
    amount = float(data.get('amount', 0))
    tax_rate = float(data.get('tax_rate', 0))
    tax_amount = amount * tax_rate / 100
    total_amount = amount + tax_amount

    invoice = PurchaseInvoice(
        invoice_no=new_no,
        invoice_code=data.get('invoice_code'),
        invoice_type=data.get('invoice_type', 1),
        purchase_order_id=order.id,
        purchase_order_no=order.order_no,
        supplier_id=order.supplier_id,
        supplier_name=order.supplier_name,
        amount=amount,
        tax_rate=tax_rate,
        tax_amount=tax_amount,
        total_amount=total_amount,
        invoice_date=data.get('invoice_date'),
        status=0,  # 待认证
        remark=data.get('remark'),
        created_by=get_jwt_identity()
    )

    # 更新采购单has_invoice
    order.has_invoice = 1

    db.session.add(invoice)
    db.session.commit()

    return jsonify({
        'code': 200,
        'message': '发票创建成功',
        'data': {'id': invoice.id, 'invoice_no': invoice.invoice_no}
    })


# ============================================
# 资产管理API
# ============================================

# 资产类型初始化数据
ASSET_TYPE_DATA = [
    {'type_code': 'network', 'type_name': '网络类设备', 'icon': 'Connection', 'sort_order': 1},
    {'type_code': 'computer', 'type_name': '电脑办公类设备', 'icon': 'Monitor', 'sort_order': 2},
    {'type_code': 'printer', 'type_name': '打印复印扫描类设备', 'icon': 'Printer', 'sort_order': 3},
    {'type_code': 'security', 'type_name': '监控安防类设备', 'icon': 'VideoCamera', 'sort_order': 4},
    {'type_code': 'server', 'type_name': '服务器机房设备', 'icon': 'Box', 'sort_order': 5},
    {'type_code': 'access', 'type_name': '门禁考勤类设备', 'icon': 'Lock', 'sort_order': 6},
    {'type_code': 'audio', 'type_name': '音响广播类设备', 'icon': 'Microphone', 'sort_order': 7},
    {'type_code': 'other', 'type_name': '其他设备', 'icon': 'More', 'sort_order': 99},
]



@app.route('/api/asset/types', methods=['GET'])
@jwt_required()
def get_asset_types():
    """获取资产类型列表"""
    try:
        types = AssetType.query.filter_by(status=1).order_by(AssetType.sort_order).all()
        data = [to_dict(t) for t in types]
        # 调试日志
        logger.info(f'[DEBUG] Asset types data: {data}')
        for t in data:
            logger.info(f'[DEBUG] type_name: {t.get("type_name")}, type_code: {t.get("type_code")}')
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': data
        })
    except Exception as e:
        logger.error(f'获取资产类型列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


def generate_asset_no():
    """生成资产编号: ZC + 年月 + 4位序号"""
    date_str = datetime.now().strftime('%Y%m')
    prefix = f'ZC{date_str}'
    # 查询本月最后一个资产编号
    last_asset = Asset.query.filter(Asset.asset_no.like(f'{prefix}%')).order_by(Asset.id.desc()).first()
    if last_asset and last_asset.asset_no:
        try:
            last_seq = int(last_asset.asset_no[-4:])
            seq = last_seq + 1
        except:
            seq = 1
    else:
        seq = 1
    return f'{prefix}{seq:04d}'


@app.route('/api/assets', methods=['GET'])
@jwt_required()
def get_assets():
    """资产列表查询"""
    try:
        # 获取查询参数
        customer_id = request.args.get('customer_id', type=int)
        office_id = request.args.get('office_id', type=int)
        asset_type_id = request.args.get('asset_type_id', type=int)
        warranty_status = request.args.get('warranty_status', type=int)
        asset_status = request.args.get('asset_status', type=int)
        keyword = request.args.get('keyword')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        # 构建查询
        query = Asset.query

        # 客户ID必填
        if customer_id:
            query = query.filter(Asset.customer_id == customer_id)

        if office_id:
            query = query.filter(Asset.office_id == office_id)
        if asset_type_id:
            query = query.filter(Asset.asset_type_id == asset_type_id)
        if warranty_status is not None:
            query = query.filter(Asset.warranty_status == warranty_status)
        if asset_status is not None:
            query = query.filter(Asset.asset_status == asset_status)
        if keyword:
            query = query.filter(
                db.or_(
                    Asset.asset_name.contains(keyword),
                    Asset.asset_no.contains(keyword),
                    Asset.sn_code.contains(keyword),
                    Asset.device_no.contains(keyword),
                    Asset.responsible_person.contains(keyword)
                )
            )
        if start_date:
            query = query.filter(Asset.created_at >= start_date)
        if end_date:
            query = query.filter(Asset.created_at <= end_date + ' 23:59:59')

        # 排除已删除的（停用状态）
        query = query.filter(Asset.asset_status != 5)

        # 排序和分页
        query = query.order_by(Asset.created_at.desc())
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()

        # 获取所有资产类型映射
        asset_types = {t.id: t for t in AssetType.query.all()}

        # 构建返回数据，添加asset_type_code
        result_items = []
        for item in items:
            item_dict = to_dict(item)
            # 根据asset_type_id获取type_code
            type_info = asset_types.get(item.asset_type_id)
            if type_info:
                item_dict['asset_type_code'] = type_info.type_code
            result_items.append(item_dict)

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'items': result_items
            }
        })
    except Exception as e:
        logger.error(f'获取资产列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


@app.route('/api/assets/<int:id>', methods=['GET'])
@jwt_required()
def get_asset(id):
    """获取资产详情"""
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'获取资产详情失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


@app.route('/api/assets', methods=['POST'])
@jwt_required()
def create_asset():
    """创建资产"""
    try:
        data = request.get_json()

        # 必填字段校验
        required_fields = ['customer_id', 'customer_name', 'asset_type_id', 'asset_name']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400

        # 获取资产类型名称
        asset_type = AssetType.query.get(data['asset_type_id'])
        if not asset_type:
            return jsonify({'code': 400, 'message': '资产类型不存在'}), 400

        # 自动生成资产编号
        asset_no = generate_asset_no()

        # 计算质保状态
        warranty_expire_date = data.get('warranty_expire_date')
        warranty_status = 1
        if warranty_expire_date and warranty_expire_date != '':
            try:
                expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                warranty_status = 0 if expire_date < datetime.now().date() else 1
            except:
                warranty_expire_date = None
        else:
            warranty_expire_date = None

        # 创建资产
        # 处理办公室名称
        office_id = data.get('office_id')
        office_name = data.get('office_name')
        if office_id and not office_name:
            office = Office.query.get(office_id)
            if office:
                office_name = office.name

        # 处理日期字段，空字符串转为 None
        def parse_date(date_val):
            if date_val and date_val != '':
                return date_val
            return None

        asset = Asset(
            asset_no=asset_no,
            customer_id=data['customer_id'],
            customer_name=data['customer_name'],
            office_id=office_id,
            office_name=office_name,
            location=data.get('location') or None,
            asset_type_id=data['asset_type_id'],
            asset_type_name=asset_type.type_name,
            asset_name=data['asset_name'],
            device_no=data.get('device_no') or None,
            sn_code=data.get('sn_code') or None,
            register_date=parse_date(data.get('register_date')),
            purchase_date=parse_date(data.get('purchase_date')),
            warranty_expire_date=warranty_expire_date,
            warranty_status=warranty_status,
            asset_status=data.get('asset_status', 1),
            responsible_person=data.get('responsible_person') or None,
            contact_phone=data.get('contact_phone') or None,
            ip_address=data.get('ip_address') or None,
            login_password=data.get('login_password') or None,
            remark=data.get('remark') or None,
            asset_data=data.get('asset_data'),
            sales_order_id=data.get('sales_order_id'),
            sales_order_no=data.get('sales_order_no'),
            created_by=get_jwt_identity(),
            created_by_name=get_current_user_name()
        )

        db.session.add(asset)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'创建资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@app.route('/api/assets/<int:id>', methods=['PUT'])
@jwt_required()
def update_asset(id):
    """更新资产"""
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404

        data = request.get_json()

        # 如果更新了办公室ID，同步更新办公室名称
        if 'office_id' in data and data['office_id']:
            office = Office.query.get(data['office_id'])
            if office:
                asset.office_id = data['office_id']
                asset.office_name = office.name
        elif 'office_name' in data:
            asset.office_name = data['office_name']

        # 更新字段（排除 office_id 和 office_name，因为上面已处理）
        updatable_fields = [
            'customer_id', 'customer_name', 'location',
            'asset_type_id', 'asset_name', 'device_no', 'sn_code',
            'register_date', 'purchase_date', 'warranty_expire_date',
            'asset_status', 'responsible_person', 'contact_phone',
            'ip_address', 'login_password', 'remark', 'asset_data',
            'sales_order_id', 'sales_order_no'
        ]

        for field in updatable_fields:
            if field in data:
                setattr(asset, field, data[field])

        # 如果更新了资产类型ID，同步更新类型名称
        if 'asset_type_id' in data:
            asset_type = AssetType.query.get(data['asset_type_id'])
            if asset_type:
                asset.asset_type_name = asset_type.type_name

        # 如果更新了质保到期日，重新计算质保状态
        if 'warranty_expire_date' in data and data['warranty_expire_date']:
            try:
                expire_date = datetime.strptime(data['warranty_expire_date'], '%Y-%m-%d').date()
                asset.warranty_status = 0 if expire_date < datetime.now().date() else 1
            except:
                pass

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '更新成功',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'更新资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新失败: {str(e)}'}), 500


@app.route('/api/assets/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_asset(id):
    """删除资产（软删除，将状态设为停用）"""
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404

        asset.asset_status = 5  # 停用状态
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '删除成功',
            'data': {'id': id}
        })
    except Exception as e:
        logger.error(f'删除资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500


@app.route('/api/assets/<int:id>/scrap', methods=['POST'])
@jwt_required()
def scrap_asset(id):
    """资产报废"""
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404

        asset.asset_status = 4  # 报废状态
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '资产已报废',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'资产报废失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'报废失败: {str(e)}'}), 500


@app.route('/api/assets/import', methods=['POST'])
@jwt_required()
def import_assets():
    """批量导入资产"""
    try:
        if 'file' not in request.files:
            return jsonify({'code': 400, 'message': '请上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'code': 400, 'message': '文件名为空'}), 400

        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # 获取表头
        headers = [cell.value for cell in ws[1]]

        # 获取所有资产类型
        asset_types = {t.type_name: t for t in AssetType.query.all()}
        asset_types_by_code = {t.type_code: t for t in AssetType.query.all()}

        success_count = 0
        error_list = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = dict(zip(headers, row))

                # 必填字段校验
                customer_id = row_data.get('客户ID')
                customer_name = row_data.get('客户名称')
                asset_type_name = row_data.get('资产类型')
                asset_name = row_data.get('资产名称')

                if not customer_id or not customer_name or not asset_name:
                    error_list.append({'row': idx, 'error': '缺少必填字段'})
                    continue

                # 查找资产类型
                asset_type = asset_types.get(asset_type_name)
                if not asset_type:
                    # 尝试通过编码查找
                    asset_type = asset_types_by_code.get(asset_type_name)
                if not asset_type:
                    error_list.append({'row': idx, 'error': '资产类型不存在'})
                    continue

                # 计算质保状态
                warranty_expire_date = row_data.get('质保到期日')
                warranty_status = 1
                if warranty_expire_date:
                    try:
                        if isinstance(warranty_expire_date, str):
                            expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                        else:
                            expire_date = warranty_expire_date
                        warranty_status = 0 if expire_date < datetime.now().date() else 1
                    except:
                        pass

                # 创建资产
                asset = Asset(
                    asset_no=generate_asset_no(),
                    customer_id=int(customer_id),
                    customer_name=customer_name,
                    office_id=row_data.get('办公室ID'),
                    office_name=row_data.get('办公室名称'),
                    location=row_data.get('存放位置'),
                    asset_type_id=asset_type.id,
                    asset_type_name=asset_type.type_name,
                    asset_name=asset_name,
                    device_no=row_data.get('设备编号'),
                    sn_code=row_data.get('SN序列号'),
                    register_date=row_data.get('登记日期'),
                    purchase_date=row_data.get('采购日期'),
                    warranty_expire_date=warranty_expire_date,
                    warranty_status=warranty_status,
                    asset_status=1,
                    responsible_person=row_data.get('责任人'),
                    contact_phone=row_data.get('联系电话'),
                    ip_address=row_data.get('IP地址'),
                    login_password=row_data.get('登录密码'),
                    remark=row_data.get('备注'),
                    created_by=get_jwt_identity(),
                    created_by_name=get_current_user_name()
                )

                db.session.add(asset)
                success_count += 1

            except Exception as e:
                error_list.append({'row': idx, 'error': str(e)})

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'导入完成，成功{success_count}条，失败{len(error_list)}条',
            'data': {
                'success_count': success_count,
                'error_count': len(error_list),
                'errors': error_list
            }
        })
    except Exception as e:
        logger.error(f'导入资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500


@app.route('/api/assets/export', methods=['GET'])
@jwt_required()
def export_assets():
    """导出资产"""
    try:
        # 获取查询参数
        customer_id = request.args.get('customer_id', type=int)
        asset_type_id = request.args.get('asset_type_id', type=int)
        keyword = request.args.get('keyword')

        # 构建查询
        query = Asset.query.filter(Asset.asset_status != 5)

        if customer_id:
            query = query.filter(Asset.customer_id == customer_id)
        if asset_type_id:
            query = query.filter(Asset.asset_type_id == asset_type_id)
        if keyword:
            query = query.filter(
                db.or_(
                    Asset.asset_name.contains(keyword),
                    Asset.asset_no.contains(keyword),
                    Asset.sn_code.contains(keyword)
                )
            )

        assets = query.order_by(Asset.created_at.desc()).all()

        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '资产台账'

        # 表头
        headers = [
            '资产编号', '资产名称', '资产类型', '客户名称', '办公室',
            '存放位置', '设备编号', 'SN序列号', '登记日期', '采购日期',
            '质保到期日', '质保状态', '资产状态', '责任人', '联系电话',
            'IP地址', '备注'
        ]
        ws.append(headers)

        # 状态映射
        warranty_status_map = {0: '过保', 1: '在保'}
        asset_status_map = {1: '正常使用', 2: '维修中', 3: '闲置', 4: '报废', 5: '停用'}

        # 数据行
        for asset in assets:
            ws.append([
                asset.asset_no,
                asset.asset_name,
                asset.asset_type_name,
                asset.customer_name,
                asset.office_name,
                asset.location,
                asset.device_no,
                asset.sn_code,
                asset.register_date.strftime('%Y-%m-%d') if asset.register_date else '',
                asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                asset.warranty_expire_date.strftime('%Y-%m-%d') if asset.warranty_expire_date else '',
                warranty_status_map.get(asset.warranty_status, '未知'),
                asset_status_map.get(asset.asset_status, '未知'),
                asset.responsible_person,
                asset.contact_phone,
                asset.ip_address,
                asset.remark
            ])

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'资产台账_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f'导出资产失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


@app.route('/api/assets/by-customer', methods=['GET'])
@jwt_required()
def get_assets_by_customer():
    """根据客户获取资产列表（用于销售模块关联）"""
    try:
        customer_id = request.args.get('customer_id', type=int)
        if not customer_id:
            return jsonify({'code': 400, 'message': '缺少客户ID参数'}), 400

        assets = Asset.query.filter(
            Asset.customer_id == customer_id,
            Asset.asset_status.in_([1, 2, 3])  # 正常使用、维修中、闲置
        ).order_by(Asset.created_at.desc()).all()

        # 返回简化字段
        data = []
        for asset in assets:
            data.append({
                'id': asset.id,
                'asset_no': asset.asset_no,
                'asset_name': asset.asset_name,
                'asset_type_name': asset.asset_type_name,
                'location': asset.location,
                'responsible_person': asset.responsible_person,
                'warranty_status': asset.warranty_status
            })

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': data
        })
    except Exception as e:
        logger.error(f'获取客户资产列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


# ============================================
# 销售模块资产关联API
# ============================================

@app.route('/api/sales/orders/<int:order_id>/assets', methods=['POST'])
@jwt_required()
def create_assets_for_sales_order(order_id):
    """销售单创建时同步创建资产"""
    try:
        order = SalesOrder.query.get(order_id)
        if not order:
            return jsonify({'code': 404, 'message': '销售单不存在'}), 404

        data = request.get_json()
        assets_data = data.get('assets', [])

        if not assets_data:
            return jsonify({'code': 400, 'message': '资产数据不能为空'}), 400

        created_assets = []
        for asset_data in assets_data:
            # 必填字段校验
            if not asset_data.get('asset_type_id') or not asset_data.get('asset_name'):
                continue

            # 获取资产类型名称
            asset_type = AssetType.query.get(asset_data['asset_type_id'])
            if not asset_type:
                continue

            # 计算质保状态
            warranty_expire_date = asset_data.get('warranty_expire_date')
            warranty_status = 1
            if warranty_expire_date:
                try:
                    expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                    warranty_status = 0 if expire_date < datetime.now().date() else 1
                except:
                    pass

            asset = Asset(
                asset_no=generate_asset_no(),
                customer_id=order.customer_id,
                customer_name=order.customer_name,
                office_id=asset_data.get('office_id'),
                office_name=asset_data.get('office_name'),
                location=asset_data.get('location'),
                asset_type_id=asset_data['asset_type_id'],
                asset_type_name=asset_type.type_name,
                asset_name=asset_data['asset_name'],
                device_no=asset_data.get('device_no'),
                sn_code=asset_data.get('sn_code'),
                purchase_date=asset_data.get('purchase_date'),
                warranty_expire_date=warranty_expire_date,
                warranty_status=warranty_status,
                asset_status=1,
                responsible_person=asset_data.get('responsible_person'),
                contact_phone=asset_data.get('contact_phone'),
                ip_address=asset_data.get('ip_address'),
                login_password=asset_data.get('login_password'),
                remark=asset_data.get('remark'),
                asset_data=asset_data.get('asset_data'),
                sales_order_id=order.id,
                sales_order_no=order.order_no,
                created_by=get_jwt_identity(),
                created_by_name=get_current_user_name()
            )

            db.session.add(asset)
            created_assets.append(asset)

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'成功创建{len(created_assets)}个资产',
            'data': [to_dict(asset) for asset in created_assets]
        })
    except Exception as e:
        logger.error(f'销售单创建资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@app.route('/api/sales/returns/<int:return_id>/unbind-assets', methods=['POST'])
@jwt_required()
def unbind_assets_for_return(return_id):
    """销售退货解绑资产"""
    try:
        return_order = ReturnOrder.query.get(return_id)
        if not return_order:
            return jsonify({'code': 404, 'message': '退货单不存在'}), 404

        # 找到原销售单
        if not return_order.source_order_id:
            return jsonify({'code': 400, 'message': '退货单未关联销售单'}), 400

        # 解除该销售单关联的所有资产的关联关系
        assets = Asset.query.filter_by(sales_order_id=return_order.source_order_id).all()
        unbind_count = 0
        for asset in assets:
            asset.sales_order_id = None
            asset.sales_order_no = None
            unbind_count += 1

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'成功解绑{unbind_count}个资产',
            'data': {'unbind_count': unbind_count}
        })
    except Exception as e:
        logger.error(f'退货解绑资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'解绑失败: {str(e)}'}), 500


@app.route('/api/sales/orders/<int:order_id>/assets', methods=['GET'])
@jwt_required()
def get_sales_order_assets(order_id):
    """获取销售单关联的资产"""
    try:
        order = SalesOrder.query.get(order_id)
        if not order:
            return jsonify({'code': 404, 'message': '销售单不存在'}), 404

        assets = Asset.query.filter_by(sales_order_id=order_id).order_by(Asset.created_at.desc()).all()

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': [to_dict(asset) for asset in assets]
        })
    except Exception as e:
        logger.error(f'获取销售单资产失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


# ============================================
# 导出功能 - 派单管理
# ============================================

@app.route('/api/dispatchorders/export', methods=['GET'])
@jwt_required()
def export_dispatch_orders():
    """导出派单记录"""
    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)

        query = DispatchRecord.query
        if keyword:
            query = query.join(WorkOrder, DispatchRecord.wo_id == WorkOrder.id).filter(
                db.or_(
                    WorkOrder.wo_no.contains(keyword),
                    WorkOrder.customer_name.contains(keyword),
                    DispatchRecord.staff_name.contains(keyword)
                )
            )
        if status is not None:
            query = query.filter(DispatchRecord.accept_status == status)

        records = query.order_by(DispatchRecord.created_at.desc()).all()

        # 接单状态映射
        accept_status_map = {0: '待接单', 1: '已接单', 2: '已拒单', 3: '已超时'}
        # 优先级映射
        priority_map = {0: '普通', 1: '紧急', 2: '特急'}

        data = []
        for r in records:
            wo = WorkOrder.query.get(r.wo_id)
            data.append({
                '派单号': f'DP{r.id:06d}',
                '工单号': wo.wo_no if wo else '',
                '客户名称': wo.customer_name if wo else '',
                '设备类型': wo.device_type if wo else '',
                '故障描述': (wo.fault_desc or '')[:200] if wo else '',
                '指派工程师': r.staff_name or '',
                '优先级': priority_map.get(wo.priority, '普通') if wo else '普通',
                '状态': accept_status_map.get(r.accept_status, '未知'),
                '创建时间': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['派单号', '工单号', '客户名称', '设备类型', '故障描述', '指派工程师', '优先级', '状态', '创建时间']
        return export_to_excel(data, columns, f'派单记录_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出派单记录失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


# ============================================
# 导出功能 - 报价管理
# ============================================

@app.route('/api/quotes/export', methods=['GET'])
@jwt_required()
def export_quotes():
    """导出报价单"""
    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)

        query = QuoteOrder.query
        if keyword:
            query = query.filter(
                db.or_(
                    QuoteOrder.quote_no.contains(keyword),
                    QuoteOrder.customer_name.contains(keyword),
                    QuoteOrder.customer_phone.contains(keyword)
                )
            )
        if status is not None:
            query = query.filter(QuoteOrder.status == status)

        orders = query.order_by(QuoteOrder.created_at.desc()).all()

        data = []
        for o in orders:
            # 获取创建者用户名
            creator = ''
            if o.created_by:
                user = User.query.get(o.created_by)
                if user:
                    creator = user.username or ''

            # 获取关联工单号或接件单号
            related_no = ''
            if o.related_type and o.related_id:
                if o.related_type == 'work_order':
                    wo = WorkOrder.query.get(o.related_id)
                    related_no = wo.wo_no if wo else ''
                elif o.related_type == 'receive_order':
                    ro = ReceiveOrder.query.get(o.related_id)
                    related_no = ro.receive_no if ro else ''

            data.append({
                '报价单号': o.quote_no or '',
                '工单号/接件单号': related_no,
                '客户名称': o.customer_name or '',
                '故障描述': '',
                '人工费': float(o.total_amount or 0) if o.total_amount else 0,
                '材料费': 0,
                '其他费': 0,
                '总计': float(o.total_amount) if o.total_amount else 0.00,
                '报价人': creator,
                '状态': QUOTE_STATUS_MAP.get(o.status, '未知'),
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['报价单号', '工单号/接件单号', '客户名称', '故障描述', '人工费', '材料费', '其他费', '总计', '报价人', '状态', '创建时间']
        return export_to_excel(data, columns, f'报价单_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出报价单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


# ============================================
# 导出功能 - 入库管理
# ============================================

@app.route('/api/inbound/orders/export', methods=['GET'])
@jwt_required()
def export_inbound_orders():
    """导出入库单"""
    try:
        keyword = request.args.get('keyword', '')
        in_type = request.args.get('in_type', type=int)
        status = request.args.get('status', type=int)

        query = InventoryIn.query
        if keyword:
            query = query.filter(
                db.or_(
                    InventoryIn.in_no.contains(keyword),
                    InventoryIn.supplier_name.contains(keyword),
                    InventoryIn.related_order_no.contains(keyword)
                )
            )
        if in_type is not None:
            query = query.filter(InventoryIn.in_type == in_type)
        if status is not None:
            query = query.filter(InventoryIn.status == status)

        orders = query.order_by(InventoryIn.created_at.desc()).all()

        # 入库类型映射
        in_type_map = {1: '采购入库', 2: '退货入库', 3: '调拨入库', 4: '组装入库', 5: '其他入库'}
        # 状态映射
        status_map = {0: '待审核', 1: '已审核', 2: '已入库'}

        data = []
        for o in orders:
            # 获取入库明细中的商品信息（取第一个商品）
            first_item = InventoryInItem.query.filter_by(in_id=o.id).first()
            product_name = first_item.product_name if first_item else ''
            specification = first_item.specification if first_item else ''
            quantity = first_item.quantity if first_item else ''
            unit = first_item.unit_name if first_item else ''

            # 获取经办人
            operator = ''
            if o.created_by:
                user = User.query.get(o.created_by)
                if user:
                    operator = user.username or ''

            data.append({
                '入库单号': o.in_no or '',
                '商品名称': product_name,
                '规格': specification,
                '数量': quantity,
                '单位': unit,
                '仓库': o.warehouse_name or '',
                '供应商': o.supplier_name or '',
                '入库类型': in_type_map.get(o.in_type, '未知'),
                '经办人': operator,
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['入库单号', '商品名称', '规格', '数量', '单位', '仓库', '供应商', '入库类型', '经办人', '创建时间']
        return export_to_excel(data, columns, f'入库单_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出入库单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


# ============================================
# 导出功能 - 出库管理
# ============================================

@app.route('/api/outbound/orders/export', methods=['GET'])
@jwt_required()
def export_outbound_orders():
    """导出出库单"""
    try:
        keyword = request.args.get('keyword', '')
        out_type = request.args.get('out_type', type=int)
        status = request.args.get('status', type=int)

        query = InventoryOut.query
        if keyword:
            query = query.filter(
                db.or_(
                    InventoryOut.out_no.contains(keyword),
                    InventoryOut.customer_name.contains(keyword),
                    InventoryOut.related_order_no.contains(keyword)
                )
            )
        if out_type is not None:
            query = query.filter(InventoryOut.out_type == out_type)
        if status is not None:
            query = query.filter(InventoryOut.status == status)

        orders = query.order_by(InventoryOut.created_at.desc()).all()

        # 出库类型映射
        out_type_map = {1: '销售出库', 2: '维修领料', 3: '调拨出库', 4: '拆卸出库', 5: '其他出库'}
        # 状态映射
        status_map = {0: '待审核', 1: '已审核', 2: '已出库'}

        data = []
        for o in orders:
            # 获取出库明细中的商品信息（取第一个商品）
            first_item = InventoryOutItem.query.filter_by(out_id=o.id).first()
            product_name = first_item.product_name if first_item else ''
            specification = first_item.specification if first_item else ''
            quantity = first_item.quantity if first_item else ''
            unit = first_item.unit_name if first_item else ''

            # 获取经办人
            operator = ''
            if o.created_by:
                user = User.query.get(o.created_by)
                if user:
                    operator = user.username or ''

            data.append({
                '出库单号': o.out_no or '',
                '商品名称': product_name,
                '规格': specification,
                '数量': quantity,
                '单位': unit,
                '仓库': o.warehouse_name or '',
                '关联单号': o.related_order_no or '',
                '出库类型': out_type_map.get(o.out_type, '未知'),
                '经办人': operator,
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['出库单号', '商品名称', '规格', '数量', '单位', '仓库', '关联单号', '出库类型', '经办人', '创建时间']
        return export_to_excel(data, columns, f'出库单_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出出库单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


# ============================================
# 导出功能 - 操作日志
# ============================================

@app.route('/api/operation-logs/export', methods=['GET'])
@jwt_required()
def export_operation_logs():
    """导出操作日志"""
    try:
        keyword = request.args.get('keyword', '')
        user_id = request.args.get('user_id', type=int)
        module = request.args.get('module', '')
        action = request.args.get('action', '')
        start_time = request.args.get('start_time', '')
        end_time = request.args.get('end_time', '')

        query = OperationLog.query
        if keyword:
            query = query.filter(
                db.or_(
                    OperationLog.user_name.contains(keyword),
                    OperationLog.content.contains(keyword),
                    OperationLog.ip.contains(keyword)
                )
            )
        if user_id:
            query = query.filter(OperationLog.user_id == user_id)
        if module:
            query = query.filter(OperationLog.module == module)
        if action:
            query = query.filter(OperationLog.action == action)
        if start_time:
            try:
                query = query.filter(OperationLog.created_at >= datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S'))
            except:
                pass
        if end_time:
            try:
                query = query.filter(OperationLog.created_at <= datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S'))
            except:
                pass

        logs = query.order_by(OperationLog.created_at.desc()).all()

        # 模块映射
        module_map = {
            'login': '登录', 'user': '用户管理', 'product': '商品管理',
            'customer': '客户管理', 'supplier': '供应商管理', 'sales': '销售管理',
            'purchase': '采购管理', 'work_order': '工单管理', 'inventory': '库存管理',
            'finance': '财务管理', 'settings': '系统设置'
        }
        # 操作类型映射
        action_map = {
            'create': '新增', 'update': '修改', 'delete': '删除',
            'query': '查询', 'export': '导出', 'login': '登录', 'logout': '登出'
        }

        data = []
        for log in logs:
            data.append({
                '操作用户': log.user_name or '',
                '操作类型': action_map.get(log.action, log.action or ''),
                '操作模块': module_map.get(log.module, log.module or ''),
                '操作描述': (log.content or '')[:500],
                'IP地址': log.ip or '',
                '创建时间': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['操作用户', '操作类型', '操作模块', '操作描述', 'IP地址', '创建时间']
        return export_to_excel(data, columns, f'操作日志_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出操作日志失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


if __name__ == '__main__':
    # 启动应用
    app.run(host='0.0.0.0', port=5000, debug=True)
