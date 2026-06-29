"""Excel 导入导出工具。"""
import io
from openpyxl import Workbook, load_workbook


def export_to_excel(rows: list, headers: dict, sheet_name: str = 'Sheet1') -> bytes:
    """导出 Excel。

    Args:
        rows: 数据列表，每项是 dict
        headers: {列名: 表头文本}，决定顺序和标题
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers.values()))
    for row in rows:
        ws.append([row.get(col, '') for col in headers.keys()])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_excel_data(file_storage) -> list:
    """读取 Excel 上传文件，返回 [{列名: 值}, ...]"""
    wb = load_workbook(filename=io.BytesIO(file_storage.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f'col{i}' for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(c is not None for c in row)]