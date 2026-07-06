"""库存管理蓝图（最大域之一）。

迁移自 source-code/app.py 中以下区段的原始路由代码：
- /api/inventory/stock  列表查询                          3539 行附近
- /api/inventory/in, /api/inventory/out  出入库           3618 / 3818 行附近
- /api/inventory/in/<id>/audit, /out/<id>/audit  审核     3726 / 3910 行附近
- /api/inventory/check  盘点                              3995 行附近
- /api/inventory/check/<id>/items  盘点明细               8043-8299 行附近
- /api/inventory/check/<id>/diff-report, /export          4099 / 4171 行附近
- /api/warehouses  仓库 CRUD                              9745-9860 行附近
- /api/inventory/logs  日志                               9875-9930 行附近
- /api/transfer/orders  调拨 CRUD                        10006-10181 行附近
- /api/cost-adjusts  成本调整 CRUD                       10407-10600 行附近
- /api/inventory/stock/export  库存导出                  10683 行附近
- /api/inventory/logs/export  日志导出                   ~9927 行附近
- /api/transfer/orders/export  调拨导出                  10346 行附近
- /api/cost-adjusts/export  成本调价导出                 ~10616 行附近
- /api/inbound/orders/export  入库单导出                 18377 行附近
- /api/outbound/orders/export  出库单导出                 18447 行附近

按子资源分 7 个 section：
  1. 库存查询  (stock + export)
  2. 入库单  (in / in/<id> / in/<id>/audit + inbound/orders/export)
  3. 出库单  (out / out/<id>/audit + outbound/orders/export)
  4. 盘点  (check + items + audit + complete + diff-report + export)
  5. 仓库  (warehouses CRUD + status)
  6. 库存变动日志  (logs + export)
  7. 调拨单  (transfer/orders CRUD + export)
  8. 成本调价  (cost-adjusts CRUD + export)

跨 blueprint 依赖：
- ProductInfo (models.product)
- User (models.system.SysUser) — 仅在 inbound/outbound export 中查操作人
- FinanceRecord (models.finance) — transfer / cost-adjust 审核时写入财务流水
- 全部以函数内 import 兼容迁移期。
"""
import io
import logging
from datetime import datetime

import openpyxl
from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity, jwt_required

from extensions import db
from app.security import permission
from app.utils import export_to_excel, generate_code, to_dict

logger = logging.getLogger(__name__)

bp = Blueprint('inventory', __name__)


# ============================================
# 工具函数
# ============================================

def _get_current_user_name():
    """获取当前登录用户姓名。迁移期兼容（app.py 顶层也定义同名函数）。"""
    from models.system import SysUser
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    if user:
        return user.real_name or user.username
    return ''


# ============================================
# 1. 库存查询
# ============================================

@bp.route('/api/inventory/stock', methods=['GET'])
@permission('inventory:view', 'inventory-in:view', 'inventory-out:view')
def get_inventory_stock():
    """获取库存列表"""
    from models.inventory.stock import InventoryStock
    from models.product.info import ProductInfo

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    warehouse_id = request.args.get('warehouse_id', type=int)
    product_id = request.args.get('product_id', type=int)

    query = InventoryStock.query

    if product_id:
        query = query.filter_by(product_id=product_id)

    if keyword:
        query = query.filter(
            db.or_(
                InventoryStock.product_name.contains(keyword),
                InventoryStock.product_code.contains(keyword)
            )
        )

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    total = query.count()
    stocks = query.order_by(InventoryStock.updated_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    stock_list = []
    for s in stocks.items:
        product = ProductInfo.query.get(s.product_id)
        purchase_price = float(product.purchase_price) if product and product.purchase_price else 0
        safe_stock = product.min_stock if product else 0
        qty = float(s.quantity) if s.quantity else 0

        if safe_stock and safe_stock > 0:
            if qty <= 0:
                stock_status = '缺货'
            elif qty < safe_stock:
                stock_status = '不足'
            elif qty > safe_stock * 3:
                stock_status = '过剩'
            else:
                stock_status = '正常'
        else:
            stock_status = '正常'

        stock_list.append({
            'id': s.id,
            'product_id': s.product_id,
            'product_code': s.product_code,
            'product_name': s.product_name,
            'warehouse_id': s.warehouse_id,
            'warehouse_name': s.warehouse_name,
            'quantity': qty,
            'frozen_quantity': float(s.frozen_quantity) if s.frozen_quantity else 0,
            'available_quantity': float(s.available_quantity) if s.available_quantity else 0,
            'cost_price': float(s.cost_price) if s.cost_price else 0,
            'batch_no': s.batch_no,
            'serial_no': s.serial_no,
            'purchase_price': purchase_price,
            'safe_stock': safe_stock,
            'stock_status': stock_status
        })

    return jsonify({
        'code': 200,
        'data': {
            'list': stock_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/inventory/stock/export', methods=['GET'])
@permission('inventory:view')
def export_inventory_stock():
    """导出库存查询结果"""
    from models.inventory.stock import InventoryStock
    from models.product.info import ProductInfo

    keyword = request.args.get('keyword', '')
    warehouse_id = request.args.get('warehouse_id', type=int)

    query = InventoryStock.query

    if keyword:
        query = query.filter(
            db.or_(
                InventoryStock.product_name.contains(keyword),
                InventoryStock.product_code.contains(keyword)
            )
        )

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    stocks = query.order_by(InventoryStock.updated_at.desc()).all()

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '库存查询'

    headers = ['商品编码', '商品名称', '仓库', '库存数量', '冻结数量', '可用数量', '成本价', '库存金额', '采购价', '安全库存', '库存状态']
    ws.append(headers)

    for s in stocks:
        product = ProductInfo.query.get(s.product_id)
        purchase_price = float(product.purchase_price) if product and product.purchase_price else 0
        safe_stock = product.min_stock if product else 0
        qty = float(s.quantity or 0)

        if safe_stock and safe_stock > 0:
            if qty <= 0:
                stock_status = '缺货'
            elif qty < safe_stock:
                stock_status = '不足'
            elif qty > safe_stock * 3:
                stock_status = '过剩'
            else:
                stock_status = '正常'
        else:
            stock_status = '正常'

        ws.append([
            s.product_code or '',
            s.product_name or '',
            s.warehouse_name or '',
            qty,
            float(s.frozen_quantity or 0),
            float(s.available_quantity or 0),
            float(s.cost_price or 0),
            round(qty * float(s.cost_price or 0), 2),
            purchase_price,
            safe_stock,
            stock_status
        ])

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'库存查询_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ============================================
# 2. 入库单
# ============================================

@bp.route('/api/inventory/in', methods=['GET'])
@permission('inventory-in:view')
def get_inventory_in_list():
    """获取入库单列表"""
    from models.inventory.flow import InventoryIn

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    in_type = request.args.get('in_type', type=int)
    status = request.args.get('status', type=int)

    query = InventoryIn.query

    if keyword:
        query = query.filter(InventoryIn.in_no.contains(keyword))

    if in_type is not None:
        query = query.filter_by(in_type=in_type)

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    orders = query.order_by(InventoryIn.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/inventory/in/<int:id>', methods=['GET'])
@permission('inventory-in:view')
def get_inventory_in(id):
    """获取入库单详情"""
    from models.inventory.flow import InventoryIn, InventoryInItem

    order = InventoryIn.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '入库单不存在'}), 404

    items = InventoryInItem.query.filter_by(in_id=id).all()

    result = to_dict(order)
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@bp.route('/api/inventory/in', methods=['POST'])
@permission('inventory-in:add')
def create_inventory_in():
    """创建入库单"""
    from models.inventory.flow import InventoryIn, InventoryInItem

    data = request.get_json()
    user_id = get_jwt_identity()

    last_order = InventoryIn.query.order_by(InventoryIn.id.desc()).first()
    in_no = generate_code('IN', last_order.id if last_order else 0)

    order = InventoryIn(
        in_no=in_no,
        in_type=data.get('in_type', 1),
        supplier_id=data.get('supplier_id'),
        supplier_name=data.get('supplier_name'),
        warehouse_id=data.get('warehouse_id', 1),
        warehouse_name=data.get('warehouse_name', '主仓库'),
        remark=data.get('remark'),
        related_order_id=data.get('related_order_id'),
        related_order_no=data.get('related_order_no'),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    items = data.get('items', [])
    total_quantity = 0
    total_amount = 0

    for item_data in items:
        item = InventoryInItem(
            in_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit_name=item_data.get('unit_name'),
            quantity=item_data.get('quantity', 0),
            unit_price=item_data.get('unit_price', 0),
            total_price=float(item_data.get('quantity', 0)) * float(item_data.get('unit_price', 0)),
            cost_price=item_data.get('cost_price', 0),
            batch_no=item_data.get('batch_no'),
            serial_no=item_data.get('serial_no'),
            remark=item_data.get('remark')
        )
        db.session.add(item)
        total_quantity += float(item_data.get('quantity', 0))
        total_amount += float(item.quantity) * float(item.unit_price or 0)

    order.total_quantity = total_quantity
    order.total_amount = total_amount
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'in_no': in_no}})


@bp.route('/api/inventory/in/<int:id>/audit', methods=['POST'])
@permission('inventory-in:edit')
def audit_inventory_in(id):
    """审核入库单"""
    from models.inventory.flow import InventoryIn, InventoryInItem
    from models.inventory.stock import InventoryLog, InventoryStock
    from models.product.info import ProductInfo

    order = InventoryIn.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '入库单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '入库单状态不正确'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    order.status = 2
    order.auditor_id = user_id
    order.auditor_name = data.get('auditor_name', '')
    order.audit_time = datetime.now()

    items = InventoryInItem.query.filter_by(in_id=id).all()
    user_name = _get_current_user_name()

    in_type_map = {1: '采购入库', 2: '退货入库', 3: '调拨入库', 4: '组装入库', 5: '其他入库'}
    order_type_text = in_type_map.get(order.in_type, '入库')

    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.warehouse_id
        ).first()

        before_qty = 0
        if stock:
            before_qty = float(stock.quantity or 0)
            stock.quantity = before_qty + float(item.quantity or 0)
            stock.available_quantity = float(stock.available_quantity or 0) + float(item.quantity or 0)
            if item.cost_price:
                stock.cost_price = item.cost_price
        else:
            stock = InventoryStock(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=order.warehouse_id,
                warehouse_name=order.warehouse_name,
                quantity=item.quantity,
                frozen_quantity=0,
                available_quantity=item.quantity,
                cost_price=item.cost_price,
                batch_no=item.batch_no,
                serial_no=item.serial_no
            )
            db.session.add(stock)

        after_qty = float(stock.quantity or 0)
        cost = float(item.cost_price or stock.cost_price or 0)
        qty = float(item.quantity or 0)

        log = InventoryLog(
            product_id=item.product_id,
            product_code=item.product_code,
            product_name=item.product_name,
            warehouse_id=order.warehouse_id,
            warehouse_name=order.warehouse_name,
            change_type='in',
            order_type=order_type_text,
            order_id=order.id,
            order_no=order.in_no,
            quantity_change=qty,
            before_quantity=before_qty,
            after_quantity=after_qty,
            cost_price=cost,
            amount=round(qty * cost, 2),
            related_party=order.supplier_name or '',
            operator_id=user_id,
            operator_name=user_name,
            remark=order.remark or ''
        )
        db.session.add(log)

        product = ProductInfo.query.get(item.product_id)
        if product:
            product.current_stock = (product.current_stock or 0) + float(item.quantity or 0)

    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，已入库'})


@bp.route('/api/inbound/orders/export', methods=['GET'])
@permission('inventory-in:view')
def export_inbound_orders():
    """导出入库单"""
    from models.inventory.flow import InventoryIn, InventoryInItem
    from models.system import SysUser

    try:
        keyword = request.args.get('keyword', '')
        in_type = request.args.get('in_type', type=int)
        status = request.args.get('status', type=int)

        query = InventoryIn.query
        if keyword:
            query = query.filter(
                db.or_(
                    InventoryIn.in_no.contains(keyword),
                    InventoryIn.supplier_name.contains(keyword),
                    InventoryIn.related_order_no.contains(keyword)
                )
            )
        if in_type is not None:
            query = query.filter(InventoryIn.in_type == in_type)
        if status is not None:
            query = query.filter(InventoryIn.status == status)

        orders = query.order_by(InventoryIn.created_at.desc()).all()

        in_type_map = {1: '采购入库', 2: '退货入库', 3: '调拨入库', 4: '组装入库', 5: '其他入库'}

        data = []
        for o in orders:
            first_item = InventoryInItem.query.filter_by(in_id=o.id).first()
            product_name = first_item.product_name if first_item else ''
            specification = first_item.specification if first_item else ''
            quantity = first_item.quantity if first_item else ''
            unit = first_item.unit_name if first_item else ''

            operator = ''
            if o.created_by:
                user = SysUser.query.get(o.created_by)
                if user:
                    operator = user.username or ''

            data.append({
                '入库单号': o.in_no or '',
                '商品名称': product_name,
                '规格': specification,
                '数量': quantity,
                '单位': unit,
                '仓库': o.warehouse_name or '',
                '供应商': o.supplier_name or '',
                '入库类型': in_type_map.get(o.in_type, '未知'),
                '经办人': operator,
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['入库单号', '商品名称', '规格', '数量', '单位', '仓库', '供应商', '入库类型', '经办人', '创建时间']
        return export_to_excel(data, columns, f'入库单_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出入库单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


# ============================================
# 3. 出库单
# ============================================

@bp.route('/api/inventory/out', methods=['GET'])
@permission('inventory-out:view')
def get_inventory_out_list():
    """获取出库单列表"""
    from models.inventory.flow import InventoryOut

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    out_type = request.args.get('out_type', type=int)
    status = request.args.get('status', type=int)

    query = InventoryOut.query

    if keyword:
        query = query.filter(InventoryOut.out_no.contains(keyword))

    if out_type is not None:
        query = query.filter_by(out_type=out_type)

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    orders = query.order_by(InventoryOut.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/inventory/out', methods=['POST'])
@permission('inventory-out:add')
def create_inventory_out():
    """创建出库单"""
    from models.inventory.flow import InventoryOut, InventoryOutItem

    data = request.get_json()
    user_id = get_jwt_identity()

    last_order = InventoryOut.query.order_by(InventoryOut.id.desc()).first()
    out_no = generate_code('OUT', last_order.id if last_order else 0)

    order = InventoryOut(
        out_no=out_no,
        out_type=data.get('out_type', 1),
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        warehouse_id=data.get('warehouse_id', 1),
        warehouse_name=data.get('warehouse_name', '主仓库'),
        remark=data.get('remark'),
        related_order_id=data.get('related_order_id'),
        related_order_no=data.get('related_order_no'),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    items = data.get('items', [])
    total_quantity = 0
    total_amount = 0

    for item_data in items:
        item = InventoryOutItem(
            out_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit_name=item_data.get('unit_name'),
            quantity=item_data.get('quantity', 0),
            unit_price=item_data.get('unit_price', 0),
            total_price=float(item_data.get('quantity', 0)) * float(item_data.get('unit_price', 0)),
            cost_price=item_data.get('cost_price', 0),
            batch_no=item_data.get('batch_no'),
            serial_no=item_data.get('serial_no'),
            remark=item_data.get('remark')
        )
        db.session.add(item)
        total_quantity += float(item_data.get('quantity', 0))
        total_amount += float(item.quantity) * float(item.unit_price or 0)

    order.total_quantity = total_quantity
    order.total_amount = total_amount
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'out_no': out_no}})


@bp.route('/api/inventory/out/<int:id>/audit', methods=['POST'])
@permission('inventory-out:edit')
def audit_inventory_out(id):
    """审核出库单"""
    from models.inventory.flow import InventoryOut, InventoryOutItem
    from models.inventory.stock import InventoryLog, InventoryStock
    from models.product.info import ProductInfo

    order = InventoryOut.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '出库单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '出库单状态不正确'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    items = InventoryOutItem.query.filter_by(out_id=id).all()
    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.warehouse_id
        ).first()

        if not stock or float(stock.available_quantity or 0) < float(item.quantity or 0):
            return jsonify({'code': 400, 'message': f'商品【{item.product_name}】库存不足'}), 400

    order.status = 2
    order.auditor_id = user_id
    order.auditor_name = data.get('auditor_name', '')
    order.audit_time = datetime.now()

    out_type_map = {1: '销售出库', 2: '维修领料', 3: '调拨出库', 4: '拆卸出库', 5: '其他出库'}
    order_type_text = out_type_map.get(order.out_type, '出库')
    user_name = _get_current_user_name()

    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.warehouse_id
        ).first()

        before_qty = 0
        if stock:
            before_qty = float(stock.quantity or 0)
            stock.quantity = before_qty - float(item.quantity or 0)
            stock.available_quantity = float(stock.available_quantity or 0) - float(item.quantity or 0)

        after_qty = float(stock.quantity or 0) if stock else 0
        cost = float(item.cost_price or (stock.cost_price if stock else 0) or 0)
        qty = float(item.quantity or 0)

        log = InventoryLog(
            product_id=item.product_id,
            product_code=item.product_code,
            product_name=item.product_name,
            warehouse_id=order.warehouse_id,
            warehouse_name=order.warehouse_name,
            change_type='out',
            order_type=order_type_text,
            order_id=order.id,
            order_no=order.out_no,
            quantity_change=-qty,
            before_quantity=before_qty,
            after_quantity=after_qty,
            cost_price=cost,
            amount=round(qty * cost, 2),
            related_party=order.customer_name or '',
            operator_id=user_id,
            operator_name=user_name,
            remark=order.remark or ''
        )
        db.session.add(log)

        product = ProductInfo.query.get(item.product_id)
        if product:
            product.current_stock = (product.current_stock or 0) - float(item.quantity or 0)

    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，已出库'})


@bp.route('/api/outbound/orders/export', methods=['GET'])
@permission('inventory-out:view')
def export_outbound_orders():
    """导出出库单"""
    from models.inventory.flow import InventoryOut, InventoryOutItem
    from models.system import SysUser

    try:
        keyword = request.args.get('keyword', '')
        out_type = request.args.get('out_type', type=int)
        status = request.args.get('status', type=int)

        query = InventoryOut.query
        if keyword:
            query = query.filter(
                db.or_(
                    InventoryOut.out_no.contains(keyword),
                    InventoryOut.customer_name.contains(keyword),
                    InventoryOut.related_order_no.contains(keyword)
                )
            )
        if out_type is not None:
            query = query.filter(InventoryOut.out_type == out_type)
        if status is not None:
            query = query.filter(InventoryOut.status == status)

        orders = query.order_by(InventoryOut.created_at.desc()).all()

        out_type_map = {1: '销售出库', 2: '维修领料', 3: '调拨出库', 4: '拆卸出库', 5: '其他出库'}

        data = []
        for o in orders:
            first_item = InventoryOutItem.query.filter_by(out_id=o.id).first()
            product_name = first_item.product_name if first_item else ''
            specification = first_item.specification if first_item else ''
            quantity = first_item.quantity if first_item else ''
            unit = first_item.unit_name if first_item else ''

            operator = ''
            if o.created_by:
                user = SysUser.query.get(o.created_by)
                if user:
                    operator = user.username or ''

            data.append({
                '出库单号': o.out_no or '',
                '商品名称': product_name,
                '规格': specification,
                '数量': quantity,
                '单位': unit,
                '仓库': o.warehouse_name or '',
                '关联单号': o.related_order_no or '',
                '出库类型': out_type_map.get(o.out_type, '未知'),
                '经办人': operator,
                '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['出库单号', '商品名称', '规格', '数量', '单位', '仓库', '关联单号', '出库类型', '经办人', '创建时间']
        return export_to_excel(data, columns, f'出库单_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        logger.error(f'导出出库单失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


# ============================================
# 4. 库存盘点
# ============================================

@bp.route('/api/inventory/check', methods=['GET'])
@jwt_required()
def get_inventory_check_list():
    """获取盘点单列表"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')

    query = InventoryCheck.query

    if status is not None and status != '':
        query = query.filter_by(status=status)

    if keyword:
        query = query.filter(
            db.or_(
                InventoryCheck.check_no.like(f'%{keyword}%'),
                InventoryCheck.warehouse_name.like(f'%{keyword}%'),
                InventoryCheck.shelf_name.like(f'%{keyword}%')
            )
        )

    total = query.count()
    orders = query.order_by(InventoryCheck.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    result_list = []
    for o in orders.items:
        item = to_dict(o)
        check_items = InventoryCheckItem.query.filter_by(check_id=o.id).all()
        profit_count = len([i for i in check_items if i.diff_quantity and float(i.diff_quantity) > 0])
        loss_count = len([i for i in check_items if i.diff_quantity and float(i.diff_quantity) < 0])
        item['total_items'] = len(check_items)
        item['profit_items'] = profit_count
        item['loss_items'] = loss_count
        result_list.append(item)

    return jsonify({
        'code': 200,
        'data': {
            'list': result_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/inventory/check', methods=['POST'])
@jwt_required()
def create_inventory_check():
    """创建盘点单（支持按仓库/货架筛选）"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem
    from models.inventory.stock import InventoryStock

    data = request.get_json()
    user_id = get_jwt_identity()

    last_order = InventoryCheck.query.order_by(InventoryCheck.id.desc()).first()
    check_no = generate_code('CHK', last_order.id if last_order else 0)

    warehouse_id = data.get('warehouse_id', 1)
    warehouse_name = data.get('warehouse_name', '主仓库')
    shelf_id = data.get('shelf_id')
    shelf_name = data.get('shelf_name', '')

    order = InventoryCheck(
        check_no=check_no,
        warehouse_id=warehouse_id,
        warehouse_name=warehouse_name,
        shelf_id=shelf_id,
        shelf_name=shelf_name,
        check_date=datetime.now().date(),
        remark=data.get('remark'),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    stocks_query = InventoryStock.query.filter_by(warehouse_id=warehouse_id)
    if shelf_id:
        stocks_query = stocks_query.filter_by(shelf_id=shelf_id)
    stocks = stocks_query.all()

    for stock in stocks:
        item = InventoryCheckItem(
            check_id=order.id,
            product_id=stock.product_id,
            product_code=stock.product_code,
            product_name=stock.product_name,
            system_quantity=stock.quantity,
            actual_quantity=stock.quantity,
            diff_quantity=0,
            cost_price=stock.cost_price,
            diff_amount=0
        )
        db.session.add(item)

    order.total_quantity = len(stocks)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'check_no': check_no}})


@bp.route('/api/inventory/check/<int:id>', methods=['GET'])
@jwt_required()
def get_inventory_check(id):
    """获取盘点单详情"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    items = InventoryCheckItem.query.filter_by(check_id=id).all()

    result = to_dict(order)
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@bp.route('/api/inventory/check/<int:id>', methods=['PUT'])
@jwt_required()
def update_inventory_check(id):
    """更新盘点单（录入实际数量）"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status != 0 and order.status != 1:
        return jsonify({'code': 400, 'message': '盘点单状态不允许修改'}), 400

    data = request.get_json()

    if 'remark' in data:
        order.remark = data['remark']

    if order.status == 0:
        order.status = 1

    if 'items' in data:
        total_diff = 0
        total_diff_amount = 0
        for item_data in data['items']:
            item_id = item_data.get('id')
            if item_id:
                item = InventoryCheckItem.query.get(item_id)
                if item and item.check_id == id:
                    actual_qty = float(item_data.get('actual_quantity', item.system_quantity or 0))
                    item.actual_quantity = actual_qty
                    item.diff_quantity = actual_qty - float(item.system_quantity or 0)
                    item.diff_amount = float(item.diff_quantity or 0) * float(item.cost_price or 0)
                    total_diff += abs(float(item.diff_quantity or 0))
                    total_diff_amount += abs(float(item.diff_amount or 0))

        order.diff_quantity = total_diff
        order.diff_amount = total_diff_amount

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@bp.route('/api/inventory/check/<int:id>/items', methods=['GET'])
@jwt_required()
def get_inventory_check_items(id):
    """获取盘点明细"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    items = InventoryCheckItem.query.filter_by(check_id=id).all()

    return jsonify({
        'code': 200,
        'data': [{
            'id': i.id,
            'product_id': i.product_id,
            'product_code': i.product_code,
            'product_name': i.product_name,
            'specification': i.specification,
            'unit_name': i.unit_name,
            'system_quantity': float(i.system_quantity) if i.system_quantity else 0,
            'actual_quantity': float(i.actual_quantity) if i.actual_quantity else 0,
            'diff_quantity': float(i.diff_quantity) if i.diff_quantity else 0,
            'cost_price': float(i.cost_price) if i.cost_price else 0,
            'diff_amount': float(i.diff_amount) if i.diff_amount else 0,
            'remark': i.remark
        } for i in items]
    })


@bp.route('/api/inventory/check/<int:id>/items', methods=['PUT'])
@jwt_required()
def update_inventory_check_items(id):
    """保存盘点明细（批量更新实际数量）"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '已完成的盘点单不能修改'}), 400

    data = request.get_json()
    items_data = data.get('items', [])

    total_diff = 0
    total_diff_amount = 0

    for item_data in items_data:
        item_id = item_data.get('id')
        if item_id:
            item = InventoryCheckItem.query.get(item_id)
            if item and item.check_id == id:
                actual_qty = float(item_data.get('actual_quantity', item.system_quantity or 0))
                item.actual_quantity = actual_qty
                item.diff_quantity = actual_qty - float(item.system_quantity or 0)
                item.diff_amount = float(item.diff_quantity or 0) * float(item.cost_price or 0)
                total_diff += abs(float(item.diff_quantity or 0))
                total_diff_amount += abs(float(item.diff_amount or 0))

    if order.status == 0:
        order.status = 1
    order.diff_quantity = total_diff
    order.diff_amount = total_diff_amount

    db.session.commit()
    return jsonify({'code': 200, 'message': '保存成功'})


@bp.route('/api/inventory/check/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_inventory_check(id):
    """审核盘点单（自动更新库存）"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem
    from models.inventory.stock import InventoryLog, InventoryStock
    from models.product.info import ProductInfo

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status != 1:
        return jsonify({'code': 400, 'message': '只有盘点中的盘点单可以审核'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    items = InventoryCheckItem.query.filter_by(check_id=id).all()
    for item in items:
        if item.diff_quantity and float(item.diff_quantity) != 0:
            stock = InventoryStock.query.filter_by(
                product_id=item.product_id,
                warehouse_id=order.warehouse_id
            ).first()

            if stock:
                diff = float(item.diff_quantity)
                before_qty = float(stock.quantity or 0)
                stock.quantity = before_qty + diff
                stock.available_quantity = float(stock.available_quantity or 0) + diff
                after_qty = stock.quantity

                log = InventoryLog(
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    warehouse_id=order.warehouse_id,
                    warehouse_name=order.warehouse_name,
                    change_type='adjust',
                    order_type='盘点调整',
                    order_id=order.id,
                    order_no=order.check_no,
                    quantity_change=diff,
                    before_quantity=before_qty,
                    after_quantity=after_qty,
                    cost_price=float(item.cost_price or 0),
                    amount=round(abs(diff) * float(item.cost_price or 0), 2),
                    related_party='',
                    operator_id=user_id,
                    operator_name=user_name,
                    remark=f'盘点调整：系统{before_qty}，实际{float(item.actual_quantity or 0)}，差异{diff}'
                )
                db.session.add(log)

                product = ProductInfo.query.get(item.product_id)
                if product:
                    product.current_stock = (product.current_stock or 0) + diff

    order.status = 2
    db.session.commit()

    return jsonify({'code': 200, 'message': '审核成功，库存已更新'})


@bp.route('/api/inventory/check/<int:id>/complete', methods=['POST'])
@jwt_required()
def complete_inventory_check(id):
    """完成盘点（审核并更新库存）"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem
    from models.inventory.stock import InventoryLog, InventoryStock
    from models.product.info import ProductInfo

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '盘点单已完成'}), 400

    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    items = InventoryCheckItem.query.filter_by(check_id=id).all()
    for item in items:
        if item.diff_quantity and float(item.diff_quantity) != 0:
            stock = InventoryStock.query.filter_by(
                product_id=item.product_id,
                warehouse_id=order.warehouse_id
            ).first()

            if stock:
                diff = float(item.diff_quantity)
                before_qty = float(stock.quantity or 0)
                stock.quantity = before_qty + diff
                stock.available_quantity = float(stock.available_quantity or 0) + diff
                after_qty = stock.quantity

                log = InventoryLog(
                    product_id=item.product_id,
                    product_code=item.product_code,
                    product_name=item.product_name,
                    warehouse_id=order.warehouse_id,
                    warehouse_name=order.warehouse_name,
                    change_type='adjust',
                    order_type='盘点调整',
                    order_id=order.id,
                    order_no=order.check_no,
                    quantity_change=diff,
                    before_quantity=before_qty,
                    after_quantity=after_qty,
                    cost_price=float(item.cost_price or 0),
                    amount=round(abs(diff) * float(item.cost_price or 0), 2),
                    related_party='',
                    operator_id=user_id,
                    operator_name=user_name,
                    remark=f'盘点调整：系统{before_qty}，实际{float(item.actual_quantity or 0)}，差异{diff}'
                )
                db.session.add(log)

                product = ProductInfo.query.get(item.product_id)
                if product:
                    product.current_stock = (product.current_stock or 0) + diff

    order.status = 2
    db.session.commit()

    return jsonify({'code': 200, 'message': '盘点完成，库存已更新'})


@bp.route('/api/inventory/check/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_inventory_check(id):
    """取消/删除盘点单"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    if order.status == 2:
        return jsonify({'code': 400, 'message': '已完成的盘点单不能取消'}), 400

    InventoryCheckItem.query.filter_by(check_id=id).delete()
    db.session.delete(order)
    db.session.commit()

    return jsonify({'code': 200, 'message': '取消成功'})


@bp.route('/api/inventory/check/<int:id>/diff-report', methods=['GET'])
@jwt_required()
def get_check_diff_report(id):
    """获取盘点差异报表"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    order = InventoryCheck.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    items = InventoryCheckItem.query.filter_by(check_id=id).all()

    profit_items = []
    loss_items = []
    normal_items = []

    total_profit_qty = 0
    total_profit_amount = 0
    total_loss_qty = 0
    total_loss_amount = 0

    for i in items:
        diff = float(i.diff_quantity or 0)
        diff_amt = float(i.diff_amount or 0)
        item_data = {
            'product_code': i.product_code,
            'product_name': i.product_name,
            'specification': i.specification,
            'unit_name': i.unit_name,
            'system_quantity': float(i.system_quantity or 0),
            'actual_quantity': float(i.actual_quantity or 0),
            'diff_quantity': diff,
            'cost_price': float(i.cost_price or 0),
            'diff_amount': diff_amt,
            'remark': i.remark
        }

        if diff > 0:
            profit_items.append(item_data)
            total_profit_qty += diff
            total_profit_amount += diff_amt
        elif diff < 0:
            loss_items.append(item_data)
            total_loss_qty += abs(diff)
            total_loss_amount += abs(diff_amt)
        else:
            normal_items.append(item_data)

    return jsonify({
        'code': 200,
        'data': {
            'order_info': {
                'check_no': order.check_no,
                'warehouse_name': order.warehouse_name,
                'shelf_name': order.shelf_name,
                'check_date': str(order.check_date) if order.check_date else '',
                'status': order.status
            },
            'summary': {
                'total_items': len(items),
                'profit_count': len(profit_items),
                'loss_count': len(loss_items),
                'normal_count': len(normal_items),
                'total_profit_qty': total_profit_qty,
                'total_profit_amount': total_profit_amount,
                'total_loss_qty': total_loss_qty,
                'total_loss_amount': total_loss_amount
            },
            'profit_items': profit_items,
            'loss_items': loss_items,
            'normal_items': normal_items
        }
    })


@bp.route('/api/inventory/check/diff-report/export', methods=['GET'])
@jwt_required()
def export_check_diff_report():
    """导出盘点差异报表为Excel"""
    from models.inventory.flow import InventoryCheck, InventoryCheckItem

    check_id = request.args.get('check_id', type=int)
    if not check_id:
        return jsonify({'code': 400, 'message': '缺少check_id参数'}), 400

    order = InventoryCheck.query.get(check_id)
    if not order:
        return jsonify({'code': 404, 'message': '盘点单不存在'}), 404

    items = InventoryCheckItem.query.filter_by(check_id=check_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '盘点差异报表'

    ws.merge_cells('A1:I1')
    ws['A1'] = f'盘点差异报表 - {order.check_no}'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

    ws.merge_cells('A2:I2')
    ws['A2'] = f'仓库: {order.warehouse_name}  货架: {order.shelf_name or "全部"}  日期: {order.check_date}'

    headers = ['商品编码', '商品名称', '规格', '单位', '账面数量', '实盘数量', '差异数', '成本价', '差异金额', '类型']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
        ws.cell(row=4, column=col).font = openpyxl.styles.Font(bold=True)

    for row_idx, item in enumerate(items, 5):
        diff = float(item.diff_quantity or 0)
        diff_type = '盘盈' if diff > 0 else ('盘亏' if diff < 0 else '正常')
        ws.cell(row=row_idx, column=1, value=item.product_code)
        ws.cell(row=row_idx, column=2, value=item.product_name)
        ws.cell(row=row_idx, column=3, value=item.specification or '')
        ws.cell(row=row_idx, column=4, value=item.unit_name or '')
        ws.cell(row=row_idx, column=5, value=float(item.system_quantity or 0))
        ws.cell(row=row_idx, column=6, value=float(item.actual_quantity or 0))
        ws.cell(row=row_idx, column=7, value=diff)
        ws.cell(row=row_idx, column=8, value=float(item.cost_price or 0))
        ws.cell(row=row_idx, column=9, value=float(item.diff_amount or 0))
        ws.cell(row=row_idx, column=10, value=diff_type)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'盘点差异报表_{order.check_no}.xlsx'
    )


# ============================================
# 5. 仓库
# ============================================

@bp.route('/api/warehouses', methods=['GET'])
@jwt_required()
def get_warehouses():
    """获取仓库列表"""
    from models.inventory.warehouse import Warehouse

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)

    query = Warehouse.query

    if keyword:
        query = query.filter(
            db.or_(
                Warehouse.warehouse_name.contains(keyword),
                Warehouse.warehouse_code.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    warehouses = query.order_by(Warehouse.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(w) for w in warehouses.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/warehouses', methods=['POST'])
@jwt_required()
def create_warehouse():
    """新增仓库"""
    from models.inventory.warehouse import Warehouse

    data = request.get_json()

    if not data.get('warehouse_name'):
        return jsonify({'code': 400, 'message': '仓库名称不能为空'}), 400

    if data.get('warehouse_code'):
        exists = Warehouse.query.filter_by(warehouse_code=data['warehouse_code']).first()
        if exists:
            return jsonify({'code': 400, 'message': '仓库编码已存在'}), 400

    warehouse = Warehouse(
        warehouse_code=data.get('warehouse_code'),
        warehouse_name=data['warehouse_name'],
        address=data.get('address', ''),
        contact_person=data.get('contact_person', ''),
        contact_phone=data.get('contact_phone', ''),
        remark=data.get('remark', ''),
        status=data.get('status', 1)
    )

    db.session.add(warehouse)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': warehouse.id}})


@bp.route('/api/warehouses/<int:id>', methods=['PUT'])
@jwt_required()
def update_warehouse(id):
    """编辑仓库"""
    from models.inventory.warehouse import Warehouse

    warehouse = Warehouse.query.get(id)
    if not warehouse:
        return jsonify({'code': 404, 'message': '仓库不存在'}), 404

    data = request.get_json()

    if data.get('warehouse_code') and data['warehouse_code'] != warehouse.warehouse_code:
        exists = Warehouse.query.filter_by(warehouse_code=data['warehouse_code']).first()
        if exists:
            return jsonify({'code': 400, 'message': '仓库编码已存在'}), 400

    for field in ['warehouse_code', 'warehouse_name', 'address', 'contact_person', 'contact_phone', 'remark', 'status']:
        if field in data:
            setattr(warehouse, field, data[field])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@bp.route('/api/warehouses/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_warehouse(id):
    """删除仓库（有库存关联的不允许删除）"""
    from models.inventory.warehouse import Warehouse
    from models.inventory.stock import InventoryStock

    warehouse = Warehouse.query.get(id)
    if not warehouse:
        return jsonify({'code': 404, 'message': '仓库不存在'}), 404

    stock_count = InventoryStock.query.filter_by(warehouse_id=id).count()
    if stock_count > 0:
        return jsonify({'code': 400, 'message': f'该仓库下有 {stock_count} 条库存记录，不允许删除'}), 400

    db.session.delete(warehouse)
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@bp.route('/api/warehouses/<int:id>/status', methods=['PUT'])
@jwt_required()
def toggle_warehouse_status(id):
    """启用/禁用仓库"""
    from models.inventory.warehouse import Warehouse

    warehouse = Warehouse.query.get(id)
    if not warehouse:
        return jsonify({'code': 404, 'message': '仓库不存在'}), 404

    data = request.get_json()
    warehouse.status = data.get('status', 0)
    db.session.commit()

    status_text = '启用' if warehouse.status == 1 else '禁用'
    return jsonify({'code': 200, 'message': f'{status_text}成功'})


# ============================================
# 6. 库存变动日志
# ============================================

@bp.route('/api/inventory/logs', methods=['GET'])
@jwt_required()
def get_inventory_logs():
    """库存变动明细列表"""
    from models.inventory.stock import InventoryLog

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    product_name = request.args.get('product_name', '')
    change_type = request.args.get('change_type', '')
    order_type = request.args.get('order_type', '')
    related_party = request.args.get('related_party', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = InventoryLog.query

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    if product_name:
        query = query.filter(InventoryLog.product_name.contains(product_name))

    if change_type:
        query = query.filter_by(change_type=change_type)

    if order_type:
        query = query.filter_by(order_type=order_type)

    if related_party:
        query = query.filter(InventoryLog.related_party.contains(related_party))

    if start_date:
        query = query.filter(InventoryLog.created_at >= start_date)

    if end_date:
        query = query.filter(InventoryLog.created_at <= end_date + ' 23:59:59')

    total = query.count()
    logs = query.order_by(InventoryLog.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(l) for l in logs.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/inventory/logs/export', methods=['GET'])
@jwt_required()
def export_inventory_logs():
    """导出库存变动明细"""
    from models.inventory.stock import InventoryLog

    warehouse_id = request.args.get('warehouse_id', type=int)
    product_name = request.args.get('product_name', '')
    change_type = request.args.get('change_type', '')
    order_type = request.args.get('order_type', '')
    related_party = request.args.get('related_party', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = InventoryLog.query

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    if product_name:
        query = query.filter(InventoryLog.product_name.contains(product_name))

    if change_type:
        query = query.filter_by(change_type=change_type)

    if order_type:
        query = query.filter_by(order_type=order_type)

    if related_party:
        query = query.filter(InventoryLog.related_party.contains(related_party))

    if start_date:
        query = query.filter(InventoryLog.created_at >= start_date)

    if end_date:
        query = query.filter(InventoryLog.created_at <= end_date + ' 23:59:59')

    logs = query.order_by(InventoryLog.created_at.desc()).all()

    change_type_map = {
        'in': '入库',
        'out': '出库',
        'adjust': '调整',
        'transfer': '调拨'
    }

    data = []
    for log in logs:
        data.append({
            '商品名称': log.product_name or '',
            '仓库': log.warehouse_name or '',
            '变动类型': change_type_map.get(log.change_type, log.change_type or ''),
            '变动数量': float(log.quantity_change) if log.quantity_change else 0,
            '变动前库存': float(log.before_quantity) if log.before_quantity else 0,
            '变动后库存': float(log.after_quantity) if log.after_quantity else 0,
            '关联单号': log.order_no or '',
            '操作人': log.operator_name or '',
            '创建时间': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '库存变动明细'
    headers = ['商品名称', '仓库', '变动类型', '变动数量', '变动前库存', '变动后库存', '关联单号', '操作人', '创建时间']
    ws.append(headers)
    for row in data:
        ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'库存变动明细_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ============================================
# 7. 调拨单
# ============================================

@bp.route('/api/transfer/orders', methods=['GET'])
@jwt_required()
def get_transfer_orders():
    """调拨单列表"""
    from models.inventory.flow import TransferOrder

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    transfer_type = request.args.get('transfer_type', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = TransferOrder.query

    if keyword:
        query = query.filter(TransferOrder.transfer_no.contains(keyword))

    if status is not None:
        query = query.filter_by(status=status)

    if transfer_type is not None:
        query = query.filter_by(transfer_type=transfer_type)

    if start_date:
        query = query.filter(TransferOrder.created_at >= start_date)

    if end_date:
        query = query.filter(TransferOrder.created_at <= end_date + ' 23:59:59')

    total = query.count()
    orders = query.order_by(TransferOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(o) for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/transfer/orders', methods=['POST'])
@jwt_required()
def create_transfer_order():
    """创建调拨单（同价/变价）"""
    from models.inventory.flow import TransferOrder, TransferOrderItem

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    date_str = datetime.now().strftime("%Y%m%d")
    last_order = TransferOrder.query.filter(
        TransferOrder.transfer_no.like(f'DB{date_str}%')
    ).order_by(TransferOrder.id.desc()).first()

    if last_order:
        last_seq = int(last_order.transfer_no[-4:]) if last_order.transfer_no else 0
        seq = str(last_seq + 1).zfill(4)
    else:
        seq = '0001'
    transfer_no = f'DB{date_str}{seq}'

    transfer_type = data.get('transfer_type', 1)

    order = TransferOrder(
        transfer_no=transfer_no,
        from_warehouse_id=data.get('from_warehouse_id'),
        from_warehouse_name=data.get('from_warehouse_name', ''),
        to_warehouse_id=data.get('to_warehouse_id'),
        to_warehouse_name=data.get('to_warehouse_name', ''),
        total_quantity=0,
        status=0,
        transfer_type=transfer_type,
        from_cost_price=data.get('from_cost_price', 0),
        to_cost_price=data.get('to_cost_price', 0),
        operator_id=user_id,
        operator_name=user_name,
        remark=data.get('remark', ''),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    items = data.get('items', [])
    total_quantity = 0

    for item_data in items:
        item = TransferOrderItem(
            transfer_id=order.id,
            product_id=item_data.get('product_id'),
            product_code=item_data.get('product_code'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification', ''),
            unit_name=item_data.get('unit_name', ''),
            quantity=item_data.get('quantity', 0),
            from_cost_price=item_data.get('from_cost_price', 0),
            to_cost_price=item_data.get('to_cost_price', 0),
            remark=item_data.get('remark', '')
        )
        db.session.add(item)
        total_quantity += float(item_data.get('quantity', 0))

    order.total_quantity = total_quantity
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'transfer_no': transfer_no}})


@bp.route('/api/transfer/orders/<int:id>', methods=['GET'])
@jwt_required()
def get_transfer_order(id):
    """调拨单详情"""
    from models.inventory.flow import TransferOrder, TransferOrderItem

    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    items = TransferOrderItem.query.filter_by(transfer_id=id).all()

    result = to_dict(order)
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@bp.route('/api/transfer/orders/<int:id>', methods=['PUT'])
@jwt_required()
def update_transfer_order(id):
    """编辑调拨单（待审核状态）"""
    from models.inventory.flow import TransferOrder, TransferOrderItem

    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调拨单可以编辑'}), 400

    data = request.get_json()

    for field in ['from_warehouse_id', 'from_warehouse_name', 'to_warehouse_id', 'to_warehouse_name',
                   'transfer_type', 'from_cost_price', 'to_cost_price', 'remark']:
        if field in data:
            setattr(order, field, data[field])

    if 'items' in data:
        TransferOrderItem.query.filter_by(transfer_id=id).delete()

        total_quantity = 0
        for item_data in data['items']:
            item = TransferOrderItem(
                transfer_id=id,
                product_id=item_data.get('product_id'),
                product_code=item_data.get('product_code'),
                product_name=item_data.get('product_name'),
                specification=item_data.get('specification', ''),
                unit_name=item_data.get('unit_name', ''),
                quantity=item_data.get('quantity', 0),
                from_cost_price=item_data.get('from_cost_price', 0),
                to_cost_price=item_data.get('to_cost_price', 0),
                remark=item_data.get('remark', '')
            )
            db.session.add(item)
            total_quantity += float(item_data.get('quantity', 0))

        order.total_quantity = total_quantity

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@bp.route('/api/transfer/orders/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_transfer_order(id):
    """审核调拨单"""
    from models.inventory.flow import TransferOrder, TransferOrderItem
    from models.inventory.stock import InventoryLog, InventoryStock
    from models.finance import FinanceRecord

    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调拨单可以审核'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    items = TransferOrderItem.query.filter_by(transfer_id=id).all()

    for item in items:
        stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.from_warehouse_id
        ).first()

        if not stock or float(stock.available_quantity or 0) < float(item.quantity or 0):
            return jsonify({'code': 400, 'message': f'商品【{item.product_name}】在调出仓库库存不足'}), 400

    order.status = 1
    order.auditor_id = user_id
    order.auditor_name = user_name
    order.audit_time = datetime.now()

    for item in items:
        qty = float(item.quantity or 0)
        from_cost = float(item.from_cost_price or 0)
        to_cost = float(item.to_cost_price or 0)

        from_stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.from_warehouse_id
        ).first()

        if from_stock:
            before_from_qty = float(from_stock.quantity or 0)
            from_stock.quantity = before_from_qty - qty
            from_stock.available_quantity = float(from_stock.available_quantity or 0) - qty

            log_out = InventoryLog(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=order.from_warehouse_id,
                warehouse_name=order.from_warehouse_name,
                change_type='transfer',
                order_type='调拨出库',
                order_id=order.id,
                order_no=order.transfer_no,
                quantity_change=-qty,
                before_quantity=before_from_qty,
                after_quantity=before_from_qty - qty,
                cost_price=from_cost,
                amount=qty * from_cost,
                related_party=order.to_warehouse_name or '',
                operator_id=user_id,
                operator_name=user_name,
                remark=f'调拨至{order.to_warehouse_name}'
            )
            db.session.add(log_out)

        to_stock = InventoryStock.query.filter_by(
            product_id=item.product_id,
            warehouse_id=order.to_warehouse_id
        ).first()

        if to_stock:
            before_to_qty = float(to_stock.quantity or 0)
            to_stock.quantity = before_to_qty + qty
            to_stock.available_quantity = float(to_stock.available_quantity or 0) + qty
            if order.transfer_type == 2 and to_cost > 0:
                to_stock.cost_price = to_cost
        else:
            before_to_qty = 0
            to_stock = InventoryStock(
                product_id=item.product_id,
                product_code=item.product_code,
                product_name=item.product_name,
                warehouse_id=order.to_warehouse_id,
                warehouse_name=order.to_warehouse_name,
                quantity=qty,
                frozen_quantity=0,
                available_quantity=qty,
                cost_price=to_cost if (order.transfer_type == 2 and to_cost > 0) else from_cost
            )
            db.session.add(to_stock)

        log_in = InventoryLog(
            product_id=item.product_id,
            product_code=item.product_code,
            product_name=item.product_name,
            warehouse_id=order.to_warehouse_id,
            warehouse_name=order.to_warehouse_name,
            change_type='transfer',
            order_type='调拨入库',
            order_id=order.id,
            order_no=order.transfer_no,
            quantity_change=qty,
            before_quantity=before_to_qty,
            after_quantity=before_to_qty + qty,
            cost_price=to_cost if (order.transfer_type == 2 and to_cost > 0) else from_cost,
            amount=qty * (to_cost if (order.transfer_type == 2 and to_cost > 0) else from_cost),
            related_party=order.from_warehouse_name or '',
            operator_id=user_id,
            operator_name=user_name,
            remark=f'从{order.from_warehouse_name}调入'
        )
        db.session.add(log_in)

        if order.transfer_type == 2 and from_cost != to_cost:
            diff_amount = qty * (to_cost - from_cost)
            finance = FinanceRecord(
                account_id=1,
                account_name='默认账户',
                record_type=2 if diff_amount > 0 else 1,
                amount=abs(diff_amount),
                balance_before=0,
                balance_after=0,
                related_type='transfer',
                related_id=order.id,
                related_no=order.transfer_no,
                remark=f'变价调拨成本差异：{item.product_name} x {qty}，原成本{from_cost}，新成本{to_cost}',
                created_by=user_id
            )
            db.session.add(finance)

    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，调拨已完成'})


@bp.route('/api/transfer/orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_transfer_order(id):
    """删除调拨单（待审核状态）"""
    from models.inventory.flow import TransferOrder, TransferOrderItem

    order = TransferOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '调拨单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调拨单可以删除'}), 400

    TransferOrderItem.query.filter_by(transfer_id=id).delete()
    db.session.delete(order)
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


@bp.route('/api/transfer/orders/export', methods=['GET'])
@jwt_required()
def export_transfer_orders():
    """导出调拨单"""
    from models.inventory.flow import TransferOrder

    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = TransferOrder.query

    if keyword:
        query = query.filter(TransferOrder.transfer_no.contains(keyword))

    if status is not None:
        query = query.filter_by(status=status)

    if start_date:
        query = query.filter(TransferOrder.created_at >= start_date)

    if end_date:
        query = query.filter(TransferOrder.created_at <= end_date + ' 23:59:59')

    orders = query.order_by(TransferOrder.created_at.desc()).all()

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '调拨单列表'

    headers = ['调拨单号', '调出仓库', '调入仓库', '调拨类型', '总数量', '状态', '经手人', '备注', '创建时间']
    ws.append(headers)

    status_map = {0: '待审核', 1: '已审核', 2: '已完成'}
    type_map = {1: '同价调拨', 2: '变价调拨'}

    for o in orders:
        ws.append([
            o.transfer_no,
            o.from_warehouse_name or '',
            o.to_warehouse_name or '',
            type_map.get(o.transfer_type, ''),
            float(o.total_quantity or 0),
            status_map.get(o.status, ''),
            o.operator_name or '',
            o.remark or '',
            o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else ''
        ])

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'调拨单列表_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ============================================
# 8. 成本调价
# ============================================

@bp.route('/api/cost-adjusts', methods=['GET'])
@jwt_required()
def get_cost_adjusts():
    """成本调价单列表"""
    from models.inventory.stock import CostAdjust

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = CostAdjust.query

    if keyword:
        query = query.filter(
            db.or_(
                CostAdjust.adjust_no.contains(keyword),
                CostAdjust.product_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if warehouse_id:
        query = query.filter_by(warehouse_id=warehouse_id)

    if start_date:
        query = query.filter(CostAdjust.created_at >= start_date)

    if end_date:
        query = query.filter(CostAdjust.created_at <= end_date + ' 23:59:59')

    total = query.count()
    adjusts = query.order_by(CostAdjust.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(a) for a in adjusts.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/cost-adjusts', methods=['POST'])
@jwt_required()
def create_cost_adjust():
    """创建成本调价单"""
    from models.inventory.stock import CostAdjust

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    date_str = datetime.now().strftime("%Y%m%d")
    last_adjust = CostAdjust.query.filter(
        CostAdjust.adjust_no.like(f'CB{date_str}%')
    ).order_by(CostAdjust.id.desc()).first()

    if last_adjust:
        last_seq = int(last_adjust.adjust_no[-4:]) if last_adjust.adjust_no else 0
        seq = str(last_seq + 1).zfill(4)
    else:
        seq = '0001'
    adjust_no = f'CB{date_str}{seq}'

    old_cost = float(data.get('old_cost_price', 0))
    new_cost = float(data.get('new_cost_price', 0))
    qty = float(data.get('adjust_quantity', 0))
    adjust_amount = round((new_cost - old_cost) * qty, 2)

    adjust = CostAdjust(
        adjust_no=adjust_no,
        warehouse_id=data.get('warehouse_id'),
        warehouse_name=data.get('warehouse_name', ''),
        product_id=data.get('product_id'),
        product_code=data.get('product_code', ''),
        product_name=data.get('product_name', ''),
        old_cost_price=old_cost,
        new_cost_price=new_cost,
        adjust_quantity=qty,
        adjust_amount=adjust_amount,
        status=0,
        remark=data.get('remark', ''),
        created_by=user_id,
        created_by_name=user_name
    )

    db.session.add(adjust)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': adjust.id, 'adjust_no': adjust_no}})


@bp.route('/api/cost-adjusts/<int:id>', methods=['GET'])
@jwt_required()
def get_cost_adjust(id):
    """调价单详情"""
    from models.inventory.stock import CostAdjust

    adjust = CostAdjust.query.get(id)
    if not adjust:
        return jsonify({'code': 404, 'message': '调价单不存在'}), 404

    return jsonify({'code': 200, 'data': to_dict(adjust)})


@bp.route('/api/cost-adjusts/<int:id>/audit', methods=['POST'])
@jwt_required()
def audit_cost_adjust(id):
    """审核成本调价单"""
    from models.inventory.stock import CostAdjust, InventoryLog, InventoryStock
    from models.finance import FinanceRecord

    adjust = CostAdjust.query.get(id)
    if not adjust:
        return jsonify({'code': 404, 'message': '调价单不存在'}), 404

    if adjust.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调价单可以审核'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    stock = InventoryStock.query.filter_by(
        product_id=adjust.product_id,
        warehouse_id=adjust.warehouse_id
    ).first()

    old_cost = float(adjust.old_cost_price or 0)
    new_cost = float(adjust.new_cost_price or 0)
    qty = float(adjust.adjust_quantity or 0)

    if stock:
        stock.cost_price = new_cost
        before_qty = float(stock.quantity or 0)
    else:
        before_qty = 0

    log = InventoryLog(
        product_id=adjust.product_id,
        product_code=adjust.product_code,
        product_name=adjust.product_name,
        warehouse_id=adjust.warehouse_id,
        warehouse_name=adjust.warehouse_name,
        change_type='adjust',
        order_type='成本调价',
        order_id=adjust.id,
        order_no=adjust.adjust_no,
        quantity_change=0,
        before_quantity=before_qty,
        after_quantity=before_qty,
        cost_price=new_cost,
        amount=adjust.adjust_amount,
        related_party='',
        operator_id=user_id,
        operator_name=user_name,
        remark=f'成本调价：{old_cost} -> {new_cost}，数量{qty}'
    )
    db.session.add(log)

    if adjust.adjust_amount and float(adjust.adjust_amount) != 0:
        diff_amount = float(adjust.adjust_amount)
        finance = FinanceRecord(
            account_id=1,
            account_name='默认账户',
            record_type=2 if diff_amount > 0 else 1,
            amount=abs(diff_amount),
            balance_before=0,
            balance_after=0,
            related_type='cost_adjust',
            related_id=adjust.id,
            related_no=adjust.adjust_no,
            remark=f'成本调价差异：{adjust.product_name} x {qty}，原成本{old_cost}，新成本{new_cost}',
            created_by=user_id
        )
        db.session.add(finance)

    adjust.status = 1
    adjust.audited_by = user_id
    adjust.audited_by_name = user_name
    adjust.audited_at = datetime.now()

    db.session.commit()
    return jsonify({'code': 200, 'message': '审核成功，成本已更新'})


@bp.route('/api/cost-adjusts/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_cost_adjust(id):
    """删除调价单（待审核状态）"""
    from models.inventory.stock import CostAdjust

    adjust = CostAdjust.query.get(id)
    if not adjust:
        return jsonify({'code': 404, 'message': '调价单不存在'}), 404

    if adjust.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核状态的调价单可以删除'}), 400

    db.session.delete(adjust)
    db.session.commit()

    return jsonify({'code': 200, 'message': '删除成功'})


@bp.route('/api/cost-adjusts/export', methods=['GET'])
@jwt_required()
def export_cost_adjusts():
    """导出成本调价单"""
    from models.inventory.stock import CostAdjust

    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = CostAdjust.query

    if keyword:
        query = query.filter(
            db.or_(
                CostAdjust.adjust_no.contains(keyword),
                CostAdjust.product_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if start_date:
        query = query.filter(CostAdjust.created_at >= start_date)

    if end_date:
        query = query.filter(CostAdjust.created_at <= end_date + ' 23:59:59')

    adjusts = query.order_by(CostAdjust.created_at.desc()).all()

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成本调价单列表'

    headers = ['调价单号', '仓库', '商品编码', '商品名称', '原成本价', '新成本价', '调整数量', '调整金额', '状态', '创建人', '创建时间']
    ws.append(headers)

    status_map = {0: '待审核', 1: '已审核', 2: '已取消'}

    for a in adjusts:
        ws.append([
            a.adjust_no,
            a.warehouse_name or '',
            a.product_code or '',
            a.product_name or '',
            float(a.old_cost_price or 0),
            float(a.new_cost_price or 0),
            float(a.adjust_quantity or 0),
            float(a.adjust_amount or 0),
            status_map.get(a.status, ''),
            a.created_by_name or '',
            a.created_at.strftime('%Y-%m-%d %H:%M:%S') if a.created_at else ''
        ])

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'成本调价单列表_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
