"""派单管理蓝图（派工状态机 + 查询 + 导出）。

迁移自 source-code/app.py 中以下区段的原始路由代码：
- /api/dispatch/manual                     手动派单                  11747 行附近
- /api/dispatch/<id>/{accept,reject,
   arrive,finish,redirect}                 派工状态机                11771-11843 行附近
- /api/dispatch/{records, records/<id>,
   logs/<id>, staff-list, pending, stats}  查询                      11876-12017 行附近
- /api/dispatchorders/export               导出                      18258 行附近

业务规则：
- 手动派单：创建 DispatchRecord + 更新工单状态为 1(已派单)
- 状态机：accept/reject/arrive/finish/redirect，附 DispatchLog
- 接单状态映射：0 待接单 / 1 已接单 / 2 已拒单 / 3 已超时
- 改派 = 旧 record accept_status=2 + 新建一条 DispatchRecord
- 统计接口：今日派单/已接/已拒/待接 + 接单率

跨子域依赖：
- WorkOrder (models.workorder)
- SysUser   (models.system)
- DispatchRecord / DispatchLog / StaffStatus (models.dispatch.record)

直接 import 需要的模型，使用 app.utils 提供 to_dict / export_to_excel。
get_current_user_id / get_current_user_name 通过本地小函数兼容迁移期。
"""
import io
import logging
from datetime import datetime

from flask import Blueprint, Response, jsonify, request, send_file
from flask_jwt_extended import get_jwt, jwt_required

from extensions import db
from app.security import permission
from app.services.permission_helpers import claims_to_perms, is_wildcard_admin
from app.utils import to_dict

logger = logging.getLogger(__name__)

bp = Blueprint('dispatch', __name__)


# ============================================
# 工具函数
# ============================================

def _get_current_user_id():
    """获取当前登录用户 ID。迁移期兼容（app.py 170 行）。"""
    from flask_jwt_extended import get_jwt_identity
    return get_jwt_identity()


def _get_current_user_name():
    """获取当前登录用户姓名。迁移期兼容（app.py 174 行）。"""
    from flask_jwt_extended import get_jwt_identity
    from models.system import SysUser
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    if user:
        return user.real_name or user.username
    return ''


# ============================================
# 1. 手动派单
# ============================================

@bp.route('/api/dispatch/manual', methods=['POST'])
@jwt_required()
@permission('dispatch:edit')
def manual_dispatch():
    """手动派单。"""
    from models.workorder import WorkOrder
    from models.system import SysUser
    from models.dispatch.record import DispatchRecord, DispatchLog

    data = request.get_json()
    wo_id = data.get('wo_id')
    staff_id = data.get('staff_id')
    remark = data.get('remark', '')

    wo = WorkOrder.query.get_or_404(wo_id)
    staff = SysUser.query.get_or_404(staff_id)

    record = DispatchRecord(
        wo_id=wo_id,
        dispatch_type='manual',
        dispatcher_id=_get_current_user_id(),
        dispatcher_name=_get_current_user_name(),
        staff_id=staff_id,
        staff_name=staff.real_name or staff.username,
        staff_phone=staff.phone,
        remark=remark,
    )
    db.session.add(record)

    wo.assigned_user_id = staff_id
    wo.assigned_user_name = staff.real_name or staff.username
    wo.assigned_time = datetime.now()
    wo.status = 1  # 已派单
    wo.status_name = '已派单'

    log = DispatchLog(
        wo_id=wo_id,
        action='派单',
        content=f"派单给{staff.real_name or staff.username}",
        operator_id=_get_current_user_id(),
        operator_name=_get_current_user_name(),
    )
    db.session.add(log)

    db.session.commit()
    return jsonify({'code': 200, 'message': '派单成功'})


# ============================================
# 2. 派工状态机：accept / reject / arrive / finish / redirect
# ============================================

@bp.route('/api/dispatch/<int:record_id>/accept', methods=['POST'])
@jwt_required()
@permission('dispatch:edit')
def accept_dispatch(record_id):
    """技术员接单。"""
    from models.workorder import WorkOrder
    from models.dispatch.record import DispatchRecord, DispatchLog

    record = DispatchRecord.query.get_or_404(record_id)
    record.accept_status = 1
    record.accept_time = datetime.now()

    wo = WorkOrder.query.get(record.wo_id)
    wo.status = 3  # 处理中
    wo.status_name = '处理中'

    log = DispatchLog(
        wo_id=record.wo_id,
        action='接单',
        content=f"{record.staff_name}已接单",
        operator_id=record.staff_id,
        operator_name=record.staff_name,
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '接单成功'})


@bp.route('/api/dispatch/<int:record_id>/reject', methods=['POST'])
@jwt_required()
@permission('dispatch:edit')
def reject_dispatch(record_id):
    """技术员拒单。"""
    from models.dispatch.record import DispatchRecord, DispatchLog

    data = request.get_json()
    record = DispatchRecord.query.get_or_404(record_id)
    record.accept_status = 2
    record.reject_reason = data.get('reason', '')

    log = DispatchLog(
        wo_id=record.wo_id,
        action='拒单',
        content=f"{record.staff_name}拒单: {data.get('reason', '')}",
        operator_id=record.staff_id,
        operator_name=record.staff_name,
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '拒单成功'})


@bp.route('/api/dispatch/<int:record_id>/arrive', methods=['POST'])
@jwt_required()
@permission('dispatch:edit')
def arrive_dispatch(record_id):
    """确认到达。"""
    from models.dispatch.record import DispatchRecord, DispatchLog

    record = DispatchRecord.query.get_or_404(record_id)
    record.arrive_time = datetime.now()
    log = DispatchLog(
        wo_id=record.wo_id,
        action='到达',
        content=f"{record.staff_name}已到达客户现场",
        operator_id=record.staff_id,
        operator_name=record.staff_name,
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '已确认到达'})


@bp.route('/api/dispatch/<int:record_id>/finish', methods=['POST'])
@jwt_required()
@permission('dispatch:edit')
def finish_dispatch(record_id):
    """确认完成：技师完成工作后，工单进入'待结算'（status=5）等待财务结算。"""
    from app.blueprints.workorder import WO_STATUS_MAP
    from models.workorder import WorkOrder
    from models.dispatch.record import DispatchRecord, DispatchLog

    record = DispatchRecord.query.get_or_404(record_id)
    record.finish_time = datetime.now()
    wo = WorkOrder.query.get(record.wo_id)
    # 派工完成 ≠ 结算完成：进入"待结算"状态，由财务/admin 走 settle 流程
    wo.status = 5
    wo.status_name = WO_STATUS_MAP.get(5, '待结算')
    wo.actual_time = datetime.now()
    log = DispatchLog(
        wo_id=record.wo_id,
        action='完成',
        content=f"{record.staff_name}已完成工单",
        operator_id=record.staff_id,
        operator_name=record.staff_name,
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '工单已完成'})


@bp.route('/api/dispatch/<int:record_id>/redirect', methods=['POST'])
@jwt_required()
@permission('dispatch:edit')
def redirect_dispatch(record_id):
    """改派。"""
    from models.workorder import WorkOrder
    from models.system import SysUser
    from models.dispatch.record import DispatchRecord, DispatchLog

    data = request.get_json()
    record = DispatchRecord.query.get_or_404(record_id)
    record.accept_status = 2
    record.reject_reason = '改派'

    new_staff_id = data.get('staff_id')
    new_staff = SysUser.query.get_or_404(new_staff_id)

    new_record = DispatchRecord(
        wo_id=record.wo_id,
        dispatch_type='manual',
        dispatcher_id=_get_current_user_id(),
        dispatcher_name=_get_current_user_name(),
        staff_id=new_staff_id,
        staff_name=new_staff.real_name or new_staff.username,
        staff_phone=new_staff.phone,
        remark=data.get('remark', '改派'),
    )
    db.session.add(new_record)

    wo = WorkOrder.query.get(record.wo_id)
    wo.assigned_user_id = new_staff_id
    wo.assigned_user_name = new_staff.real_name or new_staff.username

    log = DispatchLog(
        wo_id=record.wo_id,
        action='改派',
        content=f"从{record.staff_name}改派给{new_staff.real_name or new_staff.username}",
        operator_id=_get_current_user_id(),
        operator_name=_get_current_user_name(),
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({'code': 200, 'message': '改派成功'})


# ============================================
# 3. 查询接口
# ============================================

@bp.route('/api/dispatch/records', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def get_all_dispatch_records():
    """获取所有派单记录列表。"""
    from models.workorder import WorkOrder
    from models.dispatch.record import DispatchRecord

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', '')

    query = DispatchRecord.query
    if keyword:
        query = query.join(WorkOrder, DispatchRecord.wo_id == WorkOrder.id).filter(
            db.or_(
                WorkOrder.wo_no.contains(keyword),
                WorkOrder.customer_name.contains(keyword),
                DispatchRecord.staff_name.contains(keyword),
            )
        )
    if status:
        query = query.filter(DispatchRecord.accept_status == int(status))

    total = query.count()
    records = (
        query.order_by(DispatchRecord.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    result = []
    for r in records:
        d = to_dict(r)
        wo = WorkOrder.query.get(r.wo_id)
        if wo:
            d['wo_type'] = wo.wo_type
            d['customer_name'] = wo.customer_name
            d['customer_phone'] = wo.customer_phone
        result.append(d)

    return jsonify({
        'code': 200,
        'data': {
            'list': result,
            'total': total,
        },
    })


@bp.route('/api/dispatch/records/<int:wo_id>', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def get_dispatch_records(wo_id):
    """获取工单派单记录。"""
    from models.dispatch.record import DispatchRecord

    records = (
        DispatchRecord.query.filter_by(wo_id=wo_id)
        .order_by(DispatchRecord.created_at.desc())
        .all()
    )
    return jsonify({'code': 200, 'data': [to_dict(r) for r in records]})


@bp.route('/api/dispatch/logs/<int:wo_id>', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def get_dispatch_logs(wo_id):
    """获取工单派单日志。"""
    from models.dispatch.record import DispatchLog

    logs = (
        DispatchLog.query.filter_by(wo_id=wo_id)
        .order_by(DispatchLog.created_at.desc())
        .all()
    )
    return jsonify({'code': 200, 'data': [to_dict(l) for l in logs]})


@bp.route('/api/dispatch/staff-list', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def get_staff_list():
    """获取技术员列表（用于派单选择）。

    隐私：非 admin 角色不返回 phone 字段（技师无需看到同事手机号）。
    """
    from models.system import SysUser
    from models.dispatch.record import StaffStatus, DispatchRecord

    staff_list = SysUser.query.filter(SysUser.status == 1).all()
    include_phone = is_wildcard_admin(claims_to_perms(get_jwt()))
    result = []
    for s in staff_list:
        ss = StaffStatus.query.filter_by(staff_id=s.id).first()
        today = datetime.now().date()
        today_count = DispatchRecord.query.filter(
            DispatchRecord.staff_id == s.id,
            DispatchRecord.dispatch_time >= datetime.combine(today, datetime.min.time()),
            DispatchRecord.accept_status.in_([0, 1]),
        ).count()
        item = {
            'id': s.id,
            'name': s.real_name or s.username,
            'online_status': ss.online_status if ss else 0,
            'today_count': today_count,
            'max_daily': ss.max_daily if ss else 10,
            'skills': ss.skills if ss else '',
            'rating': float(ss.rating) if ss else 5.0,
        }
        if include_phone:
            item['phone'] = s.phone
        result.append(item)
    return jsonify({'code': 200, 'data': result})


@bp.route('/api/dispatch/pending', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def get_pending_dispatch():
    """获取待派单工单列表。"""
    from models.workorder import WorkOrder
    from models.system import SysUser

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 10, type=int)
    keyword = request.args.get('keyword', '')

    query = WorkOrder.query.filter(WorkOrder.status == 0)
    if keyword:
        query = query.filter(
            db.or_(
                WorkOrder.wo_no.contains(keyword),
                WorkOrder.customer_name.contains(keyword),
                WorkOrder.customer_phone.contains(keyword),
            )
        )

    total = query.count()
    orders = (
        query.order_by(WorkOrder.priority.desc(), WorkOrder.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    order_ids = [o.id for o in orders]
    user_ids = set()
    for o in orders:
        if o.receiver_id:
            user_ids.add(o.receiver_id)
        if o.engineer_user_id:
            user_ids.add(o.engineer_user_id)
    users_map = {}
    if user_ids:
        for u in SysUser.query.filter(SysUser.id.in_(user_ids)).all():
            users_map[u.id] = u.real_name or u.username

    result = []
    for o in orders:
        d = to_dict(o)
        d['receiver_name'] = users_map.get(o.receiver_id, '')
        d['engineer_user_name'] = users_map.get(o.engineer_user_id, '')
        d['device_brand'] = o.device_brand or ''
        d['device_model'] = o.device_model or ''
        d['device_sn'] = o.device_sn or ''
        d['device_type'] = o.device_type or ''
        result.append(d)

    return jsonify({
        'code': 200,
        'data': {
            'list': result,
            'total': total,
        },
    })


@bp.route('/api/dispatch/stats', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def get_dispatch_stats():
    """获取派单统计（今日）。"""
    from models.dispatch.record import DispatchRecord

    today = datetime.now().date()
    today_start = datetime.combine(today, datetime.min.time())

    total_dispatch = DispatchRecord.query.filter(
        DispatchRecord.dispatch_time >= today_start
    ).count()
    accepted = DispatchRecord.query.filter(
        DispatchRecord.dispatch_time >= today_start,
        DispatchRecord.accept_status == 1,
    ).count()
    rejected = DispatchRecord.query.filter(
        DispatchRecord.dispatch_time >= today_start,
        DispatchRecord.accept_status == 2,
    ).count()
    pending = DispatchRecord.query.filter(
        DispatchRecord.dispatch_time >= today_start,
        DispatchRecord.accept_status == 0,
    ).count()

    return jsonify({
        'code': 200,
        'data': {
            'total_dispatch': total_dispatch,
            'accepted': accepted,
            'rejected': rejected,
            'pending': pending,
            'accept_rate': round(accepted / total_dispatch * 100, 1) if total_dispatch > 0 else 0,
        },
    })


# ============================================
# 4. 导出
# ============================================

@bp.route('/api/dispatchorders/export', methods=['GET'])
@jwt_required()
@permission('dispatch:view')
def export_dispatch_orders():
    """导出派单记录。"""
    from models.workorder import WorkOrder
    from models.dispatch.record import DispatchRecord

    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)

        query = DispatchRecord.query
        if keyword:
            query = query.join(WorkOrder, DispatchRecord.wo_id == WorkOrder.id).filter(
                db.or_(
                    WorkOrder.wo_no.contains(keyword),
                    WorkOrder.customer_name.contains(keyword),
                    DispatchRecord.staff_name.contains(keyword),
                )
            )
        if status is not None:
            query = query.filter(DispatchRecord.accept_status == status)

        records = query.order_by(DispatchRecord.created_at.desc()).all()

        accept_status_map = {0: '待接单', 1: '已接单', 2: '已拒单', 3: '已超时'}
        priority_map = {0: '普通', 1: '紧急', 2: '特急'}

        data = []
        for r in records:
            wo = WorkOrder.query.get(r.wo_id)
            data.append({
                '派单号': f'DP{r.id:06d}',
                '工单号': wo.wo_no if wo else '',
                '客户名称': wo.customer_name if wo else '',
                '设备类型': wo.device_type if wo else '',
                '故障描述': (wo.fault_desc or '')[:200] if wo else '',
                '指派工程师': r.staff_name or '',
                '优先级': priority_map.get(wo.priority, '普通') if wo else '普通',
                '状态': accept_status_map.get(r.accept_status, '未知'),
                '创建时间': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
            })

        columns = (
            list(data[0].keys()) if data else [
                '派单号', '工单号', '客户名称', '设备类型', '故障描述',
                '指派工程师', '优先级', '状态', '创建时间',
            ]
        )

        # 用 openpyxl 直接构建，匹配原 app.py export_to_excel 的样式与返回方式
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '派单记录'

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                cell.border = thin_border

        # 调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
                12, min(50, len(str(col_name)) * 2 + 4),
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f'派单记录_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        logger.error(f'导出派单记录失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500