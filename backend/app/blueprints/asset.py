"""资产管理蓝图（资产类型 + 资产CRUD + 导入导出 + 销售单资产关联）。

迁移自 source-code/app.py 中以下区段的原始路由代码：
- /api/asset/types                          17545 行附近
- /api/assets + /api/assets/<id>            17583-17855 行附近
- /api/assets/<id>/scrap                    17859 行附近
- /api/assets/{import,export,by-customer}   17882-18112 行附近
- /api/sales/orders/<id>/assets             18119 行附近
- /api/sales/returns/<id>/unbind-assets     18199 行附近

业务规则：
- 删除：软删除（asset_status=5）
- 报废：asset_status=4
- 资产编号 ZC<YYYYMM><4位序号>，同月序号连续
- 质保状态根据 warranty_expire_date 自动判断（过保 0 / 在保 1）
- 销售单创建时同步生成关联资产；退货时清空关联字段

跨子域依赖：
- Asset / AssetType / OwnDevice          (models.asset.asset)
- Office                                (models.office)
- SysUser                               (models.system)
- SalesOrder / ReturnOrder              (models.sales.*)

直接 import 当前蓝图用得到的本地 helper，其他跨子域模型按"函数内懒加载"惯例。
"""
import io
import logging
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import jwt_required
import openpyxl

from extensions import db
from app.utils import to_dict
from app.security import permission, get_current_user_id

logger = logging.getLogger(__name__)

bp = Blueprint('asset', __name__)


# ============================================
# 工具函数
# ============================================

def _get_current_user_name():
    """获取当前登录用户姓名。迁移期兼容。"""
    from flask_jwt_extended import get_jwt_identity
    from models.system import SysUser
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    if user:
        return user.real_name or user.username
    return ''


def _generate_asset_no():
    """生成资产编号: ZC + 年月 + 4位序号。"""
    from models.asset.asset import Asset
    date_str = datetime.now().strftime('%Y%m')
    prefix = f'ZC{date_str}'
    last_asset = Asset.query.filter(Asset.asset_no.like(f'{prefix}%')).order_by(Asset.id.desc()).first()
    if last_asset and last_asset.asset_no:
        try:
            last_seq = int(last_asset.asset_no[-4:])
            seq = last_seq + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f'{prefix}{seq:04d}'


# ============================================
# 1. 资产类型
# ============================================

@bp.route('/api/asset/types', methods=['GET'])
@jwt_required()
@permission('asset:view')
def get_asset_types():
    """获取资产类型列表"""
    from models.asset.asset import AssetType
    try:
        types = AssetType.query.filter_by(status=1).order_by(AssetType.sort_order).all()
        data = [to_dict(t) for t in types]
        logger.info(f'[DEBUG] Asset types data: {data}')
        for t in data:
            logger.info(f'[DEBUG] type_name: {t.get("type_name")}, type_code: {t.get("type_code")}')
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': data
        })
    except Exception as e:
        logger.error(f'获取资产类型列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


# ============================================
# 2. 资产列表
# ============================================

@bp.route('/api/assets', methods=['GET'])
@jwt_required()
@permission('asset:view')
def get_assets():
    """资产列表查询"""
    from models.asset.asset import Asset, AssetType
    try:
        customer_id = request.args.get('customer_id', type=int)
        office_id = request.args.get('office_id', type=int)
        asset_type_id = request.args.get('asset_type_id', type=int)
        warranty_status = request.args.get('warranty_status', type=int)
        asset_status = request.args.get('asset_status', type=int)
        keyword = request.args.get('keyword')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        query = Asset.query

        if customer_id:
            query = query.filter(Asset.customer_id == customer_id)
        if office_id:
            query = query.filter(Asset.office_id == office_id)
        if asset_type_id:
            query = query.filter(Asset.asset_type_id == asset_type_id)
        if warranty_status is not None:
            query = query.filter(Asset.warranty_status == warranty_status)
        if asset_status is not None:
            query = query.filter(Asset.asset_status == asset_status)
        if keyword:
            query = query.filter(
                db.or_(
                    Asset.asset_name.contains(keyword),
                    Asset.asset_no.contains(keyword),
                    Asset.sn_code.contains(keyword),
                    Asset.device_no.contains(keyword),
                    Asset.responsible_person.contains(keyword)
                )
            )
        if start_date:
            query = query.filter(Asset.created_at >= start_date)
        if end_date:
            query = query.filter(Asset.created_at <= end_date + ' 23:59:59')

        # 排除已删除的（停用状态）
        query = query.filter(Asset.asset_status != 5)

        query = query.order_by(Asset.created_at.desc())
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()

        asset_types = {t.id: t for t in AssetType.query.all()}

        result_items = []
        for item in items:
            item_dict = to_dict(item)
            type_info = asset_types.get(item.asset_type_id)
            if type_info:
                item_dict['asset_type_code'] = type_info.type_code
            result_items.append(item_dict)

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'items': result_items
            }
        })
    except Exception as e:
        logger.error(f'获取资产列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


@bp.route('/api/assets/<int:id>', methods=['GET'])
@jwt_required()
@permission('asset:view')
def get_asset(id):
    """获取资产详情"""
    from models.asset.asset import Asset
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404
        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'获取资产详情失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


# ============================================
# 3. 资产 CRUD
# ============================================

@bp.route('/api/assets', methods=['POST'])
@jwt_required()
@permission('asset:add')
def create_asset():
    """创建资产"""
    from models.asset.asset import Asset, AssetType
    from models.office import Office
    try:
        data = request.get_json()

        required_fields = ['customer_id', 'customer_name', 'asset_type_id', 'asset_name']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'code': 400, 'message': f'缺少必填字段: {field}'}), 400

        asset_type = AssetType.query.get(data['asset_type_id'])
        if not asset_type:
            return jsonify({'code': 400, 'message': '资产类型不存在'}), 400

        asset_no = _generate_asset_no()

        warranty_expire_date = data.get('warranty_expire_date')
        warranty_status = 1
        if warranty_expire_date and warranty_expire_date != '':
            try:
                expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                warranty_status = 0 if expire_date < datetime.now().date() else 1
            except Exception:
                warranty_expire_date = None
        else:
            warranty_expire_date = None

        office_id = data.get('office_id')
        office_name = data.get('office_name')
        if office_id and not office_name:
            office = Office.query.get(office_id)
            if office:
                office_name = office.name

        def parse_date(date_val):
            if date_val and date_val != '':
                return date_val
            return None

        asset = Asset(
            asset_no=asset_no,
            customer_id=data['customer_id'],
            customer_name=data['customer_name'],
            office_id=office_id,
            office_name=office_name,
            location=data.get('location') or None,
            asset_type_id=data['asset_type_id'],
            asset_type_name=asset_type.type_name,
            asset_name=data['asset_name'],
            device_no=data.get('device_no') or None,
            sn_code=data.get('sn_code') or None,
            register_date=parse_date(data.get('register_date')),
            purchase_date=parse_date(data.get('purchase_date')),
            warranty_expire_date=warranty_expire_date,
            warranty_status=warranty_status,
            asset_status=data.get('asset_status', 1),
            responsible_person=data.get('responsible_person') or None,
            contact_phone=data.get('contact_phone') or None,
            ip_address=data.get('ip_address') or None,
            login_password=data.get('login_password') or None,
            remark=data.get('remark') or None,
            asset_data=data.get('asset_data'),
            sales_order_id=data.get('sales_order_id'),
            sales_order_no=data.get('sales_order_no'),
            created_by=get_current_user_id(),
            created_by_name=_get_current_user_name()
        )

        db.session.add(asset)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '创建成功',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'创建资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@bp.route('/api/assets/<int:id>', methods=['PUT'])
@jwt_required()
@permission('asset:edit')
def update_asset(id):
    """更新资产"""
    from models.asset.asset import Asset, AssetType
    from models.office import Office
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404

        data = request.get_json()

        if 'office_id' in data and data['office_id']:
            office = Office.query.get(data['office_id'])
            if office:
                asset.office_id = data['office_id']
                asset.office_name = office.name
        elif 'office_name' in data:
            asset.office_name = data['office_name']

        updatable_fields = [
            'customer_id', 'customer_name', 'location',
            'asset_type_id', 'asset_name', 'device_no', 'sn_code',
            'register_date', 'purchase_date', 'warranty_expire_date',
            'asset_status', 'responsible_person', 'contact_phone',
            'ip_address', 'login_password', 'remark', 'asset_data',
            'sales_order_id', 'sales_order_no'
        ]

        for field in updatable_fields:
            if field in data:
                setattr(asset, field, data[field])

        if 'asset_type_id' in data:
            asset_type = AssetType.query.get(data['asset_type_id'])
            if asset_type:
                asset.asset_type_name = asset_type.type_name

        if 'warranty_expire_date' in data and data['warranty_expire_date']:
            try:
                expire_date = datetime.strptime(data['warranty_expire_date'], '%Y-%m-%d').date()
                asset.warranty_status = 0 if expire_date < datetime.now().date() else 1
            except Exception:
                pass

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '更新成功',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'更新资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'更新失败: {str(e)}'}), 500


@bp.route('/api/assets/<int:id>', methods=['DELETE'])
@jwt_required()
@permission('asset:delete')
def delete_asset(id):
    """删除资产（软删除，将状态设为停用）"""
    from models.asset.asset import Asset
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404

        asset.asset_status = 5
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '删除成功',
            'data': {'id': id}
        })
    except Exception as e:
        logger.error(f'删除资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'删除失败: {str(e)}'}), 500


@bp.route('/api/assets/<int:id>/scrap', methods=['POST'])
@jwt_required()
@permission('asset:delete')
def scrap_asset(id):
    """资产报废"""
    from models.asset.asset import Asset
    try:
        asset = Asset.query.get(id)
        if not asset:
            return jsonify({'code': 404, 'message': '资产不存在'}), 404

        asset.asset_status = 4
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '资产已报废',
            'data': to_dict(asset)
        })
    except Exception as e:
        logger.error(f'资产报废失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'报废失败: {str(e)}'}), 500


# ============================================
# 4. 导入 / 导出 / 按客户查询
# ============================================

@bp.route('/api/assets/import', methods=['POST'])
@jwt_required()
@permission('asset:add')
def import_assets():
    """批量导入资产"""
    from models.asset.asset import Asset, AssetType
    try:
        if 'file' not in request.files:
            return jsonify({'code': 400, 'message': '请上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'code': 400, 'message': '文件名为空'}), 400

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        asset_types = {t.type_name: t for t in AssetType.query.all()}
        asset_types_by_code = {t.type_code: t for t in AssetType.query.all()}

        success_count = 0
        error_list = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                row_data = dict(zip(headers, row))

                customer_id = row_data.get('客户ID')
                customer_name = row_data.get('客户名称')
                asset_type_name = row_data.get('资产类型')
                asset_name = row_data.get('资产名称')

                if not customer_id or not customer_name or not asset_name:
                    error_list.append({'row': idx, 'error': '缺少必填字段'})
                    continue

                asset_type = asset_types.get(asset_type_name)
                if not asset_type:
                    asset_type = asset_types_by_code.get(asset_type_name)
                if not asset_type:
                    error_list.append({'row': idx, 'error': '资产类型不存在'})
                    continue

                warranty_expire_date = row_data.get('质保到期日')
                warranty_status = 1
                if warranty_expire_date:
                    try:
                        if isinstance(warranty_expire_date, str):
                            expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                        else:
                            expire_date = warranty_expire_date
                        warranty_status = 0 if expire_date < datetime.now().date() else 1
                    except Exception:
                        pass

                asset = Asset(
                    asset_no=_generate_asset_no(),
                    customer_id=int(customer_id),
                    customer_name=customer_name,
                    office_id=row_data.get('办公室ID'),
                    office_name=row_data.get('办公室名称'),
                    location=row_data.get('存放位置'),
                    asset_type_id=asset_type.id,
                    asset_type_name=asset_type.type_name,
                    asset_name=asset_name,
                    device_no=row_data.get('设备编号'),
                    sn_code=row_data.get('SN序列号'),
                    register_date=row_data.get('登记日期'),
                    purchase_date=row_data.get('采购日期'),
                    warranty_expire_date=warranty_expire_date,
                    warranty_status=warranty_status,
                    asset_status=1,
                    responsible_person=row_data.get('责任人'),
                    contact_phone=row_data.get('联系电话'),
                    ip_address=row_data.get('IP地址'),
                    login_password=row_data.get('登录密码'),
                    remark=row_data.get('备注'),
                    created_by=get_current_user_id(),
                    created_by_name=_get_current_user_name()
                )

                db.session.add(asset)
                success_count += 1

            except Exception as e:
                error_list.append({'row': idx, 'error': str(e)})

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'导入完成，成功{success_count}条，失败{len(error_list)}条',
            'data': {
                'success_count': success_count,
                'error_count': len(error_list),
                'errors': error_list
            }
        })
    except Exception as e:
        logger.error(f'导入资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500


@bp.route('/api/assets/export', methods=['GET'])
@jwt_required()
@permission('asset:view')
def export_assets():
    """导出资产"""
    from models.asset.asset import Asset
    try:
        customer_id = request.args.get('customer_id', type=int)
        asset_type_id = request.args.get('asset_type_id', type=int)
        keyword = request.args.get('keyword')

        query = Asset.query.filter(Asset.asset_status != 5)

        if customer_id:
            query = query.filter(Asset.customer_id == customer_id)
        if asset_type_id:
            query = query.filter(Asset.asset_type_id == asset_type_id)
        if keyword:
            query = query.filter(
                db.or_(
                    Asset.asset_name.contains(keyword),
                    Asset.asset_no.contains(keyword),
                    Asset.sn_code.contains(keyword)
                )
            )

        assets = query.order_by(Asset.created_at.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '资产台账'

        headers = [
            '资产编号', '资产名称', '资产类型', '客户名称', '办公室',
            '存放位置', '设备编号', 'SN序列号', '登记日期', '采购日期',
            '质保到期日', '质保状态', '资产状态', '责任人', '联系电话',
            'IP地址', '备注'
        ]
        ws.append(headers)

        warranty_status_map = {0: '过保', 1: '在保'}
        asset_status_map = {1: '正常使用', 2: '维修中', 3: '闲置', 4: '报废', 5: '停用'}

        for asset in assets:
            ws.append([
                asset.asset_no,
                asset.asset_name,
                asset.asset_type_name,
                asset.customer_name,
                asset.office_name,
                asset.location,
                asset.device_no,
                asset.sn_code,
                asset.register_date.strftime('%Y-%m-%d') if asset.register_date else '',
                asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                asset.warranty_expire_date.strftime('%Y-%m-%d') if asset.warranty_expire_date else '',
                warranty_status_map.get(asset.warranty_status, '未知'),
                asset_status_map.get(asset.asset_status, '未知'),
                asset.responsible_person,
                asset.contact_phone,
                asset.ip_address,
                asset.remark
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'资产台账_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f'导出资产失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500


@bp.route('/api/assets/by-customer', methods=['GET'])
@jwt_required()
@permission('asset:view')
def get_assets_by_customer():
    """根据客户获取资产列表（用于销售模块关联）"""
    from models.asset.asset import Asset
    try:
        customer_id = request.args.get('customer_id', type=int)
        if not customer_id:
            return jsonify({'code': 400, 'message': '缺少客户ID参数'}), 400

        assets = Asset.query.filter(
            Asset.customer_id == customer_id,
            Asset.asset_status.in_([1, 2, 3])
        ).order_by(Asset.created_at.desc()).all()

        data = []
        for asset in assets:
            data.append({
                'id': asset.id,
                'asset_no': asset.asset_no,
                'asset_name': asset.asset_name,
                'asset_type_name': asset.asset_type_name,
                'location': asset.location,
                'responsible_person': asset.responsible_person,
                'warranty_status': asset.warranty_status
            })

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': data
        })
    except Exception as e:
        logger.error(f'获取客户资产列表失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


# ============================================
# 5. 销售单资产关联
# ============================================

@bp.route('/api/sales/orders/<int:order_id>/assets', methods=['POST'])
@jwt_required()
@permission('sales:add')
def create_assets_for_sales_order(order_id):
    """销售单创建时同步创建资产"""
    from models.asset.asset import Asset, AssetType
    from models.sales.order import SalesOrder
    try:
        order = SalesOrder.query.get(order_id)
        if not order:
            return jsonify({'code': 404, 'message': '销售单不存在'}), 404

        data = request.get_json()
        assets_data = data.get('assets', [])

        if not assets_data:
            return jsonify({'code': 400, 'message': '资产数据不能为空'}), 400

        created_assets = []
        for asset_data in assets_data:
            if not asset_data.get('asset_type_id') or not asset_data.get('asset_name'):
                continue

            asset_type = AssetType.query.get(asset_data['asset_type_id'])
            if not asset_type:
                continue

            warranty_expire_date = asset_data.get('warranty_expire_date')
            warranty_status = 1
            if warranty_expire_date:
                try:
                    expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                    warranty_status = 0 if expire_date < datetime.now().date() else 1
                except Exception:
                    pass

            asset = Asset(
                asset_no=_generate_asset_no(),
                customer_id=order.customer_id,
                customer_name=order.customer_name,
                office_id=asset_data.get('office_id'),
                office_name=asset_data.get('office_name'),
                location=asset_data.get('location'),
                asset_type_id=asset_data['asset_type_id'],
                asset_type_name=asset_type.type_name,
                asset_name=asset_data['asset_name'],
                device_no=asset_data.get('device_no'),
                sn_code=asset_data.get('sn_code'),
                purchase_date=asset_data.get('purchase_date'),
                warranty_expire_date=warranty_expire_date,
                warranty_status=warranty_status,
                asset_status=1,
                responsible_person=asset_data.get('responsible_person'),
                contact_phone=asset_data.get('contact_phone'),
                ip_address=asset_data.get('ip_address'),
                login_password=asset_data.get('login_password'),
                remark=asset_data.get('remark'),
                asset_data=asset_data.get('asset_data'),
                sales_order_id=order.id,
                sales_order_no=order.order_no,
                created_by=get_current_user_id(),
                created_by_name=_get_current_user_name()
            )

            db.session.add(asset)
            created_assets.append(asset)

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'成功创建{len(created_assets)}个资产',
            'data': [to_dict(asset) for asset in created_assets]
        })
    except Exception as e:
        logger.error(f'销售单创建资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@bp.route('/api/sales/returns/<int:return_id>/unbind-assets', methods=['POST'])
@jwt_required()
@permission('sales-return:edit')
def unbind_assets_for_return(return_id):
    """销售退货解绑资产"""
    from models.asset.asset import Asset
    from models.purchase.return_order import ReturnOrder
    try:
        return_order = ReturnOrder.query.get(return_id)
        if not return_order:
            return jsonify({'code': 404, 'message': '退货单不存在'}), 404

        if not return_order.source_order_id:
            return jsonify({'code': 400, 'message': '退货单未关联销售单'}), 400

        assets = Asset.query.filter_by(sales_order_id=return_order.source_order_id).all()
        unbind_count = 0
        for asset in assets:
            asset.sales_order_id = None
            asset.sales_order_no = None
            unbind_count += 1

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'成功解绑{unbind_count}个资产',
            'data': {'unbind_count': unbind_count}
        })
    except Exception as e:
        logger.error(f'退货解绑资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'解绑失败: {str(e)}'}), 500


@bp.route('/api/sales/orders/<int:order_id>/assets', methods=['GET'])
@jwt_required()
@permission('sales:view')
def get_sales_order_assets(order_id):
    """获取销售单关联的资产"""
    from models.asset.asset import Asset
    from models.sales.order import SalesOrder
    try:
        order = SalesOrder.query.get(order_id)
        if not order:
            return jsonify({'code': 404, 'message': '销售单不存在'}), 404

        assets = Asset.query.filter_by(sales_order_id=order_id).order_by(Asset.created_at.desc()).all()

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': [to_dict(asset) for asset in assets]
        })
    except Exception as e:
        logger.error(f'获取销售单资产失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500
