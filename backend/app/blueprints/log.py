"""操作日志导出蓝图。

迁移自 source-code/app.py 中 18517 行附近的原始路由代码。

业务规则：
- 按 keyword / user_id / module / action / time range 过滤 OperationLog
- 输出 xlsx 文件，包含用户名/模块/操作/IP/内容/时间等列

跨子域依赖：
- OperationLog (models.system.user)

直接 import 当前蓝图用得到的本地 helper，其他跨子域模型按"函数内懒加载"惯例。
"""
import io
import logging
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import jwt_required

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from extensions import db
from app.utils import to_dict
from app.security import get_current_user_id

logger = logging.getLogger(__name__)

bp = Blueprint('log', __name__)


# ============================================
# 导出功能 - 操作日志
# ============================================

@bp.route('/api/operation-logs/export', methods=['GET'])
@jwt_required()
def export_operation_logs():
    """导出操作日志"""
    try:
        from models.system.user import OperationLog
        keyword = request.args.get('keyword', '')
        user_id = request.args.get('user_id', type=int)
        module = request.args.get('module', '')
        action = request.args.get('action', '')
        start_time = request.args.get('start_time', '')
        end_time = request.args.get('end_time', '')

        query = OperationLog.query
        if keyword:
            query = query.filter(
                db.or_(
                    OperationLog.user_name.contains(keyword),
                    OperationLog.content.contains(keyword),
                    OperationLog.ip.contains(keyword)
                )
            )
        if user_id:
            query = query.filter(OperationLog.user_id == user_id)
        if module:
            query = query.filter(OperationLog.module == module)
        if action:
            query = query.filter(OperationLog.action == action)
        if start_time:
            try:
                query = query.filter(OperationLog.created_at >= datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S'))
            except:
                pass
        if end_time:
            try:
                query = query.filter(OperationLog.created_at <= datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S'))
            except:
                pass

        logs = query.order_by(OperationLog.created_at.desc()).all()

        # 模块映射
        module_map = {
            'login': '登录', 'user': '用户管理', 'product': '商品管理',
            'customer': '客户管理', 'supplier': '供应商管理', 'sales': '销售管理',
            'purchase': '采购管理', 'work_order': '工单管理', 'inventory': '库存管理',
            'finance': '财务管理', 'settings': '系统设置'
        }
        # 操作类型映射
        action_map = {
            'create': '新增', 'update': '修改', 'delete': '删除',
            'query': '查询', 'export': '导出', 'login': '登录', 'logout': '登出'
        }

        data = []
        for log in logs:
            data.append({
                '操作用户': log.user_name or '',
                '操作类型': action_map.get(log.action, log.action or ''),
                '操作模块': module_map.get(log.module, log.module or ''),
                '操作描述': (log.content or '')[:500],
                'IP地址': log.ip or '',
                '创建时间': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ''
            })

        columns = list(data[0].keys()) if data else ['操作用户', '操作类型', '操作模块', '操作描述', 'IP地址', '创建时间']
        filename = f'操作日志_{datetime.now().strftime("%Y%m%d")}.xlsx'

        # 用 openpyxl 直接构建，匹配原 app.py export_to_excel 的样式与返回方式
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '数据'

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(vertical='center')
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        logger.error(f'导出操作日志失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'导出失败: {str(e)}'}), 500
