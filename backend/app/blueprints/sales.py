"""销售管理蓝图。

迁移自 source-code/app.py 中以下区段的原始路由代码：
- /api/sales/orders  CRUD + 详情                       8732-8955 行附近
- /api/sales/orders/<id>/audit  审核（自动出库+应收+流水） 8957-9132 行附近（176 行长函数）
- /api/sales/orders/<id>/print  打印                    9135-9181 行附近
- /api/sales/invoices  发票 CRUD                        9188-9390 行附近
- /api/sales/invoices/export  导出                      9393-9448 行附近
- /api/sales/receipts  收据 CRUD                        9455-9644 行附近
- /api/sales/receipts/export  导出                      9647-9698 行附近
- /api/sales/fix-prices  工具                           9705-9738 行附近
- /api/sales/orders/export  导出                        14358-14405 行附近
- /api/sales/orders/<id>/assets  资产绑定                18119-18196 行附近
- /api/sales/orders/<id>/assets  GET                    18233-18251 行附近
- /api/sales/returns/<id>/unbind-assets  退单解绑        18199-18230 行附近

按子资源分 4 个 section：
  1. 销售单   (orders CRUD + audit + print + export + assets 绑定/查询)
  2. 销售发票  (invoices CRUD + status + export)
  3. 销售收据  (receipts CRUD + print + void + export)
  4. 工具      (fix-prices, 退单解绑资产)

跨蓝图依赖：
- SalesOrder / SalesOrderItem (models.sales.order)
- SalesInvoice (models.sales.invoice)
- SalesReceipt (models.sales.receipt)
- BaseCustomer (models.customer)
- InventoryOut / InventoryOutItem / InventoryStock / InventoryLog (models.inventory)
- ProductInfo (models.product.info)
- FinanceReceivable / FinanceRecord (models.finance)
- SysUser (models.system) — 仅查询操作人
- Asset / AssetType (models.asset) — 资产绑定/解绑
- ReturnOrder (models.purchase.return_order) — 退单解绑

注意：audit 函数 ~176 行（事务 + 库存扣减 + 应收 + 财务流水）保持原样迁移，
不做任何重构以保证业务行为完全一致。
"""
import io
import json
import logging
from datetime import datetime

import openpyxl
from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity, jwt_required

from extensions import db
from app.security import permission
from app.utils import generate_code, to_dict

logger = logging.getLogger(__name__)

bp = Blueprint('sales', __name__)


# ============================================
# 常量
# ============================================

# 销售单状态映射（与原 app.py 5810-5817 行一致）
SO_STATUS_MAP = {
    0: '待审核',
    1: '已审核',
    2: '已出库',
    3: '已完成',
    4: '已取消',
}


# ============================================
# 工具函数
# ============================================

def _get_current_user_name():
    """获取当前登录用户姓名。迁移期兼容（app.py 顶层也定义同名函数）。"""
    from models.system import SysUser
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    if user:
        return user.real_name or user.username
    return ''


def _generate_asset_no():
    """生成资产编号: ZC + 年月 + 4位序号。迁移自 app.py 17566-17580 行。"""
    try:
        from models.asset import Asset
    except Exception:
        return None
    date_str = datetime.now().strftime('%Y%m')
    prefix = f'ZC{date_str}'
    last_asset = Asset.query.filter(Asset.asset_no.like(f'{prefix}%')).order_by(Asset.id.desc()).first()
    if last_asset and last_asset.asset_no:
        try:
            last_seq = int(last_asset.asset_no[-4:])
            seq = last_seq + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    return f'{prefix}{seq:04d}'


# ============================================
# 1. 销售单
# ============================================

@bp.route('/api/sales/orders', methods=['GET'])
@permission('sales:view')
def get_sales_orders():
    """获取销售单列表"""
    from models.sales.order import SalesOrder
    from models.system import SysUser

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    customer_id = request.args.get('customer_id', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    has_invoice = request.args.get('has_invoice', type=int)
    has_receipt = request.args.get('has_receipt', type=int)

    query = SalesOrder.query

    if keyword:
        query = query.filter(
            db.or_(
                SalesOrder.order_no.contains(keyword),
                SalesOrder.customer_name.contains(keyword)
            )
        )

    if status is not None:
        query = query.filter_by(status=status)

    if customer_id:
        query = query.filter_by(customer_id=customer_id)

    if start_date:
        query = query.filter(SalesOrder.order_date >= start_date)

    if end_date:
        query = query.filter(SalesOrder.order_date <= end_date)

    if has_invoice is not None:
        query = query.filter_by(has_invoice=has_invoice)

    if has_receipt is not None:
        query = query.filter_by(has_receipt=has_receipt)

    total = query.count()
    orders = query.order_by(SalesOrder.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    # 批量查询操作人名称
    user_ids = [o.created_by for o in orders.items if o.created_by]
    users_map = {}
    if user_ids:
        users = SysUser.query.filter(SysUser.id.in_(user_ids)).all()
        users_map = {u.id: (u.real_name or u.username) for u in users}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': o.id,
                'sales_no': o.order_no,
                'order_no': o.order_no,
                'customer_id': o.customer_id,
                'customer_name': o.customer_name,
                'sales_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else None,
                'order_date': o.order_date.strftime('%Y-%m-%d') if o.order_date else None,
                'total_amount': float(o.total_amount) if o.total_amount else 0,
                'received_amount': 0,
                'total_quantity': o.total_quantity,
                'status': o.status,
                'status_text': SO_STATUS_MAP.get(o.status, '未知'),
                'remark': o.remark,
                'has_invoice': o.has_invoice or 0,
                'has_receipt': o.has_receipt or 0,
                'operator_name': users_map.get(o.created_by, ''),
                'created_at': o.created_at.strftime('%Y-%m-%d %H:%M:%S') if o.created_at else None
            } for o in orders.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/sales/orders/<int:id>', methods=['GET'])
@permission('sales:view')
def get_sales_order(id):
    """获取销售单详情"""
    from models.sales.order import SalesOrder, SalesOrderItem

    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    items = SalesOrderItem.query.filter_by(order_id=id).all()

    result = to_dict(order)
    result['status_text'] = SO_STATUS_MAP.get(order.status, '未知')
    result['items'] = [to_dict(i) for i in items]

    return jsonify({'code': 200, 'data': result})


@bp.route('/api/sales/orders', methods=['POST'])
@permission('sales:add')
def create_sales_order():
    """创建销售单"""
    from models.sales.order import SalesOrder, SalesOrderItem

    data = request.get_json()
    user_id = get_jwt_identity()

    # 生成销售单号
    last_order = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
    order_no = generate_code('SO', last_order.id if last_order else 0)

    order = SalesOrder(
        order_no=order_no,
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        order_date=data.get('order_date'),
        remark=data.get('remark'),
        has_invoice=data.get('has_invoice', 0),
        has_receipt=data.get('has_receipt', 0),
        created_by=user_id
    )

    db.session.add(order)
    db.session.flush()

    # 添加明细
    items = data.get('items', [])
    total_amount = 0
    total_quantity = 0

    for item_data in items:
        qty = int(item_data.get('quantity', 0))
        price = float(item_data.get('price', 0))
        amount = qty * price
        total_amount += amount
        total_quantity += qty

        item = SalesOrderItem(
            order_id=order.id,
            product_id=item_data.get('product_id'),
            product_name=item_data.get('product_name'),
            specification=item_data.get('specification'),
            unit=item_data.get('unit'),
            quantity=qty,
            price=price,
            amount=amount,
            remark=item_data.get('remark')
        )
        db.session.add(item)

    order.total_amount = total_amount
    order.total_quantity = total_quantity
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': order.id, 'order_no': order_no}})


@bp.route('/api/sales/orders/<int:id>', methods=['PUT'])
@permission('sales:edit')
def update_sales_order(id):
    """更新销售单"""
    from models.sales.order import SalesOrder, SalesOrderItem

    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    if order.status not in [0, 1]:
        return jsonify({'code': 400, 'message': '该状态的销售单不能修改'}), 400

    data = request.get_json()

    for field in ['customer_id', 'customer_name', 'order_date', 'remark', 'has_invoice', 'has_receipt']:
        if field in data:
            setattr(order, field, data[field])

    # 更新明细
    if 'items' in data:
        # 删除旧明细
        SalesOrderItem.query.filter_by(order_id=id).delete()
        total_amount = 0
        total_quantity = 0

        for item_data in data['items']:
            qty = int(item_data.get('quantity', 0))
            price = float(item_data.get('price', 0))
            amount = qty * price
            total_amount += amount
            total_quantity += qty

            item = SalesOrderItem(
                order_id=id,
                product_id=item_data.get('product_id'),
                product_name=item_data.get('product_name'),
                specification=item_data.get('specification'),
                unit=item_data.get('unit'),
                quantity=qty,
                price=price,
                amount=amount,
                remark=item_data.get('remark')
            )
            db.session.add(item)

        order.total_amount = total_amount
        order.total_quantity = total_quantity

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@bp.route('/api/sales/orders/<int:id>', methods=['DELETE'])
@permission('sales:delete')
def delete_sales_order(id):
    """删除销售单"""
    from models.sales.order import SalesOrder

    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    if order.status == 3:
        return jsonify({'code': 400, 'message': '已完成的销售单不能删除'}), 400

    # 软删除
    order.status = 4  # 已取消
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@bp.route('/api/sales/orders/<int:id>/audit', methods=['POST'])
@permission('sales:edit')
def audit_sales_order(id):
    """审核销售单 - 自动出库、扣减库存、生成应收账款和财务流水"""
    from models.sales.order import SalesOrder, SalesOrderItem
    from models.inventory.flow import InventoryOut, InventoryOutItem
    from models.inventory.stock import InventoryStock, InventoryLog
    from models.product.info import ProductInfo
    from models.finance import FinanceReceivable, FinanceRecord

    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    if order.status != 0:
        return jsonify({'code': 400, 'message': '只有待审核的销售单可以审核'}), 400

    try:
        user_id = get_jwt_identity()
        user_name = _get_current_user_name()
        today = datetime.now().strftime('%Y%m%d')

        # 1. 更新销售单状态为已审核
        order.status = 1

        # 2. 检查是否已有出库单（避免重复创建）
        existing_out = InventoryOut.query.filter_by(related_order_id=id).first()
        if existing_out:
            order.status = 2  # 已出库
            db.session.commit()
            return jsonify({'code': 200, 'message': '审核成功（出库单已存在）'})

        # 3. 获取销售单明细
        items = SalesOrderItem.query.filter_by(order_id=id).all()
        if not items:
            order.status = 2
            db.session.commit()
            return jsonify({'code': 200, 'message': '审核成功（无明细商品）'})

        # 4. 检查库存是否充足
        for item in items:
            stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
            available = (stock.available_quantity or 0) if stock else 0
            if available < (item.quantity or 0):
                return jsonify({'code': 400, 'message': f'商品"{item.product_name}"库存不足，可用库存{available}，需要{item.quantity}'}), 400

        # 5. 生成出库单号 CK + YYYYMMDD + 4位序号
        prefix_ck = 'CK' + today
        last_out = InventoryOut.query.filter(InventoryOut.out_no.like(prefix_ck + '%')).order_by(InventoryOut.out_no.desc()).first()
        if last_out and last_out.out_no and len(last_out.out_no) > len(prefix_ck):
            seq = int(last_out.out_no[len(prefix_ck):]) + 1
        else:
            seq = 1
        out_no = prefix_ck + str(seq).zfill(4)

        # 6. 创建出库单
        inventory_out = InventoryOut(
            out_no=out_no,
            out_type=1,  # 销售出库
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            warehouse_id=1,
            warehouse_name='主仓库',
            total_quantity=sum(item.quantity or 0 for item in items),
            total_amount=float(order.total_amount or 0),
            status=2,  # 直接出库
            auditor_id=user_id,
            auditor_name=user_name,
            audit_time=datetime.now(),
            related_order_id=id,
            related_order_no=order.order_no,
            created_by=user_id
        )
        db.session.add(inventory_out)
        db.session.flush()  # 获取 out_id

        # 7. 创建出库单明细 + 扣减库存 + 写入库存日志 + 更新商品现存量
        for item in items:
            # 出库明细
            out_item = InventoryOutItem(
                out_id=inventory_out.id,
                product_id=item.product_id,
                product_code=None,
                product_name=item.product_name,
                specification=item.specification,
                unit_name=item.unit,
                quantity=item.quantity,
                unit_price=item.price,
                total_price=item.amount,
                cost_price=item.price
            )
            db.session.add(out_item)

            # 扣减库存
            stock = InventoryStock.query.filter_by(product_id=item.product_id).first()
            before_qty = stock.quantity if stock else 0
            if stock:
                stock.quantity = (stock.quantity or 0) - (item.quantity or 0)
                stock.available_quantity = (stock.available_quantity or 0) - (item.quantity or 0)
            after_qty = (stock.quantity or 0) if stock else 0

            # 写入库存日志
            inv_log = InventoryLog(
                product_id=item.product_id,
                product_code=stock.product_code if stock else None,
                product_name=item.product_name,
                warehouse_id=1,
                warehouse_name='主仓库',
                change_type='out',
                order_type='销售出库',
                order_id=inventory_out.id,
                order_no=out_no,
                quantity_change=-(item.quantity or 0),
                before_quantity=before_qty,
                after_quantity=after_qty,
                cost_price=item.price,
                amount=item.amount,
                related_party=order.customer_name,
                operator_id=user_id,
                operator_name=user_name,
                remark=f'销售单{order.order_no}审核自动出库'
            )
            db.session.add(inv_log)

            # 更新商品现存量
            if item.product_id:
                product_info = ProductInfo.query.get(item.product_id)
                if product_info:
                    product_info.current_stock = max(0, (product_info.current_stock or 0) - (item.quantity or 0))

        # 8. 生成应收编号 YS + YYYYMMDD + 4位序号
        prefix_ys = 'YS' + today
        last_receivable = FinanceReceivable.query.filter(FinanceReceivable.receivable_no.like(prefix_ys + '%')).order_by(FinanceReceivable.receivable_no.desc()).first()
        if last_receivable and last_receivable.receivable_no and len(last_receivable.receivable_no) > len(prefix_ys):
            seq_ys = int(last_receivable.receivable_no[len(prefix_ys):]) + 1
        else:
            seq_ys = 1
        receivable_no = prefix_ys + str(seq_ys).zfill(4)

        # 9. 生成应收账款
        total_amount = float(order.actual_amount or order.total_amount or 0)
        receivable = FinanceReceivable(
            receivable_no=receivable_no,
            related_type='sale',
            related_id=id,
            related_no=order.order_no,
            customer_id=order.customer_id,
            customer_name=order.customer_name,
            total_amount=total_amount,
            received_amount=float(order.paid_amount or 0),
            remaining_amount=total_amount - float(order.paid_amount or 0),
            status=1 if (order.paid_amount and float(order.paid_amount) > 0 and float(order.paid_amount) < total_amount) else (2 if (order.paid_amount and float(order.paid_amount) >= total_amount) else 0),
            remark=f'销售单{order.order_no}审核自动生成'
        )
        db.session.add(receivable)

        # 10. 生成财务流水（收入）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=1,  # 收入
            amount=total_amount,
            balance_before=0,
            balance_after=0,
            related_type='sale',
            related_id=id,
            related_no=order.order_no,
            remark=f'销售单{order.order_no}审核，应收金额{total_amount}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        # 11. 更新销售单状态为已出库
        order.status = 2

        db.session.commit()
        return jsonify({'code': 200, 'message': '审核成功，已自动出库并生成应收账款'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'审核失败：{str(e)}'}), 500


@bp.route('/api/sales/orders/<int:id>/print', methods=['GET'])
@permission('sales:view')
def print_sales_order(id):
    """打印销售单/收据"""
    from models.sales.order import SalesOrder, SalesOrderItem
    from models.customer import BaseCustomer

    order = SalesOrder.query.get(id)
    if not order:
        return jsonify({'code': 404, 'message': '销售单不存在'}), 404

    # 获取明细
    items = SalesOrderItem.query.filter_by(order_id=id).all()

    # 获取客户信息
    customer = BaseCustomer.query.get(order.customer_id) if order.customer_id else None

    # 构建打印数据
    print_data = {
        'order_no': order.order_no,
        'order_date': order.order_date.strftime('%Y-%m-%d') if order.order_date else '',
        'customer_name': order.customer_name or '',
        'customer_phone': order.customer_phone or (customer.phone if customer else ''),
        'customer_address': order.customer_address or (customer.address if customer else ''),
        'contact_name': order.contact_name or (customer.contact_name if customer else ''),
        'salesperson_name': order.salesperson_name or '',
        'payment_method': order.payment_method or '',
        'delivery_method': order.delivery_method or '',
        'items': [{
            'product_name': item.product_name,
            'specification': item.specification or '',
            'unit': item.unit or '',
            'quantity': int(item.quantity),
            'price': float(item.price),
            'amount': float(item.amount)
        } for item in items],
        'total_quantity': int(order.total_quantity),
        'total_amount': float(order.total_amount),
        'discount_amount': float(order.discount_amount) if order.discount_amount else 0,
        'freight_amount': float(order.freight_amount) if order.freight_amount else 0,
        'actual_amount': float(order.actual_amount) if order.actual_amount else float(order.total_amount),
        'paid_amount': float(order.paid_amount) if order.paid_amount else 0,
        'remark': order.remark or '',
        'print_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    return jsonify({
        'code': 200,
        'data': print_data
    })


@bp.route('/api/sales/orders/export', methods=['GET'])
@permission('sales:view')
def export_sales_orders():
    """导出销售单"""
    from models.sales.order import SalesOrder

    keyword = request.args.get('keyword', '')
    status = request.args.get('status', type=int)
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    query = SalesOrder.query
    if keyword:
        query = query.filter(db.or_(SalesOrder.order_no.contains(keyword), SalesOrder.customer_name.contains(keyword)))
    if status is not None:
        query = query.filter(SalesOrder.status == status)
    if date_start:
        query = query.filter(SalesOrder.created_at >= date_start)
    if date_end:
        query = query.filter(SalesOrder.created_at <= date_end + ' 23:59:59')

    orders = query.order_by(SalesOrder.created_at.desc()).all()

    data = []
    for o in orders:
        data.append({
            '销售单号': o.order_no,
            '客户名称': o.customer_name,
            '客户电话': getattr(o, 'customer_phone', '') or '',
            '销售日期': o.order_date.strftime('%Y-%m-%d') if o.order_date else '',
            '总数量': int(o.total_quantity or 0),
            '总金额': float(o.total_amount or 0),
            '已收金额': float(getattr(o, 'paid_amount', 0) or 0),
            '状态': {0: '待审核', 1: '已审核', 2: '已出库', 3: '已完成', 4: '已取消'}.get(o.status, str(o.status)),
            '创建时间': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '销售单列表'
    # 定义表头
    headers = ['销售单号', '客户名称', '客户电话', '销售日期', '总数量', '总金额', '已收金额', '状态', '创建时间']
    ws.append(headers)
    for row in data:
        ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'销售单列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/sales/orders/<int:order_id>/assets', methods=['POST'])
@permission('sales:edit')
def create_assets_for_sales_order(order_id):
    """销售单创建时同步创建资产"""
    from models.sales.order import SalesOrder

    try:
        # 跨蓝图依赖：Asset / AssetType（assets 蓝图尚未迁移）
        try:
            from models.asset import Asset
        except Exception:
            Asset = None
        try:
            from models.asset.asset import AssetType
        except Exception:
            AssetType = None

        order = SalesOrder.query.get(order_id)
        if not order:
            return jsonify({'code': 404, 'message': '销售单不存在'}), 404

        data = request.get_json()
        assets_data = data.get('assets', [])

        if not assets_data:
            return jsonify({'code': 400, 'message': '资产数据不能为空'}), 400

        created_assets = []
        for asset_data in assets_data:
            # 必填字段校验
            if not asset_data.get('asset_type_id') or not asset_data.get('asset_name'):
                continue

            # 获取资产类型名称
            if AssetType is None:
                asset_type_name = ''
            else:
                asset_type = AssetType.query.get(asset_data['asset_type_id'])
                if not asset_type:
                    continue
                asset_type_name = asset_type.type_name

            # 计算质保状态
            warranty_expire_date = asset_data.get('warranty_expire_date')
            warranty_status = 1
            if warranty_expire_date:
                try:
                    expire_date = datetime.strptime(warranty_expire_date, '%Y-%m-%d').date()
                    warranty_status = 0 if expire_date < datetime.now().date() else 1
                except Exception:
                    pass

            asset_no = _generate_asset_no() or f'ZC{datetime.now().strftime("%Y%m%d%H%M%S")}'

            if Asset is None:
                continue

            asset = Asset(
                asset_no=asset_no,
                customer_id=order.customer_id,
                customer_name=order.customer_name,
                office_id=asset_data.get('office_id'),
                office_name=asset_data.get('office_name'),
                location=asset_data.get('location'),
                asset_type_id=asset_data['asset_type_id'],
                asset_type_name=asset_type_name,
                asset_name=asset_data['asset_name'],
                device_no=asset_data.get('device_no'),
                sn_code=asset_data.get('sn_code'),
                purchase_date=asset_data.get('purchase_date'),
                warranty_expire_date=warranty_expire_date,
                warranty_status=warranty_status,
                asset_status=1,
                responsible_person=asset_data.get('responsible_person'),
                contact_phone=asset_data.get('contact_phone'),
                ip_address=asset_data.get('ip_address'),
                login_password=asset_data.get('login_password'),
                remark=asset_data.get('remark'),
                asset_data=asset_data.get('asset_data'),
                sales_order_id=order.id,
                sales_order_no=order.order_no,
                created_by=get_jwt_identity(),
                created_by_name=_get_current_user_name()
            )

            db.session.add(asset)
            created_assets.append(asset)

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'成功创建{len(created_assets)}个资产',
            'data': [to_dict(asset) for asset in created_assets]
        })
    except Exception as e:
        logger.error(f'销售单创建资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'创建失败: {str(e)}'}), 500


@bp.route('/api/sales/orders/<int:order_id>/assets', methods=['GET'])
@permission('sales:view')
def get_sales_order_assets(order_id):
    """获取销售单关联的资产"""
    from models.sales.order import SalesOrder

    try:
        # 跨蓝图依赖：Asset（assets 蓝图尚未迁移）
        try:
            from models.asset import Asset
        except Exception:
            return jsonify({'code': 503, 'message': '资产模块尚未启用'}), 503

        order = SalesOrder.query.get(order_id)
        if not order:
            return jsonify({'code': 404, 'message': '销售单不存在'}), 404

        assets = Asset.query.filter_by(sales_order_id=order_id).order_by(Asset.created_at.desc()).all()

        return jsonify({
            'code': 200,
            'message': '获取成功',
            'data': [to_dict(asset) for asset in assets]
        })
    except Exception as e:
        logger.error(f'获取销售单资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'获取失败: {str(e)}'}), 500


# ============================================
# 2. 销售发票
# ============================================

@bp.route('/api/sales/invoices', methods=['GET'])
@jwt_required()
def get_sales_invoices():
    """获取发票列表"""
    from models.sales.invoice import SalesInvoice

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    order_id = request.args.get('order_id', type=int)
    invoice_status = request.args.get('invoice_status', '')
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesInvoice.query

    if order_id:
        query = query.filter_by(order_id=order_id)

    if invoice_status:
        query = query.filter_by(invoice_status=invoice_status)

    if keyword:
        query = query.filter(
            db.or_(
                SalesInvoice.order_no.contains(keyword),
                SalesInvoice.buyer_name.contains(keyword),
                SalesInvoice.invoice_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= start_date)

    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= end_date)

    total = query.count()
    invoices = query.order_by(SalesInvoice.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': inv.id,
                'order_id': inv.order_id,
                'order_no': inv.order_no,
                'invoice_type': inv.invoice_type,
                'invoice_status': inv.invoice_status,
                'invoice_no': inv.invoice_no,
                'invoice_date': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else None,
                'buyer_name': inv.buyer_name,
                'total_amount': float(inv.total_amount) if inv.total_amount else 0,
                'total_tax': float(inv.total_tax) if inv.total_tax else 0,
                'total_with_tax': float(inv.total_with_tax) if inv.total_with_tax else 0,
                'tax_rate': float(inv.tax_rate) if inv.tax_rate else 0,
                'created_by_name': inv.created_by_name,
                'created_at': inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else None
            } for inv in invoices.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/sales/invoices', methods=['POST'])
@jwt_required()
def create_or_update_sales_invoice():
    """创建或更新发票"""
    from models.sales.invoice import SalesInvoice
    from models.sales.order import SalesOrder
    from models.finance import FinanceReceivable

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    order_id = data.get('order_id')
    if not order_id:
        return jsonify({'code': 400, 'message': '缺少order_id参数'}), 400

    order = SalesOrder.query.get(order_id)
    if not order:
        return jsonify({'code': 404, 'message': '关联的销售单不存在'}), 404

    # 检查是否已存在该订单的发票
    existing = SalesInvoice.query.filter_by(order_id=order_id).first()

    if existing:
        # 更新已有发票
        for field in ['invoice_type', 'invoice_no', 'invoice_date', 'buyer_name', 'buyer_tax_no',
                       'buyer_address', 'buyer_phone', 'buyer_bank', 'buyer_bank_account',
                       'items', 'total_amount', 'total_tax', 'total_with_tax', 'tax_rate',
                       'remark', 'attachment', 'invoice_status']:
            if field in data:
                setattr(existing, field, data[field])

        existing.updated_at = datetime.now()

        # 发票关联核销：更新发票状态为已开票时，关联应收账款
        if data.get('invoice_status') == '已开票':
            receivable = FinanceReceivable.query.filter_by(
                related_type='sale',
                related_id=order_id
            ).first()
            if receivable:
                receivable.invoice_id = existing.id
                invoice_amount = float(data.get('total_with_tax') or data.get('total_amount') or 0)
                receivable_amount = float(receivable.total_amount or 0)
                if invoice_amount >= receivable_amount:
                    receivable.status = 2  # 已结清
                receivable.updated_at = datetime.now()

        db.session.commit()
        return jsonify({'code': 200, 'message': '更新成功', 'data': {'id': existing.id}})
    else:
        # 创建新发票
        items_json = data.get('items', '[]')
        if isinstance(items_json, list):
            items_json = json.dumps(items_json, ensure_ascii=False)

        invoice = SalesInvoice(
            order_id=order_id,
            order_no=order.order_no,
            invoice_type=data.get('invoice_type', '普通发票'),
            invoice_status=data.get('invoice_status', '未开票'),
            invoice_no=data.get('invoice_no', ''),
            invoice_date=data.get('invoice_date'),
            buyer_name=data.get('buyer_name', ''),
            buyer_tax_no=data.get('buyer_tax_no', ''),
            buyer_address=data.get('buyer_address', ''),
            buyer_phone=data.get('buyer_phone', ''),
            buyer_bank=data.get('buyer_bank', ''),
            buyer_bank_account=data.get('buyer_bank_account', ''),
            items=items_json,
            total_amount=data.get('total_amount', 0),
            total_tax=data.get('total_tax', 0),
            total_with_tax=data.get('total_with_tax', 0),
            tax_rate=data.get('tax_rate', 0),
            remark=data.get('remark', ''),
            attachment=data.get('attachment', ''),
            created_by=user_id,
            created_by_name=user_name
        )
        db.session.add(invoice)

        # 更新销售单的has_invoice标记
        order.has_invoice = 1

        # 发票关联核销：查找关联的应收账款并更新
        if data.get('invoice_status') == '已开票':
            receivable = FinanceReceivable.query.filter_by(
                related_type='sale',
                related_id=order_id
            ).first()
            if receivable:
                receivable.invoice_id = invoice.id
                invoice_amount = float(data.get('total_with_tax') or data.get('total_amount') or 0)
                receivable_amount = float(receivable.total_amount or 0)
                if invoice_amount >= receivable_amount:
                    receivable.status = 2  # 已结清
                receivable.updated_at = datetime.now()

        db.session.commit()
        return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': invoice.id}})


@bp.route('/api/sales/invoices/<int:id>', methods=['GET'])
@jwt_required()
def get_sales_invoice(id):
    """获取发票详情"""
    from models.sales.invoice import SalesInvoice

    invoice = SalesInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    result = to_dict(invoice)
    # 解析items JSON
    if result.get('items'):
        try:
            result['items'] = json.loads(result['items'])
        except (json.JSONDecodeError, TypeError):
            pass

    return jsonify({'code': 200, 'data': result})


@bp.route('/api/sales/invoices/<int:id>/status', methods=['PUT'])
@jwt_required()
def update_sales_invoice_status(id):
    """更新发票状态"""
    from models.sales.invoice import SalesInvoice

    invoice = SalesInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    data = request.get_json()
    new_status = data.get('invoice_status')
    if not new_status or new_status not in ['未开票', '已开票', '作废']:
        return jsonify({'code': 400, 'message': '无效的开票状态，允许值：未开票/已开票/作废'}), 400

    invoice.invoice_status = new_status
    invoice.updated_at = datetime.now()
    db.session.commit()

    return jsonify({'code': 200, 'message': '状态更新成功', 'data': {'invoice_status': new_status}})


@bp.route('/api/sales/invoices/export', methods=['GET'])
@jwt_required()
def export_sales_invoices():
    """导出发票列表"""
    from models.sales.invoice import SalesInvoice

    keyword = request.args.get('keyword', '')
    invoice_status = request.args.get('invoice_status', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesInvoice.query
    if keyword:
        query = query.filter(
            db.or_(
                SalesInvoice.order_no.contains(keyword),
                SalesInvoice.buyer_name.contains(keyword),
                SalesInvoice.invoice_no.contains(keyword)
            )
        )
    if invoice_status:
        query = query.filter_by(invoice_status=invoice_status)
    if start_date:
        query = query.filter(SalesInvoice.invoice_date >= start_date)
    if end_date:
        query = query.filter(SalesInvoice.invoice_date <= end_date)

    invoices = query.order_by(SalesInvoice.created_at.desc()).all()

    data = []
    for inv in invoices:
        data.append({
            '销售单号': inv.order_no or '',
            '发票类型': inv.invoice_type or '',
            '发票编号': inv.invoice_no or '',
            '开票日期': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else '',
            '开票状态': inv.invoice_status or '',
            '客户抬头': inv.buyer_name or '',
            '税号': inv.buyer_tax_no or '',
            '金额合计': float(inv.total_amount) if inv.total_amount else 0,
            '税额合计': float(inv.total_tax) if inv.total_tax else 0,
            '价税合计': float(inv.total_with_tax) if inv.total_with_tax else 0,
            '开票人': inv.created_by_name or '',
            '创建时间': inv.created_at.strftime('%Y-%m-%d %H:%M') if inv.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '发票列表'
    if data:
        ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'发票列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# 3. 销售收据
# ============================================

@bp.route('/api/sales/receipts', methods=['GET'])
@jwt_required()
def get_sales_receipts():
    """获取收据列表"""
    from models.sales.receipt import SalesReceipt

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    order_id = request.args.get('order_id', type=int)
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesReceipt.query.filter_by(status=1)

    if order_id:
        query = query.filter_by(order_id=order_id)

    if keyword:
        query = query.filter(
            db.or_(
                SalesReceipt.order_no.contains(keyword),
                SalesReceipt.customer_name.contains(keyword),
                SalesReceipt.receipt_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(SalesReceipt.receipt_date >= start_date)

    if end_date:
        query = query.filter(SalesReceipt.receipt_date <= end_date)

    total = query.count()
    receipts = query.order_by(SalesReceipt.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': r.id,
                'order_id': r.order_id,
                'order_no': r.order_no,
                'receipt_no': r.receipt_no,
                'receipt_date': r.receipt_date.strftime('%Y-%m-%d') if r.receipt_date else None,
                'customer_name': r.customer_name,
                'customer_phone': r.customer_phone,
                'total_amount': float(r.total_amount) if r.total_amount else 0,
                'paid_amount': float(r.paid_amount) if r.paid_amount else 0,
                'payment_method': r.payment_method,
                'payee': r.payee,
                'status': r.status,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None
            } for r in receipts.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/sales/receipts', methods=['POST'])
@jwt_required()
def create_sales_receipt():
    """开具收据"""
    from models.sales.receipt import SalesReceipt
    from models.sales.order import SalesOrder

    data = request.get_json()
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    order_id = data.get('order_id')
    if not order_id:
        return jsonify({'code': 400, 'message': '缺少order_id参数'}), 400

    order = SalesOrder.query.get(order_id)
    if not order:
        return jsonify({'code': 404, 'message': '关联的销售单不存在'}), 404

    # 自动生成收据编号：SK + 年月日 + 4位序号
    today_str = datetime.now().strftime('%Y%m%d')
    prefix = f'SK{today_str}'
    last_receipt = SalesReceipt.query.filter(SalesReceipt.receipt_no.like(f'{prefix}%')).order_by(SalesReceipt.id.desc()).first()
    if last_receipt and last_receipt.receipt_no:
        last_seq = int(last_receipt.receipt_no[-4:]) if len(last_receipt.receipt_no) >= 4 else 0
    else:
        last_seq = 0
    receipt_no = f'{prefix}{str(last_seq + 1).zfill(4)}'

    items_json = data.get('items', '[]')
    if isinstance(items_json, list):
        items_json = json.dumps(items_json, ensure_ascii=False)

    receipt = SalesReceipt(
        order_id=order_id,
        order_no=order.order_no,
        receipt_no=receipt_no,
        receipt_date=data.get('receipt_date') or datetime.now().strftime('%Y-%m-%d'),
        customer_name=data.get('customer_name', order.customer_name),
        customer_phone=data.get('customer_phone', order.customer_phone),
        total_amount=data.get('total_amount', 0),
        paid_amount=data.get('paid_amount', 0),
        payment_method=data.get('payment_method', ''),
        items=items_json,
        remark=data.get('remark', ''),
        payee=data.get('payee', user_name),
        status=1,
        created_by=user_id
    )
    db.session.add(receipt)

    # 更新销售单的has_receipt标记
    order.has_receipt = 1

    db.session.commit()
    return jsonify({'code': 200, 'message': '收据开具成功', 'data': {'id': receipt.id, 'receipt_no': receipt_no}})


@bp.route('/api/sales/receipts/<int:id>', methods=['GET'])
@jwt_required()
def get_sales_receipt(id):
    """获取收据详情"""
    from models.sales.receipt import SalesReceipt

    receipt = SalesReceipt.query.get(id)
    if not receipt:
        return jsonify({'code': 404, 'message': '收据不存在'}), 404

    result = to_dict(receipt)
    # 解析items JSON
    if result.get('items'):
        try:
            result['items'] = json.loads(result['items'])
        except (json.JSONDecodeError, TypeError):
            pass

    return jsonify({'code': 200, 'data': result})


@bp.route('/api/sales/receipts/<int:id>/print', methods=['GET'])
@jwt_required()
def print_sales_receipt(id):
    """获取收据打印数据"""
    from models.sales.receipt import SalesReceipt

    receipt = SalesReceipt.query.get(id)
    if not receipt:
        return jsonify({'code': 404, 'message': '收据不存在'}), 404

    if receipt.status == 0:
        return jsonify({'code': 400, 'message': '该收据已作废，无法打印'}), 400

    # 解析items
    items = []
    if receipt.items:
        try:
            items = json.loads(receipt.items)
        except (json.JSONDecodeError, TypeError):
            pass

    print_data = {
        'receipt_no': receipt.receipt_no,
        'receipt_date': receipt.receipt_date.strftime('%Y-%m-%d') if receipt.receipt_date else '',
        'customer_name': receipt.customer_name or '',
        'customer_phone': receipt.customer_phone or '',
        'items': items,
        'total_amount': float(receipt.total_amount) if receipt.total_amount else 0,
        'paid_amount': float(receipt.paid_amount) if receipt.paid_amount else 0,
        'payment_method': receipt.payment_method or '',
        'remark': receipt.remark or '',
        'payee': receipt.payee or '',
        'print_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    return jsonify({'code': 200, 'data': print_data})


@bp.route('/api/sales/receipts/<int:id>/void', methods=['PUT'])
@jwt_required()
def void_sales_receipt(id):
    """作废收据"""
    from models.sales.receipt import SalesReceipt

    receipt = SalesReceipt.query.get(id)
    if not receipt:
        return jsonify({'code': 404, 'message': '收据不存在'}), 404

    if receipt.status == 0:
        return jsonify({'code': 400, 'message': '该收据已作废'}), 400

    receipt.status = 0
    receipt.updated_at = datetime.now()
    db.session.commit()

    return jsonify({'code': 200, 'message': '收据已作废'})


@bp.route('/api/sales/receipts/export', methods=['GET'])
@jwt_required()
def export_sales_receipts():
    """导出收据列表"""
    from models.sales.receipt import SalesReceipt

    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = SalesReceipt.query.filter_by(status=1)
    if keyword:
        query = query.filter(
            db.or_(
                SalesReceipt.order_no.contains(keyword),
                SalesReceipt.customer_name.contains(keyword),
                SalesReceipt.receipt_no.contains(keyword)
            )
        )
    if start_date:
        query = query.filter(SalesReceipt.receipt_date >= start_date)
    if end_date:
        query = query.filter(SalesReceipt.receipt_date <= end_date)

    receipts = query.order_by(SalesReceipt.created_at.desc()).all()

    data = []
    for r in receipts:
        data.append({
            '收据编号': r.receipt_no or '',
            '销售单号': r.order_no or '',
            '收款日期': r.receipt_date.strftime('%Y-%m-%d') if r.receipt_date else '',
            '客户名称': r.customer_name or '',
            '联系方式': r.customer_phone or '',
            '应收金额': float(r.total_amount) if r.total_amount else 0,
            '实收金额': float(r.paid_amount) if r.paid_amount else 0,
            '收款方式': r.payment_method or '',
            '收款人': r.payee or '',
            '备注': r.remark or '',
            '创建时间': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '收据列表'
    if data:
        ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f'收据列表_{datetime.now().strftime("%Y%m%d")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# 4. 工具
# ============================================

@bp.route('/api/sales/fix-prices', methods=['POST'])
@jwt_required()
def fix_sales_prices():
    """修复销售单明细中价格为0的数据，从商品表重新读取销售价"""
    from models.sales.order import SalesOrder, SalesOrderItem
    from models.product.info import ProductInfo

    fixed_count = 0
    fixed_order_ids = set()

    # 查找所有价格为0的明细
    items = SalesOrderItem.query.filter(
        (SalesOrderItem.price == 0) | (SalesOrderItem.price == None),
        SalesOrderItem.product_id != None
    ).all()

    for item in items:
        product = ProductInfo.query.get(item.product_id)
        if product and product.sale_price:
            item.price = float(product.sale_price)
            item.amount = int(item.quantity or 0) * float(product.sale_price)
            fixed_count += 1
            fixed_order_ids.add(item.order_id)

    # 重新计算受影响订单的总金额
    for order_id in fixed_order_ids:
        order_items = SalesOrderItem.query.filter_by(order_id=order_id).all()
        total = sum(float(i.amount or 0) for i in order_items)
        total_qty = sum(int(i.quantity or 0) for i in order_items)
        order = SalesOrder.query.get(order_id)
        if order:
            order.total_amount = total
            order.actual_amount = total - float(order.discount_amount or 0) + float(order.freight_amount or 0)
            order.total_quantity = total_qty

    db.session.commit()
    return jsonify({'code': 200, 'message': f'修复完成，共修复 {fixed_count} 条明细，更新 {len(fixed_order_ids)} 个销售单'})


@bp.route('/api/sales/returns/<int:return_id>/unbind-assets', methods=['POST'])
@jwt_required()
def unbind_assets_for_return(return_id):
    """销售退货解绑资产"""
    try:
        # 跨蓝图依赖：Asset（assets 蓝图尚未迁移）/ ReturnOrder（purchase 子包）
        try:
            from models.asset import Asset
        except Exception:
            return jsonify({'code': 503, 'message': '资产模块尚未启用'}), 503

        try:
            from models.purchase.return_order import ReturnOrder
        except Exception:
            return jsonify({'code': 503, 'message': '退货单模块尚未启用'}), 503

        return_order = ReturnOrder.query.get(return_id)
        if not return_order:
            return jsonify({'code': 404, 'message': '退货单不存在'}), 404

        # 找到原销售单
        if not return_order.source_order_id:
            return jsonify({'code': 400, 'message': '退货单未关联销售单'}), 400

        # 解除该销售单关联的所有资产的关联关系
        assets = Asset.query.filter_by(sales_order_id=return_order.source_order_id).all()
        unbind_count = 0
        for asset in assets:
            asset.sales_order_id = None
            asset.sales_order_no = None
            unbind_count += 1

        db.session.commit()

        return jsonify({
            'code': 200,
            'message': f'成功解绑{unbind_count}个资产',
            'data': {'unbind_count': unbind_count}
        })
    except Exception as e:
        logger.error(f'退货解绑资产失败: {str(e)}')
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'解绑失败: {str(e)}'}), 500
