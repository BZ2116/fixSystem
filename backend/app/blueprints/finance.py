"""财务管理蓝图（账户 + 流水 + 应收 + 应付 + 通用发票）。

迁移自 source-code/app.py 中以下区段的原始路由代码：
- /api/finance/accounts (含 transfer/adjust)     4245, 10757-10996
- /api/finance/records    (含 enhanced)         4262, 10997
- /api/finance/receivables (含 batch/print/summary/export) 13003-13170, 13814, 14042, 14172
- /api/finance/payables   (含 batch/print/summary/export) 13247-13414, 13928, 14107, 14229
- /api/invoices           (含 cancel/red-flush)  7856-7995

业务规则：
- 账户转账 (transfer) 是事务性的：源账户扣减 + 目标账户增加 + 两条 FinanceRecord
- 账户调整 (adjust) 修改余额并写入一条 FinanceRecord
- 收款 (receive) 更新 received_amount / remaining_amount / status
- 付款 (pay)  对称更新 paid_amount / remaining_amount / status
- summary: 按月统计应收/应付
- /api/invoices 是 finance_invoice 表的通用 CRUD（不含销售/采购发票）

跨子域依赖：
- FinanceAccount / Record / Receivable / Payable / Invoice  (models.finance.account)
- SysUser         (models.system)
- BaseCustomer    (models.customer) — 应收单关联客户
- BaseSupplier    (models.supplier) — 应付单关联供应商

直接 import 当前蓝图用得到的本地 helper，其他跨子域模型按"函数内懒加载"惯例。
"""
import io
import logging
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import jwt_required
import openpyxl

from extensions import db
from app.utils import to_dict
from app.security import get_current_user_id, permission

logger = logging.getLogger(__name__)

bp = Blueprint('finance', __name__)


# ============================================
# 工具函数
# ============================================

def _get_current_user_name():
    """获取当前登录用户姓名。迁移期兼容（app.py 174 行）。"""
    from flask_jwt_extended import get_jwt_identity
    from models.system import SysUser
    user_id = get_jwt_identity()
    user = SysUser.query.get(user_id)
    if user:
        return user.real_name or user.username
    return ''


# ============================================
# Section A: 财务账户 (/api/finance/accounts)
# ============================================

@bp.route('/api/finance/accounts', methods=['GET'])
@permission('finance:view')
def get_finance_accounts():
    """获取财务账户列表"""
    from models.finance.account import FinanceAccount

    accounts = FinanceAccount.query.filter_by(status=1).all()
    return jsonify({
        'code': 200,
        'data': [{
            'id': a.id,
            'account_name': a.account_name,
            'account_type': a.account_type,
            'account_no': a.account_no,
            'balance': float(a.balance) if a.balance else 0.00,
            'remark': a.remark
        } for a in accounts]
    })


@bp.route('/api/finance/accounts', methods=['POST'])
@permission('finance:view', 'finance-receivable:edit')
def create_finance_account():
    """创建财务账户"""
    from models.finance.account import FinanceAccount

    data = request.get_json()

    account = FinanceAccount(
        account_name=data.get('account_name'),
        account_type=data.get('account_type', 1),
        account_no=data.get('account_no'),
        balance=data.get('balance', 0),
        remark=data.get('remark')
    )

    db.session.add(account)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': account.id}})


@bp.route('/api/finance/accounts/<int:id>', methods=['PUT'])
@permission('finance:view', 'finance-receivable:edit')
def update_finance_account(id):
    """更新财务账户"""
    from models.finance.account import FinanceAccount

    account = FinanceAccount.query.get(id)
    if not account:
        return jsonify({'code': 404, 'message': '账户不存在'}), 404

    data = request.get_json()

    for field in ['account_name', 'account_type', 'account_no', 'remark']:
        if field in data:
            setattr(account, field, data[field])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@bp.route('/api/finance/accounts/<int:id>', methods=['DELETE'])
@permission('finance:view', 'finance-receivable:edit')
def delete_finance_account(id):
    """删除财务账户"""
    from models.finance.account import FinanceAccount

    account = FinanceAccount.query.get(id)
    if not account:
        return jsonify({'code': 404, 'message': '账户不存在'}), 404

    # 软删除
    account.status = 0
    db.session.commit()
    return jsonify({'code': 200, 'message': '删除成功'})


@bp.route('/api/finance/accounts/transfer', methods=['POST'])
@permission('finance:view', 'finance-payment:edit', 'finance-receipt:edit')
def transfer_finance_account():
    """账户转账"""
    from flask_jwt_extended import get_jwt_identity
    from models.finance.account import FinanceAccount, FinanceRecord

    try:
        current_user_id = get_jwt_identity()
        current_user_name = _get_current_user_name()
        data = request.get_json()
        if not data:
            return jsonify({'code': 400, 'message': '请求数据不能为空'}), 400

        from_account_id = data.get('from_account_id')
        to_account_id = data.get('to_account_id')
        amount = data.get('amount')
        remark = data.get('remark', '')

        if not from_account_id or not to_account_id:
            return jsonify({'code': 400, 'message': '转出账户和转入账户不能为空'}), 400
        if from_account_id == to_account_id:
            return jsonify({'code': 400, 'message': '转出账户和转入账户不能相同'}), 400
        if not amount or float(amount) <= 0:
            return jsonify({'code': 400, 'message': '转账金额必须大于0'}), 400

        amount_val = float(amount)

        from_account = FinanceAccount.query.get(from_account_id)
        to_account = FinanceAccount.query.get(to_account_id)

        if not from_account:
            return jsonify({'code': 404, 'message': '转出账户不存在'}), 404
        if not to_account:
            return jsonify({'code': 404, 'message': '转入账户不存在'}), 404
        if from_account.status != 1:
            return jsonify({'code': 400, 'message': '转出账户已停用'}), 400
        if to_account.status != 1:
            return jsonify({'code': 400, 'message': '转入账户已停用'}), 400
        if float(from_account.balance) < amount_val:
            return jsonify({'code': 400, 'message': '转出账户余额不足'}), 400

        related_no = 'TR' + str(int(datetime.now().timestamp()))

        # 扣减转出账户余额
        from_balance_before = float(from_account.balance)
        from_account.balance = from_balance_before - amount_val

        # 增加转入账户余额
        to_balance_before = float(to_account.balance)
        to_account.balance = to_balance_before + amount_val

        # 生成转出流水记录（支出）
        from_record = FinanceRecord(
            account_id=from_account.id,
            account_name=from_account.account_name,
            record_type=2,
            amount=amount_val,
            balance_before=from_balance_before,
            balance_after=from_balance_before - amount_val,
            related_type='transfer',
            related_id=to_account.id,
            related_no=related_no,
            remark=f'转账至 {to_account.account_name}' + (f'：{remark}' if remark else ''),
            created_at=datetime.now(),
            created_by=current_user_id
        )

        # 生成转入流水记录（收入）
        to_record = FinanceRecord(
            account_id=to_account.id,
            account_name=to_account.account_name,
            record_type=1,
            amount=amount_val,
            balance_before=to_balance_before,
            balance_after=to_balance_before + amount_val,
            related_type='transfer',
            related_id=from_account.id,
            related_no=related_no,
            remark=f'从 {from_account.account_name} 转入' + (f'：{remark}' if remark else ''),
            created_at=datetime.now(),
            created_by=current_user_id
        )

        db.session.add(from_record)
        db.session.add(to_record)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '转账成功',
            'data': {
                'related_no': related_no,
                'from_account': {
                    'id': from_account.id,
                    'account_name': from_account.account_name,
                    'balance_after': float(from_account.balance)
                },
                'to_account': {
                    'id': to_account.id,
                    'account_name': to_account.account_name,
                    'balance_after': float(to_account.balance)
                }
            }
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'账户转账失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'转账失败: {str(e)}'}), 500


@bp.route('/api/finance/accounts/<int:id>/adjust', methods=['POST'])
@permission('finance:view', 'finance-receivable:edit')
def adjust_finance_account(id):
    """余额调整（盘盈/盘亏）"""
    from flask_jwt_extended import get_jwt_identity
    from models.finance.account import FinanceAccount, FinanceRecord

    try:
        current_user_id = get_jwt_identity()
        current_user_name = _get_current_user_name()
        data = request.get_json()
        if not data:
            return jsonify({'code': 400, 'message': '请求数据不能为空'}), 400

        adjust_amount = data.get('adjust_amount')
        remark = data.get('remark', '')

        if adjust_amount is None:
            return jsonify({'code': 400, 'message': '调整金额不能为空'}), 400

        adjust_amount_val = float(adjust_amount)
        if adjust_amount_val == 0:
            return jsonify({'code': 400, 'message': '调整金额不能为0'}), 400

        account = FinanceAccount.query.get(id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404
        if account.status != 1:
            return jsonify({'code': 400, 'message': '账户已停用，无法调整'}), 400

        balance_before = float(account.balance)
        balance_after = balance_before + adjust_amount_val

        # 余额不能为负数
        if balance_after < 0:
            return jsonify({'code': 400, 'message': '调整后余额不能为负数'}), 400

        account.balance = balance_after

        # 确定流水类型：正数为收入（盘盈），负数为支出（盘亏）
        record_type = 1 if adjust_amount_val > 0 else 2
        adjust_type_text = '盘盈' if adjust_amount_val > 0 else '盘亏'

        record = FinanceRecord(
            account_id=account.id,
            account_name=account.account_name,
            record_type=record_type,
            amount=abs(adjust_amount_val),
            balance_before=balance_before,
            balance_after=balance_after,
            related_type='adjust',
            related_id=id,
            related_no='AD' + str(int(datetime.now().timestamp())),
            remark=f'{adjust_type_text}调整' + (f'：{remark}' if remark else ''),
            created_at=datetime.now(),
            created_by=current_user_id
        )

        db.session.add(record)
        db.session.commit()

        return jsonify({
            'code': 200,
            'message': '余额调整成功',
            'data': {
                'id': account.id,
                'account_name': account.account_name,
                'balance_before': balance_before,
                'balance_after': balance_after,
                'adjust_amount': adjust_amount_val,
                'adjust_type': adjust_type_text
            }
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f'余额调整失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'余额调整失败: {str(e)}'}), 500


# ============================================
# Section B: 财务流水 (/api/finance/records)
# ============================================

@bp.route('/api/finance/records', methods=['GET'])
@permission('finance:view', 'finance-receivable:view', 'finance-payable:view')
def get_finance_records():
    """获取财务流水"""
    from models.finance.account import FinanceRecord, FinanceReceivable, FinancePayable
    from models.system import SysUser
    from models.sales.order import SalesOrder
    from models.workorder import WorkOrder
    from models.customer import BaseCustomer

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    account_id = request.args.get('account_id', type=int)
    record_type = request.args.get('record_type', type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    keyword = request.args.get('keyword', '')

    query = FinanceRecord.query

    if account_id:
        query = query.filter_by(account_id=account_id)

    if record_type:
        query = query.filter_by(record_type=record_type)

    if start_date:
        query = query.filter(FinanceRecord.created_at >= start_date)

    if end_date:
        query = query.filter(FinanceRecord.created_at <= end_date + ' 23:59:59')

    if keyword:
        keyword_like = f'%{keyword}%'
        query = query.filter(
            db.or_(
                FinanceRecord.account_name.like(keyword_like),
                FinanceRecord.remark.like(keyword_like),
                FinanceRecord.related_no.like(keyword_like)
            )
        )

    total = query.count()
    records = query.order_by(FinanceRecord.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    # 构建返回数据，补充关联信息
    result_list = []
    for r in records.items:
        item = to_dict(r)
        # 补充操作人名称
        if r.created_by:
            creator = SysUser.query.get(r.created_by)
            item['created_by_name'] = creator.real_name if creator else ''
        else:
            item['created_by_name'] = ''

        # 补充客户/供应商名称
        item['customer_name'] = ''
        item['supplier_name'] = ''
        if r.related_type == 'sale' and r.related_id:
            # 从应收记录获取客户名称
            receivable = FinanceReceivable.query.filter_by(related_id=r.related_id, related_type='sale').first()
            if receivable and receivable.customer_name:
                item['customer_name'] = receivable.customer_name
            elif r.related_id:
                # 从备注中提取或关联查询
                sale_order = db.session.query(SalesOrder).get(r.related_id) if hasattr(db.Model, 'SaleOrder') else None
        elif r.related_type == 'purchase' and r.related_id:
            payable = FinancePayable.query.filter_by(related_id=r.related_id, related_type='purchase').first()
            if payable and payable.supplier_name:
                item['supplier_name'] = payable.supplier_name
        elif r.related_type == 'work_order' and r.related_id:
            wo = WorkOrder.query.get(r.related_id)
            if wo and wo.customer_id:
                customer = BaseCustomer.query.get(wo.customer_id)
                if customer:
                    item['customer_name'] = customer.customer_name

        result_list.append(item)

    return jsonify({
        'code': 200,
        'data': {
            'list': result_list,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/finance/records/enhanced', methods=['GET'])
@permission('finance:view', 'finance-receivable:view', 'finance-payable:view')
def get_finance_records_enhanced():
    """增强版流水查询"""
    from models.finance.account import FinanceRecord, FinanceAccount

    try:
        account_id = request.args.get('account_id', type=int)
        record_type = request.args.get('record_type', type=int)
        related_type = request.args.get('related_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        keyword = request.args.get('keyword')
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)

        if page < 1:
            page = 1
        if page_size < 1 or page_size > 100:
            page_size = 20

        query = FinanceRecord.query

        if account_id:
            query = query.filter(FinanceRecord.account_id == account_id)
        if record_type:
            query = query.filter(FinanceRecord.record_type == record_type)
        if related_type:
            query = query.filter(FinanceRecord.related_type == related_type)
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(FinanceRecord.created_at >= start_dt)
            except ValueError:
                return jsonify({'code': 400, 'message': '开始日期格式不正确，应为YYYY-MM-DD'}), 400
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                # 包含当天全天，设置到23:59:59
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
                query = query.filter(FinanceRecord.created_at <= end_dt)
            except ValueError:
                return jsonify({'code': 400, 'message': '结束日期格式不正确，应为YYYY-MM-DD'}), 400
        if keyword:
            keyword_like = f'%{keyword}%'
            query = query.filter(
                db.or_(
                    FinanceRecord.account_name.like(keyword_like),
                    FinanceRecord.remark.like(keyword_like),
                    FinanceRecord.related_no.like(keyword_like)
                )
            )

        query = query.order_by(FinanceRecord.created_at.desc())

        total = query.count()
        records = query.offset((page - 1) * page_size).limit(page_size).all()

        account_type_map = {1: '现金', 2: '银行', 3: '支付宝', 4: '微信'}
        record_type_map = {1: '收入', 2: '支出'}

        items = []
        for record in records:
            # 查找账户类型
            account = FinanceAccount.query.get(record.account_id)
            account_type_text = ''
            if account:
                account_type_text = account_type_map.get(account.account_type, '未知')

            items.append({
                'id': record.id,
                'account_id': record.account_id,
                'account_name': record.account_name,
                'account_type_text': account_type_text,
                'record_type': record.record_type,
                'record_type_text': record_type_map.get(record.record_type, '未知'),
                'amount': float(record.amount) if record.amount else 0,
                'balance_before': float(record.balance_before) if record.balance_before else 0,
                'balance_after': float(record.balance_after) if record.balance_after else 0,
                'related_type': record.related_type,
                'related_id': record.related_id,
                'related_no': record.related_no,
                'remark': record.remark,
                'created_at': record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else None,
                'created_by': record.created_by
            })

        return jsonify({
            'code': 200,
            'message': '查询成功',
            'data': {
                'total': total,
                'page': page,
                'page_size': page_size,
                'items': items
            }
        })
    except Exception as e:
        logger.error(f'增强版流水查询失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'查询失败: {str(e)}'}), 500


# ============================================
# Section C: 应收账款 (/api/finance/receivables)
# ============================================

@bp.route('/api/finance/receivables', methods=['GET'])
@permission('finance:view', 'finance-receivable:view')
def get_receivables():
    """应收账款列表"""
    from models.finance.account import FinanceReceivable

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    customer_id = request.args.get('customer_id', type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = FinanceReceivable.query

    if customer_id:
        query = query.filter_by(customer_id=customer_id)

    if status is not None:
        query = query.filter_by(status=status)

    if keyword:
        query = query.filter(
            db.or_(
                FinanceReceivable.receivable_no.contains(keyword),
                FinanceReceivable.customer_name.contains(keyword),
                FinanceReceivable.related_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(FinanceReceivable.created_at >= start_date)

    if end_date:
        query = query.filter(FinanceReceivable.created_at <= end_date + ' 23:59:59')

    total = query.count()
    receivables = query.order_by(FinanceReceivable.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    status_map = {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': r.id,
                'receivable_no': r.receivable_no,
                'related_type': r.related_type,
                'related_id': r.related_id,
                'related_no': r.related_no,
                'customer_id': r.customer_id,
                'customer_name': r.customer_name,
                'total_amount': float(r.total_amount) if r.total_amount else 0,
                'received_amount': float(r.received_amount) if r.received_amount else 0,
                'remaining_amount': float(r.remaining_amount) if r.remaining_amount else 0,
                'status': r.status,
                'status_text': status_map.get(r.status, '未知'),
                'invoice_id': r.invoice_id,
                'remark': r.remark,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None,
                'updated_at': r.updated_at.strftime('%Y-%m-%d %H:%M:%S') if r.updated_at else None
            } for r in receivables.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/finance/receivables/<int:id>', methods=['GET'])
@permission('finance:view', 'finance-receivable:view')
def get_receivable(id):
    """应收账款详情"""
    from models.finance.account import FinanceReceivable

    r = FinanceReceivable.query.get(id)
    if not r:
        return jsonify({'code': 404, 'message': '应收记录不存在'}), 404

    status_map = {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'id': r.id,
            'receivable_no': r.receivable_no,
            'related_type': r.related_type,
            'related_id': r.related_id,
            'related_no': r.related_no,
            'customer_id': r.customer_id,
            'customer_name': r.customer_name,
            'total_amount': float(r.total_amount) if r.total_amount else 0,
            'received_amount': float(r.received_amount) if r.received_amount else 0,
            'remaining_amount': float(r.remaining_amount) if r.remaining_amount else 0,
            'status': r.status,
            'status_text': status_map.get(r.status, '未知'),
            'invoice_id': r.invoice_id,
            'remark': r.remark,
            'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else None,
            'updated_at': r.updated_at.strftime('%Y-%m-%d %H:%M:%S') if r.updated_at else None
        }
    })


@bp.route('/api/finance/receivables/<int:id>/receive', methods=['POST'])
@permission('finance:view', 'finance-receipt:edit')
def receive_receivable(id):
    """收款核销"""
    from flask_jwt_extended import get_jwt_identity
    from models.finance.account import FinanceReceivable, FinanceRecord

    receivable = FinanceReceivable.query.get(id)
    if not receivable:
        return jsonify({'code': 404, 'message': '应收记录不存在'}), 404

    if receivable.status == 2:
        return jsonify({'code': 400, 'message': '该应收已结清'}), 400

    if receivable.status == 3:
        return jsonify({'code': 400, 'message': '该应收已取消'}), 400

    data = request.get_json()
    received_amount = data.get('received_amount')
    if not received_amount or float(received_amount) <= 0:
        return jsonify({'code': 400, 'message': '收款金额必须大于0'}), 400

    received_amount = float(received_amount)
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    try:
        # 更新应收金额
        receivable.received_amount = float(receivable.received_amount or 0) + received_amount
        receivable.remaining_amount = float(receivable.total_amount or 0) - float(receivable.received_amount or 0)

        if receivable.remaining_amount <= 0:
            receivable.remaining_amount = 0
            receivable.status = 2  # 已结清
        else:
            receivable.status = 1  # 部分收款

        receivable.updated_at = datetime.now()

        # 生成财务流水（收入）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=1,  # 收入
            amount=received_amount,
            balance_before=0,
            balance_after=0,
            related_type='receivable',
            related_id=id,
            related_no=receivable.receivable_no,
            remark=f'应收{receivable.receivable_no}收款{received_amount}元，操作人：{user_name}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        db.session.commit()
        return jsonify({'code': 200, 'message': '收款成功', 'data': {
            'received_amount': float(receivable.received_amount),
            'remaining_amount': float(receivable.remaining_amount),
            'status': receivable.status
        }})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'收款失败：{str(e)}'}), 500


@bp.route('/api/finance/receivables/export', methods=['GET'])
@permission('finance:view', 'finance-receivable:view')
def export_receivables():
    """导出应收账款"""
    from models.finance.account import FinanceReceivable

    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)
        customer_id = request.args.get('customer_id', type=int)
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        query = FinanceReceivable.query

        if customer_id:
            query = query.filter_by(customer_id=customer_id)
        if status is not None:
            query = query.filter_by(status=status)
        if keyword:
            query = query.filter(
                db.or_(
                    FinanceReceivable.receivable_no.contains(keyword),
                    FinanceReceivable.customer_name.contains(keyword),
                    FinanceReceivable.related_no.contains(keyword)
                )
            )
        if start_date:
            query = query.filter(FinanceReceivable.created_at >= start_date)
        if end_date:
            query = query.filter(FinanceReceivable.created_at <= end_date + ' 23:59:59')

        receivables = query.order_by(FinanceReceivable.created_at.desc()).all()

        status_map = {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '应收账款'

        headers = ['应收编号', '关联类型', '关联单号', '客户名称', '应收总额', '已收金额', '待收金额', '状态', '备注', '创建时间']
        ws.append(headers)

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for r in receivables:
            ws.append([
                r.receivable_no,
                r.related_type,
                r.related_no,
                r.customer_name,
                float(r.total_amount) if r.total_amount else 0,
                float(r.received_amount) if r.received_amount else 0,
                float(r.remaining_amount) if r.remaining_amount else 0,
                status_map.get(r.status, '未知'),
                r.remark or '',
                r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                         download_name=f'应收账款_{datetime.now().strftime("%Y%m%d")}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导出失败：{str(e)}'}), 500


@bp.route('/api/finance/receivables/batch-receive', methods=['POST'])
@permission('finance:view', 'finance-receipt:edit')
def batch_receive_receivables():
    """批量收款"""
    from flask_jwt_extended import get_jwt_identity
    from models.finance.account import FinanceReceivable, FinanceAccount, FinanceRecord

    data = request.get_json()
    items = data.get('items', [])
    account_id = data.get('account_id')
    remark = data.get('remark', '')

    if not items:
        return jsonify({'code': 400, 'message': '请选择要收款的应收记录'}), 400

    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    # 如果提供了account_id，查询账户信息
    account = None
    if account_id:
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

    success_items = []
    failed_items = []

    try:
        for item in items:
            try:
                receivable_id = item.get('id')
                amount = item.get('amount')

                if not receivable_id or not amount or float(amount) <= 0:
                    failed_items.append({'id': receivable_id, 'reason': '金额无效'})
                    continue

                amount = float(amount)
                receivable = FinanceReceivable.query.get(receivable_id)
                if not receivable:
                    failed_items.append({'id': receivable_id, 'reason': '应收记录不存在'})
                    continue

                if receivable.status == 2:
                    failed_items.append({'id': receivable_id, 'reason': '该应收已结清'})
                    continue

                if receivable.status == 3:
                    failed_items.append({'id': receivable_id, 'reason': '该应收已取消'})
                    continue

                if amount > float(receivable.remaining_amount or 0):
                    failed_items.append({'id': receivable_id, 'reason': f'收款金额不能大于待收金额{receivable.remaining_amount}'})
                    continue

                # 更新应收金额
                receivable.received_amount = float(receivable.received_amount or 0) + amount
                receivable.remaining_amount = float(receivable.total_amount or 0) - float(receivable.received_amount or 0)

                if receivable.remaining_amount <= 0:
                    receivable.remaining_amount = 0
                    receivable.status = 2  # 已结清
                else:
                    receivable.status = 1  # 部分收款

                receivable.updated_at = datetime.now()

                # 如果关联了账户，更新账户余额
                balance_before = 0
                balance_after = 0
                if account:
                    balance_before = float(account.balance or 0)
                    account.balance = balance_before + amount
                    balance_after = float(account.balance)

                # 生成财务流水（收入）
                finance_record = FinanceRecord(
                    account_id=account_id,
                    account_name=account.account_name if account else '',
                    record_type=1,  # 收入
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=balance_after,
                    related_type='receivable',
                    related_id=receivable_id,
                    related_no=receivable.receivable_no,
                    remark=f'应收{receivable.receivable_no}收款{amount}元，操作人：{user_name}{("，备注：" + remark) if remark else ""}',
                    created_at=datetime.now(),
                    created_by=user_id
                )
                db.session.add(finance_record)

                success_items.append({
                    'id': receivable_id,
                    'receivable_no': receivable.receivable_no,
                    'amount': amount,
                    'remaining_amount': float(receivable.remaining_amount),
                    'status': receivable.status
                })

            except Exception as e:
                logger.error(f'批量收款-单条处理失败: {str(e)}')
                failed_items.append({'id': item.get('id'), 'reason': str(e)})

        db.session.commit()
        return jsonify({'code': 200, 'message': f'批量收款完成，成功{len(success_items)}条，失败{len(failed_items)}条', 'data': {
            'success_items': success_items,
            'failed_items': failed_items
        }})

    except Exception as e:
        db.session.rollback()
        logger.error(f'批量收款失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'批量收款失败：{str(e)}'}), 500


@bp.route('/api/finance/receivables/<int:id>/print', methods=['GET'])
@permission('finance:view', 'finance-receivable:view')
def print_receivable(id):
    """应收账款打印数据"""
    from models.finance.account import FinanceReceivable, FinanceRecord

    receivable = FinanceReceivable.query.get(id)
    if not receivable:
        return jsonify({'code': 404, 'message': '应收记录不存在'}), 404

    try:
        # 查询收款记录
        records = FinanceRecord.query.filter_by(
            related_type='receivable',
            related_id=id
        ).order_by(FinanceRecord.created_at.asc()).all()

        record_list = []
        for r in records:
            record_list.append({
                'id': r.id,
                'amount': float(r.amount),
                'account_name': r.account_name or '',
                'remark': r.remark or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
                'created_by': r.created_by
            })

        # 计算逾期天数
        overdue_days = (datetime.now().date() - receivable.due_date).days if receivable.due_date and receivable.due_date < datetime.now().date() else 0

        # 公司信息（默认值）
        company_info = {
            'company_name': '',
            'address': '',
            'phone': '',
            'tax_no': ''
        }

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'receivable': {
                'id': receivable.id,
                'receivable_no': receivable.receivable_no,
                'related_type': receivable.related_type,
                'related_no': receivable.related_no,
                'customer_id': receivable.customer_id,
                'customer_name': receivable.customer_name,
                'total_amount': float(receivable.total_amount or 0),
                'received_amount': float(receivable.received_amount or 0),
                'remaining_amount': float(receivable.remaining_amount or 0),
                'status': receivable.status,
                'status_text': {0: '待收款', 1: '部分收款', 2: '已结清', 3: '已取消'}.get(receivable.status, ''),
                'due_date': receivable.due_date.strftime('%Y-%m-%d') if receivable.due_date else '',
                'remark': receivable.remark or '',
                'created_at': receivable.created_at.strftime('%Y-%m-%d %H:%M:%S') if receivable.created_at else '',
                'updated_at': receivable.updated_at.strftime('%Y-%m-%d %H:%M:%S') if receivable.updated_at else ''
            },
            'company_info': company_info,
            'records': record_list,
            'overdue_days': overdue_days
        }})

    except Exception as e:
        logger.error(f'获取应收打印数据失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取打印数据失败：{str(e)}'}), 500


@bp.route('/api/finance/receivables/summary', methods=['GET'])
@permission('finance:view', 'finance-receivable:view', 'finance-statistics:view')
def receivables_summary():
    """应收统计汇总"""
    from models.finance.account import FinanceReceivable, FinanceRecord

    try:
        today = datetime.now().date()

        # 应收总额：所有待收款+部分收款的remaining_amount之和
        total_receivable = db.session.query(
            db.func.coalesce(db.func.sum(FinanceReceivable.remaining_amount), 0)
        ).filter(
            FinanceReceivable.status.in_([0, 1])
        ).scalar()
        total_receivable = float(total_receivable or 0)

        # 逾期总额：due_date < 今天 且 status != 2 且 status != 3 的remaining_amount之和
        total_overdue = db.session.query(
            db.func.coalesce(db.func.sum(FinanceReceivable.remaining_amount), 0)
        ).filter(
            FinanceReceivable.status.in_([0, 1]),
            FinanceReceivable.due_date != None,
            FinanceReceivable.due_date < today
        ).scalar()
        total_overdue = float(total_overdue or 0)

        # 逾期笔数
        overdue_count = FinanceReceivable.query.filter(
            FinanceReceivable.status.in_([0, 1]),
            FinanceReceivable.due_date != None,
            FinanceReceivable.due_date < today
        ).count()

        # 今日收款总额
        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today, datetime.max.time())
        total_received_today = db.session.query(
            db.func.coalesce(db.func.sum(FinanceRecord.amount), 0)
        ).filter(
            FinanceRecord.related_type == 'receivable',
            FinanceRecord.record_type == 1,
            FinanceRecord.created_at >= today_start,
            FinanceRecord.created_at <= today_end
        ).scalar()
        total_received_today = float(total_received_today or 0)

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'total_receivable': total_receivable,
            'total_overdue': total_overdue,
            'overdue_count': overdue_count,
            'total_received_today': total_received_today
        }})

    except Exception as e:
        logger.error(f'获取应收统计汇总失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取统计汇总失败：{str(e)}'}), 500


# ============================================
# Section D: 应付账款 (/api/finance/payables)
# ============================================

@bp.route('/api/finance/payables', methods=['GET'])
@permission('finance:view', 'finance-payable:view')
def get_payables():
    """应付账款列表"""
    from models.finance.account import FinancePayable

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    supplier_id = request.args.get('supplier_id', type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    query = FinancePayable.query

    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)

    if status is not None:
        query = query.filter_by(status=status)

    if keyword:
        query = query.filter(
            db.or_(
                FinancePayable.payable_no.contains(keyword),
                FinancePayable.supplier_name.contains(keyword),
                FinancePayable.related_no.contains(keyword)
            )
        )

    if start_date:
        query = query.filter(FinancePayable.created_at >= start_date)

    if end_date:
        query = query.filter(FinancePayable.created_at <= end_date + ' 23:59:59')

    total = query.count()
    payables = query.order_by(FinancePayable.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    status_map = {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'list': [{
                'id': p.id,
                'payable_no': p.payable_no,
                'related_type': p.related_type,
                'related_id': p.related_id,
                'related_no': p.related_no,
                'supplier_id': p.supplier_id,
                'supplier_name': p.supplier_name,
                'total_amount': float(p.total_amount) if p.total_amount else 0,
                'paid_amount': float(p.paid_amount) if p.paid_amount else 0,
                'remaining_amount': float(p.remaining_amount) if p.remaining_amount else 0,
                'status': p.status,
                'status_text': status_map.get(p.status, '未知'),
                'invoice_id': p.invoice_id,
                'remark': p.remark,
                'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None,
                'updated_at': p.updated_at.strftime('%Y-%m-%d %H:%M:%S') if p.updated_at else None
            } for p in payables.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/finance/payables/<int:id>', methods=['GET'])
@permission('finance:view', 'finance-payable:view')
def get_payable(id):
    """应付账款详情"""
    from models.finance.account import FinancePayable

    p = FinancePayable.query.get(id)
    if not p:
        return jsonify({'code': 404, 'message': '应付记录不存在'}), 404

    status_map = {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}

    return jsonify({
        'code': 200,
        'data': {
            'id': p.id,
            'payable_no': p.payable_no,
            'related_type': p.related_type,
            'related_id': p.related_id,
            'related_no': p.related_no,
            'supplier_id': p.supplier_id,
            'supplier_name': p.supplier_name,
            'total_amount': float(p.total_amount) if p.total_amount else 0,
            'paid_amount': float(p.paid_amount) if p.paid_amount else 0,
            'remaining_amount': float(p.remaining_amount) if p.remaining_amount else 0,
            'status': p.status,
            'status_text': status_map.get(p.status, '未知'),
            'invoice_id': p.invoice_id,
            'remark': p.remark,
            'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None,
            'updated_at': p.updated_at.strftime('%Y-%m-%d %H:%M:%S') if p.updated_at else None
        }
    })


@bp.route('/api/finance/payables/<int:id>/pay', methods=['POST'])
@permission('finance:view', 'finance-payment:edit')
def pay_payable(id):
    """付款核销"""
    from flask_jwt_extended import get_jwt_identity
    from models.finance.account import FinancePayable, FinanceRecord

    payable = FinancePayable.query.get(id)
    if not payable:
        return jsonify({'code': 404, 'message': '应付记录不存在'}), 404

    if payable.status == 2:
        return jsonify({'code': 400, 'message': '该应付已结清'}), 400

    if payable.status == 3:
        return jsonify({'code': 400, 'message': '该应付已取消'}), 400

    data = request.get_json()
    paid_amount = data.get('paid_amount')
    if not paid_amount or float(paid_amount) <= 0:
        return jsonify({'code': 400, 'message': '付款金额必须大于0'}), 400

    paid_amount = float(paid_amount)
    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    try:
        # 更新应付金额
        payable.paid_amount = float(payable.paid_amount or 0) + paid_amount
        payable.remaining_amount = float(payable.total_amount or 0) - float(payable.paid_amount or 0)

        if payable.remaining_amount <= 0:
            payable.remaining_amount = 0
            payable.status = 2  # 已结清
        else:
            payable.status = 1  # 部分付款

        payable.updated_at = datetime.now()

        # 生成财务流水（支出）
        finance_record = FinanceRecord(
            account_id=None,
            account_name='',
            record_type=2,  # 支出
            amount=paid_amount,
            balance_before=0,
            balance_after=0,
            related_type='payable',
            related_id=id,
            related_no=payable.payable_no,
            remark=f'应付{payable.payable_no}付款{paid_amount}元，操作人：{user_name}',
            created_at=datetime.now(),
            created_by=user_id
        )
        db.session.add(finance_record)

        db.session.commit()
        return jsonify({'code': 200, 'message': '付款成功', 'data': {
            'paid_amount': float(payable.paid_amount),
            'remaining_amount': float(payable.remaining_amount),
            'status': payable.status
        }})

    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'付款失败：{str(e)}'}), 500


@bp.route('/api/finance/payables/export', methods=['GET'])
@permission('finance:view', 'finance-payable:view')
def export_payables():
    """导出应付账款"""
    from models.finance.account import FinancePayable

    try:
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', type=int)
        supplier_id = request.args.get('supplier_id', type=int)
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')

        query = FinancePayable.query

        if supplier_id:
            query = query.filter_by(supplier_id=supplier_id)
        if status is not None:
            query = query.filter_by(status=status)
        if keyword:
            query = query.filter(
                db.or_(
                    FinancePayable.payable_no.contains(keyword),
                    FinancePayable.supplier_name.contains(keyword),
                    FinancePayable.related_no.contains(keyword)
                )
            )
        if start_date:
            query = query.filter(FinancePayable.created_at >= start_date)
        if end_date:
            query = query.filter(FinancePayable.created_at <= end_date + ' 23:59:59')

        payables = query.order_by(FinancePayable.created_at.desc()).all()

        status_map = {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '应付账款'

        headers = ['应付编号', '关联类型', '关联单号', '供应商名称', '应付总额', '已付金额', '待付金额', '状态', '备注', '创建时间']
        ws.append(headers)

        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for p in payables:
            ws.append([
                p.payable_no,
                p.related_type,
                p.related_no,
                p.supplier_name,
                float(p.total_amount) if p.total_amount else 0,
                float(p.paid_amount) if p.paid_amount else 0,
                float(p.remaining_amount) if p.remaining_amount else 0,
                status_map.get(p.status, '未知'),
                p.remark or '',
                p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else ''
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True,
                         download_name=f'应付账款_{datetime.now().strftime("%Y%m%d")}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导出失败：{str(e)}'}), 500


@bp.route('/api/finance/payables/batch-pay', methods=['POST'])
@permission('finance:view', 'finance-payment:edit')
def batch_pay_payables():
    """批量付款"""
    from flask_jwt_extended import get_jwt_identity
    from models.finance.account import FinancePayable, FinanceAccount, FinanceRecord

    data = request.get_json()
    items = data.get('items', [])
    account_id = data.get('account_id')
    remark = data.get('remark', '')

    if not items:
        return jsonify({'code': 400, 'message': '请选择要付款的应付记录'}), 400

    user_id = get_jwt_identity()
    user_name = _get_current_user_name()

    # 如果提供了account_id，查询账户信息
    account = None
    if account_id:
        account = FinanceAccount.query.get(account_id)
        if not account:
            return jsonify({'code': 404, 'message': '账户不存在'}), 404

    success_items = []
    failed_items = []

    try:
        for item in items:
            try:
                payable_id = item.get('id')
                amount = item.get('amount')

                if not payable_id or not amount or float(amount) <= 0:
                    failed_items.append({'id': payable_id, 'reason': '金额无效'})
                    continue

                amount = float(amount)
                payable = FinancePayable.query.get(payable_id)
                if not payable:
                    failed_items.append({'id': payable_id, 'reason': '应付记录不存在'})
                    continue

                if payable.status == 2:
                    failed_items.append({'id': payable_id, 'reason': '该应付已结清'})
                    continue

                if payable.status == 3:
                    failed_items.append({'id': payable_id, 'reason': '该应付已取消'})
                    continue

                if amount > float(payable.remaining_amount or 0):
                    failed_items.append({'id': payable_id, 'reason': f'付款金额不能大于待付金额{payable.remaining_amount}'})
                    continue

                # 更新应付金额
                payable.paid_amount = float(payable.paid_amount or 0) + amount
                payable.remaining_amount = float(payable.total_amount or 0) - float(payable.paid_amount or 0)

                if payable.remaining_amount <= 0:
                    payable.remaining_amount = 0
                    payable.status = 2  # 已结清
                else:
                    payable.status = 1  # 部分付款

                payable.updated_at = datetime.now()

                # 如果关联了账户，更新账户余额
                balance_before = 0
                balance_after = 0
                if account:
                    balance_before = float(account.balance or 0)
                    account.balance = balance_before - amount
                    balance_after = float(account.balance)

                # 生成财务流水（支出）
                finance_record = FinanceRecord(
                    account_id=account_id,
                    account_name=account.account_name if account else '',
                    record_type=2,  # 支出
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=balance_after,
                    related_type='payable',
                    related_id=payable_id,
                    related_no=payable.payable_no,
                    remark=f'应付{payable.payable_no}付款{amount}元，操作人：{user_name}{("，备注：" + remark) if remark else ""}',
                    created_at=datetime.now(),
                    created_by=user_id
                )
                db.session.add(finance_record)

                success_items.append({
                    'id': payable_id,
                    'payable_no': payable.payable_no,
                    'amount': amount,
                    'remaining_amount': float(payable.remaining_amount),
                    'status': payable.status
                })

            except Exception as e:
                logger.error(f'批量付款-单条处理失败: {str(e)}')
                failed_items.append({'id': item.get('id'), 'reason': str(e)})

        db.session.commit()
        return jsonify({'code': 200, 'message': f'批量付款完成，成功{len(success_items)}条，失败{len(failed_items)}条', 'data': {
            'success_items': success_items,
            'failed_items': failed_items
        }})

    except Exception as e:
        db.session.rollback()
        logger.error(f'批量付款失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'批量付款失败：{str(e)}'}), 500


@bp.route('/api/finance/payables/<int:id>/print', methods=['GET'])
@permission('finance:view', 'finance-payable:view')
def print_payable(id):
    """应付账款打印数据"""
    from models.finance.account import FinancePayable, FinanceRecord

    payable = FinancePayable.query.get(id)
    if not payable:
        return jsonify({'code': 404, 'message': '应付记录不存在'}), 404

    try:
        # 查询付款记录
        records = FinanceRecord.query.filter_by(
            related_type='payable',
            related_id=id
        ).order_by(FinanceRecord.created_at.asc()).all()

        record_list = []
        for r in records:
            record_list.append({
                'id': r.id,
                'amount': float(r.amount),
                'account_name': r.account_name or '',
                'remark': r.remark or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else '',
                'created_by': r.created_by
            })

        # 计算逾期天数
        overdue_days = (datetime.now().date() - payable.due_date).days if payable.due_date and payable.due_date < datetime.now().date() else 0

        # 公司信息（默认值）
        company_info = {
            'company_name': '',
            'address': '',
            'phone': '',
            'tax_no': ''
        }

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'payable': {
                'id': payable.id,
                'payable_no': payable.payable_no,
                'related_type': payable.related_type,
                'related_no': payable.related_no,
                'supplier_id': payable.supplier_id,
                'supplier_name': payable.supplier_name,
                'total_amount': float(payable.total_amount or 0),
                'paid_amount': float(payable.paid_amount or 0),
                'remaining_amount': float(payable.remaining_amount or 0),
                'status': payable.status,
                'status_text': {0: '待付款', 1: '部分付款', 2: '已结清', 3: '已取消'}.get(payable.status, ''),
                'due_date': payable.due_date.strftime('%Y-%m-%d') if payable.due_date else '',
                'remark': payable.remark or '',
                'created_at': payable.created_at.strftime('%Y-%m-%d %H:%M:%S') if payable.created_at else '',
                'updated_at': payable.updated_at.strftime('%Y-%m-%d %H:%M:%S') if payable.updated_at else ''
            },
            'company_info': company_info,
            'records': record_list,
            'overdue_days': overdue_days
        }})

    except Exception as e:
        logger.error(f'获取应付打印数据失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取打印数据失败：{str(e)}'}), 500


@bp.route('/api/finance/payables/summary', methods=['GET'])
@permission('finance:view', 'finance-payable:view', 'finance-statistics:view')
def payables_summary():
    """应付统计汇总"""
    from models.finance.account import FinancePayable, FinanceRecord

    try:
        today = datetime.now().date()

        # 应付总额：所有待付款+部分付款的remaining_amount之和
        total_payable = db.session.query(
            db.func.coalesce(db.func.sum(FinancePayable.remaining_amount), 0)
        ).filter(
            FinancePayable.status.in_([0, 1])
        ).scalar()
        total_payable = float(total_payable or 0)

        # 逾期总额：due_date < 今天 且 status != 2 且 status != 3 的remaining_amount之和
        total_overdue = db.session.query(
            db.func.coalesce(db.func.sum(FinancePayable.remaining_amount), 0)
        ).filter(
            FinancePayable.status.in_([0, 1]),
            FinancePayable.due_date != None,
            FinancePayable.due_date < today
        ).scalar()
        total_overdue = float(total_overdue or 0)

        # 逾期笔数
        overdue_count = FinancePayable.query.filter(
            FinancePayable.status.in_([0, 1]),
            FinancePayable.due_date != None,
            FinancePayable.due_date < today
        ).count()

        # 今日付款总额
        today_start = datetime.combine(today, datetime.min.time())
        today_end = datetime.combine(today, datetime.max.time())
        total_paid_today = db.session.query(
            db.func.coalesce(db.func.sum(FinanceRecord.amount), 0)
        ).filter(
            FinanceRecord.related_type == 'payable',
            FinanceRecord.record_type == 2,
            FinanceRecord.created_at >= today_start,
            FinanceRecord.created_at <= today_end
        ).scalar()
        total_paid_today = float(total_paid_today or 0)

        return jsonify({'code': 200, 'message': '获取成功', 'data': {
            'total_payable': total_payable,
            'total_overdue': total_overdue,
            'overdue_count': overdue_count,
            'total_paid_today': total_paid_today
        }})

    except Exception as e:
        logger.error(f'获取应付统计汇总失败: {str(e)}')
        return jsonify({'code': 500, 'message': f'获取统计汇总失败：{str(e)}'}), 500


# ============================================
# Section E: 通用发票 (/api/invoices)
# ============================================

@bp.route('/api/invoices', methods=['GET'])
@permission('finance:view', 'finance-receivable:view')
def get_invoices():
    """获取发票列表"""
    from models.finance.account import FinanceInvoice

    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    keyword = request.args.get('keyword', '')
    invoice_type = request.args.get('invoice_type', type=int)
    status = request.args.get('status', type=int)

    query = FinanceInvoice.query

    if keyword:
        query = query.filter(
            db.or_(
                FinanceInvoice.invoice_no.contains(keyword),
                FinanceInvoice.customer_name.contains(keyword)
            )
        )

    if invoice_type is not None:
        query = query.filter_by(invoice_type=invoice_type)

    if status is not None:
        query = query.filter_by(status=status)

    total = query.count()
    invoices = query.order_by(FinanceInvoice.created_at.desc()).paginate(
        page=page, per_page=page_size, error_out=False
    )

    return jsonify({
        'code': 200,
        'data': {
            'list': [to_dict(i) for i in invoices.items],
            'total': total,
            'page': page,
            'page_size': page_size
        }
    })


@bp.route('/api/invoices/<int:id>', methods=['GET'])
@permission('finance:view', 'finance-receivable:view')
def get_invoice(id):
    """获取发票详情"""
    from models.finance.account import FinanceInvoice

    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    return jsonify({'code': 200, 'data': to_dict(invoice)})


@bp.route('/api/invoices', methods=['POST'])
@permission('finance:view', 'finance-receivable:edit')
def create_invoice():
    """创建发票"""
    from flask_jwt_extended import get_jwt_identity
    from app.utils import generate_code
    from models.finance.account import FinanceInvoice

    data = request.get_json()
    user_id = get_jwt_identity()

    last_invoice = FinanceInvoice.query.order_by(FinanceInvoice.id.desc()).first()
    invoice_no = generate_code('INV', last_invoice.id if last_invoice else 0)

    invoice = FinanceInvoice(
        invoice_no=invoice_no,
        invoice_type=data.get('invoice_type', 1),
        customer_id=data.get('customer_id'),
        customer_name=data.get('customer_name'),
        amount=data.get('amount', 0),
        tax_amount=data.get('tax_amount', 0),
        total_amount=data.get('total_amount', 0),
        related_type=data.get('related_type'),
        related_id=data.get('related_id'),
        related_no=data.get('related_no'),
        remark=data.get('remark'),
        created_by=user_id
    )

    db.session.add(invoice)
    db.session.commit()

    return jsonify({'code': 200, 'message': '创建成功', 'data': {'id': invoice.id, 'invoice_no': invoice_no}})


@bp.route('/api/invoices/<int:id>', methods=['PUT'])
@jwt_required()
def update_invoice(id):
    """更新发票"""
    from models.finance.account import FinanceInvoice

    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '已作废的发票不能修改'}), 400

    data = request.get_json()

    for field in ['invoice_type', 'customer_id', 'customer_name',
                  'amount', 'tax_amount', 'total_amount',
                  'related_type', 'related_id', 'related_no', 'remark']:
        if field in data:
            setattr(invoice, field, data[field])

    db.session.commit()
    return jsonify({'code': 200, 'message': '更新成功'})


@bp.route('/api/invoices/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_invoice(id):
    """作废发票"""
    from models.finance.account import FinanceInvoice

    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '发票已作废'}), 400

    invoice.status = 2
    db.session.commit()
    return jsonify({'code': 200, 'message': '发票已作废'})


@bp.route('/api/invoices/<int:id>/cancel', methods=['POST'])
@jwt_required()
def cancel_invoice(id):
    """作废发票"""
    from models.finance.account import FinanceInvoice

    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '发票已作废'}), 400

    invoice.status = 2
    db.session.commit()
    return jsonify({'code': 200, 'message': '发票已作废'})


@bp.route('/api/invoices/<int:id>/red-flush', methods=['POST'])
@jwt_required()
def red_flush_invoice(id):
    """红冲发票"""
    from flask_jwt_extended import get_jwt_identity
    from app.utils import generate_code
    from models.finance.account import FinanceInvoice

    invoice = FinanceInvoice.query.get(id)
    if not invoice:
        return jsonify({'code': 404, 'message': '发票不存在'}), 404

    if invoice.status == 2:
        return jsonify({'code': 400, 'message': '已作废的发票不能红冲'}), 400

    data = request.get_json()
    user_id = get_jwt_identity()

    # 创建红冲发票（负数金额）
    last_invoice = FinanceInvoice.query.order_by(FinanceInvoice.id.desc()).first()
    red_no = generate_code('RINV', last_invoice.id if last_invoice else 0)

    red_invoice = FinanceInvoice(
        invoice_no=red_no,
        invoice_type=invoice.invoice_type,
        customer_id=invoice.customer_id,
        customer_name=invoice.customer_name,
        amount=-(float(invoice.amount or 0)),
        tax_amount=-(float(invoice.tax_amount or 0)),
        total_amount=-(float(invoice.total_amount or 0)),
        related_type='red_flush',
        related_id=id,
        related_no=invoice.invoice_no,
        remark=f'红冲发票：{invoice.invoice_no}。{data.get("remark", "")}',
        status=1,
        created_by=user_id
    )

    db.session.add(red_invoice)

    # 原发票标记为已红冲
    invoice.status = 2
    invoice.remark = (invoice.remark or '') + f' [已红冲，红冲发票号：{red_no}]'

    db.session.commit()

    return jsonify({'code': 200, 'message': '红冲成功', 'data': {'red_invoice_id': red_invoice.id, 'red_no': red_no}})
